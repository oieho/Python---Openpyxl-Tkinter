from ScrollableNotebook import *
from tkinter import *
import tkinter.messagebox
from tkinter import ttk
# from openpyxl.styles import Font
from tkinter import font as tkFont
from datetime import datetime
from tkinter import messagebox
from openpyxl import load_workbook
import tkinter as tk
import os
import sys
# import pandas as pd
from PIL import ImageTk, Image
from copy import copy
from datetime import datetime
from copy import copy
from openpyxl.utils import range_boundaries
import openpyxl
root = Tk()
root.title('마음정원 정신건강의학과')
root.iconbitmap('icn.ico')
root.geometry("1480x900+212+54")
root.resizable(False, False)


def resource_path(relative_path):
	try:
		base_path = sys.MEIPASS
	except Exception:
		base_path = os.path.abspath(".")

	return os.path.join(base_path, relative_path)


def focus_next_widget(event):
    event.widget.tk_focusNext().focus()
    return("break")


class VerticalScrolledFrame:
    def __init__(self, master, **kwargs):
        width = kwargs.pop('width', None)
        height = kwargs.pop('height', None)
        bg = kwargs.pop('bg', kwargs.pop('background', None))
        self.outer = tk.Frame(master, **kwargs)

        self.vsb = tk.Scrollbar(self.outer, orient=tk.VERTICAL)
        self.vsb.pack(fill=tk.Y, side=tk.RIGHT)
        self.canvas = tk.Canvas(
            self.outer, highlightthickness=0, width=width, height=height, bg=bg)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.canvas['yscrollcommand'] = self.vsb.set
        # mouse scroll does not seem to work with just "bind"; You have
        # to use "bind_all". Therefore to use multiple windows you have
        # to bind_all in the current widget
        self.canvas.bind("<Enter>", self._bind_mouse)
        self.canvas.bind("<Leave>", self._unbind_mouse)
        self.vsb['command'] = self.canvas.yview

        self.inner = tk.Frame(self.canvas, bg=bg)
        # pack the inner Frame into the Canvas with the topleft corner 4 pixels offset
        self.canvas.create_window(4, 4, window=self.inner, anchor='nw')
        self.inner.bind("<Configure>", self._on_frame_configure)

        self.outer_attr = set(dir(tk.Widget))

    def __getattr__(self, item):
        if item in self.outer_attr:
            # geometry attributes etc (eg pack, destroy, tkraise) are passed on to self.outer
            return getattr(self.outer, item)
        else:
            # all other attributes (_w, children, etc) are passed to self.inner
            return getattr(self.inner, item)

    def _on_frame_configure(self, event=None):
        x1, y1, x2, y2 = self.canvas.bbox("all")
        height = self.canvas.winfo_height()
        self.canvas.config(scrollregion=(0, 0, x2, max(y2, height)))

    def _bind_mouse(self, event=None):
        self.canvas.bind_all("<4>", self._on_mousewheel)
        self.canvas.bind_all("<5>", self._on_mousewheel)
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)

    def _unbind_mouse(self, event=None):
        self.canvas.unbind_all("<4>")
        self.canvas.unbind_all("<5>")
        self.canvas.unbind_all("<MouseWheel>")

    def _on_mousewheel(self, event):
        """Linux uses event.num; Windows / Mac uses event.delta"""
        if event.num == 4 or event.delta > 0:
            self.canvas.yview_scroll(-1, "units")
        elif event.num == 5 or event.delta < 0:
            self.canvas.yview_scroll(1, "units")

    def __str__(self):
        return str(self.outer)


tabpanelStyle = ttk.Style()
tabpanelStyle.configure('TNotebook.Tab', padding=[4, 5], font=(
    'Malgun Gothic', '14', 'normal'))
tabpanelStyle.configure('TNotebook', sticky='w',
                        tabposition='sw', tabmargins=[90, 0, 0, 0])
tabpanelStyle.layout("Tab", [('Notebook.tab', {'sticky': 'nswe', 'children': [('Notebook.padding', {
                     'side': 'top', 'sticky': 'nswe', 'children': [('Notebook.label', {'side': 'top', 'sticky': ''})], })], })])
tabpanelStyle.map("TNotebook.Tab",
                  background=[
                      ("selected", '#50b049')], foreground=[("selected", 'green')])

wb = load_workbook(resource_path("mastersheet.xlsm"),keep_vba=True)
sheet0 = wb['Ptest']
# sheet9 = wb["변화"]
# sheet10 = wb["인터뷰"]
sheet1 = wb['PHQ']
sheet2 = wb['CDI']
sheet3 = wb['BDI']
sheet4 = wb['SNAP']
sheet5 = wb["ST_2"]
sheet6 = wb["ST_1"]
sheet7 = wb["HAMA"]
sheet8 = wb["HAMD"]
ws = wb.active


malgungothic9 = tkFont.Font(family='Malgun Gothic', size=13)
malgungothic13 = tkFont.Font(family='Malgun Gothic', size=19)


notebook = ScrollableNotebook(
    root, wheelscroll=True, tabmenu=True, width=1480, height=857, padding=[2, 0, 0, 0])
notebook.pack()

startframe = Frame(root)
notebook.add3(startframe, text="평가")

# frame0Alteration = VerticalScrolledFrame(root)
# notebook.add3(frame0Alteration, text="변화")
# Label(frame0Alteration, text="Alteration").pack()


def Delete(event):
    event.widget.select(event.widget.index(CURRENT))
    event.widget.hide(notebook.index(event))


notebook.bind_all('<Button-2>', Delete)


def limitnameFunc(*args):
    value = limitName.get()
    if len(value) > 2:
        limitName.set(value[:20])


def limitidFunc(*args):
    value = limitid.get()
    if len(value) > 2:
        limitid.set(value[:20])


startframebg = tk.PhotoImage(file="images/startframebg.png")
startframebgLabel1 = Label(startframe, image=startframebg)
startframebgLabel1.place(x=298, y=200)


name = Text(startframe, width=21, height=1, wrap=WORD, font=(
	'malgun gothic', 16), border=0)
name.insert(END, '이름')
name.configure(state="disabled")
name.config(cursor='arrow')
name.place(x=391, y=256)

limitName = StringVar()
limitName.trace('w', limitnameFunc)
limitid = StringVar()
limitid.trace('w', limitidFunc)


def confirm(e):
    if len(limitName.get()) == 0:
        messagebox.showinfo('마음정원', '이름을 입력해주십시오.')
        nameInput.focus_set()
        focus_next_widget(e)


def regConfirm(e):
    if len(limitid.get()) == 0:
        messagebox.showinfo('마음정원', 'ID를 입력해주십시오.')
        idInput.focus_set()
        focus_next_widget(e)
        return


nameInput = Entry(startframe, bd=0, textvariable=limitName, width=11, font=(
    'gulim', 18))
nameInput.place(x=442, y=259, height=28)
nameInput.bind("<Tab>", confirm)
nameInput.bind("<Return>", confirm)
nameInput.focus_set()


idText = Text(startframe, width=2, height=1, wrap=WORD, font=(
	'malgun gothic', 16), border=0)
idText.insert(END, 'ID')
idText.configure(state="disabled")
idText.config(cursor='arrow')
idText.place(x=615, y=256)


idInput = Entry(startframe, bd=0, textvariable=limitid, width=11, font=(
    'gulim', 18))
idInput.place(x=645, y=259, height=28)
idInput.bind("<Tab>", regConfirm)
idInput.bind("<Return>", regConfirm)
idInput.focus_set()


x = 0
xx = 0
displayPage = VerticalScrolledFrame(root)
displayPage2 = VerticalScrolledFrame(root)
phqchkImage11 = tk.PhotoImage(file='images/phqchkType11.png')
phqchkImage12 = tk.PhotoImage(file='images/phqchkType12.png')
phqchkImage21 = tk.PhotoImage(file='images/phqchkType21.png')
phqchkImage22 = tk.PhotoImage(file='images/phqchkType22.png')
phqchkImage31 = tk.PhotoImage(file='images/phqchkType31.png')
phqchkImage32 = tk.PhotoImage(file='images/phqchkType32.png')
phqchkImage41 = tk.PhotoImage(file='images/phqchkType41.png')
phqchkImage42 = tk.PhotoImage(file='images/phqchkType42.png')
phqlastchkImage11 = tk.PhotoImage(file='images/phqlastchkType11.png')
phqlastchkImage12 = tk.PhotoImage(file='images/phqlastchkType12.png')
phqlastchkImage21 = tk.PhotoImage(file='images/phqlastchkType21.png')
phqlastchkImage22 = tk.PhotoImage(file='images/phqlastchkType22.png')
phqlastchkImage31 = tk.PhotoImage(file='images/phqlastchkType31.png')
phqlastchkImage32 = tk.PhotoImage(file='images/phqlastchkType32.png')
phqlastchkImage41 = tk.PhotoImage(file='images/phqlastchkType41.png')
phqlastchkImage42 = tk.PhotoImage(file='images/phqlastchkType42.png')


addsheetImg = tk.PhotoImage(file='images/submitbtn.png')

phqchk3 = True
cdichk3 = True
bdichk3 = True
snapchk3 = True
st_2chk3 = True
st_1chk3 = True
hamachk3 = True
hamdchk3 = True



phqchkType1 = StringVar()
phqchkType2 = StringVar()
phqchkType3 = StringVar()
phqchkType4 = StringVar()
phqchkType5 = StringVar()
phqchkType6 = StringVar()
phqchkType7 = StringVar()
phqchkType8 = StringVar()
phqlastchkType = StringVar()

phqchkType1.set("")
phqchkType2.set("")
phqchkType3.set("")
phqchkType4.set("")
phqchkType5.set("")
phqchkType6.set("")
phqchkType7.set("")
phqchkType8.set("")
phqlastchkType.set("")


chkImage11 = tk.PhotoImage(file='images/btnnormal.png')
chkImage12 = tk.PhotoImage(file='images/btnclicked.png')

cdichkType1 = StringVar()
cdichkType2 = StringVar()
cdichkType3 = StringVar()
cdichkType4 = StringVar()
cdichkType5 = StringVar()
cdichkType6 = StringVar()
cdichkType7 = StringVar()
cdichkType8 = StringVar()
cdichkType9 = StringVar()
cdichkType10 = StringVar()
cdichkType11 = StringVar()
cdichkType12 = StringVar()
cdichkType13 = StringVar()
cdichkType14 = StringVar()
cdichkType15 = StringVar()
cdichkType16 = StringVar()
cdichkType17 = StringVar()
cdichkType18 = StringVar()
cdichkType19 = StringVar()
cdichkType20 = StringVar()
cdichkType21 = StringVar()
cdichkType22 = StringVar()
cdichkType23 = StringVar()
cdichkType24 = StringVar()
cdichkType25 = StringVar()
cdichkType26 = StringVar()
cdichkType27 = StringVar()

cdichkType1.set("")
cdichkType2.set("")
cdichkType3.set("")
cdichkType4.set("")
cdichkType5.set("")
cdichkType6.set("")
cdichkType7.set("")
cdichkType8.set("")
cdichkType9.set("")
cdichkType10.set("")
cdichkType11.set("")
cdichkType12.set("")
cdichkType13.set("")
cdichkType14.set("")
cdichkType15.set("")
cdichkType16.set("")
cdichkType17.set("")
cdichkType18.set("")
cdichkType19.set("")
cdichkType20.set("")
cdichkType21.set("")
cdichkType22.set("")
cdichkType23.set("")
cdichkType24.set("")
cdichkType25.set("")
cdichkType26.set("")
cdichkType27.set("")

bdichkType1 = StringVar()
bdichkType2 = StringVar()
bdichkType3 = StringVar()
bdichkType4 = StringVar()
bdichkType5 = StringVar()
bdichkType6 = StringVar()
bdichkType7 = StringVar()
bdichkType8 = StringVar()
bdichkType9 = StringVar()
bdichkType10 = StringVar()
bdichkType11 = StringVar()
bdichkType12 = StringVar()
bdichkType13 = StringVar()
bdichkType14 = StringVar()
bdichkType15 = StringVar()
bdichkType16 = StringVar()
bdichkType17 = StringVar()
bdichkType18 = StringVar()
bdichkType19 = StringVar()
bdichkTypeYesNo = StringVar()
bdichkType20 = StringVar()
bdichkType21 = StringVar()

bdichkType1.set("")
bdichkType2.set("")
bdichkType3.set("")
bdichkType4.set("")
bdichkType5.set("")
bdichkType6.set("")
bdichkType7.set("")
bdichkType8.set("")
bdichkType9.set("")
bdichkType10.set("")
bdichkType11.set("")
bdichkType12.set("")
bdichkType13.set("")
bdichkType14.set("")
bdichkType15.set("")
bdichkType16.set("")
bdichkType17.set("")
bdichkType18.set("")
bdichkType19.set("")
bdichkTypeYesNo.set("")
bdichkType20.set("")
bdichkType21.set("")

snapchkType1 = StringVar()
snapchkType2 = StringVar()
snapchkType3 = StringVar()
snapchkType4 = StringVar()
snapchkType5 = StringVar()
snapchkType6 = StringVar()
snapchkType7 = StringVar()
snapchkType8 = StringVar()
snapchkType9 = StringVar()
snapchkType10 = StringVar()
snapchkType11 = StringVar()
snapchkType12 = StringVar()
snapchkType13 = StringVar()
snapchkType14 = StringVar()
snapchkType15 = StringVar()
snapchkType16 = StringVar()
snapchkType17 = StringVar()
snapchkType18 = StringVar()

snapchkType1.set("")
snapchkType2.set("")
snapchkType3.set("")
snapchkType4.set("")
snapchkType5.set("")
snapchkType6.set("")
snapchkType7.set("")
snapchkType8.set("")
snapchkType9.set("")
snapchkType10.set("")
snapchkType11.set("")
snapchkType12.set("")
snapchkType13.set("")
snapchkType14.set("")
snapchkType15.set("")
snapchkType16.set("")
snapchkType17.set("")
snapchkType18.set("")


chkImage201 = tk.PhotoImage(file='images/chkType01.png')
chkImage202 = tk.PhotoImage(file='images/chkType02.png')
chkImage211 = tk.PhotoImage(file='images/chkType11.png')
chkImage212 = tk.PhotoImage(file='images/chkType12.png')
chkImage221 = tk.PhotoImage(file='images/chkType21.png')
chkImage222 = tk.PhotoImage(file='images/chkType22.png')
chkImage231 = tk.PhotoImage(file='images/chkType31.png')
chkImage232 = tk.PhotoImage(file='images/chkType32.png')
chkImage241 = tk.PhotoImage(file='images/chkType41.png')
chkImage242 = tk.PhotoImage(file='images/chkType42.png')

st_2chkType1 = StringVar()
st_2chkType2 = StringVar()
st_2chkType3 = StringVar()
st_2chkType4 = StringVar()
st_2chkType5 = StringVar()
st_2chkType6 = StringVar()
st_2chkType7 = StringVar()
st_2chkType8 = StringVar()
st_2chkType9 = StringVar()
st_2chkType10 = StringVar()
st_2chkType11 = StringVar()
st_2chkType12 = StringVar()
st_2chkType13 = StringVar()
st_2chkType14 = StringVar()
st_2chkType15 = StringVar()
st_2chkType16 = StringVar()
st_2chkType17 = StringVar()
st_2chkType18 = StringVar()
st_2chkType19 = StringVar()
st_2chkType20 = StringVar()

st_2chkType1.set("")
st_2chkType2.set("")
st_2chkType3.set("")
st_2chkType4.set("")
st_2chkType5.set("")
st_2chkType6.set("")
st_2chkType7.set("")
st_2chkType8.set("")
st_2chkType9.set("")
st_2chkType10.set("")
st_2chkType11.set("")
st_2chkType12.set("")
st_2chkType13.set("")
st_2chkType14.set("")
st_2chkType15.set("")
st_2chkType16.set("")
st_2chkType17.set("")
st_2chkType18.set("")
st_2chkType19.set("")
st_2chkType20.set("")

st_1chkType1 = StringVar()
st_1chkType2 = StringVar()
st_1chkType3 = StringVar()
st_1chkType4 = StringVar()
st_1chkType5 = StringVar()
st_1chkType6 = StringVar()
st_1chkType7 = StringVar()
st_1chkType8 = StringVar()
st_1chkType9 = StringVar()
st_1chkType10 = StringVar()
st_1chkType11 = StringVar()
st_1chkType12 = StringVar()
st_1chkType13 = StringVar()
st_1chkType14 = StringVar()
st_1chkType15 = StringVar()
st_1chkType16 = StringVar()
st_1chkType17 = StringVar()
st_1chkType18 = StringVar()
st_1chkType19 = StringVar()
st_1chkType20 = StringVar()

st_1chkType1.set("")
st_1chkType2.set("")
st_1chkType3.set("")
st_1chkType4.set("")
st_1chkType5.set("")
st_1chkType6.set("")
st_1chkType7.set("")
st_1chkType8.set("")
st_1chkType9.set("")
st_1chkType10.set("")
st_1chkType11.set("")
st_1chkType12.set("")
st_1chkType13.set("")
st_1chkType14.set("")
st_1chkType15.set("")
st_1chkType16.set("")
st_1chkType17.set("")
st_1chkType18.set("")
st_1chkType19.set("")
st_1chkType20.set("")

hamachkType1 = StringVar()
hamachkType2 = StringVar()
hamachkType3 = StringVar()
hamachkType4 = StringVar()
hamachkType5 = StringVar()
hamachkType6 = StringVar()
hamachkType7 = StringVar()
hamachkType8 = StringVar()
hamachkType9 = StringVar()
hamachkType10 = StringVar()
hamachkType11 = StringVar()
hamachkType12 = StringVar()
hamachkType13 = StringVar()
hamachkType14 = StringVar()

hamachkType1.set("")
hamachkType2.set("")
hamachkType3.set("")
hamachkType4.set("")
hamachkType5.set("")
hamachkType6.set("")
hamachkType7.set("")
hamachkType8.set("")
hamachkType9.set("")
hamachkType10.set("")
hamachkType11.set("")
hamachkType12.set("")
hamachkType13.set("")
hamachkType14.set("")

hamdchkType1 = StringVar()
hamdchkType2 = StringVar()
hamdchkType3 = StringVar()
hamdchkType4 = StringVar()
hamdchkType5 = StringVar()
hamdchkType6 = StringVar()
hamdchkType7 = StringVar()
hamdchkType8 = StringVar()
hamdchkType9 = StringVar()
hamdchkType10 = StringVar()
hamdchkType11 = StringVar()
hamdchkType12 = StringVar()
hamdchkType13 = StringVar()
hamdchkType14 = StringVar()
hamdchkType15 = StringVar()
hamdchkType16 = StringVar()
hamdchkType17 = StringVar()

hamdchkType1.set("")
hamdchkType2.set("")
hamdchkType3.set("")
hamdchkType4.set("")
hamdchkType5.set("")
hamdchkType6.set("")
hamdchkType7.set("")
hamdchkType8.set("")
hamdchkType9.set("")
hamdchkType10.set("")
hamdchkType11.set("")
hamdchkType12.set("")
hamdchkType13.set("")
hamdchkType14.set("")
hamdchkType15.set("")
hamdchkType16.set("")
hamdchkType17.set("")

sumscore = 0
sumscore2 = 0
class mainContent():


    def phqchkFunc():
        global now
        now = datetime.now()
        global phqchk
        phqchk = True
        sheet1['E2'].value = (" % s년 % s월 % s일 % s시 % s분" %
                              (now.year, now.month, now.day, now.hour, now.minute))
        sheet1['K2'].value = idInput.get()
        sheet1['K3'].value = nameInput.get()

        def phqchkImage11Func():
            # sumScore()
            if not str(sheet1['C6'].value).strip() == '전혀없음':
                sheet1['C6'].value = phqchk11["text"]

        def phqchkImage12Func():
            # sumScore()
            if not str(sheet1['C6'].value).strip() == '며칠동안':
                sheet1['C6'].value = phqchk12["text"]

        def phqchkImage13Func():
            # sumScore()
            if not str(sheet1['C6'].value).strip() == '일주일이상':
                sheet1['C6'].value = phqchk13["text"]

        def phqchkImage14Func():
            # sumScore()
            if not str(sheet1['C6'].value).strip() == '거의매일':
                sheet1['C6'].value = phqchk14["text"]

        def phqchkImage21Func():
            # sumScore()
            if not str(sheet1['C9'].value).strip() == '전혀없음':
                sheet1['C9'].value = phqchk21["text"]

        def phqchkImage22Func():
            # sumScore()
            if not str(sheet1['C9'].value).strip() == '며칠동안':
                sheet1['C9'].value = phqchk22["text"]

        def phqchkImage23Func():
            # sumScore()
            if not str(sheet1['C9'].value).strip() == '일주일이상':
                sheet1['C9'].value = phqchk23["text"]

        def phqchkImage24Func():
            # sumScore()
            if not str(sheet1['C9'].value).strip() == '거의매일':
                sheet1['C9'].value = phqchk24["text"]

        def phqchkImage31Func():
            # sumScore()
            if not str(sheet1['C12'].value).strip() == '전혀없음':
                sheet1['C12'].value = phqchk31["text"]

        def phqchkImage32Func():
            # sumScore()
            if not str(sheet1['C12'].value).strip() == '며칠동안':
                sheet1['C12'].value = phqchk32["text"]

        def phqchkImage33Func():
            # sumScore()
            if not str(sheet1['C12'].value).strip() == '일주일이상':
                sheet1['C12'].value = phqchk33["text"]

        def phqchkImage34Func():
            # sumScore()
            if not str(sheet1['C12'].value).strip() == '거의매일':
                sheet1['C12'].value = phqchk34["text"]

        def phqchkImage41Func():
            # sumScore()
            if not str(sheet1['C15'].value).strip() == '전혀없음':
                sheet1['C15'].value = phqchk41["text"]

        def phqchkImage42Func():
            # sumScore()
            if not str(sheet1['C15'].value).strip() == '며칠동안':
                sheet1['C15'].value = phqchk42["text"]

        def phqchkImage43Func():
            # sumScore()
            if not str(sheet1['C15'].value).strip() == '일주일이상':
                sheet1['C15'].value = phqchk43["text"]

        def phqchkImage44Func():
            # sumScore()
            if not str(sheet1['C15'].value).strip() == '거의매일':
                sheet1['C15'].value = phqchk44["text"]

        def phqchkImage51Func():
            # sumScore()
            if not str(sheet1['C18'].value).strip() == '전혀없음':
                sheet1['C18'].value = phqchk51["text"]

        def phqchkImage52Func():
            # sumScore()
            if not str(sheet1['C18'].value).strip() == '며칠동안':
                sheet1['C18'].value = phqchk52["text"]

        def phqchkImage53Func():
            # sumScore()
            if not str(sheet1['C18'].value).strip() == '일주일이상':
                sheet1['C18'].value = phqchk53["text"]

        def phqchkImage54Func():
            # sumScore()
            if not str(sheet1['C18'].value).strip() == '거의매일':
                sheet1['C18'].value = phqchk54["text"]

        def phqchkImage61Func():
            # sumScore()
            if not str(sheet1['C21'].value).strip() == '전혀없음':
                sheet1['C21'].value = phqchk61["text"]

        def phqchkImage62Func():
            # sumScore()
            if not str(sheet1['C21'].value).strip() == '며칠동안':
                sheet1['C21'].value = phqchk62["text"]

        def phqchkImage63Func():
            # sumScore()
            if not str(sheet1['C21'].value).strip() == '일주일이상':
                sheet1['C21'].value = phqchk63["text"]

        def phqchkImage64Func():
            # sumScore()
            if not str(sheet1['C21'].value).strip() == '거의매일':
                sheet1['C21'].value = phqchk64["text"]

        def phqchkImage71Func():
            # sumScore()
            if not str(sheet1['C24'].value).strip() == '전혀없음':
                sheet1['C24'].value = phqchk71["text"]

        def phqchkImage72Func():
            # sumScore()
            if not str(sheet1['C24'].value).strip() == '며칠동안':
                sheet1['C24'].value = phqchk72["text"]

        def phqchkImage73Func():
            # sumScore()
            if not str(sheet1['C24'].value).strip() == '일주일이상':
                sheet1['C24'].value = phqchk73["text"]

        def phqchkImage74Func():
            # sumScore()
            if not str(sheet1['C24'].value).strip() == '거의매일':
                sheet1['C24'].value = phqchk74["text"]

        def phqchkImage81Func():
            # sumScore()
            if not str(sheet1['C27'].value).strip() == '전혀없음':
                sheet1['C27'].value = phqchk81["text"]

        def phqchkImage82Func():
            # sumScore()
            if not str(sheet1['C27'].value).strip() == '며칠동안':
                sheet1['C27'].value = phqchk82["text"]

        def phqchkImage83Func():
            # sumScore()
            if not str(sheet1['C27'].value).strip() == '일주일이상':
                sheet1['C27'].value = phqchk83["text"]

        def phqchkImage84Func():
            # sumScore()
            if not str(sheet1['C27'].value).strip() == '거의매일':
                sheet1['C27'].value = phqchk84["text"]

        def phqlastchkImageFunc1():
            # sumScore()
            if not str(sheet1['C30'].value).strip() == '전혀 어렵지 않다':
                sheet1['C30'].value = phqlastchk1["text"]

        def phqlastchkImageFunc2():
            # sumScore()
            if not str(sheet1['C30'].value).strip() == '다소 어렵다':
                sheet1['C30'].value = phqlastchk2["text"]

        def phqlastchkImageFunc3():
            # sumScore()
            if not str(sheet1['C30'].value).strip() == '많이 어렵다':
                sheet1['C30'].value = phqlastchk3["text"]

        def phqlastchkImageFunc4():
            # sumScore()
            if not str(sheet1['C30'].value).strip() == '매우 많이 어렵다':
                sheet1['C30'].value = phqlastchk4["text"]

        frame1PhqBgImg = Image.open("images/phqbg.png")
        frame1PhqBg = ImageTk.PhotoImage(frame1PhqBgImg)
        frame1PhqBgLabel = tkinter.Label(displayPage.inner, image=frame1PhqBg)
        frame1PhqBgLabel.image = frame1PhqBg
        frame1PhqBgLabel.place(x=17, y=420, height=1111)
        frame1PhqBgLabel.pack()

        global phqchk11
        phqchk11 = Radiobutton(displayPage.inner, value=0, text="전혀없음", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage11,
                               selectimage=phqchkImage12, indicatoron=False, cursor="circle", variable=phqchkType1, font=malgungothic13, bd=0, highlightthickness=0, command=phqchkImage11Func)
        phqchk11.deselect()
        global phqchk12
        phqchk12 = Radiobutton(displayPage.inner, value=1, text="며칠동안", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage21, selectimage=phqchkImage22,
                               indicatoron=False, cursor="circle", variable=phqchkType1, font=malgungothic13, bd=0, highlightthickness=0, command=phqchkImage12Func)
        phqchk12.deselect()
        global phqchk13
        phqchk13 = Radiobutton(displayPage.inner, value=2, text="일주일이상", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage31, selectimage=phqchkImage32,
                               indicatoron=False, cursor="circle", variable=phqchkType1, font=malgungothic13, bd=0, highlightthickness=0, command=phqchkImage13Func)
        phqchk13.deselect()
        global phqchk14
        phqchk14 = Radiobutton(displayPage.inner, value=3, text="거의매일", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage41, selectimage=phqchkImage42,
                               indicatoron=False, cursor="circle", variable=phqchkType1, font=malgungothic13, bd=0, highlightthickness=0, command=phqchkImage14Func)
        phqchk14.deselect()

        global phqchk21
        phqchk21 = Radiobutton(displayPage.inner, value=0,  text="전혀없음", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage11,
                               selectimage=phqchkImage12, indicatoron=False, cursor="circle", variable=phqchkType2, font=malgungothic13, bd=0, command=phqchkImage21Func)
        phqchk21.deselect()
        global phqchk22
        phqchk22 = Radiobutton(displayPage.inner, value=1, text="며칠동안", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage21, selectimage=phqchkImage22,
                               indicatoron=False, cursor="circle", variable=phqchkType2, font=malgungothic13, bd=0, command=phqchkImage22Func)
        phqchk22.deselect()
        global phqchk23
        phqchk23 = Radiobutton(displayPage.inner, value=2, text="일주일이상", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage31, selectimage=phqchkImage32,
                               indicatoron=False, cursor="circle", variable=phqchkType2, font=malgungothic13, bd=0, command=phqchkImage23Func)
        phqchk23.deselect()
        global phqchk24
        phqchk24 = Radiobutton(displayPage.inner, value=3, text="거의매일", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage41, selectimage=phqchkImage42,
                               indicatoron=False, cursor="circle", variable=phqchkType2, font=malgungothic13, bd=0, command=phqchkImage24Func)
        phqchk24.deselect()
        global phqchk31
        phqchk31 = Radiobutton(displayPage.inner, value=0,  text="전혀없음", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage11,
                               selectimage=phqchkImage12, indicatoron=False, cursor="circle", variable=phqchkType3, font=malgungothic13, bd=0, command=phqchkImage31Func)
        phqchk31.deselect()
        global phqchk32
        phqchk32 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage21, text="며칠동안", selectimage=phqchkImage22,
                               indicatoron=False, cursor="circle", variable=phqchkType3, font=malgungothic13, bd=0, command=phqchkImage32Func)
        phqchk32.deselect()
        global phqchk33
        phqchk33 = Radiobutton(displayPage.inner, value=2, text="일주일이상", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage31, selectimage=phqchkImage32,
                               indicatoron=False, cursor="circle", variable=phqchkType3, font=malgungothic13, bd=0, command=phqchkImage33Func)
        phqchk33.deselect()
        global phqchk34
        phqchk34 = Radiobutton(displayPage.inner, value=3, text="거의매일", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage41, selectimage=phqchkImage42,
                               indicatoron=False, cursor="circle", variable=phqchkType3, font=malgungothic13, bd=0, command=phqchkImage34Func)
        phqchk34.deselect()
        global phqchk41
        phqchk41 = Radiobutton(displayPage.inner, value=0,  text="전혀없음", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage11,
                               selectimage=phqchkImage12, indicatoron=False, cursor="circle", font=malgungothic13, variable=phqchkType4, bd=0, command=phqchkImage41Func)
        phqchk41.deselect()
        global phqchk42
        phqchk42 = Radiobutton(displayPage.inner, value=1, text="며칠동안", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage21, selectimage=phqchkImage22,
                               indicatoron=False, cursor="circle", variable=phqchkType4, bd=0, font=malgungothic13, command=phqchkImage42Func)
        phqchk42.deselect()
        global phqchk43
        phqchk43 = Radiobutton(displayPage.inner, value=2, text="일주일이상", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage31, selectimage=phqchkImage32,
                               indicatoron=False, cursor="circle", variable=phqchkType4,  bd=0, font=malgungothic13, command=phqchkImage43Func)
        phqchk43.deselect()
        global phqchk44
        phqchk44 = Radiobutton(displayPage.inner, value=3, text="거의매일", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage41, selectimage=phqchkImage42,
                               indicatoron=False, cursor="circle", variable=phqchkType4,  bd=0, font=malgungothic13, command=phqchkImage44Func)
        phqchk44.deselect()
        global phqchk51
        phqchk51 = Radiobutton(displayPage.inner, value=0, text="전혀없음", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage11,
                               selectimage=phqchkImage12, indicatoron=False, cursor="circle", variable=phqchkType5, font=malgungothic13, bd=0, command=phqchkImage51Func)
        phqchk51.deselect()
        global phqchk52
        phqchk52 = Radiobutton(displayPage.inner, value=1, text="며칠동안", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage21, selectimage=phqchkImage22,
                               indicatoron=False, cursor="circle", variable=phqchkType5, font=malgungothic13, bd=0, command=phqchkImage52Func)
        phqchk52.deselect()
        global phqchk53
        phqchk53 = Radiobutton(displayPage.inner, value=2, text="일주일이상", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage31, selectimage=phqchkImage32,
                               indicatoron=False, cursor="circle", variable=phqchkType5, font=malgungothic13, bd=0, command=phqchkImage53Func)
        phqchk53.deselect()
        global phqchk54
        phqchk54 = Radiobutton(displayPage.inner, value=3, text="거의매일", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage41, selectimage=phqchkImage42,
                               indicatoron=False, cursor="circle", variable=phqchkType5, font=malgungothic13, bd=0, command=phqchkImage54Func)
        phqchk54.deselect()
        global phqchk61
        phqchk61 = Radiobutton(displayPage.inner, value=0,  text="전혀없음", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage11,
                               selectimage=phqchkImage12, indicatoron=False, cursor="circle", variable=phqchkType6, font=malgungothic13, bd=0, command=phqchkImage61Func)
        phqchk61.deselect()
        global phqchk62
        phqchk62 = Radiobutton(displayPage.inner, value=1, text="며칠동안", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage21, selectimage=phqchkImage22,
                               indicatoron=False, cursor="circle", variable=phqchkType6, font=malgungothic13, bd=0, command=phqchkImage62Func)
        phqchk62.deselect()
        global phqchk63
        phqchk63 = Radiobutton(displayPage.inner, value=2, text="일주일이상", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage31, selectimage=phqchkImage32,
                               indicatoron=False, cursor="circle", variable=phqchkType6, font=malgungothic13, bd=0, command=phqchkImage63Func)
        phqchk63.deselect()
        global phqchk64
        phqchk64 = Radiobutton(displayPage.inner, value=3, text="거의매일", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage41, selectimage=phqchkImage42,
                               indicatoron=False, cursor="circle", variable=phqchkType6, font=malgungothic13, bd=0, command=phqchkImage64Func)
        phqchk64.deselect()
        global phqchk71
        phqchk71 = Radiobutton(displayPage.inner, value=0,  text="전혀없음", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage11,
                               selectimage=phqchkImage12, indicatoron=False, cursor="circle", variable=phqchkType7, font=malgungothic13, bd=0, command=phqchkImage71Func)
        phqchk71.deselect()
        global phqchk72
        phqchk72 = Radiobutton(displayPage.inner, value=1, text="며칠동안", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage21, selectimage=phqchkImage22,
                               indicatoron=False, cursor="circle", variable=phqchkType7, font=malgungothic13, bd=0, command=phqchkImage72Func)
        phqchk72.deselect()
        global phqchk73
        phqchk73 = Radiobutton(displayPage.inner, value=2, text="일주일이상", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage31, selectimage=phqchkImage32,
                               indicatoron=False, cursor="circle", variable=phqchkType7, font=malgungothic13, bd=0, command=phqchkImage73Func)
        phqchk73.deselect()
        global phqchk74
        phqchk74 = Radiobutton(displayPage.inner, value=3, text="거의매일", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage41, selectimage=phqchkImage42,
                               indicatoron=False, cursor="circle", variable=phqchkType7, font=malgungothic13, bd=0, command=phqchkImage74Func)
        phqchk74.deselect()
        global phqchk81
        phqchk81 = Radiobutton(displayPage.inner, value=0, text="전혀없음", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage11,
                               selectimage=phqchkImage12, indicatoron=False, cursor="circle", variable=phqchkType8, font=malgungothic13, bd=0, command=phqchkImage81Func)
        phqchk81.deselect()
        global phqchk82
        phqchk82 = Radiobutton(displayPage.inner, value=1, text="며칠동안", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage21, selectimage=phqchkImage22,
                               indicatoron=False, cursor="circle", variable=phqchkType8, font=malgungothic13, bd=0, command=phqchkImage82Func)
        phqchk82.deselect()
        global phqchk83
        phqchk83 = Radiobutton(displayPage.inner, value=2, text="일주일이상", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage31, selectimage=phqchkImage32,
                               indicatoron=False, cursor="circle", variable=phqchkType8, font=malgungothic13, bd=0, command=phqchkImage83Func)
        phqchk83.deselect()
        global phqchk84
        phqchk84 = Radiobutton(displayPage.inner, value=3, text="거의매일", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage41, selectimage=phqchkImage42,
                               indicatoron=False, cursor="circle", variable=phqchkType8, font=malgungothic13, bd=0, command=phqchkImage84Func)
        phqchk84.deselect()
        global phqlastchk1
        phqlastchk1 = Radiobutton(displayPage.inner, text="전혀 어렵지 않다", value='0', background="#FFFFFF", activebackground="#FFFFFF", image=phqlastchkImage11,
                                  selectimage=phqlastchkImage12, indicatoron=False, cursor="circle", variable=phqlastchkType, font=malgungothic13, bd=0,  command=phqlastchkImageFunc1)
        phqlastchk1.deselect()
        global phqlastchk2
        phqlastchk2 = Radiobutton(displayPage.inner, text="다소 어렵다", value='1', background="#FFFFFF", activebackground="#FFFFFF", image=phqlastchkImage21, selectimage=phqlastchkImage22,
                                  indicatoron=False, cursor="circle", variable=phqlastchkType, font=malgungothic13, bd=0,  command=phqlastchkImageFunc2)
        phqlastchk2.deselect()
        global phqlastchk3
        phqlastchk3 = Radiobutton(displayPage.inner, value='2', text="많이 어렵다", background="#FFFFFF", activebackground="#FFFFFF", image=phqlastchkImage31, selectimage=phqlastchkImage32,
                                  indicatoron=False, cursor="circle", variable=phqlastchkType, font=malgungothic13, bd=0,  command=phqlastchkImageFunc3)
        phqlastchk3.deselect()
        global phqlastchk4
        phqlastchk4 = Radiobutton(displayPage.inner, value='3', text="매우 많이 어렵다", background="#FFFFFF", activebackground="#FFFFFF", image=phqlastchkImage41, selectimage=phqlastchkImage42,
                                  indicatoron=False, cursor="circle", variable=phqlastchkType, font=malgungothic13, bd=0, command=phqlastchkImageFunc4)
        phqlastchk4.deselect()

        phqchk11.place(x=900, y=278, width=106, height=29)
        phqchk12.place(x=1002, y=278, width=100, height=29)
        phqchk13.place(x=1099, y=278, width=113, height=29)
        phqchk14.place(x=1212, y=278, width=105, height=29)
        phqchk21.place(x=900, y=338, width=106, height=29)
        phqchk22.place(x=1002, y=338, width=100, height=29)
        phqchk23.place(x=1099, y=338, width=113, heigh=29)
        phqchk24.place(x=1212, y=338, width=105, heigh=29)
        phqchk31.place(x=900, y=398, width=106, height=29)
        phqchk32.place(x=1002, y=398, width=100, height=29)
        phqchk33.place(x=1099, y=398, width=113, heigh=29)
        phqchk34.place(x=1212, y=398, width=105, heigh=29)
        phqchk41.place(x=900, y=458, width=106, height=29)
        phqchk42.place(x=1002, y=458, width=100, height=29)
        phqchk43.place(x=1099, y=458, width=113, heigh=29)
        phqchk44.place(x=1212, y=458, width=105, heigh=29)
        phqchk51.place(x=900, y=518, width=106, height=29)
        phqchk52.place(x=1002, y=518, width=100, height=29)
        phqchk53.place(x=1099, y=518, width=113, heigh=29)
        phqchk54.place(x=1212, y=518, width=105, heigh=29)
        phqchk61.place(x=900, y=578, width=106, height=29)
        phqchk62.place(x=1002, y=578, width=100, height=29)
        phqchk63.place(x=1099, y=578, width=113, heigh=29)
        phqchk64.place(x=1212, y=578, width=105, heigh=29)
        phqchk71.place(x=900, y=655, width=106, height=29)
        phqchk72.place(x=1002, y=655, width=100, height=29)
        phqchk73.place(x=1099, y=655, width=113, heigh=29)
        phqchk74.place(x=1212, y=655, width=105, heigh=29)
        phqchk81.place(x=900, y=733, width=106, height=29)
        phqchk82.place(x=1002, y=733, width=100, height=29)
        phqchk83.place(x=1099, y=733, width=113, height=29)
        phqchk84.place(x=1212, y=733, width=105, height=29)
        phqlastchk1.place(x=279, y=881, width=201, height=33)
        phqlastchk2.place(x=480, y=881, width=220, height=33)
        phqlastchk3.place(x=700, y=881, width=243, height=33)
        phqlastchk4.place(x=940, y=881, width=314, height=33)

    def phqchkFunc2():
        global now
        now = datetime.now()
        global phqchk
        phqchk = True
        sheet1['E2'].value = (" % s년 % s월 % s일 % s시 % s분" %
                              (now.year, now.month, now.day, now.hour, now.minute))
        sheet1['K2'].value = idInput.get()

        def phqchkImage11Func():
            # sumScore()
            if not str(sheet1['C6'].value).strip() == '전혀없음':
                sheet1['C6'].value = phqchk11["text"]

        def phqchkImage12Func():
            # sumScore()
            if not str(sheet1['C6'].value).strip() == '며칠동안':
                sheet1['C6'].value = phqchk12["text"]

        def phqchkImage13Func():
            # sumScore()
            if not str(sheet1['C6'].value).strip() == '일주일이상':
                sheet1['C6'].value = phqchk13["text"]

        def phqchkImage14Func():
            # sumScore()
            if not str(sheet1['C6'].value).strip() == '거의매일':
                sheet1['C6'].value = phqchk14["text"]

        def phqchkImage21Func():
            # sumScore()
            if not str(sheet1['C9'].value).strip() == '전혀없음':
                sheet1['C9'].value = phqchk21["text"]

        def phqchkImage22Func():
            # sumScore()
            if not str(sheet1['C9'].value).strip() == '며칠동안':
                sheet1['C9'].value = phqchk22["text"]

        def phqchkImage23Func():
            # sumScore()
            if not str(sheet1['C9'].value).strip() == '일주일이상':
                sheet1['C9'].value = phqchk23["text"]

        def phqchkImage24Func():
            # sumScore()
            if not str(sheet1['C9'].value).strip() == '거의매일':
                sheet1['C9'].value = phqchk24["text"]

        def phqchkImage31Func():
            # sumScore()
            if not str(sheet1['C12'].value).strip() == '전혀없음':
                sheet1['C12'].value = phqchk31["text"]

        def phqchkImage32Func():
            # sumScore()
            if not str(sheet1['C12'].value).strip() == '며칠동안':
                sheet1['C12'].value = phqchk32["text"]

        def phqchkImage33Func():
            # sumScore()
            if not str(sheet1['C12'].value).strip() == '일주일이상':
                sheet1['C12'].value = phqchk33["text"]

        def phqchkImage34Func():
            # sumScore()
            if not str(sheet1['C12'].value).strip() == '거의매일':
                sheet1['C12'].value = phqchk34["text"]

        def phqchkImage41Func():
            # sumScore()
            if not str(sheet1['C15'].value).strip() == '전혀없음':
                sheet1['C15'].value = phqchk41["text"]

        def phqchkImage42Func():
            # sumScore()
            if not str(sheet1['C15'].value).strip() == '며칠동안':
                sheet1['C15'].value = phqchk42["text"]

        def phqchkImage43Func():
            # sumScore()
            if not str(sheet1['C15'].value).strip() == '일주일이상':
                sheet1['C15'].value = phqchk43["text"]

        def phqchkImage44Func():
            # sumScore()
            if not str(sheet1['C15'].value).strip() == '거의매일':
                sheet1['C15'].value = phqchk44["text"]

        def phqchkImage51Func():
            # sumScore()
            if not str(sheet1['C18'].value).strip() == '전혀없음':
                sheet1['C18'].value = phqchk51["text"]

        def phqchkImage52Func():
            # sumScore()
            if not str(sheet1['C18'].value).strip() == '며칠동안':
                sheet1['C18'].value = phqchk52["text"]

        def phqchkImage53Func():
            # sumScore()
            if not str(sheet1['C18'].value).strip() == '일주일이상':
                sheet1['C18'].value = phqchk53["text"]

        def phqchkImage54Func():
            # sumScore()
            if not str(sheet1['C18'].value).strip() == '거의매일':
                sheet1['C18'].value = phqchk54["text"]

        def phqchkImage61Func():
            # sumScore()
            if not str(sheet1['C21'].value).strip() == '전혀없음':
                sheet1['C21'].value = phqchk61["text"]

        def phqchkImage62Func():
            # sumScore()
            if not str(sheet1['C21'].value).strip() == '며칠동안':
                sheet1['C21'].value = phqchk62["text"]

        def phqchkImage63Func():
            # sumScore()
            if not str(sheet1['C21'].value).strip() == '일주일이상':
                sheet1['C21'].value = phqchk63["text"]

        def phqchkImage64Func():
            # sumScore()
            if not str(sheet1['C21'].value).strip() == '거의매일':
                sheet1['C21'].value = phqchk64["text"]

        def phqchkImage71Func():
            # sumScore()
            if not str(sheet1['C24'].value).strip() == '전혀없음':
                sheet1['C24'].value = phqchk71["text"]

        def phqchkImage72Func():
            # sumScore()
            if not str(sheet1['C24'].value).strip() == '며칠동안':
                sheet1['C24'].value = phqchk72["text"]

        def phqchkImage73Func():
            # sumScore()
            if not str(sheet1['C24'].value).strip() == '일주일이상':
                sheet1['C24'].value = phqchk73["text"]

        def phqchkImage74Func():
            # sumScore()
            if not str(sheet1['C24'].value).strip() == '거의매일':
                sheet1['C24'].value = phqchk74["text"]

        def phqchkImage81Func():
            # sumScore()
            if not str(sheet1['C27'].value).strip() == '전혀없음':
                sheet1['C27'].value = phqchk81["text"]

        def phqchkImage82Func():
            # sumScore()
            if not str(sheet1['C27'].value).strip() == '며칠동안':
                sheet1['C27'].value = phqchk82["text"]

        def phqchkImage83Func():
            # sumScore()
            if not str(sheet1['C27'].value).strip() == '일주일이상':
                sheet1['C27'].value = phqchk83["text"]

        def phqchkImage84Func():
            # sumScore()
            if not str(sheet1['C27'].value).strip() == '거의매일':
                sheet1['C27'].value = phqchk84["text"]

        def phqlastchkImageFunc1():
            # sumScore()
            if not str(sheet1['C30'].value).strip() == '전혀 어렵지 않다':
                sheet1['C30'].value = phqlastchk1["text"]

        def phqlastchkImageFunc2():
            # sumScore()
            if not str(sheet1['C30'].value).strip() == '다소 어렵다':
                sheet1['C30'].value = phqlastchk2["text"]

        def phqlastchkImageFunc3():
            # sumScore()
            if not str(sheet1['C30'].value).strip() == '많이 어렵다':
                sheet1['C30'].value = phqlastchk3["text"]

        def phqlastchkImageFunc4():
            # sumScore()
            if not str(sheet1['C30'].value).strip() == '매우 많이 어렵다':
                sheet1['C30'].value = phqlastchk4["text"]

        frame1PhqBgImg = Image.open("images/phqbg.png")
        frame1PhqBg = ImageTk.PhotoImage(frame1PhqBgImg)
        frame1PhqBgLabel = tkinter.Label(displayPage2.inner, image=frame1PhqBg)
        frame1PhqBgLabel.image = frame1PhqBg
        frame1PhqBgLabel.place(x=17, y=420, height=1111)
        frame1PhqBgLabel.pack()

        global phqchk11
        phqchk11 = Radiobutton(displayPage2.inner, value=0, text="전혀없음", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage11,
                               selectimage=phqchkImage12, indicatoron=False, cursor="circle", variable=phqchkType1, font=malgungothic13, bd=0, highlightthickness=0, command=phqchkImage11Func)
        phqchk11.deselect()
        global phqchk12
        phqchk12 = Radiobutton(displayPage2.inner, value=1, text="며칠동안", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage21, selectimage=phqchkImage22,
                               indicatoron=False, cursor="circle", variable=phqchkType1, font=malgungothic13, bd=0, highlightthickness=0, command=phqchkImage12Func)
        phqchk12.deselect()
        global phqchk13
        phqchk13 = Radiobutton(displayPage2.inner, value=2, text="일주일이상", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage31, selectimage=phqchkImage32,
                               indicatoron=False, cursor="circle", variable=phqchkType1, font=malgungothic13, bd=0, highlightthickness=0, command=phqchkImage13Func)
        phqchk13.deselect()
        global phqchk14
        phqchk14 = Radiobutton(displayPage2.inner, value=3, text="거의매일", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage41, selectimage=phqchkImage42,
                               indicatoron=False, cursor="circle", variable=phqchkType1, font=malgungothic13, bd=0, highlightthickness=0, command=phqchkImage14Func)
        phqchk14.deselect()

        global phqchk21
        phqchk21 = Radiobutton(displayPage2.inner, value=0,  text="전혀없음", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage11,
                               selectimage=phqchkImage12, indicatoron=False, cursor="circle", variable=phqchkType2, font=malgungothic13, bd=0, command=phqchkImage21Func)
        phqchk21.deselect()
        global phqchk22
        phqchk22 = Radiobutton(displayPage2.inner, value=1, text="며칠동안", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage21, selectimage=phqchkImage22,
                               indicatoron=False, cursor="circle", variable=phqchkType2, font=malgungothic13, bd=0, command=phqchkImage22Func)
        phqchk22.deselect()
        global phqchk23
        phqchk23 = Radiobutton(displayPage2.inner, value=2, text="일주일이상", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage31, selectimage=phqchkImage32,
                               indicatoron=False, cursor="circle", variable=phqchkType2, font=malgungothic13, bd=0, command=phqchkImage23Func)
        phqchk23.deselect()
        global phqchk24
        phqchk24 = Radiobutton(displayPage2.inner, value=3, text="거의매일", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage41, selectimage=phqchkImage42,
                               indicatoron=False, cursor="circle", variable=phqchkType2, font=malgungothic13, bd=0, command=phqchkImage24Func)
        phqchk24.deselect()
        global phqchk31
        phqchk31 = Radiobutton(displayPage2.inner, value=0,  text="전혀없음", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage11,
                               selectimage=phqchkImage12, indicatoron=False, cursor="circle", variable=phqchkType3, font=malgungothic13, bd=0, command=phqchkImage31Func)
        phqchk31.deselect()
        global phqchk32
        phqchk32 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage21, text="며칠동안", selectimage=phqchkImage22,
                               indicatoron=False, cursor="circle", variable=phqchkType3, font=malgungothic13, bd=0, command=phqchkImage32Func)
        phqchk32.deselect()
        global phqchk33
        phqchk33 = Radiobutton(displayPage2.inner, value=2, text="일주일이상", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage31, selectimage=phqchkImage32,
                               indicatoron=False, cursor="circle", variable=phqchkType3, font=malgungothic13, bd=0, command=phqchkImage33Func)
        phqchk33.deselect()
        global phqchk34
        phqchk34 = Radiobutton(displayPage2.inner, value=3, text="거의매일", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage41, selectimage=phqchkImage42,
                               indicatoron=False, cursor="circle", variable=phqchkType3, font=malgungothic13, bd=0, command=phqchkImage34Func)
        phqchk34.deselect()
        global phqchk41
        phqchk41 = Radiobutton(displayPage2.inner, value=0,  text="전혀없음", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage11,
                               selectimage=phqchkImage12, indicatoron=False, cursor="circle", font=malgungothic13, variable=phqchkType4, bd=0, command=phqchkImage41Func)
        phqchk41.deselect()
        global phqchk42
        phqchk42 = Radiobutton(displayPage2.inner, value=1, text="며칠동안", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage21, selectimage=phqchkImage22,
                               indicatoron=False, cursor="circle", variable=phqchkType4, bd=0, font=malgungothic13, command=phqchkImage42Func)
        phqchk42.deselect()
        global phqchk43
        phqchk43 = Radiobutton(displayPage2.inner, value=2, text="일주일이상", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage31, selectimage=phqchkImage32,
                               indicatoron=False, cursor="circle", variable=phqchkType4,  bd=0, font=malgungothic13, command=phqchkImage43Func)
        phqchk43.deselect()
        global phqchk44
        phqchk44 = Radiobutton(displayPage2.inner, value=3, text="거의매일", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage41, selectimage=phqchkImage42,
                               indicatoron=False, cursor="circle", variable=phqchkType4,  bd=0, font=malgungothic13, command=phqchkImage44Func)
        phqchk44.deselect()
        global phqchk51
        phqchk51 = Radiobutton(displayPage2.inner, value=0, text="전혀없음", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage11,
                               selectimage=phqchkImage12, indicatoron=False, cursor="circle", variable=phqchkType5, font=malgungothic13, bd=0, command=phqchkImage51Func)
        phqchk51.deselect()
        global phqchk52
        phqchk52 = Radiobutton(displayPage2.inner, value=1, text="며칠동안", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage21, selectimage=phqchkImage22,
                               indicatoron=False, cursor="circle", variable=phqchkType5, font=malgungothic13, bd=0, command=phqchkImage52Func)
        phqchk52.deselect()
        global phqchk53
        phqchk53 = Radiobutton(displayPage2.inner, value=2, text="일주일이상", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage31, selectimage=phqchkImage32,
                               indicatoron=False, cursor="circle", variable=phqchkType5, font=malgungothic13, bd=0, command=phqchkImage53Func)
        phqchk53.deselect()
        global phqchk54
        phqchk54 = Radiobutton(displayPage2.inner, value=3, text="거의매일", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage41, selectimage=phqchkImage42,
                               indicatoron=False, cursor="circle", variable=phqchkType5, font=malgungothic13, bd=0, command=phqchkImage54Func)
        phqchk54.deselect()
        global phqchk61
        phqchk61 = Radiobutton(displayPage2.inner, value=0,  text="전혀없음", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage11,
                               selectimage=phqchkImage12, indicatoron=False, cursor="circle", variable=phqchkType6, font=malgungothic13, bd=0, command=phqchkImage61Func)
        phqchk61.deselect()
        global phqchk62
        phqchk62 = Radiobutton(displayPage2.inner, value=1, text="며칠동안", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage21, selectimage=phqchkImage22,
                               indicatoron=False, cursor="circle", variable=phqchkType6, font=malgungothic13, bd=0, command=phqchkImage62Func)
        phqchk62.deselect()
        global phqchk63
        phqchk63 = Radiobutton(displayPage2.inner, value=2, text="일주일이상", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage31, selectimage=phqchkImage32,
                               indicatoron=False, cursor="circle", variable=phqchkType6, font=malgungothic13, bd=0, command=phqchkImage63Func)
        phqchk63.deselect()
        global phqchk64
        phqchk64 = Radiobutton(displayPage2.inner, value=3, text="거의매일", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage41, selectimage=phqchkImage42,
                               indicatoron=False, cursor="circle", variable=phqchkType6, font=malgungothic13, bd=0, command=phqchkImage64Func)
        phqchk64.deselect()
        global phqchk71
        phqchk71 = Radiobutton(displayPage2.inner, value=0,  text="전혀없음", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage11,
                               selectimage=phqchkImage12, indicatoron=False, cursor="circle", variable=phqchkType7, font=malgungothic13, bd=0, command=phqchkImage71Func)
        phqchk71.deselect()
        global phqchk72
        phqchk72 = Radiobutton(displayPage2.inner, value=1, text="며칠동안", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage21, selectimage=phqchkImage22,
                               indicatoron=False, cursor="circle", variable=phqchkType7, font=malgungothic13, bd=0, command=phqchkImage72Func)
        phqchk72.deselect()
        global phqchk73
        phqchk73 = Radiobutton(displayPage2.inner, value=2, text="일주일이상", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage31, selectimage=phqchkImage32,
                               indicatoron=False, cursor="circle", variable=phqchkType7, font=malgungothic13, bd=0, command=phqchkImage73Func)
        phqchk73.deselect()
        global phqchk74
        phqchk74 = Radiobutton(displayPage2.inner, value=3, text="거의매일", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage41, selectimage=phqchkImage42,
                               indicatoron=False, cursor="circle", variable=phqchkType7, font=malgungothic13, bd=0, command=phqchkImage74Func)
        phqchk74.deselect()
        global phqchk81
        phqchk81 = Radiobutton(displayPage2.inner, value=0, text="전혀없음", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage11,
                               selectimage=phqchkImage12, indicatoron=False, cursor="circle", variable=phqchkType8, font=malgungothic13, bd=0, command=phqchkImage81Func)
        phqchk81.deselect()
        global phqchk82
        phqchk82 = Radiobutton(displayPage2.inner, value=1, text="며칠동안", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage21, selectimage=phqchkImage22,
                               indicatoron=False, cursor="circle", variable=phqchkType8, font=malgungothic13, bd=0, command=phqchkImage82Func)
        phqchk82.deselect()
        global phqchk83
        phqchk83 = Radiobutton(displayPage2.inner, value=2, text="일주일이상", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage31, selectimage=phqchkImage32,
                               indicatoron=False, cursor="circle", variable=phqchkType8, font=malgungothic13, bd=0, command=phqchkImage83Func)
        phqchk83.deselect()
        global phqchk84
        phqchk84 = Radiobutton(displayPage2.inner, value=3, text="거의매일", background="#FFFFFF", activebackground="#FFFFFF", image=phqchkImage41, selectimage=phqchkImage42,
                               indicatoron=False, cursor="circle", variable=phqchkType8, font=malgungothic13, bd=0, command=phqchkImage84Func)
        phqchk84.deselect()
        global phqlastchk1
        phqlastchk1 = Radiobutton(displayPage2.inner, text="전혀 어렵지 않다", value=3, background="#FFFFFF", activebackground="#FFFFFF", image=phqlastchkImage11,
                                  selectimage=phqlastchkImage12, indicatoron=False, cursor="circle", variable=phqlastchkType, font=malgungothic13, bd=0,  command=phqlastchkImageFunc1)
        phqlastchk1.deselect()
        global phqlastchk2
        phqlastchk2 = Radiobutton(displayPage2.inner, text="다소 어렵다", value=2, background="#FFFFFF", activebackground="#FFFFFF", image=phqlastchkImage21, selectimage=phqlastchkImage22,
                                  indicatoron=False, cursor="circle", variable=phqlastchkType, font=malgungothic13, bd=0,  command=phqlastchkImageFunc2)
        phqlastchk2.deselect()
        global phqlastchk3
        phqlastchk3 = Radiobutton(displayPage2.inner, value=1, text="많이 어렵다", background="#FFFFFF", activebackground="#FFFFFF", image=phqlastchkImage31, selectimage=phqlastchkImage32,
                                  indicatoron=False, cursor="circle", variable=phqlastchkType, font=malgungothic13, bd=0,  command=phqlastchkImageFunc3)
        phqlastchk3.deselect()
        global phqlastchk4
        phqlastchk4 = Radiobutton(displayPage2.inner, value=0, text="매우 많이 어렵다", background="#FFFFFF", activebackground="#FFFFFF", image=phqlastchkImage41, selectimage=phqlastchkImage42,
                                  indicatoron=False, cursor="circle", variable=phqlastchkType, font=malgungothic13, bd=0, command=phqlastchkImageFunc4)
        phqlastchk4.deselect()

        phqchk11.place(x=900, y=278, width=106, height=29)
        phqchk12.place(x=1002, y=278, width=100, height=29)
        phqchk13.place(x=1099, y=278, width=113, height=29)
        phqchk14.place(x=1212, y=278, width=105, height=29)
        phqchk21.place(x=900, y=338, width=106, height=29)
        phqchk22.place(x=1002, y=338, width=100, height=29)
        phqchk23.place(x=1099, y=338, width=113, heigh=29)
        phqchk24.place(x=1212, y=338, width=105, heigh=29)
        phqchk31.place(x=900, y=398, width=106, height=29)
        phqchk32.place(x=1002, y=398, width=100, height=29)
        phqchk33.place(x=1099, y=398, width=113, heigh=29)
        phqchk34.place(x=1212, y=398, width=105, heigh=29)
        phqchk41.place(x=900, y=458, width=106, height=29)
        phqchk42.place(x=1002, y=458, width=100, height=29)
        phqchk43.place(x=1099, y=458, width=113, heigh=29)
        phqchk44.place(x=1212, y=458, width=105, heigh=29)
        phqchk51.place(x=900, y=518, width=106, height=29)
        phqchk52.place(x=1002, y=518, width=100, height=29)
        phqchk53.place(x=1099, y=518, width=113, heigh=29)
        phqchk54.place(x=1212, y=518, width=105, heigh=29)
        phqchk61.place(x=900, y=578, width=106, height=29)
        phqchk62.place(x=1002, y=578, width=100, height=29)
        phqchk63.place(x=1099, y=578, width=113, heigh=29)
        phqchk64.place(x=1212, y=578, width=105, heigh=29)
        phqchk71.place(x=900, y=655, width=106, height=29)
        phqchk72.place(x=1002, y=655, width=100, height=29)
        phqchk73.place(x=1099, y=655, width=113, heigh=29)
        phqchk74.place(x=1212, y=655, width=105, heigh=29)
        phqchk81.place(x=900, y=733, width=106, height=29)
        phqchk82.place(x=1002, y=733, width=100, height=29)
        phqchk83.place(x=1099, y=733, width=113, height=29)
        phqchk84.place(x=1212, y=733, width=105, height=29)
        phqlastchk1.place(x=279, y=881, width=201, height=33)
        phqlastchk2.place(x=480, y=881, width=220, height=33)
        phqlastchk3.place(x=700, y=881, width=243, height=33)
        phqlastchk4.place(x=940, y=881, width=314, height=33)

    def cdichkFunc():
        global now
        now = datetime.now()
        global cdichk
        cdichk = True
        sheet2['E2'].value = (" % s년 % s월 % s일 % s시 % s분" %
                              (now.year, now.month, now.day, now.hour, now.minute))
        sheet2['K2'].value = idInput.get()

        def cdichkImage11Func():
            if not str(sheet2['C7'].value).strip() == 0:
                sheet2['C7'].value = str(int(cdichkType1.get()) + 1) + "."

        def cdichkImage12Func():
            if not str(sheet2['C7'].value).strip() == 1:
                sheet2['C7'].value = str(int(cdichkType1.get()) + 1) + "."

        def cdichkImage13Func():
            if not str(sheet2['C7'].value).strip() == 2:
                sheet2['C7'].value = str(int(cdichkType1.get()) + 1) + "."

        def cdichkImage21Func():
            if not str(sheet2['C12'].value).strip() == 0:
                sheet2['C12'].value = str(int(cdichkType2.get()) + 1) + "."

        def cdichkImage22Func():
            if not str(sheet2['C12'].value).strip() == 1:
                sheet2['C12'].value = str(int(cdichkType2.get()) + 1) + "."

        def cdichkImage23Func():
            if not str(sheet2['C12'].value).strip() == 2:
                sheet2['C12'].value = str(int(cdichkType2.get()) + 1) + "."

        def cdichkImage31Func():
            if not str(sheet2['C17'].value).strip() == 0:
                sheet2['C17'].value = str(int(cdichkType3.get()) + 1) + "."

        def cdichkImage32Func():
            if not str(sheet2['C17'].value).strip() == 1:
                sheet2['C17'].value = str(int(cdichkType3.get()) + 1) + "."

        def cdichkImage33Func():
            if not str(sheet2['C17'].value).strip() == 2:
                sheet2['C17'].value = str(int(cdichkType3.get()) + 1) + "."

        def cdichkImage41Func():
            if not str(sheet2['C22'].value).strip() == 0:
                sheet2['C22'].value = str(int(cdichkType4.get()) + 1) + "."

        def cdichkImage42Func():
            if not str(sheet2['C22'].value).strip() == 1:
                sheet2['C22'].value = str(int(cdichkType4.get()) + 1) + "."

        def cdichkImage43Func():
            if not str(sheet2['C22'].value).strip() == 2:
                sheet2['C22'].value = str(int(cdichkType4.get()) + 1) + "."

        def cdichkImage51Func():
            if not str(sheet2['C27'].value).strip() == 0:
                sheet2['C27'].value = str(int(cdichkType5.get()) + 1) + "."

        def cdichkImage52Func():
            if not str(sheet2['C27'].value).strip() == 1:
                sheet2['C27'].value = str(int(cdichkType5.get()) + 1) + "."

        def cdichkImage53Func():
            if not str(sheet2['C27'].value).strip() == 2:
                sheet2['C27'].value = str(int(cdichkType5.get()) + 1) + "."

        def cdichkImage61Func():
            if not str(sheet2['C32'].value).strip() == 0:
                sheet2['C32'].value = str(int(cdichkType6.get()) + 1) + "."

        def cdichkImage62Func():
            if not str(sheet2['C32'].value).strip() == 1:
                sheet2['C32'].value = str(int(cdichkType6.get()) + 1) + "."

        def cdichkImage63Func():
            if not str(sheet2['C32'].value).strip() == 2:
                sheet2['C32'].value = str(int(cdichkType6.get()) + 1) + "."

        def cdichkImage71Func():
            if not str(sheet2['C37'].value).strip() == 0:
                sheet2['C37'].value = str(int(cdichkType7.get()) + 1) + "."

        def cdichkImage72Func():
            if not str(sheet2['C37'].value).strip() == 1:
                sheet2['C37'].value = str(int(cdichkType7.get()) + 1) + "."

        def cdichkImage73Func():
            if not str(sheet2['C37'].value).strip() == 2:
                sheet2['C37'].value = str(int(cdichkType7.get()) + 1) + "."

        def cdichkImage81Func():
            if not str(sheet2['C42'].value).strip() == 0:
                sheet2['C42'].value = str(int(cdichkType8.get()) + 1) + "."

        def cdichkImage82Func():
            if not str(sheet2['C42'].value).strip() == 1:
                sheet2['C42'].value = str(int(cdichkType8.get()) + 1) + "."

        def cdichkImage83Func():
            if not str(sheet2['C42'].value).strip() == 2:
                sheet2['C42'].value = str(int(cdichkType8.get()) + 1) + "."

        def cdichkImage91Func():
            if not str(sheet2['C47'].value).strip() == 0:
                sheet2['C47'].value = str(int(cdichkType9.get()) + 1) + "."

        def cdichkImage92Func():
            if not str(sheet2['C47'].value).strip() == 1:
                sheet2['C47'].value = str(int(cdichkType9.get()) + 1) + "."

        def cdichkImage93Func():
            if not str(sheet2['C47'].value).strip() == 2:
                sheet2['C47'].value = str(int(cdichkType9.get()) + 1) + "."

        def cdichkImage101Func():
            if not str(sheet2['C52'].value).strip() == 0:
                sheet2['C52'].value = str(int(cdichkType10.get()) + 1) + "."

        def cdichkImage102Func():
            if not str(sheet2['C52'].value).strip() == 1:
                sheet2['C52'].value = str(int(cdichkType10.get()) + 1) + "."

        def cdichkImage103Func():
            if not str(sheet2['C52'].value).strip() == 2:
                sheet2['C52'].value = str(int(cdichkType10.get()) + 1) + "."

        def cdichkImage111Func():
            if not str(sheet2['C57'].value).strip() == 0:
                sheet2['C57'].value = str(int(cdichkType11.get()) + 1) + "."

        def cdichkImage112Func():
            if not str(sheet2['C57'].value).strip() == 1:
                sheet2['C57'].value = str(int(cdichkType11.get()) + 1) + "."

        def cdichkImage113Func():
            if not str(sheet2['C57'].value).strip() == 2:
                sheet2['C57'].value = str(int(cdichkType11.get()) + 1) + "."

        def cdichkImage121Func():
            if not str(sheet2['C62'].value).strip() == 0:
                sheet2['C62'].value = str(int(cdichkType12.get()) + 1) + "."

        def cdichkImage122Func():
            if not str(sheet2['C62'].value).strip() == 1:
                sheet2['C62'].value = str(int(cdichkType12.get()) + 1) + "."

        def cdichkImage123Func():
            if not str(sheet2['C62'].value).strip() == 2:
                sheet2['C62'].value = str(int(cdichkType12.get()) + 1) + "."

        def cdichkImage131Func():
            if not str(sheet2['C67'].value).strip() == 0:
                sheet2['C67'].value = str(int(cdichkType13.get()) + 1) + "."

        def cdichkImage132Func():
            if not str(sheet2['C67'].value).strip() == 1:
                sheet2['C67'].value = str(int(cdichkType13.get()) + 1) + "."

        def cdichkImage133Func():
            if not str(sheet2['C67'].value).strip() == 2:
                sheet2['C67'].value = str(int(cdichkType13.get()) + 1) + "."

        def cdichkImage141Func():
            if not str(sheet2['C72'].value).strip() == 0:
                sheet2['C72'].value = str(int(cdichkType14.get()) + 1) + "."

        def cdichkImage142Func():
            if not str(sheet2['C72'].value).strip() == 1:
                sheet2['C72'].value = str(int(cdichkType14.get()) + 1) + "."

        def cdichkImage143Func():
            if not str(sheet2['C72'].value).strip() == 2:
                sheet2['C72'].value = str(int(cdichkType14.get()) + 1) + "."

        def cdichkImage151Func():
            if not str(sheet2['C77'].value).strip() == 0:
                sheet2['C77'].value = str(int(cdichkType15.get()) + 1) + "."

        def cdichkImage152Func():
            if not str(sheet2['C77'].value).strip() == 1:
                sheet2['C77'].value = str(int(cdichkType15.get()) + 1) + "."

        def cdichkImage153Func():
            if not str(sheet2['C77'].value).strip() == 2:
                sheet2['C77'].value = str(int(cdichkType15.get()) + 1) + "."

        def cdichkImage161Func():
            if not str(sheet2['C82'].value).strip() == 0:
                sheet2['C82'].value = str(int(cdichkType16.get()) + 1) + "."

        def cdichkImage162Func():
            if not str(sheet2['C82'].value).strip() == 1:
                sheet2['C82'].value = str(int(cdichkType16.get()) + 1) + "."

        def cdichkImage163Func():
            if not str(sheet2['C82'].value).strip() == 2:
                sheet2['C82'].value = str(int(cdichkType16.get()) + 1) + "."

        def cdichkImage171Func():
            if not str(sheet2['C87'].value).strip() == 0:
                sheet2['C87'].value = str(int(cdichkType17.get()) + 1) + "."

        def cdichkImage172Func():
            if not str(sheet2['C87'].value).strip() == 1:
                sheet2['C87'].value = str(int(cdichkType17.get()) + 1) + "."

        def cdichkImage173Func():
            if not str(sheet2['C87'].value).strip() == 2:
                sheet2['C87'].value = str(int(cdichkType17.get()) + 1) + "."

        def cdichkImage181Func():
            if not str(sheet2['C92'].value).strip() == 0:
                sheet2['C92'].value = str(int(cdichkType18.get()) + 1) + "."

        def cdichkImage182Func():
            if not str(sheet2['C92'].value).strip() == 1:
                sheet2['C92'].value = str(int(cdichkType18.get()) + 1) + "."

        def cdichkImage183Func():
            if not str(sheet2['C92'].value).strip() == 2:
                sheet2['C92'].value = str(int(cdichkType18.get()) + 1) + "."

        def cdichkImage191Func():
            if not str(sheet2['C97'].value).strip() == 0:
                sheet2['C97'].value = str(int(cdichkType19.get()) + 1) + "."

        def cdichkImage192Func():
            if not str(sheet2['C97'].value).strip() == 1:
                sheet2['C97'].value = str(int(cdichkType19.get()) + 1) + "."

        def cdichkImage193Func():
            if not str(sheet2['C97'].value).strip() == 2:
                sheet2['C97'].value = str(int(cdichkType19.get()) + 1) + "."

        def cdichkImage201Func():
            if not str(sheet2['C102'].value).strip() == 0:
                sheet2['C102'].value = str(int(cdichkType20.get()) + 1) + "."

        def cdichkImage202Func():
            if not str(sheet2['C102'].value).strip() == 1:
                sheet2['C102'].value = str(int(cdichkType20.get()) + 1) + "."

        def cdichkImage203Func():
            if not str(sheet2['C102'].value).strip() == 2:
                sheet2['C102'].value = str(int(cdichkType20.get()) + 1) + "."

        def cdichkImage211Func():
            if not str(sheet2['C107'].value).strip() == 0:
                sheet2['C107'].value = str(int(cdichkType21.get()) + 1) + "."

        def cdichkImage212Func():
            if not str(sheet2['C107'].value).strip() == 1:
                sheet2['C107'].value = str(int(cdichkType21.get()) + 1) + "."

        def cdichkImage213Func():
            if not str(sheet2['C107'].value).strip() == 2:
                sheet2['C107'].value = str(int(cdichkType21.get()) + 1) + "."

        def cdichkImage221Func():
            if not str(sheet2['C112'].value).strip() == 0:
                sheet2['C112'].value = str(int(cdichkType22.get()) + 1) + "."

        def cdichkImage222Func():
            if not str(sheet2['C112'].value).strip() == 1:
                sheet2['C112'].value = str(int(cdichkType22.get()) + 1) + "."

        def cdichkImage223Func():
            if not str(sheet2['C112'].value).strip() == 2:
                sheet2['C112'].value = str(int(cdichkType22.get()) + 1) + "."

        def cdichkImage231Func():
            if not str(sheet2['C117'].value).strip() == 0:
                sheet2['C117'].value = str(int(cdichkType23.get()) + 1) + "."

        def cdichkImage232Func():
            if not str(sheet2['C117'].value).strip() == 1:
                sheet2['C117'].value = str(int(cdichkType23.get()) + 1) + "."

        def cdichkImage233Func():
            if not str(sheet2['C117'].value).strip() == 2:
                sheet2['C117'].value = str(int(cdichkType23.get()) + 1) + "."

        def cdichkImage241Func():
            if not str(sheet2['C122'].value).strip() == 0:
                sheet2['C122'].value = str(int(cdichkType24.get()) + 1) + "."

        def cdichkImage242Func():
            if not str(sheet2['C122'].value).strip() == 1:
                sheet2['C122'].value = str(int(cdichkType24.get()) + 1) + "."

        def cdichkImage243Func():
            if not str(sheet2['C122'].value).strip() == 2:
                sheet2['C122'].value = str(int(cdichkType24.get()) + 1) + "."

        def cdichkImage251Func():
            if not str(sheet2['C127'].value).strip() == 0:
                sheet2['C127'].value = str(int(cdichkType25.get()) + 1) + "."

        def cdichkImage252Func():
            if not str(sheet2['C127'].value).strip() == 1:
                sheet2['C127'].value = str(int(cdichkType25.get()) + 1) + "."

        def cdichkImage253Func():
            if not str(sheet2['C127'].value).strip() == 2:
                sheet2['C127'].value = str(int(cdichkType25.get()) + 1) + "."

        def cdichkImage261Func():
            if not str(sheet2['C132'].value).strip() == 0:
                sheet2['C132'].value = str(int(cdichkType26.get()) + 1) + "."

        def cdichkImage262Func():
            if not str(sheet2['C132'].value).strip() == 1:
                sheet2['C132'].value = str(int(cdichkType26.get()) + 1) + "."

        def cdichkImage263Func():
            if not str(sheet2['C132'].value).strip() == 2:
                sheet2['C132'].value = str(int(cdichkType26.get()) + 1) + "."

        def cdichkImage271Func():
            if not str(sheet2['C137'].value).strip() == 0:
                sheet2['C137'].value = str(int(cdichkType27.get()) + 1) + "."

        def cdichkImage272Func():
            if not str(sheet2['C137'].value).strip() == 1:
                sheet2['C137'].value = str(int(cdichkType27.get()) + 1) + "."

        def cdichkImage273Func():
            if not str(sheet2['C137'].value).strip() == 2:
                sheet2['C137'].value = str(int(cdichkType27.get()) + 1) + "."

        frame2CdiBgImg = Image.open("images/cdibg.png")
        frame2CdiBg = ImageTk.PhotoImage(frame2CdiBgImg)
        frame2CdiBgLabel = tkinter.Label(displayPage.inner, image=frame2CdiBg)
        frame2CdiBgLabel.image = frame2CdiBg
        frame2CdiBgLabel.place(x=67, y=420, height=3653)
        frame2CdiBgLabel.pack()

        global cdichk11
        cdichk11 = Radiobutton(displayPage.inner, value=0,  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                               selectimage=chkImage12, indicatoron=False, cursor="circle", variable=cdichkType1, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage11Func)
        cdichk11.deselect()
        global cdichk12
        cdichk12 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=cdichkType1, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage12Func)
        cdichk12.deselect()
        global cdichk13
        cdichk13 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=cdichkType1, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage13Func)
        cdichk13.deselect()
        global cdichk21
        cdichk21 = Radiobutton(displayPage.inner, value=0,  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                               selectimage=chkImage12, indicatoron=False, cursor="circle", variable=cdichkType2, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage21Func)
        cdichk21.deselect()
        global cdichk22
        cdichk22 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=cdichkType2, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage22Func)
        cdichk22.deselect()
        global cdichk23
        cdichk23 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=cdichkType2, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage23Func)
        cdichk23.deselect()
        global cdichk31
        cdichk31 = Radiobutton(displayPage.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                               selectimage=chkImage12, indicatoron=False, cursor="circle", variable=cdichkType3, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage31Func)
        cdichk31.deselect()
        global cdichk32
        cdichk32 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=cdichkType3, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage32Func)
        cdichk32.deselect()
        global cdichk33
        cdichk33 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=cdichkType3, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage33Func)
        cdichk33.deselect()
        global cdichk41
        cdichk41 = Radiobutton(displayPage.inner, value=0,  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                               selectimage=chkImage12, indicatoron=False, cursor="circle", variable=cdichkType4, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage41Func)
        cdichk41.deselect()
        global cdichk42
        cdichk42 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=cdichkType4, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage42Func)
        cdichk42.deselect()
        global cdichk43
        cdichk43 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=cdichkType4, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage43Func)
        cdichk43.deselect()
        global cdichk51
        cdichk51 = Radiobutton(displayPage.inner, value=0,  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                               selectimage=chkImage12, indicatoron=False, cursor="circle", variable=cdichkType5, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage51Func)
        cdichk51.deselect()
        global cdichk52
        cdichk52 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=cdichkType5, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage52Func)
        cdichk52.deselect()
        global cdichk53
        cdichk53 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=cdichkType5, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage53Func)
        cdichk53.deselect()
        global cdichk61
        cdichk61 = Radiobutton(displayPage.inner, value=0,  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                               selectimage=chkImage12, indicatoron=False, cursor="circle", variable=cdichkType6, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage61Func)
        cdichk61.deselect()
        global cdichk62
        cdichk62 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=cdichkType6, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage62Func)
        cdichk62.deselect()
        global cdichk63
        cdichk63 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=cdichkType6, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage63Func)
        cdichk63.deselect()
        global cdichk71
        cdichk71 = Radiobutton(displayPage.inner, value=0,  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                               selectimage=chkImage12, indicatoron=False, cursor="circle", variable=cdichkType7, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage71Func)
        cdichk71.deselect()
        global cdichk72
        cdichk72 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=cdichkType7, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage72Func)
        cdichk72.deselect()
        global cdichk73
        cdichk73 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=cdichkType7, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage73Func)
        cdichk73.deselect()
        global cdichk81
        cdichk81 = Radiobutton(displayPage.inner, value=0,  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                               selectimage=chkImage12, indicatoron=False, cursor="circle", variable=cdichkType8, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage81Func)
        cdichk81.deselect()
        global cdichk82
        cdichk82 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=cdichkType8, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage82Func)
        cdichk82.deselect()
        global cdichk83
        cdichk83 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=cdichkType8, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage83Func)
        cdichk83.deselect()
        global cdichk91
        cdichk91 = Radiobutton(displayPage.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                               selectimage=chkImage12, indicatoron=False, cursor="circle", variable=cdichkType9, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage91Func)
        cdichk91.deselect()
        global cdichk92
        cdichk92 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=cdichkType9, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage92Func)
        cdichk92.deselect()
        global cdichk93
        cdichk93 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=cdichkType9, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage93Func)
        cdichk93.deselect()
        global cdichk101
        cdichk101 = Radiobutton(displayPage.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType10, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage101Func)
        cdichk101.deselect()
        global cdichk102
        cdichk102 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType10, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage102Func)
        cdichk102.deselect()
        global cdichk103
        cdichk103 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType10, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage103Func)
        cdichk103.deselect()
        global cdichk111
        cdichk111 = Radiobutton(displayPage.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType11, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage111Func)
        cdichk111.deselect()
        global cdichk112
        cdichk112 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType11, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage112Func)
        cdichk112.deselect()
        global cdichk113
        cdichk113 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType11, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage113Func)
        cdichk113.deselect()
        global cdichk121
        cdichk121 = Radiobutton(displayPage.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType12, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage121Func)
        cdichk121.deselect()
        global cdichk122
        cdichk122 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType12, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage122Func)
        cdichk122.deselect()
        global cdichk123
        cdichk123 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType12, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage123Func)
        cdichk123.deselect()
        global cdichk131
        cdichk131 = Radiobutton(displayPage.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType13, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage131Func)
        cdichk131.deselect()
        global cdichk132
        cdichk132 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType13, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage132Func)
        cdichk132.deselect()
        global cdichk133
        cdichk133 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType13, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage133Func)
        cdichk133.deselect()
        global cdichk141
        cdichk141 = Radiobutton(displayPage.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType14, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage141Func)
        cdichk141.deselect()
        global cdichk142
        cdichk142 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType14, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage142Func)
        cdichk142.deselect()
        global cdichk143
        cdichk143 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType14, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage143Func)
        cdichk143.deselect()
        global cdichk151
        cdichk151 = Radiobutton(displayPage.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType15, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage151Func)
        cdichk151.deselect()
        global cdichk152
        cdichk152 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType15, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage152Func)
        cdichk152.deselect()
        global cdichk153
        cdichk153 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType15, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage153Func)
        cdichk153.deselect()
        global cdichk161
        cdichk161 = Radiobutton(displayPage.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType16, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage161Func)
        cdichk161.deselect()
        global cdichk162
        cdichk162 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType16, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage162Func)
        cdichk162.deselect()
        global cdichk163
        cdichk163 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType16, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage163Func)
        cdichk163.deselect()
        global cdichk171
        cdichk171 = Radiobutton(displayPage.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType17, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage171Func)
        cdichk171.deselect()
        global cdichk172
        cdichk172 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType17, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage172Func)
        cdichk172.deselect()
        global cdichk173
        cdichk173 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType17, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage173Func)
        cdichk173.deselect()
        global cdichk181
        cdichk181 = Radiobutton(displayPage.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType18, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage181Func)
        cdichk181.deselect()
        global cdichk182
        cdichk182 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType18, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage182Func)
        cdichk182.deselect()
        global cdichk183
        cdichk183 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType18, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage183Func)
        cdichk183.deselect()
        global cdichk191
        cdichk191 = Radiobutton(displayPage.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType19, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage191Func)
        cdichk191.deselect()
        global cdichk192
        cdichk192 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType19, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage192Func)
        cdichk192.deselect()
        global cdichk193
        cdichk193 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType19, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage193Func)
        cdichk193.deselect()
        global cdichk201
        cdichk201 = Radiobutton(displayPage.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType20, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage201Func)
        cdichk201.deselect()
        global cdichk202
        cdichk202 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType20, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage202Func)
        cdichk202.deselect()
        global cdichk203
        cdichk203 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType20, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage203Func)
        cdichk203.deselect()
        global cdichk211
        cdichk211 = Radiobutton(displayPage.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType21, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage211Func)
        cdichk211.deselect()
        global cdichk212
        cdichk212 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType21, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage212Func)
        cdichk212.deselect()
        global cdichk213
        cdichk213 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType21, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage213Func)
        cdichk213.deselect()
        global cdichk221
        cdichk221 = Radiobutton(displayPage.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType22, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage221Func)
        cdichk221.deselect()
        global cdichk222
        cdichk222 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType22, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage222Func)
        cdichk222.deselect()
        global cdichk223
        cdichk223 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType22, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage223Func)
        cdichk223.deselect()
        global cdichk231
        cdichk231 = Radiobutton(displayPage.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType23, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage231Func)
        cdichk231.deselect()
        global cdichk232
        cdichk232 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType23, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage232Func)
        cdichk232.deselect()
        global cdichk233
        cdichk233 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType23, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage233Func)
        cdichk233.deselect()
        global cdichk241
        cdichk241 = Radiobutton(displayPage.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType24, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage241Func)
        cdichk241.deselect()
        global cdichk242
        cdichk242 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType24, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage242Func)
        cdichk242.deselect()
        global cdichk243
        cdichk243 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType24, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage243Func)
        cdichk243.deselect()
        global cdichk251
        cdichk251 = Radiobutton(displayPage.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType25, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage251Func)
        cdichk251.deselect()
        global cdichk252
        cdichk252 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType25, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage252Func)
        cdichk252.deselect()
        global cdichk253
        cdichk253 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType25, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage253Func)
        cdichk253.deselect()
        global cdichk261
        cdichk261 = Radiobutton(displayPage.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType26, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage261Func)
        cdichk261.deselect()
        global cdichk262
        cdichk262 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType26, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage262Func)
        cdichk262.deselect()
        global cdichk263
        cdichk263 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType26, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage263Func)
        cdichk263.deselect()
        global cdichk271
        cdichk271 = Radiobutton(displayPage.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType27, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage271Func)
        cdichk271.deselect()
        global cdichk272
        cdichk272 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType27, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage272Func)
        cdichk272.deselect()
        global cdichk273
        cdichk273 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType27, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage273Func)
        cdichk273.deselect()

        cdichk11.place(x=428, y=264, width=21, height=21)
        cdichk12.place(x=428, y=303, width=21, height=21)
        cdichk13.place(x=428, y=340, width=21, height=21)
        cdichk21.place(x=428, y=380, width=21, height=21)
        cdichk22.place(x=428, y=419, width=21, height=21)
        cdichk23.place(x=428, y=459, width=21, height=21)
        cdichk31.place(x=428, y=496, width=21, height=21)
        cdichk32.place(x=428, y=535, width=21, height=21)
        cdichk33.place(x=428, y=575, width=21, height=21)
        cdichk41.place(x=428, y=614, width=21, height=21)
        cdichk42.place(x=428, y=651, width=21, height=21)
        cdichk43.place(x=428, y=691, width=21, height=21)
        cdichk51.place(x=428, y=730, width=21, height=21)
        cdichk52.place(x=428, y=767, width=21, height=21)
        cdichk53.place(x=428, y=807, width=21, height=21)
        cdichk61.place(x=428, y=846, width=21, height=21)
        cdichk62.place(x=428, y=883, width=21, height=21)
        cdichk63.place(x=428, y=923, width=21, height=21)
        cdichk71.place(x=428, y=965, width=21, height=21)
        cdichk72.place(x=428, y=1002, width=21, height=21)
        cdichk73.place(x=428, y=1042, width=21, height=21)
        cdichk81.place(x=428, y=1079, width=21, height=21)
        cdichk82.place(x=428, y=1118, width=21, height=21)
        cdichk83.place(x=428, y=1158, width=21, height=21)
        cdichk91.place(x=428, y=1198, width=21, height=21)
        cdichk92.place(x=428, y=1237, width=21, height=21)
        cdichk93.place(x=428, y=1277, width=21, height=21)
        cdichk101.place(x=428, y=1315, width=21, height=21)
        cdichk102.place(x=428, y=1354, width=21, height=21)
        cdichk103.place(x=428, y=1394, width=21, height=21)
        cdichk111.place(x=428, y=1432, width=21, height=21)
        cdichk112.place(x=428, y=1470, width=21, height=21)
        cdichk113.place(x=428, y=1510, width=21, height=21)
        cdichk121.place(x=428, y=1548, width=21, height=21)
        cdichk122.place(x=428, y=1588, width=21, height=21)
        cdichk123.place(x=428, y=1626, width=21, height=21)
        cdichk131.place(x=428, y=1664, width=21, height=21)
        cdichk132.place(x=428, y=1704, width=21, height=21)
        cdichk133.place(x=428, y=1742, width=21, height=21)
        cdichk141.place(x=428, y=1783, width=21, height=21)
        cdichk142.place(x=428, y=1823, width=21, height=21)
        cdichk143.place(x=428, y=1861, width=21, height=21)
        cdichk151.place(x=428, y=1899, width=21, height=21)
        cdichk152.place(x=428, y=1939, width=21, height=21)
        cdichk153.place(x=428, y=1977, width=21, height=21)
        cdichk161.place(x=428, y=2017, width=21, height=21)
        cdichk162.place(x=428, y=2057, width=21, height=21)
        cdichk163.place(x=428, y=2095, width=21, height=21)
        cdichk171.place(x=428, y=2133, width=21, height=21)
        cdichk172.place(x=428, y=2172, width=21, height=21)
        cdichk173.place(x=428, y=2211, width=21, height=21)
        cdichk181.place(x=428, y=2251, width=21, height=21)
        cdichk182.place(x=428, y=2289, width=21, height=21)
        cdichk183.place(x=428, y=2328, width=21, height=21)
        cdichk191.place(x=428, y=2367, width=21, height=21)
        cdichk192.place(x=428, y=2407, width=21, height=21)
        cdichk193.place(x=428, y=2445, width=21, height=21)
        cdichk201.place(x=428, y=2484, width=21, height=21)
        cdichk202.place(x=428, y=2523, width=21, height=21)
        cdichk203.place(x=428, y=2560, width=21, height=21)
        cdichk211.place(x=428, y=2600, width=21, height=21)
        cdichk212.place(x=428, y=2640, width=21, height=21)
        cdichk213.place(x=428, y=2679, width=21, height=21)
        cdichk221.place(x=428, y=2718, width=21, height=21)
        cdichk222.place(x=428, y=2758, width=21, height=21)
        cdichk223.place(x=428, y=2797, width=21, height=21)
        cdichk231.place(x=428, y=2835, width=21, height=21)
        cdichk232.place(x=428, y=2875, width=21, height=21)
        cdichk233.place(x=428, y=2912, width=21, height=21)
        cdichk241.place(x=428, y=2953, width=21, height=21)
        cdichk242.place(x=428, y=2992, width=21, height=21)
        cdichk243.place(x=428, y=3030, width=21, height=21)
        cdichk251.place(x=428, y=3070, width=21, height=21)
        cdichk252.place(x=428, y=3108, width=21, height=21)
        cdichk253.place(x=428, y=3147, width=21, height=21)
        cdichk261.place(x=428, y=3186, width=21, height=21)
        cdichk262.place(x=428, y=3226, width=21, height=21)
        cdichk263.place(x=428, y=3263, width=21, height=21)
        cdichk271.place(x=428, y=3304, width=21, height=21)
        cdichk272.place(x=428, y=3342, width=21, height=21)
        cdichk273.place(x=428, y=3381, width=21, height=21)

    def cdichkFunc2():
        global now
        now = datetime.now()
        global cdichk
        cdichk = True
        sheet2['E2'].value = (" % s년 % s월 % s일 % s시 % s분" %
                              (now.year, now.month, now.day, now.hour, now.minute))
        sheet2['K2'].value = idInput.get()

        def cdichkImage11Func():
            if not str(sheet2['C7'].value).strip() == 0:
                sheet2['C7'].value = str(int(cdichkType1.get()) + 1) + "."

        def cdichkImage12Func():
            if not str(sheet2['C7'].value).strip() == 1:
                sheet2['C7'].value = str(int(cdichkType1.get()) + 1) + "."

        def cdichkImage13Func():
            if not str(sheet2['C7'].value).strip() == 2:
                sheet2['C7'].value = str(int(cdichkType1.get()) + 1) + "."

        def cdichkImage21Func():
            if not str(sheet2['C12'].value).strip() == 0:
                sheet2['C12'].value = str(int(cdichkType2.get()) + 1) + "."

        def cdichkImage22Func():
            if not str(sheet2['C12'].value).strip() == 1:
                sheet2['C12'].value = str(int(cdichkType2.get()) + 1) + "."

        def cdichkImage23Func():
            if not str(sheet2['C12'].value).strip() == 2:
                sheet2['C12'].value = str(int(cdichkType2.get()) + 1) + "."

        def cdichkImage31Func():
            if not str(sheet2['C17'].value).strip() == 0:
                sheet2['C17'].value = str(int(cdichkType3.get()) + 1) + "."

        def cdichkImage32Func():
            if not str(sheet2['C17'].value).strip() == 1:
                sheet2['C17'].value = str(int(cdichkType3.get()) + 1) + "."

        def cdichkImage33Func():
            if not str(sheet2['C17'].value).strip() == 2:
                sheet2['C17'].value = str(int(cdichkType3.get()) + 1) + "."

        def cdichkImage41Func():
            if not str(sheet2['C22'].value).strip() == 0:
                sheet2['C22'].value = str(int(cdichkType4.get()) + 1) + "."

        def cdichkImage42Func():
            if not str(sheet2['C22'].value).strip() == 1:
                sheet2['C22'].value = str(int(cdichkType4.get()) + 1) + "."

        def cdichkImage43Func():
            if not str(sheet2['C22'].value).strip() == 2:
                sheet2['C22'].value = str(int(cdichkType4.get()) + 1) + "."

        def cdichkImage51Func():
            if not str(sheet2['C27'].value).strip() == 0:
                sheet2['C27'].value = str(int(cdichkType5.get()) + 1) + "."

        def cdichkImage52Func():
            if not str(sheet2['C27'].value).strip() == 1:
                sheet2['C27'].value = str(int(cdichkType5.get()) + 1) + "."

        def cdichkImage53Func():
            if not str(sheet2['C27'].value).strip() == 2:
                sheet2['C27'].value = str(int(cdichkType5.get()) + 1) + "."

        def cdichkImage61Func():
            if not str(sheet2['C32'].value).strip() == 0:
                sheet2['C32'].value = str(int(cdichkType6.get()) + 1) + "."

        def cdichkImage62Func():
            if not str(sheet2['C32'].value).strip() == 1:
                sheet2['C32'].value = str(int(cdichkType6.get()) + 1) + "."

        def cdichkImage63Func():
            if not str(sheet2['C32'].value).strip() == 2:
                sheet2['C32'].value = str(int(cdichkType6.get()) + 1) + "."

        def cdichkImage71Func():
            if not str(sheet2['C37'].value).strip() == 0:
                sheet2['C37'].value = str(int(cdichkType7.get()) + 1) + "."

        def cdichkImage72Func():
            if not str(sheet2['C37'].value).strip() == 1:
                sheet2['C37'].value = str(int(cdichkType7.get()) + 1) + "."

        def cdichkImage73Func():
            if not str(sheet2['C37'].value).strip() == 2:
                sheet2['C37'].value = str(int(cdichkType7.get()) + 1) + "."

        def cdichkImage81Func():
            if not str(sheet2['C42'].value).strip() == 0:
                sheet2['C42'].value = str(int(cdichkType8.get()) + 1) + "."

        def cdichkImage82Func():
            if not str(sheet2['C42'].value).strip() == 1:
                sheet2['C42'].value = str(int(cdichkType8.get()) + 1) + "."

        def cdichkImage83Func():
            if not str(sheet2['C42'].value).strip() == 2:
                sheet2['C42'].value = str(int(cdichkType8.get()) + 1) + "."

        def cdichkImage91Func():
            if not str(sheet2['C47'].value).strip() == 0:
                sheet2['C47'].value = str(int(cdichkType9.get()) + 1) + "."

        def cdichkImage92Func():
            if not str(sheet2['C47'].value).strip() == 1:
                sheet2['C47'].value = str(int(cdichkType9.get()) + 1) + "."

        def cdichkImage93Func():
            if not str(sheet2['C47'].value).strip() == 2:
                sheet2['C47'].value = str(int(cdichkType9.get()) + 1) + "."

        def cdichkImage101Func():
            if not str(sheet2['C52'].value).strip() == 0:
                sheet2['C52'].value = str(int(cdichkType10.get()) + 1) + "."

        def cdichkImage102Func():
            if not str(sheet2['C52'].value).strip() == 1:
                sheet2['C52'].value = str(int(cdichkType10.get()) + 1) + "."

        def cdichkImage103Func():
            if not str(sheet2['C52'].value).strip() == 2:
                sheet2['C52'].value = str(int(cdichkType10.get()) + 1) + "."

        def cdichkImage111Func():
            if not str(sheet2['C57'].value).strip() == 0:
                sheet2['C57'].value = str(int(cdichkType11.get()) + 1) + "."

        def cdichkImage112Func():
            if not str(sheet2['C57'].value).strip() == 1:
                sheet2['C57'].value = str(int(cdichkType11.get()) + 1) + "."

        def cdichkImage113Func():
            if not str(sheet2['C57'].value).strip() == 2:
                sheet2['C57'].value = str(int(cdichkType11.get()) + 1) + "."

        def cdichkImage121Func():
            if not str(sheet2['C62'].value).strip() == 0:
                sheet2['C62'].value = str(int(cdichkType12.get()) + 1) + "."

        def cdichkImage122Func():
            if not str(sheet2['C62'].value).strip() == 1:
                sheet2['C62'].value = str(int(cdichkType12.get()) + 1) + "."

        def cdichkImage123Func():
            if not str(sheet2['C62'].value).strip() == 2:
                sheet2['C62'].value = str(int(cdichkType12.get()) + 1) + "."

        def cdichkImage131Func():
            if not str(sheet2['C67'].value).strip() == 0:
                sheet2['C67'].value = str(int(cdichkType13.get()) + 1) + "."

        def cdichkImage132Func():
            if not str(sheet2['C67'].value).strip() == 1:
                sheet2['C67'].value = str(int(cdichkType13.get()) + 1) + "."

        def cdichkImage133Func():
            if not str(sheet2['C67'].value).strip() == 2:
                sheet2['C67'].value = str(int(cdichkType13.get()) + 1) + "."

        def cdichkImage141Func():
            if not str(sheet2['C72'].value).strip() == 0:
                sheet2['C72'].value = str(int(cdichkType14.get()) + 1) + "."

        def cdichkImage142Func():
            if not str(sheet2['C72'].value).strip() == 1:
                sheet2['C72'].value = str(int(cdichkType14.get()) + 1) + "."

        def cdichkImage143Func():
            if not str(sheet2['C72'].value).strip() == 2:
                sheet2['C72'].value = str(int(cdichkType14.get()) + 1) + "."

        def cdichkImage151Func():
            if not str(sheet2['C77'].value).strip() == 0:
                sheet2['C77'].value = str(int(cdichkType15.get()) + 1) + "."

        def cdichkImage152Func():
            if not str(sheet2['C77'].value).strip() == 1:
                sheet2['C77'].value = str(int(cdichkType15.get()) + 1) + "."

        def cdichkImage153Func():
            if not str(sheet2['C77'].value).strip() == 2:
                sheet2['C77'].value = str(int(cdichkType15.get()) + 1) + "."

        def cdichkImage161Func():
            if not str(sheet2['C82'].value).strip() == 0:
                sheet2['C82'].value = str(int(cdichkType16.get()) + 1) + "."

        def cdichkImage162Func():
            if not str(sheet2['C82'].value).strip() == 1:
                sheet2['C82'].value = str(int(cdichkType16.get()) + 1) + "."

        def cdichkImage163Func():
            if not str(sheet2['C82'].value).strip() == 2:
                sheet2['C82'].value = str(int(cdichkType16.get()) + 1) + "."

        def cdichkImage171Func():
            if not str(sheet2['C87'].value).strip() == 0:
                sheet2['C87'].value = str(int(cdichkType17.get()) + 1) + "."

        def cdichkImage172Func():
            if not str(sheet2['C87'].value).strip() == 1:
                sheet2['C87'].value = str(int(cdichkType17.get()) + 1) + "."

        def cdichkImage173Func():
            if not str(sheet2['C87'].value).strip() == 2:
                sheet2['C87'].value = str(int(cdichkType17.get()) + 1) + "."

        def cdichkImage181Func():
            if not str(sheet2['C92'].value).strip() == 0:
                sheet2['C92'].value = str(int(cdichkType18.get()) + 1) + "."

        def cdichkImage182Func():
            if not str(sheet2['C92'].value).strip() == 1:
                sheet2['C92'].value = str(int(cdichkType18.get()) + 1) + "."

        def cdichkImage183Func():
            if not str(sheet2['C92'].value).strip() == 2:
                sheet2['C92'].value = str(int(cdichkType18.get()) + 1) + "."

        def cdichkImage191Func():
            if not str(sheet2['C97'].value).strip() == 0:
                sheet2['C97'].value = str(int(cdichkType19.get()) + 1) + "."

        def cdichkImage192Func():
            if not str(sheet2['C97'].value).strip() == 1:
                sheet2['C97'].value = str(int(cdichkType19.get()) + 1) + "."

        def cdichkImage193Func():
            if not str(sheet2['C97'].value).strip() == 2:
                sheet2['C97'].value = str(int(cdichkType19.get()) + 1) + "."

        def cdichkImage201Func():
            if not str(sheet2['C102'].value).strip() == 0:
                sheet2['C102'].value = str(int(cdichkType20.get()) + 1) + "."

        def cdichkImage202Func():
            if not str(sheet2['C102'].value).strip() == 1:
                sheet2['C102'].value = str(int(cdichkType20.get()) + 1) + "."

        def cdichkImage203Func():
            if not str(sheet2['C102'].value).strip() == 2:
                sheet2['C102'].value = str(int(cdichkType20.get()) + 1) + "."

        def cdichkImage211Func():
            if not str(sheet2['C107'].value).strip() == 0:
                sheet2['C107'].value = str(int(cdichkType21.get()) + 1) + "."

        def cdichkImage212Func():
            if not str(sheet2['C107'].value).strip() == 1:
                sheet2['C107'].value = str(int(cdichkType21.get()) + 1) + "."

        def cdichkImage213Func():
            if not str(sheet2['C107'].value).strip() == 2:
                sheet2['C107'].value = str(int(cdichkType21.get()) + 1) + "."

        def cdichkImage221Func():
            if not str(sheet2['C112'].value).strip() == 0:
                sheet2['C112'].value = str(int(cdichkType22.get()) + 1) + "."

        def cdichkImage222Func():
            if not str(sheet2['C112'].value).strip() == 1:
                sheet2['C112'].value = str(int(cdichkType22.get()) + 1) + "."

        def cdichkImage223Func():
            if not str(sheet2['C112'].value).strip() == 2:
                sheet2['C112'].value = str(int(cdichkType22.get()) + 1) + "."

        def cdichkImage231Func():
            if not str(sheet2['C117'].value).strip() == 0:
                sheet2['C117'].value = str(int(cdichkType23.get()) + 1) + "."

        def cdichkImage232Func():
            if not str(sheet2['C117'].value).strip() == 1:
                sheet2['C117'].value = str(int(cdichkType23.get()) + 1) + "."

        def cdichkImage233Func():
            if not str(sheet2['C117'].value).strip() == 2:
                sheet2['C117'].value = str(int(cdichkType23.get()) + 1) + "."

        def cdichkImage241Func():
            if not str(sheet2['C122'].value).strip() == 0:
                sheet2['C122'].value = str(int(cdichkType24.get()) + 1) + "."

        def cdichkImage242Func():
            if not str(sheet2['C122'].value).strip() == 1:
                sheet2['C122'].value = str(int(cdichkType24.get()) + 1) + "."

        def cdichkImage243Func():
            if not str(sheet2['C122'].value).strip() == 2:
                sheet2['C122'].value = str(int(cdichkType24.get()) + 1) + "."

        def cdichkImage251Func():
            if not str(sheet2['C127'].value).strip() == 0:
                sheet2['C127'].value = str(int(cdichkType25.get()) + 1) + "."

        def cdichkImage252Func():
            if not str(sheet2['C127'].value).strip() == 1:
                sheet2['C127'].value = str(int(cdichkType25.get()) + 1) + "."

        def cdichkImage253Func():
            if not str(sheet2['C127'].value).strip() == 2:
                sheet2['C127'].value = str(int(cdichkType25.get()) + 1) + "."

        def cdichkImage261Func():
            if not str(sheet2['C132'].value).strip() == 0:
                sheet2['C132'].value = str(int(cdichkType26.get()) + 1) + "."

        def cdichkImage262Func():
            if not str(sheet2['C132'].value).strip() == 1:
                sheet2['C132'].value = str(int(cdichkType26.get()) + 1) + "."

        def cdichkImage263Func():
            if not str(sheet2['C132'].value).strip() == 2:
                sheet2['C132'].value = str(int(cdichkType26.get()) + 1) + "."

        def cdichkImage271Func():
            if not str(sheet2['C137'].value).strip() == 0:
                sheet2['C137'].value = str(int(cdichkType27.get()) + 1) + "."

        def cdichkImage272Func():
            if not str(sheet2['C137'].value).strip() == 1:
                sheet2['C137'].value = str(int(cdichkType27.get()) + 1) + "."

        def cdichkImage273Func():
            if not str(sheet2['C137'].value).strip() == 2:
                sheet2['C137'].value = str(int(cdichkType27.get()) + 1) + "."

        frame2CdiBgImg = Image.open("images/cdibg.png")
        frame2CdiBg = ImageTk.PhotoImage(frame2CdiBgImg)
        frame2CdiBgLabel = tkinter.Label(displayPage2.inner, image=frame2CdiBg)
        frame2CdiBgLabel.image = frame2CdiBg
        frame2CdiBgLabel.place(x=47, y=420, height=3653)
        frame2CdiBgLabel.pack()

        global cdichk11
        cdichk11 = Radiobutton(displayPage2.inner, value=0,  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                               selectimage=chkImage12, indicatoron=False, cursor="circle", variable=cdichkType1, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage11Func)
        cdichk11.deselect()
        global cdichk12
        cdichk12 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=cdichkType1, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage12Func)
        cdichk12.deselect()
        global cdichk13
        cdichk13 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=cdichkType1, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage13Func)
        cdichk13.deselect()
        global cdichk21
        cdichk21 = Radiobutton(displayPage2.inner, value=0,  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                               selectimage=chkImage12, indicatoron=False, cursor="circle", variable=cdichkType2, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage21Func)
        cdichk21.deselect()
        global cdichk22
        cdichk22 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=cdichkType2, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage22Func)
        cdichk22.deselect()
        global cdichk23
        cdichk23 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=cdichkType2, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage23Func)
        cdichk23.deselect()
        global cdichk31
        cdichk31 = Radiobutton(displayPage2.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                               selectimage=chkImage12, indicatoron=False, cursor="circle", variable=cdichkType3, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage31Func)
        cdichk31.deselect()
        global cdichk32
        cdichk32 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=cdichkType3, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage32Func)
        cdichk32.deselect()
        global cdichk33
        cdichk33 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=cdichkType3, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage33Func)
        cdichk33.deselect()
        global cdichk41
        cdichk41 = Radiobutton(displayPage2.inner, value=0,  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                               selectimage=chkImage12, indicatoron=False, cursor="circle", variable=cdichkType4, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage41Func)
        cdichk41.deselect()
        global cdichk42
        cdichk42 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=cdichkType4, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage42Func)
        cdichk42.deselect()
        global cdichk43
        cdichk43 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=cdichkType4, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage43Func)
        cdichk43.deselect()
        global cdichk51
        cdichk51 = Radiobutton(displayPage2.inner, value=0,  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                               selectimage=chkImage12, indicatoron=False, cursor="circle", variable=cdichkType5, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage51Func)
        cdichk51.deselect()
        global cdichk52
        cdichk52 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=cdichkType5, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage52Func)
        cdichk52.deselect()
        global cdichk53
        cdichk53 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=cdichkType5, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage53Func)
        cdichk53.deselect()
        global cdichk61
        cdichk61 = Radiobutton(displayPage2.inner, value=0,  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                               selectimage=chkImage12, indicatoron=False, cursor="circle", variable=cdichkType6, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage61Func)
        cdichk61.deselect()
        global cdichk62
        cdichk62 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=cdichkType6, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage62Func)
        cdichk62.deselect()
        global cdichk63
        cdichk63 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=cdichkType6, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage63Func)
        cdichk63.deselect()
        global cdichk71
        cdichk71 = Radiobutton(displayPage2.inner, value=0,  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                               selectimage=chkImage12, indicatoron=False, cursor="circle", variable=cdichkType7, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage71Func)
        cdichk71.deselect()
        global cdichk72
        cdichk72 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=cdichkType7, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage72Func)
        cdichk72.deselect()
        global cdichk73
        cdichk73 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=cdichkType7, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage73Func)
        cdichk73.deselect()
        global cdichk81
        cdichk81 = Radiobutton(displayPage2.inner, value=0,  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                               selectimage=chkImage12, indicatoron=False, cursor="circle", variable=cdichkType8, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage81Func)
        cdichk81.deselect()
        global cdichk82
        cdichk82 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=cdichkType8, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage82Func)
        cdichk82.deselect()
        global cdichk83
        cdichk83 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=cdichkType8, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage83Func)
        cdichk83.deselect()
        global cdichk91
        cdichk91 = Radiobutton(displayPage2.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                               selectimage=chkImage12, indicatoron=False, cursor="circle", variable=cdichkType9, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage91Func)
        cdichk91.deselect()
        global cdichk92
        cdichk92 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=cdichkType9, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage92Func)
        cdichk92.deselect()
        global cdichk93
        cdichk93 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=cdichkType9, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage93Func)
        cdichk93.deselect()
        global cdichk101
        cdichk101 = Radiobutton(displayPage2.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType10, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage101Func)
        cdichk101.deselect()
        global cdichk102
        cdichk102 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType10, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage102Func)
        cdichk102.deselect()
        global cdichk103
        cdichk103 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType10, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage103Func)
        cdichk103.deselect()
        global cdichk111
        cdichk111 = Radiobutton(displayPage2.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType11, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage111Func)
        cdichk111.deselect()
        global cdichk112
        cdichk112 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType11, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage112Func)
        cdichk112.deselect()
        global cdichk113
        cdichk113 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType11, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage113Func)
        cdichk113.deselect()
        global cdichk121
        cdichk121 = Radiobutton(displayPage2.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType12, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage121Func)
        cdichk121.deselect()
        global cdichk122
        cdichk122 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType12, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage122Func)
        cdichk122.deselect()
        global cdichk123
        cdichk123 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType12, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage123Func)
        cdichk123.deselect()
        global cdichk131
        cdichk131 = Radiobutton(displayPage2.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType13, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage131Func)
        cdichk131.deselect()
        global cdichk132
        cdichk132 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType13, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage132Func)
        cdichk132.deselect()
        global cdichk133
        cdichk133 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType13, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage133Func)
        cdichk133.deselect()
        global cdichk141
        cdichk141 = Radiobutton(displayPage2.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType14, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage141Func)
        cdichk141.deselect()
        global cdichk142
        cdichk142 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType14, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage142Func)
        cdichk142.deselect()
        global cdichk143
        cdichk143 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType14, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage143Func)
        cdichk143.deselect()
        global cdichk151
        cdichk151 = Radiobutton(displayPage2.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType15, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage151Func)
        cdichk151.deselect()
        global cdichk152
        cdichk152 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType15, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage152Func)
        cdichk152.deselect()
        global cdichk153
        cdichk153 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType15, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage153Func)
        cdichk153.deselect()
        global cdichk161
        cdichk161 = Radiobutton(displayPage2.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType16, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage161Func)
        cdichk161.deselect()
        global cdichk162
        cdichk162 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType16, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage162Func)
        cdichk162.deselect()
        global cdichk163
        cdichk163 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType16, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage163Func)
        cdichk163.deselect()
        global cdichk171
        cdichk171 = Radiobutton(displayPage2.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType17, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage171Func)
        cdichk171.deselect()
        global cdichk172
        cdichk172 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType17, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage172Func)
        cdichk172.deselect()
        global cdichk173
        cdichk173 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType17, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage173Func)
        cdichk173.deselect()
        global cdichk181
        cdichk181 = Radiobutton(displayPage2.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType18, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage181Func)
        cdichk181.deselect()
        global cdichk182
        cdichk182 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType18, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage182Func)
        cdichk182.deselect()
        global cdichk183
        cdichk183 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType18, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage183Func)
        cdichk183.deselect()
        global cdichk191
        cdichk191 = Radiobutton(displayPage2.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType19, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage191Func)
        cdichk191.deselect()
        global cdichk192
        cdichk192 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType19, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage192Func)
        cdichk192.deselect()
        global cdichk193
        cdichk193 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType19, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage193Func)
        cdichk193.deselect()
        global cdichk201
        cdichk201 = Radiobutton(displayPage2.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType20, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage201Func)
        cdichk201.deselect()
        global cdichk202
        cdichk202 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType20, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage202Func)
        cdichk202.deselect()
        global cdichk203
        cdichk203 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType20, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage203Func)
        cdichk203.deselect()
        global cdichk211
        cdichk211 = Radiobutton(displayPage2.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType21, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage211Func)
        cdichk211.deselect()
        global cdichk212
        cdichk212 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType21, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage212Func)
        cdichk212.deselect()
        global cdichk213
        cdichk213 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType21, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage213Func)
        cdichk213.deselect()
        global cdichk221
        cdichk221 = Radiobutton(displayPage2.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType22, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage221Func)
        cdichk221.deselect()
        global cdichk222
        cdichk222 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType22, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage222Func)
        cdichk222.deselect()
        global cdichk223
        cdichk223 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType22, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage223Func)
        cdichk223.deselect()
        global cdichk231
        cdichk231 = Radiobutton(displayPage2.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType23, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage231Func)
        cdichk231.deselect()
        global cdichk232
        cdichk232 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType23, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage232Func)
        cdichk232.deselect()
        global cdichk233
        cdichk233 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType23, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage233Func)
        cdichk233.deselect()
        global cdichk241
        cdichk241 = Radiobutton(displayPage2.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType24, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage241Func)
        cdichk241.deselect()
        global cdichk242
        cdichk242 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType24, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage242Func)
        cdichk242.deselect()
        global cdichk243
        cdichk243 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType24, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage243Func)
        cdichk243.deselect()
        global cdichk251
        cdichk251 = Radiobutton(displayPage2.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType25, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage251Func)
        cdichk251.deselect()
        global cdichk252
        cdichk252 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType25, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage252Func)
        cdichk252.deselect()
        global cdichk253
        cdichk253 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType25, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage253Func)
        cdichk253.deselect()
        global cdichk261
        cdichk261 = Radiobutton(displayPage2.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType26, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage261Func)
        cdichk261.deselect()
        global cdichk262
        cdichk262 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType26, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage262Func)
        cdichk262.deselect()
        global cdichk263
        cdichk263 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType26, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage263Func)
        cdichk263.deselect()
        global cdichk271
        cdichk271 = Radiobutton(displayPage2.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType27, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage271Func)
        cdichk271.deselect()
        global cdichk272
        cdichk272 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType27, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage272Func)
        cdichk272.deselect()
        global cdichk273
        cdichk273 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=cdichkType27, font=malgungothic13, bd=0, highlightthickness=0, command=cdichkImage273Func)
        cdichk273.deselect()

        cdichk11.place(x=428, y=264+addheight, width=21, height=21)
        cdichk12.place(x=428, y=303+addheight, width=21, height=21)
        cdichk13.place(x=428, y=343+addheight, width=21, height=21)
        cdichk21.place(x=428, y=380+addheight, width=21, height=21)
        cdichk22.place(x=428, y=419+addheight, width=21, height=21)
        cdichk23.place(x=428, y=459+addheight, width=21, height=21)
        cdichk31.place(x=428, y=496+addheight, width=21, height=21)
        cdichk32.place(x=428, y=535+addheight, width=21, height=21)
        cdichk33.place(x=428, y=575+addheight, width=21, height=21)
        cdichk41.place(x=428, y=614+addheight, width=21, height=21)
        cdichk42.place(x=428, y=651+addheight, width=21, height=21)
        cdichk43.place(x=428, y=691+addheight, width=21, height=21)
        cdichk51.place(x=428, y=730+addheight, width=21, height=21)
        cdichk52.place(x=428, y=767+addheight, width=21, height=21)
        cdichk53.place(x=428, y=807+addheight, width=21, height=21)
        cdichk61.place(x=428, y=846+addheight, width=21, height=21)
        cdichk62.place(x=428, y=883+addheight, width=21, height=21)
        cdichk63.place(x=428, y=923+addheight, width=21, height=21)
        cdichk71.place(x=428, y=965+addheight, width=21, height=21)
        cdichk72.place(x=428, y=1002+addheight, width=21, height=21)
        cdichk73.place(x=428, y=1042+addheight, width=21, height=21)
        cdichk81.place(x=428, y=1079+addheight, width=21, height=21)
        cdichk82.place(x=428, y=1118+addheight, width=21, height=21)
        cdichk83.place(x=428, y=1158+addheight, width=21, height=21)
        cdichk91.place(x=428, y=1198+addheight, width=21, height=21)
        cdichk92.place(x=428, y=1237+addheight, width=21, height=21)
        cdichk93.place(x=428, y=1277+addheight, width=21, height=21)
        cdichk101.place(x=428, y=1315+addheight, width=21, height=21)
        cdichk102.place(x=428, y=1354+addheight, width=21, height=21)
        cdichk103.place(x=428, y=1394+addheight, width=21, height=21)
        cdichk111.place(x=428, y=1432+addheight, width=21, height=21)
        cdichk112.place(x=428, y=1470+addheight, width=21, height=21)
        cdichk113.place(x=428, y=1510+addheight, width=21, height=21)
        cdichk121.place(x=428, y=1548+addheight, width=21, height=21)
        cdichk122.place(x=428, y=1588+addheight, width=21, height=21)
        cdichk123.place(x=428, y=1626+addheight, width=21, height=21)
        cdichk131.place(x=428, y=1664+addheight, width=21, height=21)
        cdichk132.place(x=428, y=1704+addheight, width=21, height=21)
        cdichk133.place(x=428, y=1742+addheight, width=21, height=21)
        cdichk141.place(x=428, y=1783+addheight, width=21, height=21)
        cdichk142.place(x=428, y=1823+addheight, width=21, height=21)
        cdichk143.place(x=428, y=1861+addheight, width=21, height=21)
        cdichk151.place(x=428, y=1899+addheight, width=21, height=21)
        cdichk152.place(x=428, y=1939+addheight, width=21, height=21)
        cdichk153.place(x=428, y=1977+addheight, width=21, height=21)
        cdichk161.place(x=428, y=2017+addheight, width=21, height=21)
        cdichk162.place(x=428, y=2057+addheight, width=21, height=21)
        cdichk163.place(x=428, y=2095+addheight, width=21, height=21)
        cdichk171.place(x=428, y=2133+addheight, width=21, height=21)
        cdichk172.place(x=428, y=2172+addheight, width=21, height=21)
        cdichk173.place(x=428, y=2211+addheight, width=21, height=21)
        cdichk181.place(x=428, y=2251+addheight, width=21, height=21)
        cdichk182.place(x=428, y=2289+addheight, width=21, height=21)
        cdichk183.place(x=428, y=2328+addheight, width=21, height=21)
        cdichk191.place(x=428, y=2367+addheight, width=21, height=21)
        cdichk192.place(x=428, y=2407+addheight, width=21, height=21)
        cdichk193.place(x=428, y=2445+addheight, width=21, height=21)
        cdichk201.place(x=428, y=2484+addheight, width=21, height=21)
        cdichk202.place(x=428, y=2523+addheight, width=21, height=21)
        cdichk203.place(x=428, y=2560+addheight, width=21, height=21)
        cdichk211.place(x=428, y=2600+addheight, width=21, height=21)
        cdichk212.place(x=428, y=2640+addheight, width=21, height=21)
        cdichk213.place(x=428, y=2679+addheight, width=21, height=21)
        cdichk221.place(x=428, y=2718+addheight, width=21, height=21)
        cdichk222.place(x=428, y=2758+addheight, width=21, height=21)
        cdichk223.place(x=428, y=2797+addheight, width=21, height=21)
        cdichk231.place(x=428, y=2835+addheight, width=21, height=21)
        cdichk232.place(x=428, y=2875+addheight, width=21, height=21)
        cdichk233.place(x=428, y=2912+addheight, width=21, height=21)
        cdichk241.place(x=428, y=2953+addheight, width=21, height=21)
        cdichk242.place(x=428, y=2992+addheight, width=21, height=21)
        cdichk243.place(x=428, y=3030+addheight, width=21, height=21)
        cdichk251.place(x=428, y=3070+addheight, width=21, height=21)
        cdichk252.place(x=428, y=3108+addheight, width=21, height=21)
        cdichk253.place(x=428, y=3147+addheight, width=21, height=21)
        cdichk261.place(x=428, y=3186+addheight, width=21, height=21)
        cdichk262.place(x=428, y=3226+addheight, width=21, height=21)
        cdichk263.place(x=428, y=3263+addheight, width=21, height=21)
        cdichk271.place(x=428, y=3304+addheight, width=21, height=21)
        cdichk272.place(x=428, y=3342+addheight, width=21, height=21)
        cdichk273.place(x=428, y=3381+addheight, width=21, height=21)

    def bdichkFunc():
        global now
        now = datetime.now()
        global bdichk
        bdichk = True
        sheet3['E2'].value = (" % s년 % s월 % s일 % s시 % s분" %
                              (now.year, now.month, now.day, now.hour, now.minute))
        sheet3['K2'].value = idInput.get()

        def bdichkImage11Func():
            if not str(sheet3['C9'].value).strip() == 1:
                sheet3['C9'].value = str(int(bdichkType1.get())) + "."

        def bdichkImage12Func():
            if not str(sheet3['C9'].value).strip() == 2:
                sheet3['C9'].value = str(int(bdichkType1.get())) + "."

        def bdichkImage13Func():
            if not str(sheet3['C9'].value).strip() == 3:
                sheet3['C9'].value = str(int(bdichkType1.get())) + "."

        def bdichkImage14Func():
            if not str(sheet3['C9'].value).strip() == 4:
                sheet3['C9'].value = str(int(bdichkType1.get())) + "."

        def bdichkImage21Func():
            if not str(sheet3['C15'].value).strip() == 1:
                sheet3['C15'].value = str(int(bdichkType2.get())) + "."

        def bdichkImage22Func():
            if not str(sheet3['C15'].value).strip() == 2:
                sheet3['C15'].value = str(int(bdichkType2.get())) + "."

        def bdichkImage23Func():
            if not str(sheet3['C15'].value).strip() == 3:
                sheet3['C15'].value = str(int(bdichkType2.get())) + "."

        def bdichkImage24Func():
            if not str(sheet3['C15'].value).strip() == 4:
                sheet3['C15'].value = str(int(bdichkType2.get())) + "."

        def bdichkImage31Func():
            if not str(sheet3['C21'].value).strip() == 1:
                sheet3['C21'].value = str(int(bdichkType3.get())) + "."

        def bdichkImage32Func():
            if not str(sheet3['C21'].value).strip() == 2:
                sheet3['C21'].value = str(int(bdichkType3.get())) + "."

        def bdichkImage33Func():
            if not str(sheet3['C21'].value).strip() == 3:
                sheet3['C21'].value = str(int(bdichkType3.get())) + "."

        def bdichkImage34Func():
            if not str(sheet3['C21'].value).strip() == 4:
                sheet3['C21'].value = str(int(bdichkType3.get())) + "."

        def bdichkImage41Func():
            if not str(sheet3['C27'].value).strip() == 1:
                sheet3['C27'].value = str(int(bdichkType4.get())) + "."

        def bdichkImage42Func():
            if not str(sheet3['C27'].value).strip() == 2:
                sheet3['C27'].value = str(int(bdichkType4.get())) + "."

        def bdichkImage43Func():
            if not str(sheet3['C27'].value).strip() == 3:
                sheet3['C27'].value = str(int(bdichkType4.get())) + "."

        def bdichkImage44Func():
            if not str(sheet3['C27'].value).strip() == 4:
                sheet3['C27'].value = str(int(bdichkType4.get())) + "."

        def bdichkImage51Func():
            if not str(sheet3['C33'].value).strip() == 1:
                sheet3['C33'].value = str(int(bdichkType5.get())) + "."

        def bdichkImage52Func():
            if not str(sheet3['C33'].value).strip() == 2:
                sheet3['C33'].value = str(int(bdichkType5.get())) + "."

        def bdichkImage53Func():
            if not str(sheet3['C33'].value).strip() == 3:
                sheet3['C33'].value = str(int(bdichkType5.get())) + "."

        def bdichkImage54Func():
            if not str(sheet3['C33'].value).strip() == 4:
                sheet3['C33'].value = str(int(bdichkType5.get())) + "."

        def bdichkImage61Func():
            if not str(sheet3['C39'].value).strip() == 1:
                sheet3['C39'].value = str(int(bdichkType6.get())) + "."

        def bdichkImage62Func():
            if not str(sheet3['C39'].value).strip() == 2:
                sheet3['C39'].value = str(int(bdichkType6.get())) + "."

        def bdichkImage63Func():
            if not str(sheet3['C39'].value).strip() == 3:
                sheet3['C39'].value = str(int(bdichkType6.get())) + "."

        def bdichkImage64Func():
            if not str(sheet3['C39'].value).strip() == 4:
                sheet3['C39'].value = str(int(bdichkType6.get())) + "."

        def bdichkImage71Func():
            if not str(sheet3['C45'].value).strip() == 1:
                sheet3['C45'].value = str(int(bdichkType7.get())) + "."

        def bdichkImage72Func():
            if not str(sheet3['C45'].value).strip() == 2:
                sheet3['C45'].value = str(int(bdichkType7.get())) + "."

        def bdichkImage73Func():
            if not str(sheet3['C45'].value).strip() == 3:
                sheet3['C45'].value = str(int(bdichkType7.get())) + "."

        def bdichkImage74Func():
            if not str(sheet3['C45'].value).strip() == 4:
                sheet3['C45'].value = str(int(bdichkType7.get())) + "."

        def bdichkImage81Func():
            if not str(sheet3['C51'].value).strip() == 1:
                sheet3['C51'].value = str(int(bdichkType8.get())) + "."

        def bdichkImage82Func():
            if not str(sheet3['C51'].value).strip() == 2:
                sheet3['C51'].value = str(int(bdichkType8.get())) + "."

        def bdichkImage83Func():
            if not str(sheet3['C51'].value).strip() == 3:
                sheet3['C51'].value = str(int(bdichkType8.get())) + "."

        def bdichkImage84Func():
            if not str(sheet3['C51'].value).strip() == 4:
                sheet3['C51'].value = str(int(bdichkType8.get())) + "."

        def bdichkImage91Func():
            if not str(sheet3['C57'].value).strip() == 1:
                sheet3['C57'].value = str(int(bdichkType9.get())) + "."

        def bdichkImage92Func():
            if not str(sheet3['C57'].value).strip() == 2:
                sheet3['C57'].value = str(int(bdichkType9.get())) + "."

        def bdichkImage93Func():
            if not str(sheet3['C57'].value).strip() == 3:
                sheet3['C57'].value = str(int(bdichkType9.get())) + "."

        def bdichkImage94Func():
            if not str(sheet3['C57'].value).strip() == 4:
                sheet3['C57'].value = str(int(bdichkType9.get())) + "."

        def bdichkImage101Func():
            if not str(sheet3['C63'].value).strip() == 1:
                sheet3['C63'].value = str(int(bdichkType10.get())) + "."

        def bdichkImage102Func():
            if not str(sheet3['C63'].value).strip() == 2:
                sheet3['C63'].value = str(int(bdichkType10.get())) + "."

        def bdichkImage103Func():
            if not str(sheet3['C63'].value).strip() == 3:
                sheet3['C63'].value = str(int(bdichkType10.get())) + "."

        def bdichkImage104Func():
            if not str(sheet3['C63'].value).strip() == 4:
                sheet3['C63'].value = str(int(bdichkType10.get())) + "."

        def bdichkImage111Func():
            if not str(sheet3['C69'].value).strip() == 1:
                sheet3['C69'].value = str(int(bdichkType11.get())) + "."

        def bdichkImage112Func():
            if not str(sheet3['C69'].value).strip() == 2:
                sheet3['C69'].value = str(int(bdichkType11.get())) + "."

        def bdichkImage113Func():
            if not str(sheet3['C69'].value).strip() == 3:
                sheet3['C69'].value = str(int(bdichkType11.get())) + "."

        def bdichkImage114Func():
            if not str(sheet3['C69'].value).strip() == 4:
                sheet3['C69'].value = str(int(bdichkType11.get())) + "."

        def bdichkImage121Func():
            if not str(sheet3['C75'].value).strip() == 1:
                sheet3['C75'].value = str(int(bdichkType12.get())) + "."

        def bdichkImage122Func():
            if not str(sheet3['C75'].value).strip() == 2:
                sheet3['C75'].value = str(int(bdichkType12.get())) + "."

        def bdichkImage123Func():
            if not str(sheet3['C75'].value).strip() == 3:
                sheet3['C75'].value = str(int(bdichkType12.get())) + "."

        def bdichkImage124Func():
            if not str(sheet3['C75'].value).strip() == 4:
                sheet3['C75'].value = str(int(bdichkType12.get())) + "."

        def bdichkImage131Func():
            if not str(sheet3['C81'].value).strip() == 1:
                sheet3['C81'].value = str(int(bdichkType13.get())) + "."

        def bdichkImage132Func():
            if not str(sheet3['C81'].value).strip() == 2:
                sheet3['C81'].value = str(int(bdichkType13.get())) + "."

        def bdichkImage133Func():
            if not str(sheet3['C81'].value).strip() == 3:
                sheet3['C81'].value = str(int(bdichkType13.get())) + "."

        def bdichkImage134Func():
            if not str(sheet3['C81'].value).strip() == 4:
                sheet3['C81'].value = str(int(bdichkType13.get())) + "."

        def bdichkImage141Func():
            if not str(sheet3['C87'].value).strip() == 1:
                sheet3['C87'].value = str(int(bdichkType14.get())) + "."

        def bdichkImage142Func():
            if not str(sheet3['C87'].value).strip() == 2:
                sheet3['C87'].value = str(int(bdichkType14.get())) + "."

        def bdichkImage143Func():
            if not str(sheet3['C87'].value).strip() == 3:
                sheet3['C87'].value = str(int(bdichkType14.get())) + "."

        def bdichkImage144Func():
            if not str(sheet3['C87'].value).strip() == 4:
                sheet3['C87'].value = str(int(bdichkType14.get())) + "."

        def bdichkImage151Func():
            if not str(sheet3['C93'].value).strip() == 1:
                sheet3['C93'].value = str(int(bdichkType15.get())) + "."

        def bdichkImage152Func():
            if not str(sheet3['C93'].value).strip() == 2:
                sheet3['C93'].value = str(int(bdichkType15.get())) + "."

        def bdichkImage153Func():
            if not str(sheet3['C93'].value).strip() == 3:
                sheet3['C93'].value = str(int(bdichkType15.get())) + "."

        def bdichkImage154Func():
            if not str(sheet3['C93'].value).strip() == 4:
                sheet3['C93'].value = str(int(bdichkType15.get())) + "."

        def bdichkImage161Func():
            if not str(sheet3['C99'].value).strip() == 1:
                sheet3['C99'].value = str(int(bdichkType16.get())) + "."

        def bdichkImage162Func():
            if not str(sheet3['C99'].value).strip() == 2:
                sheet3['C99'].value = str(int(bdichkType16.get())) + "."

        def bdichkImage163Func():
            if not str(sheet3['C99'].value).strip() == 3:
                sheet3['C99'].value = str(int(bdichkType16.get())) + "."

        def bdichkImage164Func():
            if not str(sheet3['C99'].value).strip() == 4:
                sheet3['C99'].value = str(int(bdichkType16.get())) + "."

        def bdichkImage171Func():
            if not str(sheet3['C105'].value).strip() == 1:
                sheet3['C105'].value = str(int(bdichkType17.get())) + "."

        def bdichkImage172Func():
            if not str(sheet3['C105'].value).strip() == 2:
                sheet3['C105'].value = str(int(bdichkType17.get())) + "."

        def bdichkImage173Func():
            if not str(sheet3['C105'].value).strip() == 3:
                sheet3['C105'].value = str(int(bdichkType17.get())) + "."

        def bdichkImage174Func():
            if not str(sheet3['C105'].value).strip() == 4:
                sheet3['C105'].value = str(int(bdichkType17.get())) + "."

        def bdichkImage181Func():
            if not str(sheet3['C111'].value).strip() == 1:
                sheet3['C111'].value = str(int(bdichkType18.get())) + "."

        def bdichkImage182Func():
            if not str(sheet3['C111'].value).strip() == 2:
                sheet3['C111'].value = str(int(bdichkType18.get())) + "."

        def bdichkImage183Func():
            if not str(sheet3['C111'].value).strip() == 3:
                sheet3['C111'].value = str(int(bdichkType18.get())) + "."

        def bdichkImage184Func():
            if not str(sheet3['C111'].value).strip() == 4:
                sheet3['C111'].value = str(int(bdichkType18.get())) + "."

        def bdichkImage191Func():
            if not str(sheet3['C118'].value).strip() == 1:
                sheet3['C118'].value = str(int(bdichkType19.get())) + "."

        def bdichkImage192Func():
            if not str(sheet3['C118'].value).strip() == 2:
                sheet3['C118'].value = str(int(bdichkType19.get())) + "."

        def bdichkImage193Func():
            if not str(sheet3['C118'].value).strip() == 3:
                sheet3['C118'].value = str(int(bdichkType19.get())) + "."

        def bdichkImage194Func():
            if not str(sheet3['C118'].value).strip() == 4:
                sheet3['C118'].value = str(int(bdichkType19.get())) + "."

        def bdichkYesFunc():
            if not str(sheet3['G118'].value).strip() == '예':
                sheet3['G118'].value = str(bdichkTypeYesNo.get())

        def bdichkNoFunc():
            if not str(sheet3['G118'].value).strip() == "아니오":
                sheet3['G118'].value = str(bdichkTypeYesNo.get())

        def bdichkImage201Func():
            if not str(sheet3['C124'].value).strip() == 1:
                sheet3['C124'].value = str(int(bdichkType20.get())) + "."

        def bdichkImage202Func():
            if not str(sheet3['C124'].value).strip() == 2:
                sheet3['C124'].value = str(int(bdichkType20.get())) + "."

        def bdichkImage203Func():
            if not str(sheet3['C124'].value).strip() == 3:
                sheet3['C124'].value = str(int(bdichkType20.get())) + "."

        def bdichkImage204Func():
            if not str(sheet3['C124'].value).strip() == 4:
                sheet3['C124'].value = str(int(bdichkType20.get())) + "."

        def bdichkImage211Func():
            if not str(sheet3['C130'].value).strip() == 1:
                sheet3['C130'].value = str(int(bdichkType21.get())) + "."

        def bdichkImage212Func():
            if not str(sheet3['C130'].value).strip() == 2:
                sheet3['C130'].value = str(int(bdichkType21.get())) + "."

        def bdichkImage213Func():
            if not str(sheet3['C130'].value).strip() == 3:
                sheet3['C130'].value = str(int(bdichkType21.get())) + "."

        def bdichkImage214Func():
            if not str(sheet3['C130'].value).strip() == 4:
                sheet3['C130'].value = str(int(bdichkType21.get())) + "."

        frame3BdiBgImg = Image.open("images/bdibg.png")
        frame3BdiBg = ImageTk.PhotoImage(frame3BdiBgImg)
        frame3BdiBgLabel = tkinter.Label(displayPage.inner, image=frame3BdiBg)
        frame3BdiBgLabel.image = frame3BdiBg
        frame3BdiBgLabel.place(x=47, y=420, height=3653)
        frame3BdiBgLabel.pack()

        global bdichk11
        bdichk11 = Radiobutton(displayPage.inner, value=1,  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                               selectimage=chkImage12, indicatoron=False, cursor="circle", variable=bdichkType1, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage11Func)
        bdichk11.deselect()
        global bdichk12
        bdichk12 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType1, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage12Func)
        bdichk12.deselect()
        global bdichk13
        bdichk13 = Radiobutton(displayPage.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType1, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage13Func)
        bdichk13.deselect()
        global bdichk14
        bdichk14 = Radiobutton(displayPage.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType1, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage14Func)
        bdichk14.deselect()
        global bdichk21
        bdichk21 = Radiobutton(displayPage.inner, value=1,  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                               selectimage=chkImage12, indicatoron=False, cursor="circle", variable=bdichkType2, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage21Func)
        bdichk21.deselect()
        global bdichk22
        bdichk22 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType2, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage22Func)
        bdichk22.deselect()
        global bdichk23
        bdichk23 = Radiobutton(displayPage.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType2, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage23Func)
        bdichk23.deselect()
        global bdichk24
        bdichk24 = Radiobutton(displayPage.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType2, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage24Func)
        bdichk24.deselect()
        global bdichk31
        bdichk31 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                               selectimage=chkImage12, indicatoron=False, cursor="circle", variable=bdichkType3, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage31Func)
        bdichk31.deselect()
        global bdichk32
        bdichk32 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType3, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage32Func)
        bdichk32.deselect()
        global bdichk33
        bdichk33 = Radiobutton(displayPage.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType3, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage33Func)
        bdichk33.deselect()
        global bdichk34
        bdichk34 = Radiobutton(displayPage.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType3, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage34Func)
        bdichk34.deselect()
        global bdichk41
        bdichk41 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                               selectimage=chkImage12, indicatoron=False, cursor="circle", variable=bdichkType4, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage41Func)
        bdichk41.deselect()
        global bdichk42
        bdichk42 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType4, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage42Func)
        bdichk42.deselect()
        global bdichk43
        bdichk43 = Radiobutton(displayPage.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType4, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage43Func)
        bdichk43.deselect()
        global bdichk44
        bdichk44 = Radiobutton(displayPage.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType4, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage44Func)
        bdichk44.deselect()
        global bdichk51
        bdichk51 = Radiobutton(displayPage.inner, value=1,  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                               selectimage=chkImage12, indicatoron=False, cursor="circle", variable=bdichkType5, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage51Func)
        bdichk51.deselect()
        global bdichk52
        bdichk52 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType5, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage52Func)
        bdichk52.deselect()
        global bdichk53
        bdichk53 = Radiobutton(displayPage.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType5, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage53Func)
        bdichk53.deselect()
        global bdichk54
        bdichk54 = Radiobutton(displayPage.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType5, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage54Func)
        bdichk54.deselect()
        global bdichk61
        bdichk61 = Radiobutton(displayPage.inner, value=1,  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                               selectimage=chkImage12, indicatoron=False, cursor="circle", variable=bdichkType6, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage61Func)
        bdichk61.deselect()
        global bdichk62
        bdichk62 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType6, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage62Func)
        bdichk62.deselect()
        global bdichk63
        bdichk63 = Radiobutton(displayPage.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType6, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage63Func)
        bdichk63.deselect()
        global bdichk64
        bdichk64 = Radiobutton(displayPage.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType6, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage64Func)
        bdichk64.deselect()
        global bdichk71
        bdichk71 = Radiobutton(displayPage.inner, value=1,  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                               selectimage=chkImage12, indicatoron=False, cursor="circle", variable=bdichkType7, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage71Func)
        bdichk71.deselect()
        global bdichk72
        bdichk72 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType7, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage72Func)
        bdichk72.deselect()
        global bdichk73
        bdichk73 = Radiobutton(displayPage.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType7, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage73Func)
        bdichk73.deselect()
        global bdichk74
        bdichk74 = Radiobutton(displayPage.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType7, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage74Func)
        bdichk74.deselect()
        global bdichk81
        bdichk81 = Radiobutton(displayPage.inner, value=1,  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                               selectimage=chkImage12, indicatoron=False, cursor="circle", variable=bdichkType8, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage81Func)
        bdichk81.deselect()
        global bdichk82
        bdichk82 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType8, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage82Func)
        bdichk82.deselect()
        global bdichk83
        bdichk83 = Radiobutton(displayPage.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType8, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage83Func)
        bdichk83.deselect()
        global bdichk84
        bdichk84 = Radiobutton(displayPage.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType8, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage84Func)
        bdichk84.deselect()
        global bdichk91
        bdichk91 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                               selectimage=chkImage12, indicatoron=False, cursor="circle", variable=bdichkType9, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage91Func)
        bdichk91.deselect()
        global bdichk92
        bdichk92 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType9, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage92Func)
        bdichk92.deselect()
        global bdichk93
        bdichk93 = Radiobutton(displayPage.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType9, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage93Func)
        bdichk93.deselect()
        global bdichk94
        bdichk94 = Radiobutton(displayPage.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType9, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage94Func)
        bdichk94.deselect()
        global bdichk101
        bdichk101 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType10, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage101Func)
        bdichk101.deselect()
        global bdichk102
        bdichk102 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType10, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage102Func)
        bdichk102.deselect()
        global bdichk103
        bdichk103 = Radiobutton(displayPage.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType10, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage103Func)
        bdichk103.deselect()
        global bdichk104
        bdichk104 = Radiobutton(displayPage.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType10, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage104Func)
        bdichk104.deselect()
        global bdichk111
        bdichk111 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType11, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage111Func)
        bdichk111.deselect()
        global bdichk112
        bdichk112 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType11, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage112Func)
        bdichk112.deselect()
        global bdichk113
        bdichk113 = Radiobutton(displayPage.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType11, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage113Func)
        bdichk113.deselect()
        global bdichk114
        bdichk114 = Radiobutton(displayPage.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType11, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage114Func)
        bdichk114.deselect()
        global bdichk121
        bdichk121 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType12, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage121Func)
        bdichk121.deselect()
        global bdichk122
        bdichk122 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType12, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage122Func)
        bdichk122.deselect()
        global bdichk123
        bdichk123 = Radiobutton(displayPage.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType12, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage123Func)
        bdichk123.deselect()
        global bdichk124
        bdichk124 = Radiobutton(displayPage.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType12, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage124Func)
        bdichk124.deselect()
        global bdichk131
        bdichk131 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType13, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage131Func)
        bdichk131.deselect()
        global bdichk132
        bdichk132 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType13, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage132Func)
        bdichk132.deselect()
        global bdichk133
        bdichk133 = Radiobutton(displayPage.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType13, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage133Func)
        bdichk133.deselect()
        global bdichk134
        bdichk134 = Radiobutton(displayPage.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType13, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage134Func)
        bdichk134.deselect()
        global bdichk141
        bdichk141 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType14, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage141Func)
        bdichk141.deselect()
        global bdichk142
        bdichk142 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType14, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage142Func)
        bdichk142.deselect()
        global bdichk143
        bdichk143 = Radiobutton(displayPage.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType14, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage143Func)
        bdichk143.deselect()
        global bdichk144
        bdichk144 = Radiobutton(displayPage.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType14, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage144Func)
        bdichk144.deselect()
        global bdichk151
        bdichk151 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType15, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage151Func)
        bdichk151.deselect()
        global bdichk152
        bdichk152 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType15, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage152Func)
        bdichk152.deselect()
        global bdichk153
        bdichk153 = Radiobutton(displayPage.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType15, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage153Func)
        bdichk153.deselect()
        global bdichk154
        bdichk154 = Radiobutton(displayPage.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType15, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage154Func)
        bdichk154.deselect()
        global bdichk161
        bdichk161 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType16, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage161Func)
        bdichk161.deselect()
        global bdichk162
        bdichk162 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType16, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage162Func)
        bdichk162.deselect()
        global bdichk163
        bdichk163 = Radiobutton(displayPage.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType16, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage163Func)
        bdichk163.deselect()
        global bdichk164
        bdichk164 = Radiobutton(displayPage.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType16, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage164Func)
        bdichk164.deselect()
        global bdichk171
        bdichk171 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType17, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage171Func)
        bdichk171.deselect()
        global bdichk172
        bdichk172 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType17, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage172Func)
        bdichk172.deselect()
        global bdichk173
        bdichk173 = Radiobutton(displayPage.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType17, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage173Func)
        bdichk173.deselect()
        global bdichk174
        bdichk174 = Radiobutton(displayPage.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType17, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage174Func)
        bdichk174.deselect()
        global bdichk181
        bdichk181 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType18, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage181Func)
        bdichk181.deselect()
        global bdichk182
        bdichk182 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType18, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage182Func)
        bdichk182.deselect()
        global bdichk183
        bdichk183 = Radiobutton(displayPage.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType18, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage183Func)
        bdichk183.deselect()
        global bdichk184
        bdichk184 = Radiobutton(displayPage.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType18, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage184Func)
        bdichk184.deselect()
        global bdichk191
        bdichk191 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType19, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage191Func)
        bdichk191.deselect()
        global bdichk192
        bdichk192 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType19, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage192Func)
        bdichk192.deselect()
        global bdichk19yes
        bdichk19yes = Radiobutton(displayPage.inner, value='예', background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                  indicatoron=False, cursor="circle", variable=bdichkTypeYesNo, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkYesFunc)
        bdichk19yes.deselect()
        global bdichk19no
        bdichk19no = Radiobutton(displayPage.inner, value='아니오', background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=bdichkTypeYesNo, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkNoFunc)
        bdichk19no.deselect()
        global bdichk193
        bdichk193 = Radiobutton(displayPage.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType19, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage193Func)
        bdichk193.deselect()
        global bdichk194
        bdichk194 = Radiobutton(displayPage.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType19, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage194Func)
        bdichk194.deselect()
        global bdichk201
        bdichk201 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType20, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage201Func)
        bdichk201.deselect()
        global bdichk202
        bdichk202 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType20, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage202Func)
        bdichk202.deselect()
        global bdichk203
        bdichk203 = Radiobutton(displayPage.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType20, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage203Func)
        bdichk203.deselect()
        global bdichk204
        bdichk204 = Radiobutton(displayPage.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType20, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage204Func)
        bdichk204.deselect()
        global bdichk211
        bdichk211 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType21, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage211Func)
        bdichk211.deselect()
        global bdichk212
        bdichk212 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType21, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage212Func)
        bdichk212.deselect()
        global bdichk213
        bdichk213 = Radiobutton(displayPage.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType21, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage213Func)
        bdichk213.deselect()
        global bdichk214
        bdichk214 = Radiobutton(displayPage.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType21, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage214Func)
        bdichk214.deselect()

        bdichk11.place(x=160, y=262, width=21, height=21)
        bdichk12.place(x=160, y=292, width=21, height=21)
        bdichk13.place(x=160, y=321, width=21, height=21)
        bdichk14.place(x=160, y=351, width=21, height=21)
        bdichk21.place(x=160, y=411, width=21, height=21)
        bdichk22.place(x=160, y=441, width=21, height=21)
        bdichk23.place(x=160, y=471, width=21, height=21)
        bdichk24.place(x=160, y=501, width=21, height=21)
        bdichk31.place(x=160, y=561, width=21, height=21)
        bdichk32.place(x=160, y=591, width=21, height=21)
        bdichk33.place(x=160, y=621, width=21, height=21)
        bdichk34.place(x=160, y=651, width=21, height=21)
        bdichk41.place(x=160, y=710, width=21, height=21)
        bdichk42.place(x=160, y=739, width=21, height=21)
        bdichk43.place(x=160, y=769, width=21, height=21)
        bdichk44.place(x=160, y=799, width=21, height=21)
        bdichk51.place(x=160, y=858, width=21, height=21)
        bdichk52.place(x=160, y=889, width=21, height=21)
        bdichk53.place(x=160, y=919, width=21, height=21)
        bdichk54.place(x=160, y=948, width=21, height=21)
        bdichk61.place(x=160, y=1008, width=21, height=21)
        bdichk62.place(x=160, y=1038, width=21, height=21)
        bdichk63.place(x=160, y=1068, width=21, height=21)
        bdichk64.place(x=160, y=1097, width=21, height=21)
        bdichk71.place(x=160, y=1156, width=21, height=21)
        bdichk72.place(x=160, y=1186, width=21, height=21)
        bdichk73.place(x=160, y=1215, width=21, height=21)
        bdichk74.place(x=160, y=1245, width=21, height=21)
        bdichk81.place(x=160, y=1305, width=21, height=21)
        bdichk82.place(x=160, y=1335, width=21, height=21)
        bdichk83.place(x=160, y=1365, width=21, height=21)
        bdichk84.place(x=160, y=1395, width=21, height=21)
        bdichk91.place(x=160, y=1460, width=21, height=21)
        bdichk92.place(x=160, y=1490, width=21, height=21)
        bdichk93.place(x=160, y=1520, width=21, height=21)
        bdichk94.place(x=160, y=1550, width=21, height=21)
        bdichk101.place(x=160, y=1609, width=21, height=21)
        bdichk102.place(x=160, y=1639, width=21, height=21)
        bdichk103.place(x=160, y=1669, width=21, height=21)
        bdichk104.place(x=160, y=1699, width=21, height=21)
        bdichk111.place(x=160, y=1758, width=21, height=21)
        bdichk112.place(x=160, y=1788, width=21, height=21)
        bdichk113.place(x=160, y=1818, width=21, height=21)
        bdichk114.place(x=160, y=1848, width=21, height=21)
        bdichk121.place(x=160, y=1906, width=21, height=21)
        bdichk122.place(x=160, y=1936, width=21, height=21)
        bdichk123.place(x=160, y=1966, width=21, height=21)
        bdichk124.place(x=160, y=1996, width=21, height=21)
        bdichk131.place(x=160, y=2055, width=21, height=21)
        bdichk132.place(x=160, y=2085, width=21, height=21)
        bdichk133.place(x=160, y=2115, width=21, height=21)
        bdichk134.place(x=160, y=2145, width=21, height=21)
        bdichk141.place(x=160, y=2204, width=21, height=21)
        bdichk142.place(x=160, y=2234, width=21, height=21)
        bdichk143.place(x=160, y=2264, width=21, height=21)
        bdichk144.place(x=160, y=2294, width=21, height=21)
        bdichk151.place(x=160, y=2352, width=21, height=21)
        bdichk152.place(x=160, y=2382, width=21, height=21)
        bdichk153.place(x=160, y=2412, width=21, height=21)
        bdichk154.place(x=160, y=2442, width=21, height=21)
        bdichk161.place(x=160, y=2507, width=21, height=21)
        bdichk162.place(x=160, y=2537, width=21, height=21)
        bdichk163.place(x=160, y=2567, width=21, height=21)
        bdichk164.place(x=160, y=2597, width=21, height=21)
        bdichk171.place(x=160, y=2655, width=21, height=21)
        bdichk172.place(x=160, y=2685, width=21, height=21)
        bdichk173.place(x=160, y=2715, width=21, height=21)
        bdichk174.place(x=160, y=2745, width=21, height=21)
        bdichk181.place(x=160, y=2805, width=21, height=21)
        bdichk182.place(x=160, y=2835, width=21, height=21)
        bdichk183.place(x=160, y=2865, width=21, height=21)
        bdichk184.place(x=160, y=2895, width=21, height=21)
        bdichk191.place(x=160, y=2954, width=21, height=21)
        bdichk192.place(x=160, y=2984, width=21, height=21)
        bdichk193.place(x=160, y=3014, width=21, height=21)
        bdichk194.place(x=160, y=3045, width=21, height=21)
        bdichk19yes.place(x=730, y=3092, width=21, height=21)
        bdichk19no.place(x=838, y=3092, width=21, height=21)
        bdichk201.place(x=160, y=3162, width=21, height=21)
        bdichk202.place(x=160, y=3192, width=21, height=21)
        bdichk203.place(x=160, y=3222, width=21, height=21)
        bdichk204.place(x=160, y=3251, width=21, height=21)
        bdichk211.place(x=160, y=3311, width=21, height=21)
        bdichk212.place(x=160, y=3340, width=21, height=21)
        bdichk213.place(x=160, y=3370, width=21, height=21)
        bdichk214.place(x=160, y=3400, width=21, height=21)

    def bdichkFunc2():
        global now
        now = datetime.now()
        global bdichk
        bdichk = True
        sheet3['E2'].value = (" % s년 % s월 % s일 % s시 % s분" %
                              (now.year, now.month, now.day, now.hour, now.minute))
        sheet3['K2'].value = idInput.get()

        def bdichkImage11Func():
            if not str(sheet3['C9'].value).strip() == 1:
                sheet3['C9'].value = str(int(bdichkType1.get())) + "."

        def bdichkImage12Func():
            if not str(sheet3['C9'].value).strip() == 2:
                sheet3['C9'].value = str(int(bdichkType1.get())) + "."

        def bdichkImage13Func():
            if not str(sheet3['C9'].value).strip() == 3:
                sheet3['C9'].value = str(int(bdichkType1.get())) + "."

        def bdichkImage14Func():
            if not str(sheet3['C9'].value).strip() == 4:
                sheet3['C9'].value = str(int(bdichkType1.get())) + "."

        def bdichkImage21Func():
            if not str(sheet3['C15'].value).strip() == 1:
                sheet3['C15'].value = str(int(bdichkType2.get())) + "."

        def bdichkImage22Func():
            if not str(sheet3['C15'].value).strip() == 2:
                sheet3['C15'].value = str(int(bdichkType2.get())) + "."

        def bdichkImage23Func():
            if not str(sheet3['C15'].value).strip() == 3:
                sheet3['C15'].value = str(int(bdichkType2.get())) + "."

        def bdichkImage24Func():
            if not str(sheet3['C15'].value).strip() == 4:
                sheet3['C15'].value = str(int(bdichkType2.get())) + "."

        def bdichkImage31Func():
            if not str(sheet3['C21'].value).strip() == 1:
                sheet3['C21'].value = str(int(bdichkType3.get())) + "."

        def bdichkImage32Func():
            if not str(sheet3['C21'].value).strip() == 2:
                sheet3['C21'].value = str(int(bdichkType3.get())) + "."

        def bdichkImage33Func():
            if not str(sheet3['C21'].value).strip() == 3:
                sheet3['C21'].value = str(int(bdichkType3.get())) + "."

        def bdichkImage34Func():
            if not str(sheet3['C21'].value).strip() == 4:
                sheet3['C21'].value = str(int(bdichkType3.get())) + "."

        def bdichkImage41Func():
            if not str(sheet3['C27'].value).strip() == 1:
                sheet3['C27'].value = str(int(bdichkType4.get())) + "."

        def bdichkImage42Func():
            if not str(sheet3['C27'].value).strip() == 2:
                sheet3['C27'].value = str(int(bdichkType4.get())) + "."

        def bdichkImage43Func():
            if not str(sheet3['C27'].value).strip() == 3:
                sheet3['C27'].value = str(int(bdichkType4.get())) + "."

        def bdichkImage44Func():
            if not str(sheet3['C27'].value).strip() == 4:
                sheet3['C27'].value = str(int(bdichkType4.get())) + "."

        def bdichkImage51Func():
            if not str(sheet3['C33'].value).strip() == 1:
                sheet3['C33'].value = str(int(bdichkType5.get())) + "."

        def bdichkImage52Func():
            if not str(sheet3['C33'].value).strip() == 2:
                sheet3['C33'].value = str(int(bdichkType5.get())) + "."

        def bdichkImage53Func():
            if not str(sheet3['C33'].value).strip() == 3:
                sheet3['C33'].value = str(int(bdichkType5.get())) + "."

        def bdichkImage54Func():
            if not str(sheet3['C33'].value).strip() == 4:
                sheet3['C33'].value = str(int(bdichkType5.get())) + "."

        def bdichkImage61Func():
            if not str(sheet3['C39'].value).strip() == 1:
                sheet3['C39'].value = str(int(bdichkType6.get())) + "."

        def bdichkImage62Func():
            if not str(sheet3['C39'].value).strip() == 2:
                sheet3['C39'].value = str(int(bdichkType6.get())) + "."

        def bdichkImage63Func():
            if not str(sheet3['C39'].value).strip() == 3:
                sheet3['C39'].value = str(int(bdichkType6.get())) + "."

        def bdichkImage64Func():
            if not str(sheet3['C39'].value).strip() == 4:
                sheet3['C39'].value = str(int(bdichkType6.get())) + "."

        def bdichkImage71Func():
            if not str(sheet3['C45'].value).strip() == 1:
                sheet3['C45'].value = str(int(bdichkType7.get())) + "."

        def bdichkImage72Func():
            if not str(sheet3['C45'].value).strip() == 2:
                sheet3['C45'].value = str(int(bdichkType7.get())) + "."

        def bdichkImage73Func():
            if not str(sheet3['C45'].value).strip() == 3:
                sheet3['C45'].value = str(int(bdichkType7.get())) + "."

        def bdichkImage74Func():
            if not str(sheet3['C45'].value).strip() == 4:
                sheet3['C45'].value = str(int(bdichkType7.get())) + "."

        def bdichkImage81Func():
            if not str(sheet3['C51'].value).strip() == 1:
                sheet3['C51'].value = str(int(bdichkType8.get())) + "."

        def bdichkImage82Func():
            if not str(sheet3['C51'].value).strip() == 2:
                sheet3['C51'].value = str(int(bdichkType8.get())) + "."

        def bdichkImage83Func():
            if not str(sheet3['C51'].value).strip() == 3:
                sheet3['C51'].value = str(int(bdichkType8.get())) + "."

        def bdichkImage84Func():
            if not str(sheet3['C51'].value).strip() == 4:
                sheet3['C51'].value = str(int(bdichkType8.get())) + "."

        def bdichkImage91Func():
            if not str(sheet3['C57'].value).strip() == 1:
                sheet3['C57'].value = str(int(bdichkType9.get())) + "."

        def bdichkImage92Func():
            if not str(sheet3['C57'].value).strip() == 2:
                sheet3['C57'].value = str(int(bdichkType9.get())) + "."

        def bdichkImage93Func():
            if not str(sheet3['C57'].value).strip() == 3:
                sheet3['C57'].value = str(int(bdichkType9.get())) + "."

        def bdichkImage94Func():
            if not str(sheet3['C57'].value).strip() == 4:
                sheet3['C57'].value = str(int(bdichkType9.get())) + "."

        def bdichkImage101Func():
            if not str(sheet3['C63'].value).strip() == 1:
                sheet3['C63'].value = str(int(bdichkType10.get())) + "."

        def bdichkImage102Func():
            if not str(sheet3['C63'].value).strip() == 2:
                sheet3['C63'].value = str(int(bdichkType10.get())) + "."

        def bdichkImage103Func():
            if not str(sheet3['C63'].value).strip() == 3:
                sheet3['C63'].value = str(int(bdichkType10.get())) + "."

        def bdichkImage104Func():
            if not str(sheet3['C63'].value).strip() == 4:
                sheet3['C63'].value = str(int(bdichkType10.get())) + "."

        def bdichkImage111Func():
            if not str(sheet3['C69'].value).strip() == 1:
                sheet3['C69'].value = str(int(bdichkType11.get())) + "."

        def bdichkImage112Func():
            if not str(sheet3['C69'].value).strip() == 2:
                sheet3['C69'].value = str(int(bdichkType11.get())) + "."

        def bdichkImage113Func():
            if not str(sheet3['C69'].value).strip() == 3:
                sheet3['C69'].value = str(int(bdichkType11.get())) + "."

        def bdichkImage114Func():
            if not str(sheet3['C69'].value).strip() == 4:
                sheet3['C69'].value = str(int(bdichkType11.get())) + "."

        def bdichkImage121Func():
            if not str(sheet3['C75'].value).strip() == 1:
                sheet3['C75'].value = str(int(bdichkType12.get())) + "."

        def bdichkImage122Func():
            if not str(sheet3['C75'].value).strip() == 2:
                sheet3['C75'].value = str(int(bdichkType12.get())) + "."

        def bdichkImage123Func():
            if not str(sheet3['C75'].value).strip() == 3:
                sheet3['C75'].value = str(int(bdichkType12.get())) + "."

        def bdichkImage124Func():
            if not str(sheet3['C75'].value).strip() == 4:
                sheet3['C75'].value = str(int(bdichkType12.get())) + "."

        def bdichkImage131Func():
            if not str(sheet3['C81'].value).strip() == 1:
                sheet3['C81'].value = str(int(bdichkType13.get())) + "."

        def bdichkImage132Func():
            if not str(sheet3['C81'].value).strip() == 2:
                sheet3['C81'].value = str(int(bdichkType13.get())) + "."

        def bdichkImage133Func():
            if not str(sheet3['C81'].value).strip() == 3:
                sheet3['C81'].value = str(int(bdichkType13.get())) + "."

        def bdichkImage134Func():
            if not str(sheet3['C81'].value).strip() == 4:
                sheet3['C81'].value = str(int(bdichkType13.get())) + "."

        def bdichkImage141Func():
            if not str(sheet3['C87'].value).strip() == 1:
                sheet3['C87'].value = str(int(bdichkType14.get())) + "."

        def bdichkImage142Func():
            if not str(sheet3['C87'].value).strip() == 2:
                sheet3['C87'].value = str(int(bdichkType14.get())) + "."

        def bdichkImage143Func():
            if not str(sheet3['C87'].value).strip() == 3:
                sheet3['C87'].value = str(int(bdichkType14.get())) + "."

        def bdichkImage144Func():
            if not str(sheet3['C87'].value).strip() == 4:
                sheet3['C87'].value = str(int(bdichkType14.get())) + "."

        def bdichkImage151Func():
            if not str(sheet3['C93'].value).strip() == 1:
                sheet3['C93'].value = str(int(bdichkType15.get())) + "."

        def bdichkImage152Func():
            if not str(sheet3['C93'].value).strip() == 2:
                sheet3['C93'].value = str(int(bdichkType15.get())) + "."

        def bdichkImage153Func():
            if not str(sheet3['C93'].value).strip() == 3:
                sheet3['C93'].value = str(int(bdichkType15.get())) + "."

        def bdichkImage154Func():
            if not str(sheet3['C93'].value).strip() == 4:
                sheet3['C93'].value = str(int(bdichkType15.get())) + "."

        def bdichkImage161Func():
            if not str(sheet3['C99'].value).strip() == 1:
                sheet3['C99'].value = str(int(bdichkType16.get())) + "."

        def bdichkImage162Func():
            if not str(sheet3['C99'].value).strip() == 2:
                sheet3['C99'].value = str(int(bdichkType16.get())) + "."

        def bdichkImage163Func():
            if not str(sheet3['C99'].value).strip() == 3:
                sheet3['C99'].value = str(int(bdichkType16.get())) + "."

        def bdichkImage164Func():
            if not str(sheet3['C99'].value).strip() == 4:
                sheet3['C99'].value = str(int(bdichkType16.get())) + "."

        def bdichkImage171Func():
            if not str(sheet3['C105'].value).strip() == 1:
                sheet3['C105'].value = str(int(bdichkType17.get())) + "."

        def bdichkImage172Func():
            if not str(sheet3['C105'].value).strip() == 2:
                sheet3['C105'].value = str(int(bdichkType17.get())) + "."

        def bdichkImage173Func():
            if not str(sheet3['C105'].value).strip() == 3:
                sheet3['C105'].value = str(int(bdichkType17.get())) + "."

        def bdichkImage174Func():
            if not str(sheet3['C105'].value).strip() == 4:
                sheet3['C105'].value = str(int(bdichkType17.get())) + "."

        def bdichkImage181Func():
            if not str(sheet3['C111'].value).strip() == 1:
                sheet3['C111'].value = str(int(bdichkType18.get())) + "."

        def bdichkImage182Func():
            if not str(sheet3['C111'].value).strip() == 2:
                sheet3['C111'].value = str(int(bdichkType18.get())) + "."

        def bdichkImage183Func():
            if not str(sheet3['C111'].value).strip() == 3:
                sheet3['C111'].value = str(int(bdichkType18.get())) + "."

        def bdichkImage184Func():
            if not str(sheet3['C111'].value).strip() == 4:
                sheet3['C111'].value = str(int(bdichkType18.get())) + "."

        def bdichkImage191Func():
            if not str(sheet3['C118'].value).strip() == 1:
                sheet3['C118'].value = str(int(bdichkType19.get())) + "."

        def bdichkImage192Func():
            if not str(sheet3['C118'].value).strip() == 2:
                sheet3['C118'].value = str(int(bdichkType19.get())) + "."

        def bdichkImage193Func():
            if not str(sheet3['C118'].value).strip() == 3:
                sheet3['C118'].value = str(int(bdichkType19.get())) + "."

        def bdichkImage194Func():
            if not str(sheet3['C118'].value).strip() == 4:
                sheet3['C118'].value = str(int(bdichkType19.get())) + "."

        def bdichkYesFunc():
            if not str(sheet3['G118'].value).strip() == '예':
                sheet3['G118'].value = str(bdichkTypeYesNo.get())

        def bdichkNoFunc():
            if not str(sheet3['G118'].value).strip() == "아니오":
                sheet3['G118'].value = str(bdichkTypeYesNo.get())

        def bdichkImage201Func():
            if not str(sheet3['C124'].value).strip() == 1:
                sheet3['C124'].value = str(int(bdichkType20.get())) + "."

        def bdichkImage202Func():
            if not str(sheet3['C124'].value).strip() == 2:
                sheet3['C124'].value = str(int(bdichkType20.get())) + "."

        def bdichkImage203Func():
            if not str(sheet3['C124'].value).strip() == 3:
                sheet3['C124'].value = str(int(bdichkType20.get())) + "."

        def bdichkImage204Func():
            if not str(sheet3['C124'].value).strip() == 4:
                sheet3['C124'].value = str(int(bdichkType20.get())) + "."

        def bdichkImage211Func():
            if not str(sheet3['C130'].value).strip() == 1:
                sheet3['C130'].value = str(int(bdichkType21.get())) + "."

        def bdichkImage212Func():
            if not str(sheet3['C130'].value).strip() == 2:
                sheet3['C130'].value = str(int(bdichkType21.get())) + "."

        def bdichkImage213Func():
            if not str(sheet3['C130'].value).strip() == 3:
                sheet3['C130'].value = str(int(bdichkType21.get())) + "."

        def bdichkImage214Func():
            if not str(sheet3['C130'].value).strip() == 4:
                sheet3['C130'].value = str(int(bdichkType21.get())) + "."

        frame3BdiBgImg = Image.open("images/bdibg.png")
        frame3BdiBg = ImageTk.PhotoImage(frame3BdiBgImg)
        frame3BdiBgLabel = tkinter.Label(displayPage2.inner, image=frame3BdiBg)
        frame3BdiBgLabel.image = frame3BdiBg
        frame3BdiBgLabel.place(x=47, y=420, height=3653)
        frame3BdiBgLabel.pack()

        global bdichk11
        bdichk11 = Radiobutton(displayPage2.inner, value=1,  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                               selectimage=chkImage12, indicatoron=False, cursor="circle", variable=bdichkType1, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage11Func)
        bdichk11.deselect()
        global bdichk12
        bdichk12 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType1, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage12Func)
        bdichk12.deselect()
        global bdichk13
        bdichk13 = Radiobutton(displayPage2.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType1, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage13Func)
        bdichk13.deselect()
        global bdichk14
        bdichk14 = Radiobutton(displayPage2.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType1, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage14Func)
        bdichk14.deselect()
        global bdichk21
        bdichk21 = Radiobutton(displayPage2.inner, value=1,  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                               selectimage=chkImage12, indicatoron=False, cursor="circle", variable=bdichkType2, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage21Func)
        bdichk21.deselect()
        global bdichk22
        bdichk22 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType2, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage22Func)
        bdichk22.deselect()
        global bdichk23
        bdichk23 = Radiobutton(displayPage2.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType2, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage23Func)
        bdichk23.deselect()
        global bdichk24
        bdichk24 = Radiobutton(displayPage2.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType2, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage24Func)
        bdichk24.deselect()
        global bdichk31
        bdichk31 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                               selectimage=chkImage12, indicatoron=False, cursor="circle", variable=bdichkType3, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage31Func)
        bdichk31.deselect()
        global bdichk32
        bdichk32 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType3, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage32Func)
        bdichk32.deselect()
        global bdichk33
        bdichk33 = Radiobutton(displayPage2.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType3, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage33Func)
        bdichk33.deselect()
        global bdichk34
        bdichk34 = Radiobutton(displayPage2.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType3, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage34Func)
        bdichk34.deselect()
        global bdichk41
        bdichk41 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                               selectimage=chkImage12, indicatoron=False, cursor="circle", variable=bdichkType4, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage41Func)
        bdichk41.deselect()
        global bdichk42
        bdichk42 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType4, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage42Func)
        bdichk42.deselect()
        global bdichk43
        bdichk43 = Radiobutton(displayPage2.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType4, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage43Func)
        bdichk43.deselect()
        global bdichk44
        bdichk44 = Radiobutton(displayPage2.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType4, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage44Func)
        bdichk44.deselect()
        global bdichk51
        bdichk51 = Radiobutton(displayPage2.inner, value=1,  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                               selectimage=chkImage12, indicatoron=False, cursor="circle", variable=bdichkType5, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage51Func)
        bdichk51.deselect()
        global bdichk52
        bdichk52 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType5, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage52Func)
        bdichk52.deselect()
        global bdichk53
        bdichk53 = Radiobutton(displayPage2.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType5, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage53Func)
        bdichk53.deselect()
        global bdichk54
        bdichk54 = Radiobutton(displayPage2.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType5, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage54Func)
        bdichk54.deselect()
        global bdichk61
        bdichk61 = Radiobutton(displayPage2.inner, value=1,  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                               selectimage=chkImage12, indicatoron=False, cursor="circle", variable=bdichkType6, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage61Func)
        bdichk61.deselect()
        global bdichk62
        bdichk62 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType6, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage62Func)
        bdichk62.deselect()
        global bdichk63
        bdichk63 = Radiobutton(displayPage2.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType6, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage63Func)
        bdichk63.deselect()
        global bdichk64
        bdichk64 = Radiobutton(displayPage2.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType6, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage64Func)
        bdichk64.deselect()
        global bdichk71
        bdichk71 = Radiobutton(displayPage2.inner, value=1,  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                               selectimage=chkImage12, indicatoron=False, cursor="circle", variable=bdichkType7, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage71Func)
        bdichk71.deselect()
        global bdichk72
        bdichk72 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType7, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage72Func)
        bdichk72.deselect()
        global bdichk73
        bdichk73 = Radiobutton(displayPage2.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType7, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage73Func)
        bdichk73.deselect()
        global bdichk74
        bdichk74 = Radiobutton(displayPage2.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType7, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage74Func)
        bdichk74.deselect()
        global bdichk81
        bdichk81 = Radiobutton(displayPage2.inner, value=1,  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                               selectimage=chkImage12, indicatoron=False, cursor="circle", variable=bdichkType8, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage81Func)
        bdichk81.deselect()
        global bdichk82
        bdichk82 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType8, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage82Func)
        bdichk82.deselect()
        global bdichk83
        bdichk83 = Radiobutton(displayPage2.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType8, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage83Func)
        bdichk83.deselect()
        global bdichk84
        bdichk84 = Radiobutton(displayPage2.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType8, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage84Func)
        bdichk84.deselect()
        global bdichk91
        bdichk91 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                               selectimage=chkImage12, indicatoron=False, cursor="circle", variable=bdichkType9, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage91Func)
        bdichk91.deselect()
        global bdichk92
        bdichk92 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType9, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage92Func)
        bdichk92.deselect()
        global bdichk93
        bdichk93 = Radiobutton(displayPage2.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType9, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage93Func)
        bdichk93.deselect()
        global bdichk94
        bdichk94 = Radiobutton(displayPage2.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                               indicatoron=False, cursor="circle", variable=bdichkType9, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage94Func)
        bdichk94.deselect()
        global bdichk101
        bdichk101 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType10, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage101Func)
        bdichk101.deselect()
        global bdichk102
        bdichk102 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType10, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage102Func)
        bdichk102.deselect()
        global bdichk103
        bdichk103 = Radiobutton(displayPage2.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType10, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage103Func)
        bdichk103.deselect()
        global bdichk104
        bdichk104 = Radiobutton(displayPage2.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType10, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage104Func)
        bdichk104.deselect()
        global bdichk111
        bdichk111 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType11, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage111Func)
        bdichk111.deselect()
        global bdichk112
        bdichk112 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType11, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage112Func)
        bdichk112.deselect()
        global bdichk113
        bdichk113 = Radiobutton(displayPage2.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType11, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage113Func)
        bdichk113.deselect()
        global bdichk114
        bdichk114 = Radiobutton(displayPage2.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType11, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage114Func)
        bdichk114.deselect()
        global bdichk121
        bdichk121 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType12, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage121Func)
        bdichk121.deselect()
        global bdichk122
        bdichk122 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType12, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage122Func)
        bdichk122.deselect()
        global bdichk123
        bdichk123 = Radiobutton(displayPage2.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType12, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage123Func)
        bdichk123.deselect()
        global bdichk124
        bdichk124 = Radiobutton(displayPage2.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType12, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage124Func)
        bdichk124.deselect()
        global bdichk131
        bdichk131 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType13, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage131Func)
        bdichk131.deselect()
        global bdichk132
        bdichk132 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType13, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage132Func)
        bdichk132.deselect()
        global bdichk133
        bdichk133 = Radiobutton(displayPage2.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType13, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage133Func)
        bdichk133.deselect()
        global bdichk134
        bdichk134 = Radiobutton(displayPage2.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType13, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage134Func)
        bdichk134.deselect()
        global bdichk141
        bdichk141 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType14, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage141Func)
        bdichk141.deselect()
        global bdichk142
        bdichk142 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType14, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage142Func)
        bdichk142.deselect()
        global bdichk143
        bdichk143 = Radiobutton(displayPage2.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType14, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage143Func)
        bdichk143.deselect()
        global bdichk144
        bdichk144 = Radiobutton(displayPage2.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType14, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage144Func)
        bdichk144.deselect()
        global bdichk151
        bdichk151 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType15, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage151Func)
        bdichk151.deselect()
        global bdichk152
        bdichk152 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType15, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage152Func)
        bdichk152.deselect()
        global bdichk153
        bdichk153 = Radiobutton(displayPage2.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType15, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage153Func)
        bdichk153.deselect()
        global bdichk154
        bdichk154 = Radiobutton(displayPage2.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType15, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage154Func)
        bdichk154.deselect()
        global bdichk161
        bdichk161 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType16, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage161Func)
        bdichk161.deselect()
        global bdichk162
        bdichk162 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType16, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage162Func)
        bdichk162.deselect()
        global bdichk163
        bdichk163 = Radiobutton(displayPage2.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType16, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage163Func)
        bdichk163.deselect()
        global bdichk164
        bdichk164 = Radiobutton(displayPage2.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType16, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage164Func)
        bdichk164.deselect()
        global bdichk171
        bdichk171 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType17, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage171Func)
        bdichk171.deselect()
        global bdichk172
        bdichk172 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType17, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage172Func)
        bdichk172.deselect()
        global bdichk173
        bdichk173 = Radiobutton(displayPage2.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType17, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage173Func)
        bdichk173.deselect()
        global bdichk174
        bdichk174 = Radiobutton(displayPage2.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType17, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage174Func)
        bdichk174.deselect()
        global bdichk181
        bdichk181 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType18, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage181Func)
        bdichk181.deselect()
        global bdichk182
        bdichk182 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType18, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage182Func)
        bdichk182.deselect()
        global bdichk183
        bdichk183 = Radiobutton(displayPage2.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType18, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage183Func)
        bdichk183.deselect()
        global bdichk184
        bdichk184 = Radiobutton(displayPage2.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType18, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage184Func)
        bdichk184.deselect()
        global bdichk191
        bdichk191 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType19, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage191Func)
        bdichk191.deselect()
        global bdichk192
        bdichk192 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType19, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage192Func)
        bdichk192.deselect()
        global bdichk19yes
        bdichk19yes = Radiobutton(displayPage2.inner, value='예', background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                  indicatoron=False, cursor="circle", variable=bdichkTypeYesNo, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkYesFunc)
        bdichk19yes.deselect()
        global bdichk19no
        bdichk19no = Radiobutton(displayPage2.inner, value='아니오', background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=bdichkTypeYesNo, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkNoFunc)
        bdichk19no.deselect()
        global bdichk193
        bdichk193 = Radiobutton(displayPage2.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType19, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage193Func)
        bdichk193.deselect()
        global bdichk194
        bdichk194 = Radiobutton(displayPage2.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType19, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage194Func)
        bdichk194.deselect()
        global bdichk201
        bdichk201 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType20, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage201Func)
        bdichk201.deselect()
        global bdichk202
        bdichk202 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType20, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage202Func)
        bdichk202.deselect()
        global bdichk203
        bdichk203 = Radiobutton(displayPage2.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType20, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage203Func)
        bdichk203.deselect()
        global bdichk204
        bdichk204 = Radiobutton(displayPage2.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType20, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage204Func)
        bdichk204.deselect()
        global bdichk211
        bdichk211 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType21, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage211Func)
        bdichk211.deselect()
        global bdichk212
        bdichk212 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType21, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage212Func)
        bdichk212.deselect()
        global bdichk213
        bdichk213 = Radiobutton(displayPage2.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType21, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage213Func)
        bdichk213.deselect()
        global bdichk214
        bdichk214 = Radiobutton(displayPage2.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=bdichkType21, font=malgungothic13, bd=0, highlightthickness=0, command=bdichkImage214Func)
        bdichk214.deselect()

        bdichk11.place(x=160, y=260+addheight, width=21, height=21)
        bdichk12.place(x=160, y=290+addheight, width=21, height=21)
        bdichk13.place(x=160, y=319+addheight, width=21, height=21)
        bdichk14.place(x=160, y=349+addheight, width=21, height=21)
        bdichk21.place(x=160, y=409+addheight, width=21, height=21)
        bdichk22.place(x=160, y=439+addheight, width=21, height=21)
        bdichk23.place(x=160, y=469+addheight, width=21, height=21)
        bdichk24.place(x=160, y=499+addheight, width=21, height=21)
        bdichk31.place(x=160, y=559+addheight, width=21, height=21)
        bdichk32.place(x=160, y=589+addheight, width=21, height=21)
        bdichk33.place(x=160, y=619+addheight, width=21, height=21)
        bdichk34.place(x=160, y=649+addheight, width=21, height=21)
        bdichk41.place(x=160, y=708+addheight, width=21, height=21)
        bdichk42.place(x=160, y=737+addheight, width=21, height=21)
        bdichk43.place(x=160, y=767+addheight, width=21, height=21)
        bdichk44.place(x=160, y=797+addheight, width=21, height=21)
        bdichk51.place(x=160, y=856+addheight, width=21, height=21)
        bdichk52.place(x=160, y=887+addheight, width=21, height=21)
        bdichk53.place(x=160, y=917+addheight, width=21, height=21)
        bdichk54.place(x=160, y=946+addheight, width=21, height=21)
        bdichk61.place(x=160, y=1006+addheight, width=21, height=21)
        bdichk62.place(x=160, y=1036+addheight, width=21, height=21)
        bdichk63.place(x=160, y=1066+addheight, width=21, height=21)
        bdichk64.place(x=160, y=1095+addheight, width=21, height=21)
        bdichk71.place(x=160, y=1154+addheight, width=21, height=21)
        bdichk72.place(x=160, y=1184+addheight, width=21, height=21)
        bdichk73.place(x=160, y=1213+addheight, width=21, height=21)
        bdichk74.place(x=160, y=1243+addheight, width=21, height=21)
        bdichk81.place(x=160, y=1303+addheight, width=21, height=21)
        bdichk82.place(x=160, y=1333+addheight, width=21, height=21)
        bdichk83.place(x=160, y=1363+addheight, width=21, height=21)
        bdichk84.place(x=160, y=1393+addheight, width=21, height=21)
        bdichk91.place(x=160, y=1458+addheight, width=21, height=21)
        bdichk92.place(x=160, y=1488+addheight, width=21, height=21)
        bdichk93.place(x=160, y=1518+addheight, width=21, height=21)
        bdichk94.place(x=160, y=1548+addheight, width=21, height=21)
        bdichk101.place(x=160, y=1607+addheight, width=21, height=21)
        bdichk102.place(x=160, y=1637+addheight, width=21, height=21)
        bdichk103.place(x=160, y=1667+addheight, width=21, height=21)
        bdichk104.place(x=160, y=1697+addheight, width=21, height=21)
        bdichk111.place(x=160, y=1756+addheight, width=21, height=21)
        bdichk112.place(x=160, y=1786+addheight, width=21, height=21)
        bdichk113.place(x=160, y=1816+addheight, width=21, height=21)
        bdichk114.place(x=160, y=1846+addheight, width=21, height=21)
        bdichk121.place(x=160, y=1904+addheight, width=21, height=21)
        bdichk122.place(x=160, y=1934+addheight, width=21, height=21)
        bdichk123.place(x=160, y=1964+addheight, width=21, height=21)
        bdichk124.place(x=160, y=1994+addheight, width=21, height=21)
        bdichk131.place(x=160, y=2053+addheight, width=21, height=21)
        bdichk132.place(x=160, y=2083+addheight, width=21, height=21)
        bdichk133.place(x=160, y=2113+addheight, width=21, height=21)
        bdichk134.place(x=160, y=2143+addheight, width=21, height=21)
        bdichk141.place(x=160, y=2202+addheight, width=21, height=21)
        bdichk142.place(x=160, y=2232+addheight, width=21, height=21)
        bdichk143.place(x=160, y=2262+addheight, width=21, height=21)
        bdichk144.place(x=160, y=2292+addheight, width=21, height=21)
        bdichk151.place(x=160, y=2350+addheight, width=21, height=21)
        bdichk152.place(x=160, y=2380+addheight, width=21, height=21)
        bdichk153.place(x=160, y=2410+addheight, width=21, height=21)
        bdichk154.place(x=160, y=2440+addheight, width=21, height=21)
        bdichk161.place(x=160, y=2505+addheight, width=21, height=21)
        bdichk162.place(x=160, y=2535+addheight, width=21, height=21)
        bdichk163.place(x=160, y=2565+addheight, width=21, height=21)
        bdichk164.place(x=160, y=2595+addheight, width=21, height=21)
        bdichk171.place(x=160, y=2653+addheight, width=21, height=21)
        bdichk172.place(x=160, y=2683+addheight, width=21, height=21)
        bdichk173.place(x=160, y=2713+addheight, width=21, height=21)
        bdichk174.place(x=160, y=2743+addheight, width=21, height=21)
        bdichk181.place(x=160, y=2803+addheight, width=21, height=21)
        bdichk182.place(x=160, y=2833+addheight, width=21, height=21)
        bdichk183.place(x=160, y=2863+addheight, width=21, height=21)
        bdichk184.place(x=160, y=2893+addheight, width=21, height=21)
        bdichk191.place(x=160, y=2952+addheight, width=21, height=21)
        bdichk192.place(x=160, y=2982+addheight, width=21, height=21)
        bdichk193.place(x=160, y=3012+addheight, width=21, height=21)
        bdichk194.place(x=160, y=3043+addheight, width=21, height=21)
        bdichk19yes.place(x=730, y=3085+addheight, width=21, height=21)
        bdichk19no.place(x=838, y=3085+addheight, width=21, height=21)
        bdichk201.place(x=160, y=3160+addheight, width=21, height=21)
        bdichk202.place(x=160, y=3190+addheight, width=21, height=21)
        bdichk203.place(x=160, y=3220+addheight, width=21, height=21)
        bdichk204.place(x=160, y=3249+addheight, width=21, height=21)
        bdichk211.place(x=160, y=3309+addheight, width=21, height=21)
        bdichk212.place(x=160, y=3338+addheight, width=21, height=21)
        bdichk213.place(x=160, y=3368+addheight, width=21, height=21)
        bdichk214.place(x=160, y=3398+addheight, width=21, height=21)

    def snapchkFunc():
        global now
        now = datetime.now()
        global snapchk
        snapchk = True
        sheet4['E2'].value = (" % s년 % s월 % s일 % s시 % s분" %
                              (now.year, now.month, now.day, now.hour, now.minute))
        sheet4['K2'].value = idInput.get()

        def snapchkImage11Func():
            if not str(sheet4['C6'].value).strip() == '전혀 그렇지 않다':
                sheet4['C6'].value = snapchk11["text"]

        def snapchkImage12Func():
            if not str(sheet4['C6'].value).strip() == '약간 그렇다':
                sheet4['C6'].value = snapchk12["text"]

        def snapchkImage13Func():
            if not str(sheet4['C6'].value).strip() == '꽤 그렇다':
                sheet4['C6'].value = snapchk13["text"]

        def snapchkImage14Func():
            if not str(sheet4['C6'].value).strip() == '아주 많이 그렇다':
                sheet4['C6'].value = snapchk14["text"]

        def snapchkImage21Func():
            if not str(sheet4['C9'].value).strip() == '전혀 그렇지 않다':
                sheet4['C9'].value = snapchk21["text"]

        def snapchkImage22Func():
            if not str(sheet4['C9'].value).strip() == '약간 그렇다':
                sheet4['C9'].value = snapchk22["text"]

        def snapchkImage23Func():
            if not str(sheet4['C9'].value).strip() == '꽤 그렇다':
                sheet4['C9'].value = snapchk23["text"]

        def snapchkImage24Func():
            if not str(sheet4['C9'].value).strip() == '아주 많이 그렇다':
                sheet4['C9'].value = snapchk24["text"]

        def snapchkImage31Func():
            if not str(sheet4['C12'].value).strip() == '전혀 그렇지 않다':
                sheet4['C12'].value = snapchk31["text"]

        def snapchkImage32Func():
            if not str(sheet4['C12'].value).strip() == '약간 그렇다':
                sheet4['C12'].value = snapchk32["text"]

        def snapchkImage33Func():
            if not str(sheet4['C12'].value).strip() == '꽤 그렇다':
                sheet4['C12'].value = snapchk33["text"]

        def snapchkImage34Func():
            if not str(sheet4['C12'].value).strip() == '아주 많이 그렇다':
                sheet4['C12'].value = snapchk34["text"]

        def snapchkImage41Func():
            if not str(sheet4['C15'].value).strip() == '전혀 그렇지 않다':
                sheet4['C15'].value = snapchk41["text"]

        def snapchkImage42Func():
            if not str(sheet4['C15'].value).strip() == '약간 그렇다':
                sheet4['C15'].value = snapchk42["text"]

        def snapchkImage43Func():
            if not str(sheet4['C15'].value).strip() == '꽤 그렇다':
                sheet4['C15'].value = snapchk43["text"]

        def snapchkImage44Func():
            if not str(sheet4['C15'].value).strip() == '아주 많이 그렇다':
                sheet4['C15'].value = snapchk44["text"]

        def snapchkImage51Func():
            if not str(sheet4['C18'].value).strip() == '전혀 그렇지 않다':
                sheet4['C18'].value = snapchk51["text"]

        def snapchkImage52Func():
            if not str(sheet4['C18'].value).strip() == '약간 그렇다':
                sheet4['C18'].value = snapchk52["text"]

        def snapchkImage53Func():
            if not str(sheet4['C18'].value).strip() == '꽤 그렇다':
                sheet4['C18'].value = snapchk53["text"]

        def snapchkImage54Func():
            if not str(sheet4['C18'].value).strip() == '아주 많이 그렇다':
                sheet4['C18'].value = snapchk54["text"]

        def snapchkImage61Func():
            if not str(sheet4['C21'].value).strip() == '전혀 그렇지 않다':
                sheet4['C21'].value = snapchk61["text"]

        def snapchkImage62Func():
            if not str(sheet4['C21'].value).strip() == '약간 그렇다':
                sheet4['C21'].value = snapchk62["text"]

        def snapchkImage63Func():
            if not str(sheet4['C21'].value).strip() == '꽤 그렇다':
                sheet4['C21'].value = snapchk63["text"]

        def snapchkImage64Func():
            if not str(sheet4['C21'].value).strip() == '아주 많이 그렇다':
                sheet4['C21'].value = snapchk64["text"]

        def snapchkImage71Func():
            if not str(sheet4['C24'].value).strip() == '전혀 그렇지 않다':
                sheet4['C24'].value = snapchk71["text"]

        def snapchkImage72Func():
            if not str(sheet4['C24'].value).strip() == '약간 그렇다':
                sheet4['C24'].value = snapchk72["text"]

        def snapchkImage73Func():
            if not str(sheet4['C24'].value).strip() == '꽤 그렇다':
                sheet4['C24'].value = snapchk73["text"]

        def snapchkImage74Func():
            if not str(sheet4['C24'].value).strip() == '아주 많이 그렇다':
                sheet4['C24'].value = snapchk74["text"]

        def snapchkImage81Func():
            if not str(sheet4['C27'].value).strip() == '전혀 그렇지 않다':
                sheet4['C27'].value = snapchk81["text"]

        def snapchkImage82Func():
            if not str(sheet4['C27'].value).strip() == '약간 그렇다':
                sheet4['C27'].value = snapchk82["text"]

        def snapchkImage83Func():
            if not str(sheet4['C27'].value).strip() == '꽤 그렇다':
                sheet4['C27'].value = snapchk83["text"]

        def snapchkImage84Func():
            if not str(sheet4['C27'].value).strip() == '아주 많이 그렇다':
                sheet4['C27'].value = snapchk84["text"]

        def snapchkImage91Func():
            if not str(sheet4['C30'].value).strip() == '전혀 그렇지 않다':
                sheet4['C30'].value = snapchk91["text"]

        def snapchkImage92Func():
            if not str(sheet4['C30'].value).strip() == '약간 그렇다':
                sheet4['C30'].value = snapchk92["text"]

        def snapchkImage93Func():
            if not str(sheet4['C30'].value).strip() == '꽤 그렇다':
                sheet4['C30'].value = snapchk93["text"]

        def snapchkImage94Func():
            if not str(sheet4['C30'].value).strip() == '아주 많이 그렇다':
                sheet4['C30'].value = snapchk94["text"]

        def snapchkImage101Func():
            if not str(sheet4['C33'].value).strip() == '전혀 그렇지 않다':
                sheet4['C33'].value = snapchk101["text"]

        def snapchkImage102Func():
            if not str(sheet4['C33'].value).strip() == '약간 그렇다':
                sheet4['C33'].value = snapchk102["text"]

        def snapchkImage103Func():
            if not str(sheet4['C33'].value).strip() == '꽤 그렇다':
                sheet4['C33'].value = snapchk103["text"]

        def snapchkImage104Func():
            if not str(sheet4['C33'].value).strip() == '아주 많이 그렇다':
                sheet4['C33'].value = snapchk104["text"]

        def snapchkImage111Func():
            if not str(sheet4['C36'].value).strip() == '전혀 그렇지 않다':
                sheet4['C36'].value = snapchk111["text"]

        def snapchkImage112Func():
            if not str(sheet4['C36'].value).strip() == '약간 그렇다':
                sheet4['C36'].value = snapchk112["text"]

        def snapchkImage113Func():
            if not str(sheet4['C36'].value).strip() == '꽤 그렇다':
                sheet4['C36'].value = snapchk113["text"]

        def snapchkImage114Func():
            if not str(sheet4['C36'].value).strip() == '아주 많이 그렇다':
                sheet4['C36'].value = snapchk114["text"]

        def snapchkImage121Func():
            if not str(sheet4['C39'].value).strip() == '전혀 그렇지 않다':
                sheet4['C39'].value = snapchk121["text"]

        def snapchkImage122Func():
            if not str(sheet4['C39'].value).strip() == '약간 그렇다':
                sheet4['C39'].value = snapchk122["text"]

        def snapchkImage123Func():
            if not str(sheet4['C39'].value).strip() == '꽤 그렇다':
                sheet4['C39'].value = snapchk123["text"]

        def snapchkImage124Func():
            if not str(sheet4['C39'].value).strip() == '아주 많이 그렇다':
                sheet4['C39'].value = snapchk124["text"]

        def snapchkImage131Func():
            if not str(sheet4['C42'].value).strip() == '전혀 그렇지 않다':
                sheet4['C42'].value = snapchk131["text"]

        def snapchkImage132Func():
            if not str(sheet4['C42'].value).strip() == '약간 그렇다':
                sheet4['C42'].value = snapchk132["text"]

        def snapchkImage133Func():
            if not str(sheet4['C42'].value).strip() == '꽤 그렇다':
                sheet4['C42'].value = snapchk133["text"]

        def snapchkImage134Func():
            if not str(sheet4['C42'].value).strip() == '아주 많이 그렇다':
                sheet4['C42'].value = snapchk134["text"]

        def snapchkImage141Func():
            if not str(sheet4['C45'].value).strip() == '전혀 그렇지 않다':
                sheet4['C45'].value = snapchk141["text"]

        def snapchkImage142Func():
            if not str(sheet4['C45'].value).strip() == '약간 그렇다':
                sheet4['C45'].value = snapchk142["text"]

        def snapchkImage143Func():
            if not str(sheet4['C45'].value).strip() == '꽤 그렇다':
                sheet4['C45'].value = snapchk143["text"]

        def snapchkImage144Func():
            if not str(sheet4['C45'].value).strip() == '아주 많이 그렇다':
                sheet4['C45'].value = snapchk144["text"]

        def snapchkImage151Func():
            if not str(sheet4['C48'].value).strip() == '전혀 그렇지 않다':
                sheet4['C48'].value = snapchk151["text"]

        def snapchkImage152Func():
            if not str(sheet4['C48'].value).strip() == '약간 그렇다':
                sheet4['C48'].value = snapchk152["text"]

        def snapchkImage153Func():
            if not str(sheet4['C48'].value).strip() == '꽤 그렇다':
                sheet4['C48'].value = snapchk153["text"]

        def snapchkImage154Func():
            if not str(sheet4['C48'].value).strip() == '아주 많이 그렇다':
                sheet4['C48'].value = snapchk154["text"]

        def snapchkImage161Func():
            if not str(sheet4['C51'].value).strip() == '전혀 그렇지 않다':
                sheet4['C51'].value = snapchk161["text"]

        def snapchkImage162Func():
            if not str(sheet4['C51'].value).strip() == '약간 그렇다':
                sheet4['C51'].value = snapchk162["text"]

        def snapchkImage163Func():
            if not str(sheet4['C51'].value).strip() == '꽤 그렇다':
                sheet4['C51'].value = snapchk163["text"]

        def snapchkImage164Func():
            if not str(sheet4['C51'].value).strip() == '아주 많이 그렇다':
                sheet4['C51'].value = snapchk164["text"]

        def snapchkImage171Func():
            if not str(sheet4['C54'].value).strip() == '전혀 그렇지 않다':
                sheet4['C54'].value = snapchk171["text"]

        def snapchkImage172Func():
            if not str(sheet4['C54'].value).strip() == '약간 그렇다':
                sheet4['C54'].value = snapchk172["text"]

        def snapchkImage173Func():
            if not str(sheet4['C54'].value).strip() == '꽤 그렇다':
                sheet4['C54'].value = snapchk173["text"]

        def snapchkImage174Func():
            if not str(sheet4['C54'].value).strip() == '아주 많이 그렇다':
                sheet4['C54'].value = snapchk174["text"]

        def snapchkImage181Func():
            if not str(sheet4['C57'].value).strip() == '전혀 그렇지 않다':
                sheet4['C57'].value = snapchk181["text"]

        def snapchkImage182Func():
            if not str(sheet4['C57'].value).strip() == '약간 그렇다':
                sheet4['C57'].value = snapchk182["text"]

        def snapchkImage183Func():
            if not str(sheet4['C57'].value).strip() == '꽤 그렇다':
                sheet4['C57'].value = snapchk183["text"]

        def snapchkImage184Func():
            if not str(sheet4['C57'].value).strip() == '아주 많이 그렇다':
                sheet4['C57'].value = snapchk184["text"]

        frame4SnapBgImg = Image.open("images/snapbg.png")
        frame4SnapBg = ImageTk.PhotoImage(frame4SnapBgImg)
        frame4SnapBgLabel = tkinter.Label(
            displayPage.inner, image=frame4SnapBg)
        frame4SnapBgLabel.image = frame4SnapBg
        frame4SnapBgLabel.place(x=47, y=420, height=3653)
        frame4SnapBgLabel.pack()

        global snapchk11
        snapchk11 = Radiobutton(displayPage.inner, value=0, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                selectimage=chkImage12, indicatoron=False, cursor="circle", variable=snapchkType1, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage11Func)
        snapchk11.deselect()
        global snapchk12
        snapchk12 = Radiobutton(displayPage.inner, value=1, text="약간 그렇다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType1, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage12Func)
        snapchk12.deselect()
        global snapchk13
        snapchk13 = Radiobutton(displayPage.inner, value=2, text="꽤 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType1, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage13Func)
        snapchk13.deselect()
        global snapchk14
        snapchk14 = Radiobutton(displayPage.inner, value=3, text="아주 많이 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType1, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage14Func)
        snapchk14.deselect()

        global snapchk21
        snapchk21 = Radiobutton(displayPage.inner, value=0, text="전혀 그렇지 않다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                selectimage=chkImage12, indicatoron=False, cursor="circle", variable=snapchkType2, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage21Func)
        snapchk21.deselect()
        global snapchk22
        snapchk22 = Radiobutton(displayPage.inner, value=1, text="약간 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType2, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage22Func)
        snapchk22.deselect()
        global snapchk23
        snapchk23 = Radiobutton(displayPage.inner, value=2, text="꽤 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType2, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage23Func)
        snapchk23.deselect()
        global snapchk24
        snapchk24 = Radiobutton(displayPage.inner, value=3, text="아주 많이 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType2, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage24Func)
        snapchk24.deselect()

        global snapchk31
        snapchk31 = Radiobutton(displayPage.inner, value=0, text="전혀 그렇지 않다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                selectimage=chkImage12, indicatoron=False, cursor="circle", variable=snapchkType3, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage31Func)
        snapchk31.deselect()
        global snapchk32
        snapchk32 = Radiobutton(displayPage.inner, value=1, text="약간 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType3, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage32Func)
        snapchk32.deselect()
        global snapchk33
        snapchk33 = Radiobutton(displayPage.inner, value=2, text="꽤 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType3, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage33Func)
        snapchk33.deselect()
        global snapchk34
        snapchk34 = Radiobutton(displayPage.inner, value=3, text="아주 많이 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType3, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage34Func)
        snapchk34.deselect()

        global snapchk41
        snapchk41 = Radiobutton(displayPage.inner, value=0, text="전혀 그렇지 않다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                selectimage=chkImage12, indicatoron=False, cursor="circle", variable=snapchkType4, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage41Func)
        snapchk41.deselect()
        global snapchk42
        snapchk42 = Radiobutton(displayPage.inner, value=1, text="약간 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType4, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage42Func)
        snapchk42.deselect()
        global snapchk43
        snapchk43 = Radiobutton(displayPage.inner, value=2, text="꽤 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType4, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage43Func)
        snapchk43.deselect()
        global snapchk44
        snapchk44 = Radiobutton(displayPage.inner, value=3, text="아주 많이 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType4, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage44Func)
        snapchk44.deselect()

        global snapchk51
        snapchk51 = Radiobutton(displayPage.inner, value=0, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                selectimage=chkImage12, indicatoron=False, cursor="circle", variable=snapchkType5, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage51Func)
        snapchk51.deselect()
        global snapchk52
        snapchk52 = Radiobutton(displayPage.inner, value=1, text="약간 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType5, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage52Func)
        snapchk52.deselect()
        global snapchk53
        snapchk53 = Radiobutton(displayPage.inner, value=2, text="꽤 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType5, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage53Func)
        snapchk53.deselect()
        global snapchk54
        snapchk54 = Radiobutton(displayPage.inner, value=3, text="아주 많이 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType5, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage54Func)
        snapchk54.deselect()

        global snapchk61
        snapchk61 = Radiobutton(displayPage.inner, value=0, text="전혀 그렇지 않다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                selectimage=chkImage12, indicatoron=False, cursor="circle", variable=snapchkType6, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage61Func)
        snapchk61.deselect()
        global snapchk62
        snapchk62 = Radiobutton(displayPage.inner, value=1, text="약간 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType6, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage62Func)
        snapchk62.deselect()
        global snapchk63
        snapchk63 = Radiobutton(displayPage.inner, value=2, text="꽤 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType6, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage63Func)
        snapchk63.deselect()
        global snapchk64
        snapchk64 = Radiobutton(displayPage.inner, value=3, text="아주 많이 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType6, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage64Func)
        snapchk64.deselect()

        global snapchk71
        snapchk71 = Radiobutton(displayPage.inner, value=0, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                selectimage=chkImage12, indicatoron=False, cursor="circle", variable=snapchkType7, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage71Func)
        snapchk71.deselect()
        global snapchk72
        snapchk72 = Radiobutton(displayPage.inner, value=1, text="약간 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType7, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage72Func)
        snapchk72.deselect()
        global snapchk73
        snapchk73 = Radiobutton(displayPage.inner, value=2, text="꽤 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType7, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage73Func)
        snapchk73.deselect()
        global snapchk74
        snapchk74 = Radiobutton(displayPage.inner, value=3, text="아주 많이 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType7, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage74Func)
        snapchk74.deselect()

        global snapchk81
        snapchk81 = Radiobutton(displayPage.inner, value=0, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                selectimage=chkImage12, indicatoron=False, cursor="circle", variable=snapchkType8, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage81Func)
        snapchk81.deselect()
        global snapchk82
        snapchk82 = Radiobutton(displayPage.inner, value=1, text="약간 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType8, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage82Func)
        snapchk82.deselect()
        global snapchk83
        snapchk83 = Radiobutton(displayPage.inner, value=2, text="꽤 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType8, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage83Func)
        snapchk83.deselect()
        global snapchk84
        snapchk84 = Radiobutton(displayPage.inner, value=3, text="아주 많이 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType8, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage84Func)
        snapchk84.deselect()

        global snapchk91
        snapchk91 = Radiobutton(displayPage.inner, value=0, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                selectimage=chkImage12, indicatoron=False, cursor="circle", variable=snapchkType9, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage91Func)
        snapchk91.deselect()
        global snapchk92
        snapchk92 = Radiobutton(displayPage.inner, value=1, text="약간 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType9, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage92Func)
        snapchk92.deselect()
        global snapchk93
        snapchk93 = Radiobutton(displayPage.inner, value=2, text="꽤 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType9, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage93Func)
        snapchk93.deselect()
        global snapchk94
        snapchk94 = Radiobutton(displayPage.inner, value=3, text="아주 많이 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType9, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage94Func)
        snapchk94.deselect()

        global snapchk101
        snapchk101 = Radiobutton(displayPage.inner, value=0, text="전혀 그렇지 않다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                 selectimage=chkImage12, indicatoron=False, cursor="circle", variable=snapchkType10, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage101Func)
        snapchk101.deselect()
        global snapchk102
        snapchk102 = Radiobutton(displayPage.inner, value=1, text="약간 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType10, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage102Func)
        snapchk102.deselect()
        global snapchk103
        snapchk103 = Radiobutton(displayPage.inner, value=2, text="꽤 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType10, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage103Func)
        snapchk103.deselect()
        global snapchk104
        snapchk104 = Radiobutton(displayPage.inner, value=3, text="아주 많이 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType10, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage104Func)
        snapchk104.deselect()

        global snapchk111
        snapchk111 = Radiobutton(displayPage.inner, value=0, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                 selectimage=chkImage12, indicatoron=False, cursor="circle", variable=snapchkType11, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage111Func)
        snapchk111.deselect()
        global snapchk112
        snapchk112 = Radiobutton(displayPage.inner, value=1, text="약간 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType11, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage112Func)
        snapchk112.deselect()
        global snapchk113
        snapchk113 = Radiobutton(displayPage.inner, value=2, text="꽤 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType11, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage113Func)
        snapchk113.deselect()
        global snapchk114
        snapchk114 = Radiobutton(displayPage.inner, value=3, text="아주 많이 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType11, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage114Func)
        snapchk114.deselect()

        global snapchk121
        snapchk121 = Radiobutton(displayPage.inner, value=0, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                 selectimage=chkImage12, indicatoron=False, cursor="circle", variable=snapchkType12, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage121Func)
        snapchk121.deselect()
        global snapchk122
        snapchk122 = Radiobutton(displayPage.inner, value=1, text="약간 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType12, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage122Func)
        snapchk122.deselect()
        global snapchk123
        snapchk123 = Radiobutton(displayPage.inner, value=2, text="꽤 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType12, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage123Func)
        snapchk123.deselect()
        global snapchk124
        snapchk124 = Radiobutton(displayPage.inner, value=3, text="아주 많이 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType12, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage124Func)
        snapchk124.deselect()

        global snapchk131
        snapchk131 = Radiobutton(displayPage.inner, value=0, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                 selectimage=chkImage12, indicatoron=False, cursor="circle", variable=snapchkType13, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage131Func)
        snapchk131.deselect()
        global snapchk132
        snapchk132 = Radiobutton(displayPage.inner, value=1, text="약간 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType13, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage132Func)
        snapchk132.deselect()
        global snapchk133
        snapchk133 = Radiobutton(displayPage.inner, value=2, text="꽤 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType13, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage133Func)
        snapchk133.deselect()
        global snapchk134
        snapchk134 = Radiobutton(displayPage.inner, value=3, text="아주 많이 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType13, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage134Func)
        snapchk134.deselect()

        global snapchk141
        snapchk141 = Radiobutton(displayPage.inner, value=0, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                 selectimage=chkImage12, indicatoron=False, cursor="circle", variable=snapchkType14, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage141Func)
        snapchk141.deselect()
        global snapchk142
        snapchk142 = Radiobutton(displayPage.inner, value=1, text="약간 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType14, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage142Func)
        snapchk142.deselect()
        global snapchk143
        snapchk143 = Radiobutton(displayPage.inner, value=2, text="꽤 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType14, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage143Func)
        snapchk143.deselect()
        global snapchk144
        snapchk144 = Radiobutton(displayPage.inner, value=3, text="아주 많이 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType14, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage144Func)
        snapchk144.deselect()

        global snapchk151
        snapchk151 = Radiobutton(displayPage.inner, value=0, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                 selectimage=chkImage12, indicatoron=False, cursor="circle", variable=snapchkType15, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage151Func)
        snapchk151.deselect()
        global snapchk152
        snapchk152 = Radiobutton(displayPage.inner, value=1, text="약간 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType15, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage152Func)
        snapchk152.deselect()
        global snapchk153
        snapchk153 = Radiobutton(displayPage.inner, value=2, text="꽤 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType15, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage153Func)
        snapchk153.deselect()
        global snapchk154
        snapchk154 = Radiobutton(displayPage.inner, value=3, text="아주 많이 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType15, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage154Func)
        snapchk154.deselect()

        global snapchk161
        snapchk161 = Radiobutton(displayPage.inner, value=0, text="전혀 그렇지 않다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                 selectimage=chkImage12, indicatoron=False, cursor="circle", variable=snapchkType16, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage161Func)
        snapchk161.deselect()
        global snapchk162
        snapchk162 = Radiobutton(displayPage.inner, value=1, text="약간 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType16, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage162Func)
        snapchk162.deselect()
        global snapchk163
        snapchk163 = Radiobutton(displayPage.inner, value=2, text="꽤 그렇다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType16, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage163Func)
        snapchk163.deselect()
        global snapchk164
        snapchk164 = Radiobutton(displayPage.inner, value=3, text="아주 많이 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType16, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage164Func)
        snapchk164.deselect()

        global snapchk171
        snapchk171 = Radiobutton(displayPage.inner, value=0, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                 selectimage=chkImage12, indicatoron=False, cursor="circle", variable=snapchkType17, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage171Func)
        snapchk171.deselect()
        global snapchk172
        snapchk172 = Radiobutton(displayPage.inner, value=1, text="약간 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType17, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage172Func)
        snapchk172.deselect()
        global snapchk173
        snapchk173 = Radiobutton(displayPage.inner, value=2, text="꽤 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType17, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage173Func)
        snapchk173.deselect()
        global snapchk174
        snapchk174 = Radiobutton(displayPage.inner, value=3, text="아주 많이 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType17, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage174Func)
        snapchk174.deselect()

        global snapchk181
        snapchk181 = Radiobutton(displayPage.inner, value=0, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                 selectimage=chkImage12, indicatoron=False, cursor="circle", variable=snapchkType18, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage181Func)
        snapchk181.deselect()
        global snapchk182
        snapchk182 = Radiobutton(displayPage.inner, value=1, text="약간 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType18, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage182Func)
        snapchk182.deselect()
        global snapchk183
        snapchk183 = Radiobutton(displayPage.inner, value=2, text="꽤 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType18, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage183Func)
        snapchk183.deselect()
        global snapchk184
        snapchk184 = Radiobutton(displayPage.inner, value=3, text="아주 많이 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType18, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage184Func)
        snapchk184.deselect()

        snapchk11.place(x=1024, y=345, width=21, height=21)
        snapchk12.place(x=1117, y=345, width=21, height=21)
        snapchk13.place(x=1204, y=345, width=21, height=21)
        snapchk14.place(x=1304, y=345, width=21, height=21)
        snapchk21.place(x=1024, y=385, width=21, height=21)
        snapchk22.place(x=1117, y=385, width=21, height=21)
        snapchk23.place(x=1204, y=385, width=21, height=21)
        snapchk24.place(x=1304, y=385, width=21, height=21)
        snapchk31.place(x=1024, y=424, width=21, height=21)
        snapchk32.place(x=1117, y=424, width=21, height=21)
        snapchk33.place(x=1204, y=424, width=21, height=21)
        snapchk34.place(x=1304, y=424, width=21, height=21)
        snapchk41.place(x=1024, y=463, width=21, height=21)
        snapchk42.place(x=1117, y=463, width=21, height=21)
        snapchk43.place(x=1204, y=463, width=21, height=21)
        snapchk44.place(x=1304, y=463, width=21, height=21)
        snapchk51.place(x=1024, y=501, width=21, height=21)
        snapchk52.place(x=1117, y=501, width=21, height=21)
        snapchk53.place(x=1204, y=501, width=21, height=21)
        snapchk54.place(x=1304, y=501, width=21, height=21)
        snapchk61.place(x=1024, y=540, width=21, height=21)
        snapchk62.place(x=1117, y=540, width=21, height=21)
        snapchk63.place(x=1204, y=540, width=21, height=21)
        snapchk64.place(x=1304, y=540, width=21, height=21)
        snapchk71.place(x=1024, y=579, width=21, height=21)
        snapchk72.place(x=1117, y=579, width=21, height=21)
        snapchk73.place(x=1204, y=579, width=21, height=21)
        snapchk74.place(x=1304, y=579, width=21, height=21)
        snapchk81.place(x=1024, y=619, width=21, height=21)
        snapchk82.place(x=1117, y=619, width=21, height=21)
        snapchk83.place(x=1204, y=619, width=21, height=21)
        snapchk84.place(x=1304, y=619, width=21, height=21)
        snapchk91.place(x=1024, y=658, width=21, height=21)
        snapchk92.place(x=1117, y=658, width=21, height=21)
        snapchk93.place(x=1204, y=658, width=21, height=21)
        snapchk94.place(x=1304, y=658, width=21, height=21)
        snapchk101.place(x=1024, y=696, width=21, height=21)
        snapchk102.place(x=1117, y=696, width=21, height=21)
        snapchk103.place(x=1204, y=696, width=21, height=21)
        snapchk104.place(x=1304, y=696, width=21, height=21)
        snapchk111.place(x=1024, y=736, width=21, height=21)
        snapchk112.place(x=1117, y=736, width=21, height=21)
        snapchk113.place(x=1204, y=736, width=21, height=21)
        snapchk114.place(x=1304, y=736, width=21, height=21)
        snapchk121.place(x=1024, y=774, width=21, height=21)
        snapchk122.place(x=1117, y=774, width=21, height=21)
        snapchk123.place(x=1204, y=774, width=21, height=21)
        snapchk124.place(x=1304, y=774, width=21, height=21)
        snapchk131.place(x=1024, y=813, width=21, height=21)
        snapchk132.place(x=1117, y=813, width=21, height=21)
        snapchk133.place(x=1204, y=813, width=21, height=21)
        snapchk134.place(x=1304, y=813, width=21, height=21)
        snapchk141.place(x=1024, y=851, width=21, height=21)
        snapchk142.place(x=1117, y=851, width=21, height=21)
        snapchk143.place(x=1204, y=851, width=21, height=21)
        snapchk144.place(x=1304, y=851, width=21, height=21)
        snapchk151.place(x=1024, y=890, width=21, height=21)
        snapchk152.place(x=1117, y=890, width=21, height=21)
        snapchk153.place(x=1204, y=890, width=21, height=21)
        snapchk154.place(x=1304, y=890, width=21, height=21)
        snapchk161.place(x=1024, y=929, width=21, height=21)
        snapchk162.place(x=1117, y=929, width=21, height=21)
        snapchk163.place(x=1204, y=929, width=21, height=21)
        snapchk164.place(x=1304, y=929, width=21, height=21)
        snapchk171.place(x=1024, y=969, width=21, height=21)
        snapchk172.place(x=1117, y=969, width=21, height=21)
        snapchk173.place(x=1204, y=969, width=21, height=21)
        snapchk174.place(x=1304, y=969, width=21, height=21)
        snapchk181.place(x=1024, y=1008, width=21, height=21)
        snapchk182.place(x=1117, y=1008, width=21, height=21)
        snapchk183.place(x=1204, y=1008, width=21, height=21)
        snapchk184.place(x=1304, y=1008, width=21, height=21)

    def snapchkFunc2():
        global now
        now = datetime.now()
        global snapchk
        snapchk = True
        sheet4['E2'].value = (" % s년 % s월 % s일 % s시 % s분" %
                              (now.year, now.month, now.day, now.hour, now.minute))
        sheet4['K2'].value = idInput.get()

        def snapchkImage11Func():
            if not str(sheet4['C6'].value).strip() == '전혀 그렇지 않다':
                sheet4['C6'].value = snapchk11["text"]

        def snapchkImage12Func():
            if not str(sheet4['C6'].value).strip() == '약간 그렇다':
                sheet4['C6'].value = snapchk12["text"]

        def snapchkImage13Func():
            if not str(sheet4['C6'].value).strip() == '꽤 그렇다':
                sheet4['C6'].value = snapchk13["text"]

        def snapchkImage14Func():
            if not str(sheet4['C6'].value).strip() == '아주 많이 그렇다':
                sheet4['C6'].value = snapchk14["text"]

        def snapchkImage21Func():
            if not str(sheet4['C9'].value).strip() == '전혀 그렇지 않다':
                sheet4['C9'].value = snapchk21["text"]

        def snapchkImage22Func():
            if not str(sheet4['C9'].value).strip() == '약간 그렇다':
                sheet4['C9'].value = snapchk22["text"]

        def snapchkImage23Func():
            if not str(sheet4['C9'].value).strip() == '꽤 그렇다':
                sheet4['C9'].value = snapchk23["text"]

        def snapchkImage24Func():
            if not str(sheet4['C9'].value).strip() == '아주 많이 그렇다':
                sheet4['C9'].value = snapchk24["text"]

        def snapchkImage31Func():
            if not str(sheet4['C12'].value).strip() == '전혀 그렇지 않다':
                sheet4['C12'].value = snapchk31["text"]

        def snapchkImage32Func():
            if not str(sheet4['C12'].value).strip() == '약간 그렇다':
                sheet4['C12'].value = snapchk32["text"]

        def snapchkImage33Func():
            if not str(sheet4['C12'].value).strip() == '꽤 그렇다':
                sheet4['C12'].value = snapchk33["text"]

        def snapchkImage34Func():
            if not str(sheet4['C12'].value).strip() == '아주 많이 그렇다':
                sheet4['C12'].value = snapchk34["text"]

        def snapchkImage41Func():
            if not str(sheet4['C15'].value).strip() == '전혀 그렇지 않다':
                sheet4['C15'].value = snapchk41["text"]

        def snapchkImage42Func():
            if not str(sheet4['C15'].value).strip() == '약간 그렇다':
                sheet4['C15'].value = snapchk42["text"]

        def snapchkImage43Func():
            if not str(sheet4['C15'].value).strip() == '꽤 그렇다':
                sheet4['C15'].value = snapchk43["text"]

        def snapchkImage44Func():
            if not str(sheet4['C15'].value).strip() == '아주 많이 그렇다':
                sheet4['C15'].value = snapchk44["text"]

        def snapchkImage51Func():
            if not str(sheet4['C18'].value).strip() == '전혀 그렇지 않다':
                sheet4['C18'].value = snapchk51["text"]

        def snapchkImage52Func():
            if not str(sheet4['C18'].value).strip() == '약간 그렇다':
                sheet4['C18'].value = snapchk52["text"]

        def snapchkImage53Func():
            if not str(sheet4['C18'].value).strip() == '꽤 그렇다':
                sheet4['C18'].value = snapchk53["text"]

        def snapchkImage54Func():
            if not str(sheet4['C18'].value).strip() == '아주 많이 그렇다':
                sheet4['C18'].value = snapchk54["text"]

        def snapchkImage61Func():
            if not str(sheet4['C21'].value).strip() == '전혀 그렇지 않다':
                sheet4['C21'].value = snapchk61["text"]

        def snapchkImage62Func():
            if not str(sheet4['C21'].value).strip() == '약간 그렇다':
                sheet4['C21'].value = snapchk62["text"]

        def snapchkImage63Func():
            if not str(sheet4['C21'].value).strip() == '꽤 그렇다':
                sheet4['C21'].value = snapchk63["text"]

        def snapchkImage64Func():
            if not str(sheet4['C21'].value).strip() == '아주 많이 그렇다':
                sheet4['C21'].value = snapchk64["text"]

        def snapchkImage71Func():
            if not str(sheet4['C24'].value).strip() == '전혀 그렇지 않다':
                sheet4['C24'].value = snapchk71["text"]

        def snapchkImage72Func():
            if not str(sheet4['C24'].value).strip() == '약간 그렇다':
                sheet4['C24'].value = snapchk72["text"]

        def snapchkImage73Func():
            if not str(sheet4['C24'].value).strip() == '꽤 그렇다':
                sheet4['C24'].value = snapchk73["text"]

        def snapchkImage74Func():
            if not str(sheet4['C24'].value).strip() == '아주 많이 그렇다':
                sheet4['C24'].value = snapchk74["text"]

        def snapchkImage81Func():
            if not str(sheet4['C27'].value).strip() == '전혀 그렇지 않다':
                sheet4['C27'].value = snapchk81["text"]

        def snapchkImage82Func():
            if not str(sheet4['C27'].value).strip() == '약간 그렇다':
                sheet4['C27'].value = snapchk82["text"]

        def snapchkImage83Func():
            if not str(sheet4['C27'].value).strip() == '꽤 그렇다':
                sheet4['C27'].value = snapchk83["text"]

        def snapchkImage84Func():
            if not str(sheet4['C27'].value).strip() == '아주 많이 그렇다':
                sheet4['C27'].value = snapchk84["text"]

        def snapchkImage91Func():
            if not str(sheet4['C30'].value).strip() == '전혀 그렇지 않다':
                sheet4['C30'].value = snapchk91["text"]

        def snapchkImage92Func():
            if not str(sheet4['C30'].value).strip() == '약간 그렇다':
                sheet4['C30'].value = snapchk92["text"]

        def snapchkImage93Func():
            if not str(sheet4['C30'].value).strip() == '꽤 그렇다':
                sheet4['C30'].value = snapchk93["text"]

        def snapchkImage94Func():
            if not str(sheet4['C30'].value).strip() == '아주 많이 그렇다':
                sheet4['C30'].value = snapchk94["text"]

        def snapchkImage101Func():
            if not str(sheet4['C33'].value).strip() == '전혀 그렇지 않다':
                sheet4['C33'].value = snapchk101["text"]

        def snapchkImage102Func():
            if not str(sheet4['C33'].value).strip() == '약간 그렇다':
                sheet4['C33'].value = snapchk102["text"]

        def snapchkImage103Func():
            if not str(sheet4['C33'].value).strip() == '꽤 그렇다':
                sheet4['C33'].value = snapchk103["text"]

        def snapchkImage104Func():
            if not str(sheet4['C33'].value).strip() == '아주 많이 그렇다':
                sheet4['C33'].value = snapchk104["text"]

        def snapchkImage111Func():
            if not str(sheet4['C36'].value).strip() == '전혀 그렇지 않다':
                sheet4['C36'].value = snapchk111["text"]

        def snapchkImage112Func():
            if not str(sheet4['C36'].value).strip() == '약간 그렇다':
                sheet4['C36'].value = snapchk112["text"]

        def snapchkImage113Func():
            if not str(sheet4['C36'].value).strip() == '꽤 그렇다':
                sheet4['C36'].value = snapchk113["text"]

        def snapchkImage114Func():
            if not str(sheet4['C36'].value).strip() == '아주 많이 그렇다':
                sheet4['C36'].value = snapchk114["text"]

        def snapchkImage121Func():
            if not str(sheet4['C39'].value).strip() == '전혀 그렇지 않다':
                sheet4['C39'].value = snapchk121["text"]

        def snapchkImage122Func():
            if not str(sheet4['C39'].value).strip() == '약간 그렇다':
                sheet4['C39'].value = snapchk122["text"]

        def snapchkImage123Func():
            if not str(sheet4['C39'].value).strip() == '꽤 그렇다':
                sheet4['C39'].value = snapchk123["text"]

        def snapchkImage124Func():
            if not str(sheet4['C39'].value).strip() == '아주 많이 그렇다':
                sheet4['C39'].value = snapchk124["text"]

        def snapchkImage131Func():
            if not str(sheet4['C42'].value).strip() == '전혀 그렇지 않다':
                sheet4['C42'].value = snapchk131["text"]

        def snapchkImage132Func():
            if not str(sheet4['C42'].value).strip() == '약간 그렇다':
                sheet4['C42'].value = snapchk132["text"]

        def snapchkImage133Func():
            if not str(sheet4['C42'].value).strip() == '꽤 그렇다':
                sheet4['C42'].value = snapchk133["text"]

        def snapchkImage134Func():
            if not str(sheet4['C42'].value).strip() == '아주 많이 그렇다':
                sheet4['C42'].value = snapchk134["text"]

        def snapchkImage141Func():
            if not str(sheet4['C45'].value).strip() == '전혀 그렇지 않다':
                sheet4['C45'].value = snapchk141["text"]

        def snapchkImage142Func():
            if not str(sheet4['C45'].value).strip() == '약간 그렇다':
                sheet4['C45'].value = snapchk142["text"]

        def snapchkImage143Func():
            if not str(sheet4['C45'].value).strip() == '꽤 그렇다':
                sheet4['C45'].value = snapchk143["text"]

        def snapchkImage144Func():
            if not str(sheet4['C45'].value).strip() == '아주 많이 그렇다':
                sheet4['C45'].value = snapchk144["text"]

        def snapchkImage151Func():
            if not str(sheet4['C48'].value).strip() == '전혀 그렇지 않다':
                sheet4['C48'].value = snapchk151["text"]

        def snapchkImage152Func():
            if not str(sheet4['C48'].value).strip() == '약간 그렇다':
                sheet4['C48'].value = snapchk152["text"]

        def snapchkImage153Func():
            if not str(sheet4['C48'].value).strip() == '꽤 그렇다':
                sheet4['C48'].value = snapchk153["text"]

        def snapchkImage154Func():
            if not str(sheet4['C48'].value).strip() == '아주 많이 그렇다':
                sheet4['C48'].value = snapchk154["text"]

        def snapchkImage161Func():
            if not str(sheet4['C51'].value).strip() == '전혀 그렇지 않다':
                sheet4['C51'].value = snapchk161["text"]

        def snapchkImage162Func():
            if not str(sheet4['C51'].value).strip() == '약간 그렇다':
                sheet4['C51'].value = snapchk162["text"]

        def snapchkImage163Func():
            if not str(sheet4['C51'].value).strip() == '꽤 그렇다':
                sheet4['C51'].value = snapchk163["text"]

        def snapchkImage164Func():
            if not str(sheet4['C51'].value).strip() == '아주 많이 그렇다':
                sheet4['C51'].value = snapchk164["text"]

        def snapchkImage171Func():
            if not str(sheet4['C54'].value).strip() == '전혀 그렇지 않다':
                sheet4['C54'].value = snapchk171["text"]

        def snapchkImage172Func():
            if not str(sheet4['C54'].value).strip() == '약간 그렇다':
                sheet4['C54'].value = snapchk172["text"]

        def snapchkImage173Func():
            if not str(sheet4['C54'].value).strip() == '꽤 그렇다':
                sheet4['C54'].value = snapchk173["text"]

        def snapchkImage174Func():
            if not str(sheet4['C54'].value).strip() == '아주 많이 그렇다':
                sheet4['C54'].value = snapchk174["text"]

        def snapchkImage181Func():
            if not str(sheet4['C57'].value).strip() == '전혀 그렇지 않다':
                sheet4['C57'].value = snapchk181["text"]

        def snapchkImage182Func():
            if not str(sheet4['C57'].value).strip() == '약간 그렇다':
                sheet4['C57'].value = snapchk182["text"]

        def snapchkImage183Func():
            if not str(sheet4['C57'].value).strip() == '꽤 그렇다':
                sheet4['C57'].value = snapchk183["text"]

        def snapchkImage184Func():
            if not str(sheet4['C57'].value).strip() == '아주 많이 그렇다':
                sheet4['C57'].value = snapchk184["text"]

        frame4SnapBgImg = Image.open("images/snapbg.png")
        frame4SnapBg = ImageTk.PhotoImage(frame4SnapBgImg)
        frame4SnapBgLabel = tkinter.Label(
            displayPage2.inner, image=frame4SnapBg)
        frame4SnapBgLabel.image = frame4SnapBg
        frame4SnapBgLabel.place(x=47, y=420, height=3653)
        frame4SnapBgLabel.pack()

        global snapchk11
        snapchk11 = Radiobutton(displayPage2.inner, value=0, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                selectimage=chkImage12, indicatoron=False, cursor="circle", variable=snapchkType1, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage11Func)
        snapchk11.deselect()
        global snapchk12
        snapchk12 = Radiobutton(displayPage2.inner, value=1, text="약간 그렇다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType1, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage12Func)
        snapchk12.deselect()
        global snapchk13
        snapchk13 = Radiobutton(displayPage2.inner, value=2, text="꽤 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType1, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage13Func)
        snapchk13.deselect()
        global snapchk14
        snapchk14 = Radiobutton(displayPage2.inner, value=3, text="아주 많이 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType1, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage14Func)
        snapchk14.deselect()

        global snapchk21
        snapchk21 = Radiobutton(displayPage2.inner, value=0, text="전혀 그렇지 않다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                selectimage=chkImage12, indicatoron=False, cursor="circle", variable=snapchkType2, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage21Func)
        snapchk21.deselect()
        global snapchk22
        snapchk22 = Radiobutton(displayPage2.inner, value=1, text="약간 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType2, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage22Func)
        snapchk22.deselect()
        global snapchk23
        snapchk23 = Radiobutton(displayPage2.inner, value=2, text="꽤 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType2, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage23Func)
        snapchk23.deselect()
        global snapchk24
        snapchk24 = Radiobutton(displayPage2.inner, value=3, text="아주 많이 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType2, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage24Func)
        snapchk24.deselect()

        global snapchk31
        snapchk31 = Radiobutton(displayPage2.inner, value=0, text="전혀 그렇지 않다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                selectimage=chkImage12, indicatoron=False, cursor="circle", variable=snapchkType3, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage31Func)
        snapchk31.deselect()
        global snapchk32
        snapchk32 = Radiobutton(displayPage2.inner, value=1, text="약간 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType3, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage32Func)
        snapchk32.deselect()
        global snapchk33
        snapchk33 = Radiobutton(displayPage2.inner, value=2, text="꽤 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType3, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage33Func)
        snapchk33.deselect()
        global snapchk34
        snapchk34 = Radiobutton(displayPage2.inner, value=3, text="아주 많이 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType3, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage34Func)
        snapchk34.deselect()

        global snapchk41
        snapchk41 = Radiobutton(displayPage2.inner, value=0, text="전혀 그렇지 않다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                selectimage=chkImage12, indicatoron=False, cursor="circle", variable=snapchkType4, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage41Func)
        snapchk41.deselect()
        global snapchk42
        snapchk42 = Radiobutton(displayPage2.inner, value=1, text="약간 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType4, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage42Func)
        snapchk42.deselect()
        global snapchk43
        snapchk43 = Radiobutton(displayPage2.inner, value=2, text="꽤 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType4, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage43Func)
        snapchk43.deselect()
        global snapchk44
        snapchk44 = Radiobutton(displayPage2.inner, value=3, text="아주 많이 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType4, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage44Func)
        snapchk44.deselect()

        global snapchk51
        snapchk51 = Radiobutton(displayPage2.inner, value=0, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                selectimage=chkImage12, indicatoron=False, cursor="circle", variable=snapchkType5, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage51Func)
        snapchk51.deselect()
        global snapchk52
        snapchk52 = Radiobutton(displayPage2.inner, value=1, text="약간 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType5, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage52Func)
        snapchk52.deselect()
        global snapchk53
        snapchk53 = Radiobutton(displayPage2.inner, value=2, text="꽤 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType5, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage53Func)
        snapchk53.deselect()
        global snapchk54
        snapchk54 = Radiobutton(displayPage2.inner, value=3, text="아주 많이 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType5, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage54Func)
        snapchk54.deselect()

        global snapchk61
        snapchk61 = Radiobutton(displayPage2.inner, value=0, text="전혀 그렇지 않다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                selectimage=chkImage12, indicatoron=False, cursor="circle", variable=snapchkType6, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage61Func)
        snapchk61.deselect()
        global snapchk62
        snapchk62 = Radiobutton(displayPage2.inner, value=1, text="약간 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType6, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage62Func)
        snapchk62.deselect()
        global snapchk63
        snapchk63 = Radiobutton(displayPage2.inner, value=2, text="꽤 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType6, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage63Func)
        snapchk63.deselect()
        global snapchk64
        snapchk64 = Radiobutton(displayPage2.inner, value=3, text="아주 많이 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType6, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage64Func)
        snapchk64.deselect()

        global snapchk71
        snapchk71 = Radiobutton(displayPage2.inner, value=0, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                selectimage=chkImage12, indicatoron=False, cursor="circle", variable=snapchkType7, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage71Func)
        snapchk71.deselect()
        global snapchk72
        snapchk72 = Radiobutton(displayPage2.inner, value=1, text="약간 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType7, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage72Func)
        snapchk72.deselect()
        global snapchk73
        snapchk73 = Radiobutton(displayPage2.inner, value=2, text="꽤 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType7, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage73Func)
        snapchk73.deselect()
        global snapchk74
        snapchk74 = Radiobutton(displayPage2.inner, value=3, text="아주 많이 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType7, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage74Func)
        snapchk74.deselect()

        global snapchk81
        snapchk81 = Radiobutton(displayPage2.inner, value=0, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                selectimage=chkImage12, indicatoron=False, cursor="circle", variable=snapchkType8, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage81Func)
        snapchk81.deselect()
        global snapchk82
        snapchk82 = Radiobutton(displayPage2.inner, value=1, text="약간 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType8, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage82Func)
        snapchk82.deselect()
        global snapchk83
        snapchk83 = Radiobutton(displayPage2.inner, value=2, text="꽤 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType8, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage83Func)
        snapchk83.deselect()
        global snapchk84
        snapchk84 = Radiobutton(displayPage2.inner, value=3, text="아주 많이 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType8, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage84Func)
        snapchk84.deselect()

        global snapchk91
        snapchk91 = Radiobutton(displayPage2.inner, value=0, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                selectimage=chkImage12, indicatoron=False, cursor="circle", variable=snapchkType9, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage91Func)
        snapchk91.deselect()
        global snapchk92
        snapchk92 = Radiobutton(displayPage2.inner, value=1, text="약간 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType9, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage92Func)
        snapchk92.deselect()
        global snapchk93
        snapchk93 = Radiobutton(displayPage2.inner, value=2, text="꽤 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType9, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage93Func)
        snapchk93.deselect()
        global snapchk94
        snapchk94 = Radiobutton(displayPage2.inner, value=3, text="아주 많이 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=snapchkType9, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage94Func)
        snapchk94.deselect()

        global snapchk101
        snapchk101 = Radiobutton(displayPage2.inner, value=0, text="전혀 그렇지 않다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                 selectimage=chkImage12, indicatoron=False, cursor="circle", variable=snapchkType10, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage101Func)
        snapchk101.deselect()
        global snapchk102
        snapchk102 = Radiobutton(displayPage2.inner, value=1, text="약간 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType10, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage102Func)
        snapchk102.deselect()
        global snapchk103
        snapchk103 = Radiobutton(displayPage2.inner, value=2, text="꽤 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType10, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage103Func)
        snapchk103.deselect()
        global snapchk104
        snapchk104 = Radiobutton(displayPage2.inner, value=3, text="아주 많이 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType10, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage104Func)
        snapchk104.deselect()

        global snapchk111
        snapchk111 = Radiobutton(displayPage2.inner, value=0, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                 selectimage=chkImage12, indicatoron=False, cursor="circle", variable=snapchkType11, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage111Func)
        snapchk111.deselect()
        global snapchk112
        snapchk112 = Radiobutton(displayPage2.inner, value=1, text="약간 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType11, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage112Func)
        snapchk112.deselect()
        global snapchk113
        snapchk113 = Radiobutton(displayPage2.inner, value=2, text="꽤 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType11, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage113Func)
        snapchk113.deselect()
        global snapchk114
        snapchk114 = Radiobutton(displayPage2.inner, value=3, text="아주 많이 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType11, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage114Func)
        snapchk114.deselect()

        global snapchk121
        snapchk121 = Radiobutton(displayPage2.inner, value=0, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                 selectimage=chkImage12, indicatoron=False, cursor="circle", variable=snapchkType12, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage121Func)
        snapchk121.deselect()
        global snapchk122
        snapchk122 = Radiobutton(displayPage2.inner, value=1, text="약간 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType12, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage122Func)
        snapchk122.deselect()
        global snapchk123
        snapchk123 = Radiobutton(displayPage2.inner, value=2, text="꽤 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType12, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage123Func)
        snapchk123.deselect()
        global snapchk124
        snapchk124 = Radiobutton(displayPage2.inner, value=3, text="아주 많이 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType12, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage124Func)
        snapchk124.deselect()

        global snapchk131
        snapchk131 = Radiobutton(displayPage2.inner, value=0, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                 selectimage=chkImage12, indicatoron=False, cursor="circle", variable=snapchkType13, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage131Func)
        snapchk131.deselect()
        global snapchk132
        snapchk132 = Radiobutton(displayPage2.inner, value=1, text="약간 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType13, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage132Func)
        snapchk132.deselect()
        global snapchk133
        snapchk133 = Radiobutton(displayPage2.inner, value=2, text="꽤 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType13, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage133Func)
        snapchk133.deselect()
        global snapchk134
        snapchk134 = Radiobutton(displayPage2.inner, value=3, text="아주 많이 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType13, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage134Func)
        snapchk134.deselect()

        global snapchk141
        snapchk141 = Radiobutton(displayPage2.inner, value=0, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                 selectimage=chkImage12, indicatoron=False, cursor="circle", variable=snapchkType14, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage141Func)
        snapchk141.deselect()
        global snapchk142
        snapchk142 = Radiobutton(displayPage2.inner, value=1, text="약간 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType14, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage142Func)
        snapchk142.deselect()
        global snapchk143
        snapchk143 = Radiobutton(displayPage2.inner, value=2, text="꽤 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType14, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage143Func)
        snapchk143.deselect()
        global snapchk144
        snapchk144 = Radiobutton(displayPage2.inner, value=3, text="아주 많이 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType14, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage144Func)
        snapchk144.deselect()

        global snapchk151
        snapchk151 = Radiobutton(displayPage2.inner, value=0, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                 selectimage=chkImage12, indicatoron=False, cursor="circle", variable=snapchkType15, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage151Func)
        snapchk151.deselect()
        global snapchk152
        snapchk152 = Radiobutton(displayPage2.inner, value=1, text="약간 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType15, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage152Func)
        snapchk152.deselect()
        global snapchk153
        snapchk153 = Radiobutton(displayPage2.inner, value=2, text="꽤 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType15, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage153Func)
        snapchk153.deselect()
        global snapchk154
        snapchk154 = Radiobutton(displayPage2.inner, value=3, text="아주 많이 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType15, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage154Func)
        snapchk154.deselect()

        global snapchk161
        snapchk161 = Radiobutton(displayPage2.inner, value=0, text="전혀 그렇지 않다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                 selectimage=chkImage12, indicatoron=False, cursor="circle", variable=snapchkType16, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage161Func)
        snapchk161.deselect()
        global snapchk162
        snapchk162 = Radiobutton(displayPage2.inner, value=1, text="약간 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType16, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage162Func)
        snapchk162.deselect()
        global snapchk163
        snapchk163 = Radiobutton(displayPage2.inner, value=2, text="꽤 그렇다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType16, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage163Func)
        snapchk163.deselect()
        global snapchk164
        snapchk164 = Radiobutton(displayPage2.inner, value=3, text="아주 많이 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType16, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage164Func)
        snapchk164.deselect()

        global snapchk171
        snapchk171 = Radiobutton(displayPage2.inner, value=0, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                 selectimage=chkImage12, indicatoron=False, cursor="circle", variable=snapchkType17, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage171Func)
        snapchk171.deselect()
        global snapchk172
        snapchk172 = Radiobutton(displayPage2.inner, value=1, text="약간 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType17, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage172Func)
        snapchk172.deselect()
        global snapchk173
        snapchk173 = Radiobutton(displayPage2.inner, value=2, text="꽤 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType17, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage173Func)
        snapchk173.deselect()
        global snapchk174
        snapchk174 = Radiobutton(displayPage2.inner, value=3, text="아주 많이 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType17, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage174Func)
        snapchk174.deselect()

        global snapchk181
        snapchk181 = Radiobutton(displayPage2.inner, value=0, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                 selectimage=chkImage12, indicatoron=False, cursor="circle", variable=snapchkType18, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage181Func)
        snapchk181.deselect()
        global snapchk182
        snapchk182 = Radiobutton(displayPage2.inner, value=1, text="약간 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType18, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage182Func)
        snapchk182.deselect()
        global snapchk183
        snapchk183 = Radiobutton(displayPage2.inner, value=2, text="꽤 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType18, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage183Func)
        snapchk183.deselect()
        global snapchk184
        snapchk184 = Radiobutton(displayPage2.inner, value=3, text="아주 많이 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=snapchkType18, font=malgungothic13, bd=0, highlightthickness=0, command=snapchkImage184Func)
        snapchk184.deselect()

        snapchk11.place(x=1024, y=345+addheight, width=21, height=21)
        snapchk12.place(x=1117, y=345+addheight, width=21, height=21)
        snapchk13.place(x=1204, y=345+addheight, width=21, height=21)
        snapchk14.place(x=1304, y=345+addheight, width=21, height=21)
        snapchk21.place(x=1024, y=385+addheight, width=21, height=21)
        snapchk22.place(x=1117, y=385+addheight, width=21, height=21)
        snapchk23.place(x=1204, y=385+addheight, width=21, height=21)
        snapchk24.place(x=1304, y=385+addheight, width=21, height=21)
        snapchk31.place(x=1024, y=424+addheight, width=21, height=21)
        snapchk32.place(x=1117, y=424+addheight, width=21, height=21)
        snapchk33.place(x=1204, y=424+addheight, width=21, height=21)
        snapchk34.place(x=1304, y=424+addheight, width=21, height=21)
        snapchk41.place(x=1024, y=463+addheight, width=21, height=21)
        snapchk42.place(x=1117, y=463+addheight, width=21, height=21)
        snapchk43.place(x=1204, y=463+addheight, width=21, height=21)
        snapchk44.place(x=1304, y=463+addheight, width=21, height=21)
        snapchk51.place(x=1024, y=501+addheight, width=21, height=21)
        snapchk52.place(x=1117, y=501+addheight, width=21, height=21)
        snapchk53.place(x=1204, y=501+addheight, width=21, height=21)
        snapchk54.place(x=1304, y=501+addheight, width=21, height=21)
        snapchk61.place(x=1024, y=540+addheight, width=21, height=21)
        snapchk62.place(x=1117, y=540+addheight, width=21, height=21)
        snapchk63.place(x=1204, y=540+addheight, width=21, height=21)
        snapchk64.place(x=1304, y=540+addheight, width=21, height=21)
        snapchk71.place(x=1024, y=579+addheight, width=21, height=21)
        snapchk72.place(x=1117, y=579+addheight, width=21, height=21)
        snapchk73.place(x=1204, y=579+addheight, width=21, height=21)
        snapchk74.place(x=1304, y=579+addheight, width=21, height=21)
        snapchk81.place(x=1024, y=619+addheight, width=21, height=21)
        snapchk82.place(x=1117, y=619+addheight, width=21, height=21)
        snapchk83.place(x=1204, y=619+addheight, width=21, height=21)
        snapchk84.place(x=1304, y=619+addheight, width=21, height=21)
        snapchk91.place(x=1024, y=658+addheight, width=21, height=21)
        snapchk92.place(x=1117, y=658+addheight, width=21, height=21)
        snapchk93.place(x=1204, y=658+addheight, width=21, height=21)
        snapchk94.place(x=1304, y=658+addheight, width=21, height=21)
        snapchk101.place(x=1024, y=696+addheight, width=21, height=21)
        snapchk102.place(x=1117, y=696+addheight, width=21, height=21)
        snapchk103.place(x=1204, y=696+addheight, width=21, height=21)
        snapchk104.place(x=1304, y=696+addheight, width=21, height=21)
        snapchk111.place(x=1024, y=736+addheight, width=21, height=21)
        snapchk112.place(x=1117, y=736+addheight, width=21, height=21)
        snapchk113.place(x=1204, y=736+addheight, width=21, height=21)
        snapchk114.place(x=1304, y=736+addheight, width=21, height=21)
        snapchk121.place(x=1024, y=774+addheight, width=21, height=21)
        snapchk122.place(x=1117, y=774+addheight, width=21, height=21)
        snapchk123.place(x=1204, y=774+addheight, width=21, height=21)
        snapchk124.place(x=1304, y=774+addheight, width=21, height=21)
        snapchk131.place(x=1024, y=813+addheight, width=21, height=21)
        snapchk132.place(x=1117, y=813+addheight, width=21, height=21)
        snapchk133.place(x=1204, y=813+addheight, width=21, height=21)
        snapchk134.place(x=1304, y=813+addheight, width=21, height=21)
        snapchk141.place(x=1024, y=851+addheight, width=21, height=21)
        snapchk142.place(x=1117, y=851+addheight, width=21, height=21)
        snapchk143.place(x=1204, y=851+addheight, width=21, height=21)
        snapchk144.place(x=1304, y=851+addheight, width=21, height=21)
        snapchk151.place(x=1024, y=890+addheight, width=21, height=21)
        snapchk152.place(x=1117, y=890+addheight, width=21, height=21)
        snapchk153.place(x=1204, y=890+addheight, width=21, height=21)
        snapchk154.place(x=1304, y=890+addheight, width=21, height=21)
        snapchk161.place(x=1024, y=929+addheight, width=21, height=21)
        snapchk162.place(x=1117, y=929+addheight, width=21, height=21)
        snapchk163.place(x=1204, y=929+addheight, width=21, height=21)
        snapchk164.place(x=1304, y=929+addheight, width=21, height=21)
        snapchk171.place(x=1024, y=969+addheight, width=21, height=21)
        snapchk172.place(x=1117, y=969+addheight, width=21, height=21)
        snapchk173.place(x=1204, y=969+addheight, width=21, height=21)
        snapchk174.place(x=1304, y=969+addheight, width=21, height=21)
        snapchk181.place(x=1024, y=1008+addheight, width=21, height=21)
        snapchk182.place(x=1117, y=1008+addheight, width=21, height=21)
        snapchk183.place(x=1204, y=1008+addheight, width=21, height=21)
        snapchk184.place(x=1304, y=1008+addheight, width=21, height=21)

    def st_2chkFunc():
        global now
        now = datetime.now()
        global st_2chk
        st_2chk = True
        sheet5['E2'].value = (" % s년 % s월 % s일 % s시 % s분" %
                              (now.year, now.month, now.day, now.hour, now.minute))
        sheet5['K2'].value = idInput.get()

        def st_2chkImage11Func():
            # sumScore()
            if not str(sheet5['C6'].value).strip() == '거의 그렇지 않다':
                sheet5['C6'].value = st_2chk11["text"]

        def st_2chkImage12Func():
            # sumScore()
            if not str(sheet5['C6'].value).strip() == '가끔 그렇다':
                sheet5['C6'].value = st_2chk12["text"]

        def st_2chkImage13Func():
            # sumScore()
            if not str(sheet5['C6'].value).strip() == '자주 그렇다':
                sheet5['C6'].value = st_2chk13["text"]

        def st_2chkImage14Func():
            # sumScore()
            if not str(sheet5['C6'].value).strip() == '거의 언제나 그렇다':
                sheet5['C6'].value = st_2chk14["text"]

        def st_2chkImage21Func():
            if not str(sheet5['C9'].value).strip() == '거의 그렇지 않다':
                sheet5['C9'].value = st_2chk21["text"]

        def st_2chkImage22Func():
            if not str(sheet5['C9'].value).strip() == '가끔 그렇다':
                sheet5['C9'].value = st_2chk22["text"]

        def st_2chkImage23Func():
            if not str(sheet5['C9'].value).strip() == '자주 그렇다':
                sheet5['C9'].value = st_2chk23["text"]

        def st_2chkImage24Func():
            if not str(sheet5['C9'].value).strip() == '거의 언제나 그렇다':
                sheet5['C9'].value = st_2chk24["text"]

        def st_2chkImage31Func():
            if not str(sheet5['C12'].value).strip() == '거의 그렇지 않다':
                sheet5['C12'].value = st_2chk31["text"]

        def st_2chkImage32Func():
            if not str(sheet5['C12'].value).strip() == '가끔 그렇다':
                sheet5['C12'].value = st_2chk32["text"]

        def st_2chkImage33Func():
            if not str(sheet5['C12'].value).strip() == '자주 그렇다':
                sheet5['C12'].value = st_2chk33["text"]

        def st_2chkImage34Func():
            if not str(sheet5['C12'].value).strip() == '거의 언제나 그렇다':
                sheet5['C12'].value = st_2chk34["text"]

        def st_2chkImage41Func():
            if not str(sheet5['C15'].value).strip() == '거의 그렇지 않다':
                sheet5['C15'].value = st_2chk41["text"]

        def st_2chkImage42Func():
            if not str(sheet5['C15'].value).strip() == '가끔 그렇다':
                sheet5['C15'].value = st_2chk42["text"]

        def st_2chkImage43Func():
            if not str(sheet5['C15'].value).strip() == '자주 그렇다':
                sheet5['C15'].value = st_2chk43["text"]

        def st_2chkImage44Func():
            if not str(sheet5['C15'].value).strip() == '거의 언제나 그렇다':
                sheet5['C15'].value = st_2chk44["text"]

        def st_2chkImage51Func():
            if not str(sheet5['C18'].value).strip() == '거의 그렇지 않다':
                sheet5['C18'].value = st_2chk51["text"]

        def st_2chkImage52Func():
            if not str(sheet5['C18'].value).strip() == '가끔 그렇다':
                sheet5['C18'].value = st_2chk52["text"]

        def st_2chkImage53Func():
            if not str(sheet5['C18'].value).strip() == '자주 그렇다':
                sheet5['C18'].value = st_2chk53["text"]

        def st_2chkImage54Func():
            if not str(sheet5['C18'].value).strip() == '거의 언제나 그렇다':
                sheet5['C18'].value = st_2chk54["text"]

        def st_2chkImage61Func():
            if not str(sheet5['C21'].value).strip() == '거의 그렇지 않다':
                sheet5['C21'].value = st_2chk61["text"]

        def st_2chkImage62Func():
            if not str(sheet5['C21'].value).strip() == '가끔 그렇다':
                sheet5['C21'].value = st_2chk62["text"]

        def st_2chkImage63Func():
            if not str(sheet5['C21'].value).strip() == '자주 그렇다':
                sheet5['C21'].value = st_2chk63["text"]

        def st_2chkImage64Func():
            if not str(sheet5['C21'].value).strip() == '거의 언제나 그렇다':
                sheet5['C21'].value = st_2chk64["text"]

        def st_2chkImage71Func():
            if not str(sheet5['C24'].value).strip() == '거의 그렇지 않다':
                sheet5['C24'].value = st_2chk71["text"]

        def st_2chkImage72Func():
            if not str(sheet5['C24'].value).strip() == '가끔 그렇다':
                sheet5['C24'].value = st_2chk72["text"]

        def st_2chkImage73Func():
            if not str(sheet5['C24'].value).strip() == '자주 그렇다':
                sheet5['C24'].value = st_2chk73["text"]

        def st_2chkImage74Func():
            if not str(sheet5['C24'].value).strip() == '거의 언제나 그렇다':
                sheet5['C24'].value = st_2chk74["text"]

        def st_2chkImage81Func():
            if not str(sheet5['C27'].value).strip() == '거의 그렇지 않다':
                sheet5['C27'].value = st_2chk81["text"]

        def st_2chkImage82Func():
            if not str(sheet5['C27'].value).strip() == '가끔 그렇다':
                sheet5['C27'].value = st_2chk82["text"]

        def st_2chkImage83Func():
            if not str(sheet5['C27'].value).strip() == '자주 그렇다':
                sheet5['C27'].value = st_2chk83["text"]

        def st_2chkImage84Func():
            if not str(sheet5['C27'].value).strip() == '거의 언제나 그렇다':
                sheet5['C27'].value = st_2chk84["text"]

        def st_2chkImage91Func():
            if not str(sheet5['C30'].value).strip() == '거의 그렇지 않다':
                sheet5['C30'].value = st_2chk91["text"]

        def st_2chkImage92Func():
            if not str(sheet5['C30'].value).strip() == '가끔 그렇다':
                sheet5['C30'].value = st_2chk92["text"]

        def st_2chkImage93Func():
            if not str(sheet5['C30'].value).strip() == '자주 그렇다':
                sheet5['C30'].value = st_2chk93["text"]

        def st_2chkImage94Func():
            if not str(sheet5['C30'].value).strip() == '거의 언제나 그렇다':
                sheet5['C30'].value = st_2chk94["text"]

        def st_2chkImage101Func():
            if not str(sheet5['C33'].value).strip() == '거의 그렇지 않다':
                sheet5['C33'].value = st_2chk101["text"]

        def st_2chkImage102Func():
            if not str(sheet5['C33'].value).strip() == '가끔 그렇다':
                sheet5['C33'].value = st_2chk102["text"]

        def st_2chkImage103Func():
            if not str(sheet5['C33'].value).strip() == '자주 그렇다':
                sheet5['C33'].value = st_2chk103["text"]

        def st_2chkImage104Func():
            if not str(sheet5['C33'].value).strip() == '거의 언제나 그렇다':
                sheet5['C33'].value = st_2chk104["text"]

        def st_2chkImage111Func():
            if not str(sheet5['C36'].value).strip() == '거의 그렇지 않다':
                sheet5['C36'].value = st_2chk111["text"]

        def st_2chkImage112Func():
            if not str(sheet5['C36'].value).strip() == '가끔 그렇다':
                sheet5['C36'].value = st_2chk112["text"]

        def st_2chkImage113Func():
            if not str(sheet5['C36'].value).strip() == '자주 그렇다':
                sheet5['C36'].value = st_2chk113["text"]

        def st_2chkImage114Func():
            if not str(sheet5['C36'].value).strip() == '거의 언제나 그렇다':
                sheet5['C36'].value = st_2chk114["text"]

        def st_2chkImage121Func():
            if not str(sheet5['C39'].value).strip() == '거의 그렇지 않다':
                sheet5['C39'].value = st_2chk121["text"]

        def st_2chkImage122Func():
            if not str(sheet5['C39'].value).strip() == '가끔 그렇다':
                sheet5['C39'].value = st_2chk122["text"]

        def st_2chkImage123Func():
            if not str(sheet5['C39'].value).strip() == '자주 그렇다':
                sheet5['C39'].value = st_2chk123["text"]

        def st_2chkImage124Func():
            if not str(sheet5['C39'].value).strip() == '거의 언제나 그렇다':
                sheet5['C39'].value = st_2chk124["text"]

        def st_2chkImage131Func():
            if not str(sheet5['C42'].value).strip() == '거의 그렇지 않다':
                sheet5['C42'].value = st_2chk131["text"]

        def st_2chkImage132Func():
            if not str(sheet5['C42'].value).strip() == '가끔 그렇다':
                sheet5['C42'].value = st_2chk132["text"]

        def st_2chkImage133Func():
            if not str(sheet5['C42'].value).strip() == '자주 그렇다':
                sheet5['C42'].value = st_2chk133["text"]

        def st_2chkImage134Func():
            if not str(sheet5['C42'].value).strip() == '거의 언제나 그렇다':
                sheet5['C42'].value = st_2chk134["text"]

        def st_2chkImage141Func():
            if not str(sheet5['C45'].value).strip() == '거의 그렇지 않다':
                sheet5['C45'].value = st_2chk141["text"]

        def st_2chkImage142Func():
            if not str(sheet5['C45'].value).strip() == '가끔 그렇다':
                sheet5['C45'].value = st_2chk142["text"]

        def st_2chkImage143Func():
            if not str(sheet5['C45'].value).strip() == '자주 그렇다':
                sheet5['C45'].value = st_2chk143["text"]

        def st_2chkImage144Func():
            if not str(sheet5['C45'].value).strip() == '거의 언제나 그렇다':
                sheet5['C45'].value = st_2chk144["text"]

        def st_2chkImage151Func():
            if not str(sheet5['C48'].value).strip() == '거의 그렇지 않다':
                sheet5['C48'].value = st_2chk151["text"]

        def st_2chkImage152Func():
            if not str(sheet5['C48'].value).strip() == '가끔 그렇다':
                sheet5['C48'].value = st_2chk152["text"]

        def st_2chkImage153Func():
            if not str(sheet5['C48'].value).strip() == '자주 그렇다':
                sheet5['C48'].value = st_2chk153["text"]

        def st_2chkImage154Func():
            if not str(sheet5['C48'].value).strip() == '거의 언제나 그렇다':
                sheet5['C48'].value = st_2chk154["text"]

        def st_2chkImage161Func():
            if not str(sheet5['C51'].value).strip() == '거의 그렇지 않다':
                sheet5['C51'].value = st_2chk161["text"]

        def st_2chkImage162Func():
            if not str(sheet5['C51'].value).strip() == '가끔 그렇다':
                sheet5['C51'].value = st_2chk162["text"]

        def st_2chkImage163Func():
            if not str(sheet5['C51'].value).strip() == '자주 그렇다':
                sheet5['C51'].value = st_2chk163["text"]

        def st_2chkImage164Func():
            if not str(sheet5['C51'].value).strip() == '거의 언제나 그렇다':
                sheet5['C51'].value = st_2chk164["text"]

        def st_2chkImage171Func():
            if not str(sheet5['C54'].value).strip() == '거의 그렇지 않다':
                sheet5['C54'].value = st_2chk171["text"]

        def st_2chkImage172Func():
            if not str(sheet5['C54'].value).strip() == '가끔 그렇다':
                sheet5['C54'].value = st_2chk172["text"]

        def st_2chkImage173Func():
            if not str(sheet5['C54'].value).strip() == '자주 그렇다':
                sheet5['C54'].value = st_2chk173["text"]

        def st_2chkImage174Func():
            if not str(sheet5['C54'].value).strip() == '거의 언제나 그렇다':
                sheet5['C54'].value = st_2chk174["text"]

        def st_2chkImage181Func():
            if not str(sheet5['C57'].value).strip() == '거의 그렇지 않다':
                sheet5['C57'].value = st_2chk181["text"]

        def st_2chkImage182Func():
            if not str(sheet5['C57'].value).strip() == '가끔 그렇다':
                sheet5['C57'].value = st_2chk182["text"]

        def st_2chkImage183Func():
            if not str(sheet5['C57'].value).strip() == '자주 그렇다':
                sheet5['C57'].value = st_2chk183["text"]

        def st_2chkImage184Func():
            if not str(sheet5['C57'].value).strip() == '거의 언제나 그렇다':
                sheet5['C57'].value = st_2chk184["text"]

        def st_2chkImage191Func():
            if not str(sheet5['C60'].value).strip() == '거의 그렇지 않다':
                sheet5['C60'].value = st_2chk181["text"]

        def st_2chkImage192Func():
            if not str(sheet5['C60'].value).strip() == '가끔 그렇다':
                sheet5['C60'].value = st_2chk182["text"]

        def st_2chkImage193Func():
            if not str(sheet5['C60'].value).strip() == '자주 그렇다':
                sheet5['C60'].value = st_2chk183["text"]

        def st_2chkImage194Func():
            if not str(sheet5['C60'].value).strip() == '거의 언제나 그렇다':
                sheet5['C60'].value = st_2chk184["text"]

        def st_2chkImage201Func():
            if not str(sheet5['C63'].value).strip() == '거의 그렇지 않다':
                sheet5['C63'].value = st_2chk181["text"]

        def st_2chkImage202Func():
            if not str(sheet5['C63'].value).strip() == '가끔 그렇다':
                sheet5['C63'].value = st_2chk182["text"]

        def st_2chkImage203Func():
            if not str(sheet5['C63'].value).strip() == '자주 그렇다':
                sheet5['C63'].value = st_2chk183["text"]

        def st_2chkImage204Func():
            if not str(sheet5['C63'].value).strip() == '거의 언제나 그렇다':
                sheet5['C63'].value = st_2chk184["text"]

        frame5St_2BgImg = Image.open("images/st_2bg.png")
        frame5St_2Bg = ImageTk.PhotoImage(frame5St_2BgImg)
        frame5St_2BgLabel = tkinter.Label(
            displayPage.inner, image=frame5St_2Bg)
        frame5St_2BgLabel.image = frame5St_2Bg
        frame5St_2BgLabel.place(x=47, y=420, height=1330)
        frame5St_2BgLabel.pack()

        global st_2chk11
        st_2chk11 = Radiobutton(displayPage.inner, value=4, text="거의 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType1, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage11Func)
        st_2chk11.deselect()
        global st_2chk12
        st_2chk12 = Radiobutton(displayPage.inner, value=3, text="가끔 그렇다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=st_2chkType1, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage12Func)
        st_2chk12.deselect()
        global st_2chk13
        st_2chk13 = Radiobutton(displayPage.inner, value=2, text="자주 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=st_2chkType1, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage13Func)
        st_2chk13.deselect()
        global st_2chk14
        st_2chk14 = Radiobutton(displayPage.inner, value=1, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=st_2chkType1, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage14Func)
        st_2chk14.deselect()
        global st_2chk21
        st_2chk21 = Radiobutton(displayPage.inner, value=1, text="거의 그렇지 않다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType2, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage21Func)
        st_2chk21.deselect()
        global st_2chk22
        st_2chk22 = Radiobutton(displayPage.inner, value=2, text="가끔 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=st_2chkType2, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage22Func)
        st_2chk22.deselect()
        global st_2chk23
        st_2chk23 = Radiobutton(displayPage.inner, value=3, text="자주 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=st_2chkType2, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage23Func)
        st_2chk23.deselect()
        global st_2chk24
        st_2chk24 = Radiobutton(displayPage.inner, value=4, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=st_2chkType2, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage24Func)
        st_2chk24.deselect()
        global st_2chk31
        st_2chk31 = Radiobutton(displayPage.inner, value=1, text="거의 그렇지 않다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType3, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage31Func)
        st_2chk31.deselect()
        global st_2chk32
        st_2chk32 = Radiobutton(displayPage.inner, value=2, text="가끔 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=st_2chkType3, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage32Func)
        st_2chk32.deselect()
        global st_2chk33
        st_2chk33 = Radiobutton(displayPage.inner, value=3, text="자주 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=st_2chkType3, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage33Func)
        st_2chk33.deselect()
        global st_2chk34
        st_2chk34 = Radiobutton(displayPage.inner, value=4, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=st_2chkType3, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage34Func)
        st_2chk34.deselect()
        global st_2chk41
        st_2chk41 = Radiobutton(displayPage.inner, value=1, text="거의 그렇지 않다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType4, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage41Func)
        st_2chk41.deselect()
        global st_2chk42
        st_2chk42 = Radiobutton(displayPage.inner, value=2, text="가끔 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=st_2chkType4, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage42Func)
        st_2chk42.deselect()
        global st_2chk43
        st_2chk43 = Radiobutton(displayPage.inner, value=3, text="자주 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=st_2chkType4, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage43Func)
        st_2chk43.deselect()
        global st_2chk44
        st_2chk44 = Radiobutton(displayPage.inner, value=4, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=st_2chkType4, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage44Func)
        st_2chk44.deselect()
        global st_2chk51
        st_2chk51 = Radiobutton(displayPage.inner, value=1, text="거의 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType5, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage51Func)
        st_2chk51.deselect()
        global st_2chk52
        st_2chk52 = Radiobutton(displayPage.inner, value=2, text="가끔 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=st_2chkType5, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage52Func)
        st_2chk52.deselect()
        global st_2chk53
        st_2chk53 = Radiobutton(displayPage.inner, value=3, text="자주 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=st_2chkType5, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage53Func)
        st_2chk53.deselect()
        global st_2chk54
        st_2chk54 = Radiobutton(displayPage.inner, value=4, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=st_2chkType5, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage54Func)
        st_2chk54.deselect()
        global st_2chk61
        st_2chk61 = Radiobutton(displayPage.inner, value=4, text="거의 그렇지 않다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType6, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage61Func)
        st_2chk61.deselect()
        global st_2chk62
        st_2chk62 = Radiobutton(displayPage.inner, value=3, text="가끔 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=st_2chkType6, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage62Func)
        st_2chk62.deselect()
        global st_2chk63
        st_2chk63 = Radiobutton(displayPage.inner, value=2, text="자주 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=st_2chkType6, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage63Func)
        st_2chk63.deselect()
        global st_2chk64
        st_2chk64 = Radiobutton(displayPage.inner, value=1, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=st_2chkType6, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage64Func)
        st_2chk64.deselect()

        global st_2chk71
        st_2chk71 = Radiobutton(displayPage.inner, value=4, text="거의 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType7, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage71Func)
        st_2chk71.deselect()
        global st_2chk72
        st_2chk72 = Radiobutton(displayPage.inner, value=3, text="가끔 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=st_2chkType7, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage72Func)
        st_2chk72.deselect()
        global st_2chk73
        st_2chk73 = Radiobutton(displayPage.inner, value=2, text="자주 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=st_2chkType7, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage73Func)
        st_2chk73.deselect()
        global st_2chk74
        st_2chk74 = Radiobutton(displayPage.inner, value=1, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=st_2chkType7, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage74Func)
        st_2chk74.deselect()

        global st_2chk81
        st_2chk81 = Radiobutton(displayPage.inner, value=1, text="거의 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType8, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage81Func)
        st_2chk81.deselect()
        global st_2chk82
        st_2chk82 = Radiobutton(displayPage.inner, value=2, text="가끔 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=st_2chkType8, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage82Func)
        st_2chk82.deselect()
        global st_2chk83
        st_2chk83 = Radiobutton(displayPage.inner, value=3, text="자주 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=st_2chkType8, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage83Func)
        st_2chk83.deselect()
        global st_2chk84
        st_2chk84 = Radiobutton(displayPage.inner, value=4, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=st_2chkType8, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage84Func)
        st_2chk84.deselect()

        global st_2chk91
        st_2chk91 = Radiobutton(displayPage.inner, value=1, text="거의 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType9, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage91Func)
        st_2chk91.deselect()
        global st_2chk92
        st_2chk92 = Radiobutton(displayPage.inner, value=2, text="가끔 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=st_2chkType9, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage92Func)
        st_2chk92.deselect()
        global st_2chk93
        st_2chk93 = Radiobutton(displayPage.inner, value=3, text="자주 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=st_2chkType9, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage93Func)
        st_2chk93.deselect()
        global st_2chk94
        st_2chk94 = Radiobutton(displayPage.inner, value=4, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=st_2chkType9, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage94Func)
        st_2chk94.deselect()
        global st_2chk101
        st_2chk101 = Radiobutton(displayPage.inner, value=4, text="거의 그렇지 않다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType10, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage101Func)
        st_2chk101.deselect()
        global st_2chk102
        st_2chk102 = Radiobutton(displayPage.inner, value=3, text="가끔 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_2chkType10, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage102Func)
        st_2chk102.deselect()
        global st_2chk103
        st_2chk103 = Radiobutton(displayPage.inner, value=2, text="자주 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_2chkType10, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage103Func)
        st_2chk103.deselect()
        global st_2chk104
        st_2chk104 = Radiobutton(displayPage.inner, value=1, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_2chkType10, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage104Func)
        st_2chk104.deselect()
        global st_2chk111
        st_2chk111 = Radiobutton(displayPage.inner, value=1, text="거의 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType11, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage111Func)
        st_2chk111.deselect()
        global st_2chk112
        st_2chk112 = Radiobutton(displayPage.inner, value=2, text="가끔 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_2chkType11, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage112Func)
        st_2chk112.deselect()
        global st_2chk113
        st_2chk113 = Radiobutton(displayPage.inner, value=3, text="자주 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_2chkType11, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage113Func)
        st_2chk113.deselect()
        global st_2chk114
        st_2chk114 = Radiobutton(displayPage.inner, value=4, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_2chkType11, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage114Func)
        st_2chk114.deselect()
        global st_2chk121
        st_2chk121 = Radiobutton(displayPage.inner, value=1, text="거의 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType12, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage121Func)
        st_2chk121.deselect()
        global st_2chk122
        st_2chk122 = Radiobutton(displayPage.inner, value=2, text="가끔 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_2chkType12, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage122Func)
        st_2chk122.deselect()
        global st_2chk123
        st_2chk123 = Radiobutton(displayPage.inner, value=3, text="자주 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_2chkType12, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage123Func)
        st_2chk123.deselect()
        global st_2chk124
        st_2chk124 = Radiobutton(displayPage.inner, value=4, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_2chkType12, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage124Func)
        st_2chk124.deselect()
        global st_2chk131
        st_2chk131 = Radiobutton(displayPage.inner, value=4, text="거의 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType13, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage131Func)
        st_2chk131.deselect()
        global st_2chk132
        st_2chk132 = Radiobutton(displayPage.inner, value=3, text="가끔 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_2chkType13, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage132Func)
        st_2chk132.deselect()
        global st_2chk133
        st_2chk133 = Radiobutton(displayPage.inner, value=2, text="자주 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_2chkType13, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage133Func)
        st_2chk133.deselect()
        global st_2chk134
        st_2chk134 = Radiobutton(displayPage.inner, value=1, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_2chkType13, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage134Func)
        st_2chk134.deselect()
        global st_2chk141
        st_2chk141 = Radiobutton(displayPage.inner, value=1, text="거의 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType14, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage141Func)
        st_2chk141.deselect()
        global st_2chk142
        st_2chk142 = Radiobutton(displayPage.inner, value=2, text="가끔 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_2chkType14, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage142Func)
        st_2chk142.deselect()
        global st_2chk143
        st_2chk143 = Radiobutton(displayPage.inner, value=3, text="자주 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_2chkType14, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage143Func)
        st_2chk143.deselect()
        global st_2chk144
        st_2chk144 = Radiobutton(displayPage.inner, value=4, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_2chkType14, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage144Func)
        st_2chk144.deselect()
        global st_2chk151
        st_2chk151 = Radiobutton(displayPage.inner, value=1, text="거의 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType15, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage151Func)
        st_2chk151.deselect()
        global st_2chk152
        st_2chk152 = Radiobutton(displayPage.inner, value=2, text="가끔 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_2chkType15, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage152Func)
        st_2chk152.deselect()
        global st_2chk153
        st_2chk153 = Radiobutton(displayPage.inner, value=3, text="자주 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_2chkType15, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage153Func)
        st_2chk153.deselect()
        global st_2chk154
        st_2chk154 = Radiobutton(displayPage.inner, value=4, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_2chkType15, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage154Func)
        st_2chk154.deselect()
        global st_2chk161
        st_2chk161 = Radiobutton(displayPage.inner, value=4, text="거의 그렇지 않다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType16, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage161Func)
        st_2chk161.deselect()
        global st_2chk162
        st_2chk162 = Radiobutton(displayPage.inner, value=3, text="가끔 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_2chkType16, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage162Func)
        st_2chk162.deselect()
        global st_2chk163
        st_2chk163 = Radiobutton(displayPage.inner, value=2, text="자주 그렇다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_2chkType16, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage163Func)
        st_2chk163.deselect()
        global st_2chk164
        st_2chk164 = Radiobutton(displayPage.inner, value=1, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_2chkType16, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage164Func)
        st_2chk164.deselect()
        global st_2chk171
        st_2chk171 = Radiobutton(displayPage.inner, value=1, text="거의 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType17, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage171Func)
        st_2chk171.deselect()
        global st_2chk172
        st_2chk172 = Radiobutton(displayPage.inner, value=2, text="가끔 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_2chkType17, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage172Func)
        st_2chk172.deselect()
        global st_2chk173
        st_2chk173 = Radiobutton(displayPage.inner, value=3, text="자주 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_2chkType17, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage173Func)
        st_2chk173.deselect()
        global st_2chk174
        st_2chk174 = Radiobutton(displayPage.inner, value=4, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_2chkType17, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage174Func)
        st_2chk174.deselect()
        global st_2chk181
        st_2chk181 = Radiobutton(displayPage.inner, value=1, text="거의 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType18, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage181Func)
        st_2chk181.deselect()
        global st_2chk182
        st_2chk182 = Radiobutton(displayPage.inner, value=2, text="가끔 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_2chkType18, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage182Func)
        st_2chk182.deselect()
        global st_2chk183
        st_2chk183 = Radiobutton(displayPage.inner, value=3, text="자주 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_2chkType18, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage183Func)
        st_2chk183.deselect()
        global st_2chk184
        st_2chk184 = Radiobutton(displayPage.inner, value=4, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_2chkType18, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage184Func)
        st_2chk184.deselect()
        global st_2chk191
        st_2chk191 = Radiobutton(displayPage.inner, value=4, text="거의 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType19, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage191Func)
        st_2chk191.deselect()
        global st_2chk192
        st_2chk192 = Radiobutton(displayPage.inner, value=3, text="가끔 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_2chkType19, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage192Func)
        st_2chk192.deselect()
        global st_2chk193
        st_2chk193 = Radiobutton(displayPage.inner, value=2, text="자주 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_2chkType19, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage193Func)
        st_2chk193.deselect()
        global st_2chk194
        st_2chk194 = Radiobutton(displayPage.inner, value=1, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_2chkType19, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage194Func)
        st_2chk194.deselect()
        global st_2chk201
        st_2chk201 = Radiobutton(displayPage.inner, value=1, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType20, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage201Func)
        st_2chk201.deselect()
        global st_2chk202
        st_2chk202 = Radiobutton(displayPage.inner, value=2, text="가끔 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_2chkType20, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage202Func)
        st_2chk202.deselect()
        global st_2chk203
        st_2chk203 = Radiobutton(displayPage.inner, value=3, text="자주 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_2chkType20, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage203Func)
        st_2chk203.deselect()
        global st_2chk204
        st_2chk204 = Radiobutton(displayPage.inner, value=4, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_2chkType20, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage204Func)
        st_2chk204.deselect()

        st_2chk11.place(x=1000, y=472, width=60, height=21)
        st_2chk12.place(x=1076, y=472, width=60, height=21)
        st_2chk13.place(x=1150, y=472, width=60, height=21)
        st_2chk14.place(x=1224, y=472, width=60, height=21)
        st_2chk21.place(x=1000, y=512, width=60, height=21)
        st_2chk22.place(x=1076, y=512, width=60, height=21)
        st_2chk23.place(x=1150, y=512, width=60, height=21)
        st_2chk24.place(x=1224, y=512, width=60, height=21)
        st_2chk31.place(x=1000, y=551, width=60, height=21)
        st_2chk32.place(x=1076, y=551, width=60, height=21)
        st_2chk33.place(x=1150, y=551, width=60, height=21)
        st_2chk34.place(x=1224, y=551, width=60, height=21)
        st_2chk41.place(x=1000, y=589, width=60, height=21)
        st_2chk42.place(x=1076, y=589, width=60, height=21)
        st_2chk43.place(x=1150, y=589, width=60, height=21)
        st_2chk44.place(x=1224, y=589, width=60, height=21)
        st_2chk51.place(x=1000, y=628, width=60, height=21)
        st_2chk52.place(x=1076, y=628, width=60, height=21)
        st_2chk53.place(x=1150, y=628, width=60, height=21)
        st_2chk54.place(x=1224, y=628, width=60, height=21)
        st_2chk61.place(x=1000, y=667, width=60, height=21)
        st_2chk62.place(x=1076, y=667, width=60, height=21)
        st_2chk63.place(x=1150, y=667, width=60, height=21)
        st_2chk64.place(x=1224, y=667, width=60, height=21)
        st_2chk71.place(x=1000, y=704, width=60, height=21)
        st_2chk72.place(x=1076, y=704, width=60, height=21)
        st_2chk73.place(x=1150, y=704, width=60, height=21)
        st_2chk74.place(x=1224, y=704, width=60, height=21)
        st_2chk81.place(x=1000, y=742, width=60, height=21)
        st_2chk82.place(x=1076, y=742, width=60, height=21)
        st_2chk83.place(x=1150, y=742, width=60, height=21)
        st_2chk84.place(x=1224, y=742, width=60, height=21)
        st_2chk91.place(x=1000, y=781, width=60, height=21)
        st_2chk92.place(x=1076, y=781, width=60, height=21)
        st_2chk93.place(x=1150, y=781, width=60, height=21)
        st_2chk94.place(x=1224, y=781, width=60, height=21)
        st_2chk101.place(x=1000, y=821, width=60, height=21)
        st_2chk102.place(x=1076, y=821, width=60, height=21)
        st_2chk103.place(x=1150, y=821, width=60, height=21)
        st_2chk104.place(x=1224, y=821, width=60, height=21)
        st_2chk111.place(x=1000, y=860, width=60, height=21)
        st_2chk112.place(x=1076, y=860, width=60, height=21)
        st_2chk113.place(x=1150, y=860, width=60, height=21)
        st_2chk114.place(x=1224, y=860, width=60, height=21)
        st_2chk121.place(x=1000, y=899, width=60, height=21)
        st_2chk122.place(x=1076, y=899, width=60, height=21)
        st_2chk123.place(x=1150, y=899, width=60, height=21)
        st_2chk124.place(x=1224, y=899, width=60, height=21)
        st_2chk131.place(x=1000, y=938, width=60, height=21)
        st_2chk132.place(x=1076, y=938, width=60, height=21)
        st_2chk133.place(x=1150, y=938, width=60, height=21)
        st_2chk134.place(x=1224, y=938, width=60, height=21)
        st_2chk141.place(x=1000, y=977, width=60, height=21)
        st_2chk142.place(x=1076, y=977, width=60, height=21)
        st_2chk143.place(x=1150, y=977, width=60, height=21)
        st_2chk144.place(x=1224, y=977, width=60, height=21)
        st_2chk151.place(x=1000, y=1016, width=60, height=21)
        st_2chk152.place(x=1076, y=1016, width=60, height=21)
        st_2chk153.place(x=1150, y=1016, width=60, height=21)
        st_2chk154.place(x=1224, y=1016, width=60, height=21)
        st_2chk161.place(x=1000, y=1054, width=60, height=21)
        st_2chk162.place(x=1076, y=1054, width=60, height=21)
        st_2chk163.place(x=1150, y=1054, width=60, height=21)
        st_2chk164.place(x=1224, y=1054, width=60, height=21)
        st_2chk171.place(x=1000, y=1092, width=60, height=21)
        st_2chk172.place(x=1076, y=1092, width=60, height=21)
        st_2chk173.place(x=1150, y=1092, width=60, height=21)
        st_2chk174.place(x=1224, y=1092, width=60, height=21)
        st_2chk181.place(x=1000, y=1133, width=60, height=21)
        st_2chk182.place(x=1076, y=1133, width=60, height=21)
        st_2chk183.place(x=1150, y=1133, width=60, height=21)
        st_2chk184.place(x=1224, y=1133, width=60, height=21)
        st_2chk191.place(x=1000, y=1173, width=60, height=21)
        st_2chk192.place(x=1076, y=1173, width=60, height=21)
        st_2chk193.place(x=1150, y=1173, width=60, height=21)
        st_2chk194.place(x=1224, y=1173, width=60, height=21)
        st_2chk201.place(x=1000, y=1213, width=60, height=21)
        st_2chk202.place(x=1076, y=1213, width=60, height=21)
        st_2chk203.place(x=1150, y=1213, width=60, height=21)
        st_2chk204.place(x=1224, y=1213, width=60, height=21)

    def st_2chkFunc2():
        global now
        now = datetime.now()
        global st_2chk
        st_2chk = True
        sheet5['E2'].value = (" % s년 % s월 % s일 % s시 % s분" %
                              (now.year, now.month, now.day, now.hour, now.minute))
        sheet5['K2'].value = idInput.get()

        def st_2chkImage11Func():
            # sumScore()
            if not str(sheet5['C6'].value).strip() == '거의 그렇지 않다':
                sheet5['C6'].value = st_2chk11["text"]

        def st_2chkImage12Func():
            # sumScore()
            if not str(sheet5['C6'].value).strip() == '가끔 그렇다':
                sheet5['C6'].value = st_2chk12["text"]

        def st_2chkImage13Func():
            # sumScore()
            if not str(sheet5['C6'].value).strip() == '자주 그렇다':
                sheet5['C6'].value = st_2chk13["text"]

        def st_2chkImage14Func():
            # sumScore()
            if not str(sheet5['C6'].value).strip() == '거의 언제나 그렇다':
                sheet5['C6'].value = st_2chk14["text"]

        def st_2chkImage21Func():
            if not str(sheet5['C9'].value).strip() == '거의 그렇지 않다':
                sheet5['C9'].value = st_2chk21["text"]

        def st_2chkImage22Func():
            if not str(sheet5['C9'].value).strip() == '가끔 그렇다':
                sheet5['C9'].value = st_2chk22["text"]

        def st_2chkImage23Func():
            if not str(sheet5['C9'].value).strip() == '자주 그렇다':
                sheet5['C9'].value = st_2chk23["text"]

        def st_2chkImage24Func():
            if not str(sheet5['C9'].value).strip() == '거의 언제나 그렇다':
                sheet5['C9'].value = st_2chk24["text"]

        def st_2chkImage31Func():
            if not str(sheet5['C12'].value).strip() == '거의 그렇지 않다':
                sheet5['C12'].value = st_2chk31["text"]

        def st_2chkImage32Func():
            if not str(sheet5['C12'].value).strip() == '가끔 그렇다':
                sheet5['C12'].value = st_2chk32["text"]

        def st_2chkImage33Func():
            if not str(sheet5['C12'].value).strip() == '자주 그렇다':
                sheet5['C12'].value = st_2chk33["text"]

        def st_2chkImage34Func():
            if not str(sheet5['C12'].value).strip() == '거의 언제나 그렇다':
                sheet5['C12'].value = st_2chk34["text"]

        def st_2chkImage41Func():
            if not str(sheet5['C15'].value).strip() == '거의 그렇지 않다':
                sheet5['C15'].value = st_2chk41["text"]

        def st_2chkImage42Func():
            if not str(sheet5['C15'].value).strip() == '가끔 그렇다':
                sheet5['C15'].value = st_2chk42["text"]

        def st_2chkImage43Func():
            if not str(sheet5['C15'].value).strip() == '자주 그렇다':
                sheet5['C15'].value = st_2chk43["text"]

        def st_2chkImage44Func():
            if not str(sheet5['C15'].value).strip() == '거의 언제나 그렇다':
                sheet5['C15'].value = st_2chk44["text"]

        def st_2chkImage51Func():
            if not str(sheet5['C18'].value).strip() == '거의 그렇지 않다':
                sheet5['C18'].value = st_2chk51["text"]

        def st_2chkImage52Func():
            if not str(sheet5['C18'].value).strip() == '가끔 그렇다':
                sheet5['C18'].value = st_2chk52["text"]

        def st_2chkImage53Func():
            if not str(sheet5['C18'].value).strip() == '자주 그렇다':
                sheet5['C18'].value = st_2chk53["text"]

        def st_2chkImage54Func():
            if not str(sheet5['C18'].value).strip() == '거의 언제나 그렇다':
                sheet5['C18'].value = st_2chk54["text"]

        def st_2chkImage61Func():
            if not str(sheet5['C21'].value).strip() == '거의 그렇지 않다':
                sheet5['C21'].value = st_2chk61["text"]

        def st_2chkImage62Func():
            if not str(sheet5['C21'].value).strip() == '가끔 그렇다':
                sheet5['C21'].value = st_2chk62["text"]

        def st_2chkImage63Func():
            if not str(sheet5['C21'].value).strip() == '자주 그렇다':
                sheet5['C21'].value = st_2chk63["text"]

        def st_2chkImage64Func():
            if not str(sheet5['C21'].value).strip() == '거의 언제나 그렇다':
                sheet5['C21'].value = st_2chk64["text"]

        def st_2chkImage71Func():
            if not str(sheet5['C24'].value).strip() == '거의 그렇지 않다':
                sheet5['C24'].value = st_2chk71["text"]

        def st_2chkImage72Func():
            if not str(sheet5['C24'].value).strip() == '가끔 그렇다':
                sheet5['C24'].value = st_2chk72["text"]

        def st_2chkImage73Func():
            if not str(sheet5['C24'].value).strip() == '자주 그렇다':
                sheet5['C24'].value = st_2chk73["text"]

        def st_2chkImage74Func():
            if not str(sheet5['C24'].value).strip() == '거의 언제나 그렇다':
                sheet5['C24'].value = st_2chk74["text"]

        def st_2chkImage81Func():
            if not str(sheet5['C27'].value).strip() == '거의 그렇지 않다':
                sheet5['C27'].value = st_2chk81["text"]

        def st_2chkImage82Func():
            if not str(sheet5['C27'].value).strip() == '가끔 그렇다':
                sheet5['C27'].value = st_2chk82["text"]

        def st_2chkImage83Func():
            if not str(sheet5['C27'].value).strip() == '자주 그렇다':
                sheet5['C27'].value = st_2chk83["text"]

        def st_2chkImage84Func():
            if not str(sheet5['C27'].value).strip() == '거의 언제나 그렇다':
                sheet5['C27'].value = st_2chk84["text"]

        def st_2chkImage91Func():
            if not str(sheet5['C30'].value).strip() == '거의 그렇지 않다':
                sheet5['C30'].value = st_2chk91["text"]

        def st_2chkImage92Func():
            if not str(sheet5['C30'].value).strip() == '가끔 그렇다':
                sheet5['C30'].value = st_2chk92["text"]

        def st_2chkImage93Func():
            if not str(sheet5['C30'].value).strip() == '자주 그렇다':
                sheet5['C30'].value = st_2chk93["text"]

        def st_2chkImage94Func():
            if not str(sheet5['C30'].value).strip() == '거의 언제나 그렇다':
                sheet5['C30'].value = st_2chk94["text"]

        def st_2chkImage101Func():
            if not str(sheet5['C33'].value).strip() == '거의 그렇지 않다':
                sheet5['C33'].value = st_2chk101["text"]

        def st_2chkImage102Func():
            if not str(sheet5['C33'].value).strip() == '가끔 그렇다':
                sheet5['C33'].value = st_2chk102["text"]

        def st_2chkImage103Func():
            if not str(sheet5['C33'].value).strip() == '자주 그렇다':
                sheet5['C33'].value = st_2chk103["text"]

        def st_2chkImage104Func():
            if not str(sheet5['C33'].value).strip() == '거의 언제나 그렇다':
                sheet5['C33'].value = st_2chk104["text"]

        def st_2chkImage111Func():
            if not str(sheet5['C36'].value).strip() == '거의 그렇지 않다':
                sheet5['C36'].value = st_2chk111["text"]

        def st_2chkImage112Func():
            if not str(sheet5['C36'].value).strip() == '가끔 그렇다':
                sheet5['C36'].value = st_2chk112["text"]

        def st_2chkImage113Func():
            if not str(sheet5['C36'].value).strip() == '자주 그렇다':
                sheet5['C36'].value = st_2chk113["text"]

        def st_2chkImage114Func():
            if not str(sheet5['C36'].value).strip() == '거의 언제나 그렇다':
                sheet5['C36'].value = st_2chk114["text"]

        def st_2chkImage121Func():
            if not str(sheet5['C39'].value).strip() == '거의 그렇지 않다':
                sheet5['C39'].value = st_2chk121["text"]

        def st_2chkImage122Func():
            if not str(sheet5['C39'].value).strip() == '가끔 그렇다':
                sheet5['C39'].value = st_2chk122["text"]

        def st_2chkImage123Func():
            if not str(sheet5['C39'].value).strip() == '자주 그렇다':
                sheet5['C39'].value = st_2chk123["text"]

        def st_2chkImage124Func():
            if not str(sheet5['C39'].value).strip() == '거의 언제나 그렇다':
                sheet5['C39'].value = st_2chk124["text"]

        def st_2chkImage131Func():
            if not str(sheet5['C42'].value).strip() == '거의 그렇지 않다':
                sheet5['C42'].value = st_2chk131["text"]

        def st_2chkImage132Func():
            if not str(sheet5['C42'].value).strip() == '가끔 그렇다':
                sheet5['C42'].value = st_2chk132["text"]

        def st_2chkImage133Func():
            if not str(sheet5['C42'].value).strip() == '자주 그렇다':
                sheet5['C42'].value = st_2chk133["text"]

        def st_2chkImage134Func():
            if not str(sheet5['C42'].value).strip() == '거의 언제나 그렇다':
                sheet5['C42'].value = st_2chk134["text"]

        def st_2chkImage141Func():
            if not str(sheet5['C45'].value).strip() == '거의 그렇지 않다':
                sheet5['C45'].value = st_2chk141["text"]

        def st_2chkImage142Func():
            if not str(sheet5['C45'].value).strip() == '가끔 그렇다':
                sheet5['C45'].value = st_2chk142["text"]

        def st_2chkImage143Func():
            if not str(sheet5['C45'].value).strip() == '자주 그렇다':
                sheet5['C45'].value = st_2chk143["text"]

        def st_2chkImage144Func():
            if not str(sheet5['C45'].value).strip() == '거의 언제나 그렇다':
                sheet5['C45'].value = st_2chk144["text"]

        def st_2chkImage151Func():
            if not str(sheet5['C48'].value).strip() == '거의 그렇지 않다':
                sheet5['C48'].value = st_2chk151["text"]

        def st_2chkImage152Func():
            if not str(sheet5['C48'].value).strip() == '가끔 그렇다':
                sheet5['C48'].value = st_2chk152["text"]

        def st_2chkImage153Func():
            if not str(sheet5['C48'].value).strip() == '자주 그렇다':
                sheet5['C48'].value = st_2chk153["text"]

        def st_2chkImage154Func():
            if not str(sheet5['C48'].value).strip() == '거의 언제나 그렇다':
                sheet5['C48'].value = st_2chk154["text"]

        def st_2chkImage161Func():
            if not str(sheet5['C51'].value).strip() == '거의 그렇지 않다':
                sheet5['C51'].value = st_2chk161["text"]

        def st_2chkImage162Func():
            if not str(sheet5['C51'].value).strip() == '가끔 그렇다':
                sheet5['C51'].value = st_2chk162["text"]

        def st_2chkImage163Func():
            if not str(sheet5['C51'].value).strip() == '자주 그렇다':
                sheet5['C51'].value = st_2chk163["text"]

        def st_2chkImage164Func():
            if not str(sheet5['C51'].value).strip() == '거의 언제나 그렇다':
                sheet5['C51'].value = st_2chk164["text"]

        def st_2chkImage171Func():
            if not str(sheet5['C54'].value).strip() == '거의 그렇지 않다':
                sheet5['C54'].value = st_2chk171["text"]

        def st_2chkImage172Func():
            if not str(sheet5['C54'].value).strip() == '가끔 그렇다':
                sheet5['C54'].value = st_2chk172["text"]

        def st_2chkImage173Func():
            if not str(sheet5['C54'].value).strip() == '자주 그렇다':
                sheet5['C54'].value = st_2chk173["text"]

        def st_2chkImage174Func():
            if not str(sheet5['C54'].value).strip() == '거의 언제나 그렇다':
                sheet5['C54'].value = st_2chk174["text"]

        def st_2chkImage181Func():
            if not str(sheet5['C57'].value).strip() == '거의 그렇지 않다':
                sheet5['C57'].value = st_2chk181["text"]

        def st_2chkImage182Func():
            if not str(sheet5['C57'].value).strip() == '가끔 그렇다':
                sheet5['C57'].value = st_2chk182["text"]

        def st_2chkImage183Func():
            if not str(sheet5['C57'].value).strip() == '자주 그렇다':
                sheet5['C57'].value = st_2chk183["text"]

        def st_2chkImage184Func():
            if not str(sheet5['C57'].value).strip() == '거의 언제나 그렇다':
                sheet5['C57'].value = st_2chk184["text"]

        def st_2chkImage191Func():
            if not str(sheet5['C60'].value).strip() == '거의 그렇지 않다':
                sheet5['C60'].value = st_2chk181["text"]

        def st_2chkImage192Func():
            if not str(sheet5['C60'].value).strip() == '가끔 그렇다':
                sheet5['C60'].value = st_2chk182["text"]

        def st_2chkImage193Func():
            if not str(sheet5['C60'].value).strip() == '자주 그렇다':
                sheet5['C60'].value = st_2chk183["text"]

        def st_2chkImage194Func():
            if not str(sheet5['C60'].value).strip() == '거의 언제나 그렇다':
                sheet5['C60'].value = st_2chk184["text"]

        def st_2chkImage201Func():
            if not str(sheet5['C63'].value).strip() == '거의 그렇지 않다':
                sheet5['C63'].value = st_2chk181["text"]

        def st_2chkImage202Func():
            if not str(sheet5['C63'].value).strip() == '가끔 그렇다':
                sheet5['C63'].value = st_2chk182["text"]

        def st_2chkImage203Func():
            if not str(sheet5['C63'].value).strip() == '자주 그렇다':
                sheet5['C63'].value = st_2chk183["text"]

        def st_2chkImage204Func():
            if not str(sheet5['C63'].value).strip() == '거의 언제나 그렇다':
                sheet5['C63'].value = st_2chk184["text"]

        frame5St_2BgImg = Image.open("images/st_2bg.png")
        frame5St_2Bg = ImageTk.PhotoImage(frame5St_2BgImg)
        frame5St_2BgLabel = tkinter.Label(
            displayPage2.inner, image=frame5St_2Bg)
        frame5St_2BgLabel.image = frame5St_2Bg
        frame5St_2BgLabel.place(x=47, y=420, height=1330)
        frame5St_2BgLabel.pack()

        global st_2chk11
        st_2chk11 = Radiobutton(displayPage2.inner, value=4, text="거의 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType1, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage11Func)
        st_2chk11.deselect()
        global st_2chk12
        st_2chk12 = Radiobutton(displayPage2.inner, value=3, text="가끔 그렇다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=st_2chkType1, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage12Func)
        st_2chk12.deselect()
        global st_2chk13
        st_2chk13 = Radiobutton(displayPage2.inner, value=2, text="자주 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=st_2chkType1, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage13Func)
        st_2chk13.deselect()
        global st_2chk14
        st_2chk14 = Radiobutton(displayPage2.inner, value=1, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=st_2chkType1, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage14Func)
        st_2chk14.deselect()
        global st_2chk21
        st_2chk21 = Radiobutton(displayPage2.inner, value=1, text="거의 그렇지 않다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType2, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage21Func)
        st_2chk21.deselect()
        global st_2chk22
        st_2chk22 = Radiobutton(displayPage2.inner, value=2, text="가끔 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=st_2chkType2, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage22Func)
        st_2chk22.deselect()
        global st_2chk23
        st_2chk23 = Radiobutton(displayPage2.inner, value=3, text="자주 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=st_2chkType2, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage23Func)
        st_2chk23.deselect()
        global st_2chk24
        st_2chk24 = Radiobutton(displayPage2.inner, value=4, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=st_2chkType2, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage24Func)
        st_2chk24.deselect()
        global st_2chk31
        st_2chk31 = Radiobutton(displayPage2.inner, value=1, text="거의 그렇지 않다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType3, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage31Func)
        st_2chk31.deselect()
        global st_2chk32
        st_2chk32 = Radiobutton(displayPage2.inner, value=2, text="가끔 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=st_2chkType3, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage32Func)
        st_2chk32.deselect()
        global st_2chk33
        st_2chk33 = Radiobutton(displayPage2.inner, value=3, text="자주 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=st_2chkType3, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage33Func)
        st_2chk33.deselect()
        global st_2chk34
        st_2chk34 = Radiobutton(displayPage2.inner, value=4, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=st_2chkType3, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage34Func)
        st_2chk34.deselect()
        global st_2chk41
        st_2chk41 = Radiobutton(displayPage2.inner, value=1, text="거의 그렇지 않다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType4, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage41Func)
        st_2chk41.deselect()
        global st_2chk42
        st_2chk42 = Radiobutton(displayPage2.inner, value=2, text="가끔 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=st_2chkType4, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage42Func)
        st_2chk42.deselect()
        global st_2chk43
        st_2chk43 = Radiobutton(displayPage2.inner, value=3, text="자주 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=st_2chkType4, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage43Func)
        st_2chk43.deselect()
        global st_2chk44
        st_2chk44 = Radiobutton(displayPage2.inner, value=4, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=st_2chkType4, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage44Func)
        st_2chk44.deselect()
        global st_2chk51
        st_2chk51 = Radiobutton(displayPage2.inner, value=1, text="거의 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType5, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage51Func)
        st_2chk51.deselect()
        global st_2chk52
        st_2chk52 = Radiobutton(displayPage2.inner, value=2, text="가끔 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=st_2chkType5, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage52Func)
        st_2chk52.deselect()
        global st_2chk53
        st_2chk53 = Radiobutton(displayPage2.inner, value=3, text="자주 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=st_2chkType5, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage53Func)
        st_2chk53.deselect()
        global st_2chk54
        st_2chk54 = Radiobutton(displayPage2.inner, value=4, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=st_2chkType5, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage54Func)
        st_2chk54.deselect()
        global st_2chk61
        st_2chk61 = Radiobutton(displayPage2.inner, value=4, text="거의 그렇지 않다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType6, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage61Func)
        st_2chk61.deselect()
        global st_2chk62
        st_2chk62 = Radiobutton(displayPage2.inner, value=3, text="가끔 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=st_2chkType6, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage62Func)
        st_2chk62.deselect()
        global st_2chk63
        st_2chk63 = Radiobutton(displayPage2.inner, value=2, text="자주 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=st_2chkType6, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage63Func)
        st_2chk63.deselect()
        global st_2chk64
        st_2chk64 = Radiobutton(displayPage2.inner, value=1, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=st_2chkType6, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage64Func)
        st_2chk64.deselect()

        global st_2chk71
        st_2chk71 = Radiobutton(displayPage2.inner, value=4, text="거의 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType7, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage71Func)
        st_2chk71.deselect()
        global st_2chk72
        st_2chk72 = Radiobutton(displayPage2.inner, value=3, text="가끔 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=st_2chkType7, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage72Func)
        st_2chk72.deselect()
        global st_2chk73
        st_2chk73 = Radiobutton(displayPage2.inner, value=2, text="자주 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=st_2chkType7, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage73Func)
        st_2chk73.deselect()
        global st_2chk74
        st_2chk74 = Radiobutton(displayPage2.inner, value=1, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=st_2chkType7, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage74Func)
        st_2chk74.deselect()

        global st_2chk81
        st_2chk81 = Radiobutton(displayPage2.inner, value=1, text="거의 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType8, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage81Func)
        st_2chk81.deselect()
        global st_2chk82
        st_2chk82 = Radiobutton(displayPage2.inner, value=2, text="가끔 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=st_2chkType8, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage82Func)
        st_2chk82.deselect()
        global st_2chk83
        st_2chk83 = Radiobutton(displayPage2.inner, value=3, text="자주 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=st_2chkType8, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage83Func)
        st_2chk83.deselect()
        global st_2chk84
        st_2chk84 = Radiobutton(displayPage2.inner, value=4, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=st_2chkType8, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage84Func)
        st_2chk84.deselect()

        global st_2chk91
        st_2chk91 = Radiobutton(displayPage2.inner, value=1, text="거의 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType9, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage91Func)
        st_2chk91.deselect()
        global st_2chk92
        st_2chk92 = Radiobutton(displayPage2.inner, value=2, text="가끔 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=st_2chkType9, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage92Func)
        st_2chk92.deselect()
        global st_2chk93
        st_2chk93 = Radiobutton(displayPage2.inner, value=3, text="자주 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=st_2chkType9, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage93Func)
        st_2chk93.deselect()
        global st_2chk94
        st_2chk94 = Radiobutton(displayPage2.inner, value=4, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=st_2chkType9, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage94Func)
        st_2chk94.deselect()
        global st_2chk101
        st_2chk101 = Radiobutton(displayPage2.inner, value=4, text="거의 그렇지 않다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType10, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage101Func)
        st_2chk101.deselect()
        global st_2chk102
        st_2chk102 = Radiobutton(displayPage2.inner, value=3, text="가끔 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_2chkType10, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage102Func)
        st_2chk102.deselect()
        global st_2chk103
        st_2chk103 = Radiobutton(displayPage2.inner, value=2, text="자주 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_2chkType10, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage103Func)
        st_2chk103.deselect()
        global st_2chk104
        st_2chk104 = Radiobutton(displayPage2.inner, value=1, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_2chkType10, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage104Func)
        st_2chk104.deselect()
        global st_2chk111
        st_2chk111 = Radiobutton(displayPage2.inner, value=1, text="거의 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType11, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage111Func)
        st_2chk111.deselect()
        global st_2chk112
        st_2chk112 = Radiobutton(displayPage2.inner, value=2, text="가끔 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_2chkType11, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage112Func)
        st_2chk112.deselect()
        global st_2chk113
        st_2chk113 = Radiobutton(displayPage2.inner, value=3, text="자주 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_2chkType11, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage113Func)
        st_2chk113.deselect()
        global st_2chk114
        st_2chk114 = Radiobutton(displayPage2.inner, value=4, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_2chkType11, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage114Func)
        st_2chk114.deselect()
        global st_2chk121
        st_2chk121 = Radiobutton(displayPage2.inner, value=1, text="거의 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType12, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage121Func)
        st_2chk121.deselect()
        global st_2chk122
        st_2chk122 = Radiobutton(displayPage2.inner, value=2, text="가끔 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_2chkType12, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage122Func)
        st_2chk122.deselect()
        global st_2chk123
        st_2chk123 = Radiobutton(displayPage2.inner, value=3, text="자주 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_2chkType12, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage123Func)
        st_2chk123.deselect()
        global st_2chk124
        st_2chk124 = Radiobutton(displayPage2.inner, value=4, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_2chkType12, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage124Func)
        st_2chk124.deselect()
        global st_2chk131
        st_2chk131 = Radiobutton(displayPage2.inner, value=4, text="거의 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType13, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage131Func)
        st_2chk131.deselect()
        global st_2chk132
        st_2chk132 = Radiobutton(displayPage2.inner, value=3, text="가끔 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_2chkType13, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage132Func)
        st_2chk132.deselect()
        global st_2chk133
        st_2chk133 = Radiobutton(displayPage2.inner, value=2, text="자주 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_2chkType13, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage133Func)
        st_2chk133.deselect()
        global st_2chk134
        st_2chk134 = Radiobutton(displayPage2.inner, value=1, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_2chkType13, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage134Func)
        st_2chk134.deselect()
        global st_2chk141
        st_2chk141 = Radiobutton(displayPage2.inner, value=1, text="거의 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType14, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage141Func)
        st_2chk141.deselect()
        global st_2chk142
        st_2chk142 = Radiobutton(displayPage2.inner, value=2, text="가끔 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_2chkType14, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage142Func)
        st_2chk142.deselect()
        global st_2chk143
        st_2chk143 = Radiobutton(displayPage2.inner, value=3, text="자주 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_2chkType14, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage143Func)
        st_2chk143.deselect()
        global st_2chk144
        st_2chk144 = Radiobutton(displayPage2.inner, value=4, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_2chkType14, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage144Func)
        st_2chk144.deselect()
        global st_2chk151
        st_2chk151 = Radiobutton(displayPage2.inner, value=1, text="거의 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType15, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage151Func)
        st_2chk151.deselect()
        global st_2chk152
        st_2chk152 = Radiobutton(displayPage2.inner, value=2, text="가끔 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_2chkType15, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage152Func)
        st_2chk152.deselect()
        global st_2chk153
        st_2chk153 = Radiobutton(displayPage2.inner, value=3, text="자주 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_2chkType15, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage153Func)
        st_2chk153.deselect()
        global st_2chk154
        st_2chk154 = Radiobutton(displayPage2.inner, value=4, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_2chkType15, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage154Func)
        st_2chk154.deselect()
        global st_2chk161
        st_2chk161 = Radiobutton(displayPage2.inner, value=4, text="거의 그렇지 않다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType16, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage161Func)
        st_2chk161.deselect()
        global st_2chk162
        st_2chk162 = Radiobutton(displayPage2.inner, value=3, text="가끔 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_2chkType16, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage162Func)
        st_2chk162.deselect()
        global st_2chk163
        st_2chk163 = Radiobutton(displayPage2.inner, value=2, text="자주 그렇다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_2chkType16, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage163Func)
        st_2chk163.deselect()
        global st_2chk164
        st_2chk164 = Radiobutton(displayPage2.inner, value=1, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_2chkType16, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage164Func)
        st_2chk164.deselect()
        global st_2chk171
        st_2chk171 = Radiobutton(displayPage2.inner, value=1, text="거의 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType17, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage171Func)
        st_2chk171.deselect()
        global st_2chk172
        st_2chk172 = Radiobutton(displayPage2.inner, value=2, text="가끔 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_2chkType17, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage172Func)
        st_2chk172.deselect()
        global st_2chk173
        st_2chk173 = Radiobutton(displayPage2.inner, value=3, text="자주 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_2chkType17, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage173Func)
        st_2chk173.deselect()
        global st_2chk174
        st_2chk174 = Radiobutton(displayPage2.inner, value=4, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_2chkType17, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage174Func)
        st_2chk174.deselect()
        global st_2chk181
        st_2chk181 = Radiobutton(displayPage2.inner, value=1, text="거의 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType18, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage181Func)
        st_2chk181.deselect()
        global st_2chk182
        st_2chk182 = Radiobutton(displayPage2.inner, value=2, text="가끔 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_2chkType18, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage182Func)
        st_2chk182.deselect()
        global st_2chk183
        st_2chk183 = Radiobutton(displayPage2.inner, value=3, text="자주 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_2chkType18, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage183Func)
        st_2chk183.deselect()
        global st_2chk184
        st_2chk184 = Radiobutton(displayPage2.inner, value=4, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_2chkType18, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage184Func)
        st_2chk184.deselect()
        global st_2chk191
        st_2chk191 = Radiobutton(displayPage2.inner, value=4, text="거의 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType19, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage191Func)
        st_2chk191.deselect()
        global st_2chk192
        st_2chk192 = Radiobutton(displayPage2.inner, value=3, text="가끔 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_2chkType19, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage192Func)
        st_2chk192.deselect()
        global st_2chk193
        st_2chk193 = Radiobutton(displayPage2.inner, value=2, text="자주 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_2chkType19, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage193Func)
        st_2chk193.deselect()
        global st_2chk194
        st_2chk194 = Radiobutton(displayPage2.inner, value=1, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_2chkType19, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage194Func)
        st_2chk194.deselect()
        global st_2chk201
        st_2chk201 = Radiobutton(displayPage2.inner, value=1, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_2chkType20, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage201Func)
        st_2chk201.deselect()
        global st_2chk202
        st_2chk202 = Radiobutton(displayPage2.inner, value=2, text="가끔 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_2chkType20, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage202Func)
        st_2chk202.deselect()
        global st_2chk203
        st_2chk203 = Radiobutton(displayPage2.inner, value=3, text="자주 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_2chkType20, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage203Func)
        st_2chk203.deselect()
        global st_2chk204
        st_2chk204 = Radiobutton(displayPage2.inner, value=4, text="거의 언제나 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_2chkType20, font=malgungothic13, bd=0, highlightthickness=0, command=st_2chkImage204Func)
        st_2chk204.deselect()
        st_2chk11.place(x=1000, y=472+addheight, width=60, height=21)
        st_2chk12.place(x=1076, y=472+addheight, width=60, height=21)
        st_2chk13.place(x=1150, y=472+addheight, width=60, height=21)
        st_2chk14.place(x=1224, y=472+addheight, width=60, height=21)
        st_2chk21.place(x=1000, y=512+addheight, width=60, height=21)
        st_2chk22.place(x=1076, y=512+addheight, width=60, height=21)
        st_2chk23.place(x=1150, y=512+addheight, width=60, height=21)
        st_2chk24.place(x=1224, y=512+addheight, width=60, height=21)
        st_2chk31.place(x=1000, y=551+addheight, width=60, height=21)
        st_2chk32.place(x=1076, y=551+addheight, width=60, height=21)
        st_2chk33.place(x=1150, y=551+addheight, width=60, height=21)
        st_2chk34.place(x=1224, y=551+addheight, width=60, height=21)
        st_2chk41.place(x=1000, y=589+addheight, width=60, height=21)
        st_2chk42.place(x=1076, y=589+addheight, width=60, height=21)
        st_2chk43.place(x=1150, y=589+addheight, width=60, height=21)
        st_2chk44.place(x=1224, y=589+addheight, width=60, height=21)
        st_2chk51.place(x=1000, y=628+addheight, width=60, height=21)
        st_2chk52.place(x=1076, y=628+addheight, width=60, height=21)
        st_2chk53.place(x=1150, y=628+addheight, width=60, height=21)
        st_2chk54.place(x=1224, y=628+addheight, width=60, height=21)
        st_2chk61.place(x=1000, y=667+addheight, width=60, height=21)
        st_2chk62.place(x=1076, y=667+addheight, width=60, height=21)
        st_2chk63.place(x=1150, y=667+addheight, width=60, height=21)
        st_2chk64.place(x=1224, y=667+addheight, width=60, height=21)
        st_2chk71.place(x=1000, y=704+addheight, width=60, height=21)
        st_2chk72.place(x=1076, y=704+addheight, width=60, height=21)
        st_2chk73.place(x=1150, y=704+addheight, width=60, height=21)
        st_2chk74.place(x=1224, y=704+addheight, width=60, height=21)
        st_2chk81.place(x=1000, y=742+addheight, width=60, height=21)
        st_2chk82.place(x=1076, y=742+addheight, width=60, height=21)
        st_2chk83.place(x=1150, y=742+addheight, width=60, height=21)
        st_2chk84.place(x=1224, y=742+addheight, width=60, height=21)
        st_2chk91.place(x=1000, y=781+addheight, width=60, height=21)
        st_2chk92.place(x=1076, y=781+addheight, width=60, height=21)
        st_2chk93.place(x=1150, y=781+addheight, width=60, height=21)
        st_2chk94.place(x=1224, y=781+addheight, width=60, height=21)
        st_2chk101.place(x=1000, y=821+addheight, width=60, height=21)
        st_2chk102.place(x=1076, y=821+addheight, width=60, height=21)
        st_2chk103.place(x=1150, y=821+addheight, width=60, height=21)
        st_2chk104.place(x=1224, y=821+addheight, width=60, height=21)
        st_2chk111.place(x=1000, y=860+addheight, width=60, height=21)
        st_2chk112.place(x=1076, y=860+addheight, width=60, height=21)
        st_2chk113.place(x=1150, y=860+addheight, width=60, height=21)
        st_2chk114.place(x=1224, y=860+addheight, width=60, height=21)
        st_2chk121.place(x=1000, y=899+addheight, width=60, height=21)
        st_2chk122.place(x=1076, y=899+addheight, width=60, height=21)
        st_2chk123.place(x=1150, y=899+addheight, width=60, height=21)
        st_2chk124.place(x=1224, y=899+addheight, width=60, height=21)
        st_2chk131.place(x=1000, y=938+addheight, width=60, height=21)
        st_2chk132.place(x=1076, y=938+addheight, width=60, height=21)
        st_2chk133.place(x=1150, y=938+addheight, width=60, height=21)
        st_2chk134.place(x=1224, y=938+addheight, width=60, height=21)
        st_2chk141.place(x=1000, y=977+addheight, width=60, height=21)
        st_2chk142.place(x=1076, y=977+addheight, width=60, height=21)
        st_2chk143.place(x=1150, y=977+addheight, width=60, height=21)
        st_2chk144.place(x=1224, y=977+addheight, width=60, height=21)
        st_2chk151.place(x=1000, y=1016+addheight, width=60, height=21)
        st_2chk152.place(x=1076, y=1016+addheight, width=60, height=21)
        st_2chk153.place(x=1150, y=1016+addheight, width=60, height=21)
        st_2chk154.place(x=1224, y=1016+addheight, width=60, height=21)
        st_2chk161.place(x=1000, y=1054+addheight, width=60, height=21)
        st_2chk162.place(x=1076, y=1054+addheight, width=60, height=21)
        st_2chk163.place(x=1150, y=1054+addheight, width=60, height=21)
        st_2chk164.place(x=1224, y=1054+addheight, width=60, height=21)
        st_2chk171.place(x=1000, y=1092+addheight, width=60, height=21)
        st_2chk172.place(x=1076, y=1092+addheight, width=60, height=21)
        st_2chk173.place(x=1150, y=1092+addheight, width=60, height=21)
        st_2chk174.place(x=1224, y=1092+addheight, width=60, height=21)
        st_2chk181.place(x=1000, y=1133+addheight, width=60, height=21)
        st_2chk182.place(x=1076, y=1133+addheight, width=60, height=21)
        st_2chk183.place(x=1150, y=1133+addheight, width=60, height=21)
        st_2chk184.place(x=1224, y=1133+addheight, width=60, height=21)
        st_2chk191.place(x=1000, y=1173+addheight, width=60, height=21)
        st_2chk192.place(x=1076, y=1173+addheight, width=60, height=21)
        st_2chk193.place(x=1150, y=1173+addheight, width=60, height=21)
        st_2chk194.place(x=1224, y=1173+addheight, width=60, height=21)
        st_2chk201.place(x=1000, y=1213+addheight, width=60, height=21)
        st_2chk202.place(x=1076, y=1213+addheight, width=60, height=21)
        st_2chk203.place(x=1150, y=1213+addheight, width=60, height=21)
        st_2chk204.place(x=1224, y=1213+addheight, width=60, height=21)

    def st_1chkFunc():
        global now
        now = datetime.now()
        global st_1chk
        st_1chk = True
        sheet6['E2'].value = (" % s년 % s월 % s일 % s시 % s분" %
                              (now.year, now.month, now.day, now.hour, now.minute))
        sheet6['K2'].value = idInput.get()

        def st_1chkImage11Func():
            # sumScore()
            if not str(sheet6['C6'].value).strip() == '전혀 그렇지 않다':
                sheet6['C6'].value = st_1chk11["text"]

        def st_1chkImage12Func():
            # sumScore()
            if not str(sheet6['C6'].value).strip() == '조금 그렇다':
                sheet6['C6'].value = st_1chk12["text"]

        def st_1chkImage13Func():
            # sumScore()
            if not str(sheet6['C6'].value).strip() == '보통으로 그렇다':
                sheet6['C6'].value = st_1chk13["text"]

        def st_1chkImage14Func():
            # sumScore()
            if not str(sheet6['C6'].value).strip() == '대단히 그렇다':
                sheet6['C6'].value = st_1chk14["text"]

        def st_1chkImage21Func():
            if not str(sheet6['C9'].value).strip() == '전혀 그렇지 않다':
                sheet6['C9'].value = st_1chk21["text"]

        def st_1chkImage22Func():
            if not str(sheet6['C9'].value).strip() == '조금 그렇다':
                sheet6['C9'].value = st_1chk22["text"]

        def st_1chkImage23Func():
            if not str(sheet6['C9'].value).strip() == '보통으로 그렇다':
                sheet6['C9'].value = st_1chk23["text"]

        def st_1chkImage24Func():
            if not str(sheet6['C9'].value).strip() == '대단히 그렇다':
                sheet6['C9'].value = st_1chk24["text"]

        def st_1chkImage31Func():
            if not str(sheet6['C12'].value).strip() == '전혀 그렇지 않다':
                sheet6['C12'].value = st_1chk31["text"]

        def st_1chkImage32Func():
            if not str(sheet6['C12'].value).strip() == '조금 그렇다':
                sheet6['C12'].value = st_1chk32["text"]

        def st_1chkImage33Func():
            if not str(sheet6['C12'].value).strip() == '보통으로 그렇다':
                sheet6['C12'].value = st_1chk33["text"]

        def st_1chkImage34Func():
            if not str(sheet6['C12'].value).strip() == '대단히 그렇다':
                sheet6['C12'].value = st_1chk34["text"]

        def st_1chkImage41Func():
            if not str(sheet6['C15'].value).strip() == '전혀 그렇지 않다':
                sheet6['C15'].value = st_1chk41["text"]

        def st_1chkImage42Func():
            if not str(sheet6['C15'].value).strip() == '조금 그렇다':
                sheet6['C15'].value = st_1chk42["text"]

        def st_1chkImage43Func():
            if not str(sheet6['C15'].value).strip() == '보통으로 그렇다':
                sheet6['C15'].value = st_1chk43["text"]

        def st_1chkImage44Func():
            if not str(sheet6['C15'].value).strip() == '대단히 그렇다':
                sheet6['C15'].value = st_1chk44["text"]

        def st_1chkImage51Func():
            if not str(sheet6['C18'].value).strip() == '전혀 그렇지 않다':
                sheet6['C18'].value = st_1chk51["text"]

        def st_1chkImage52Func():
            if not str(sheet6['C18'].value).strip() == '조금 그렇다':
                sheet6['C18'].value = st_1chk52["text"]

        def st_1chkImage53Func():
            if not str(sheet6['C18'].value).strip() == '보통으로 그렇다':
                sheet6['C18'].value = st_1chk53["text"]

        def st_1chkImage54Func():
            if not str(sheet6['C18'].value).strip() == '대단히 그렇다':
                sheet6['C18'].value = st_1chk54["text"]

        def st_1chkImage61Func():
            if not str(sheet6['C21'].value).strip() == '전혀 그렇지 않다':
                sheet6['C21'].value = st_1chk61["text"]

        def st_1chkImage62Func():
            if not str(sheet6['C21'].value).strip() == '조금 그렇다':
                sheet6['C21'].value = st_1chk62["text"]

        def st_1chkImage63Func():
            if not str(sheet6['C21'].value).strip() == '보통으로 그렇다':
                sheet6['C21'].value = st_1chk63["text"]

        def st_1chkImage64Func():
            if not str(sheet6['C21'].value).strip() == '대단히 그렇다':
                sheet6['C21'].value = st_1chk64["text"]

        def st_1chkImage71Func():
            if not str(sheet6['C24'].value).strip() == '전혀 그렇지 않다':
                sheet6['C24'].value = st_1chk71["text"]

        def st_1chkImage72Func():
            if not str(sheet6['C24'].value).strip() == '조금 그렇다':
                sheet6['C24'].value = st_1chk72["text"]

        def st_1chkImage73Func():
            if not str(sheet6['C24'].value).strip() == '보통으로 그렇다':
                sheet6['C24'].value = st_1chk73["text"]

        def st_1chkImage74Func():
            if not str(sheet6['C24'].value).strip() == '대단히 그렇다':
                sheet6['C24'].value = st_1chk74["text"]

        def st_1chkImage81Func():
            if not str(sheet6['C27'].value).strip() == '전혀 그렇지 않다':
                sheet6['C27'].value = st_1chk81["text"]

        def st_1chkImage82Func():
            if not str(sheet6['C27'].value).strip() == '조금 그렇다':
                sheet6['C27'].value = st_1chk82["text"]

        def st_1chkImage83Func():
            if not str(sheet6['C27'].value).strip() == '보통으로 그렇다':
                sheet6['C27'].value = st_1chk83["text"]

        def st_1chkImage84Func():
            if not str(sheet6['C27'].value).strip() == '대단히 그렇다':
                sheet6['C27'].value = st_1chk84["text"]

        def st_1chkImage91Func():
            if not str(sheet6['C30'].value).strip() == '전혀 그렇지 않다':
                sheet6['C30'].value = st_1chk91["text"]

        def st_1chkImage92Func():
            if not str(sheet6['C30'].value).strip() == '조금 그렇다':
                sheet6['C30'].value = st_1chk92["text"]

        def st_1chkImage93Func():
            if not str(sheet6['C30'].value).strip() == '보통으로 그렇다':
                sheet6['C30'].value = st_1chk93["text"]

        def st_1chkImage94Func():
            if not str(sheet6['C30'].value).strip() == '대단히 그렇다':
                sheet6['C30'].value = st_1chk94["text"]

        def st_1chkImage101Func():
            if not str(sheet6['C33'].value).strip() == '전혀 그렇지 않다':
                sheet6['C33'].value = st_1chk101["text"]

        def st_1chkImage102Func():
            if not str(sheet6['C33'].value).strip() == '조금 그렇다':
                sheet6['C33'].value = st_1chk102["text"]

        def st_1chkImage103Func():
            if not str(sheet6['C33'].value).strip() == '보통으로 그렇다':
                sheet6['C33'].value = st_1chk103["text"]

        def st_1chkImage104Func():
            if not str(sheet6['C33'].value).strip() == '대단히 그렇다':
                sheet6['C33'].value = st_1chk104["text"]

        def st_1chkImage111Func():
            if not str(sheet6['C36'].value).strip() == '전혀 그렇지 않다':
                sheet6['C36'].value = st_1chk111["text"]

        def st_1chkImage112Func():
            if not str(sheet6['C36'].value).strip() == '조금 그렇다':
                sheet6['C36'].value = st_1chk112["text"]

        def st_1chkImage113Func():
            if not str(sheet6['C36'].value).strip() == '보통으로 그렇다':
                sheet6['C36'].value = st_1chk113["text"]

        def st_1chkImage114Func():
            if not str(sheet6['C36'].value).strip() == '대단히 그렇다':
                sheet6['C36'].value = st_1chk114["text"]

        def st_1chkImage121Func():
            if not str(sheet6['C39'].value).strip() == '전혀 그렇지 않다':
                sheet6['C39'].value = st_1chk121["text"]

        def st_1chkImage122Func():
            if not str(sheet6['C39'].value).strip() == '조금 그렇다':
                sheet6['C39'].value = st_1chk122["text"]

        def st_1chkImage123Func():
            if not str(sheet6['C39'].value).strip() == '보통으로 그렇다':
                sheet6['C39'].value = st_1chk123["text"]

        def st_1chkImage124Func():
            if not str(sheet6['C39'].value).strip() == '대단히 그렇다':
                sheet6['C39'].value = st_1chk124["text"]

        def st_1chkImage131Func():
            if not str(sheet6['C42'].value).strip() == '전혀 그렇지 않다':
                sheet6['C42'].value = st_1chk131["text"]

        def st_1chkImage132Func():
            if not str(sheet6['C42'].value).strip() == '조금 그렇다':
                sheet6['C42'].value = st_1chk132["text"]

        def st_1chkImage133Func():
            if not str(sheet6['C42'].value).strip() == '보통으로 그렇다':
                sheet6['C42'].value = st_1chk133["text"]

        def st_1chkImage134Func():
            if not str(sheet6['C42'].value).strip() == '대단히 그렇다':
                sheet6['C42'].value = st_1chk134["text"]

        def st_1chkImage141Func():
            if not str(sheet6['C45'].value).strip() == '전혀 그렇지 않다':
                sheet6['C45'].value = st_1chk141["text"]

        def st_1chkImage142Func():
            if not str(sheet6['C45'].value).strip() == '조금 그렇다':
                sheet6['C45'].value = st_1chk142["text"]

        def st_1chkImage143Func():
            if not str(sheet6['C45'].value).strip() == '보통으로 그렇다':
                sheet6['C45'].value = st_1chk143["text"]

        def st_1chkImage144Func():
            if not str(sheet6['C45'].value).strip() == '대단히 그렇다':
                sheet6['C45'].value = st_1chk144["text"]

        def st_1chkImage151Func():
            if not str(sheet6['C48'].value).strip() == '전혀 그렇지 않다':
                sheet6['C48'].value = st_1chk151["text"]

        def st_1chkImage152Func():
            if not str(sheet6['C48'].value).strip() == '조금 그렇다':
                sheet6['C48'].value = st_1chk152["text"]

        def st_1chkImage153Func():
            if not str(sheet6['C48'].value).strip() == '보통으로 그렇다':
                sheet6['C48'].value = st_1chk153["text"]

        def st_1chkImage154Func():
            if not str(sheet6['C48'].value).strip() == '대단히 그렇다':
                sheet6['C48'].value = st_1chk154["text"]

        def st_1chkImage161Func():
            if not str(sheet6['C51'].value).strip() == '전혀 그렇지 않다':
                sheet6['C51'].value = st_1chk161["text"]

        def st_1chkImage162Func():
            if not str(sheet6['C51'].value).strip() == '조금 그렇다':
                sheet6['C51'].value = st_1chk162["text"]

        def st_1chkImage163Func():
            if not str(sheet6['C51'].value).strip() == '보통으로 그렇다':
                sheet6['C51'].value = st_1chk163["text"]

        def st_1chkImage164Func():
            if not str(sheet6['C51'].value).strip() == '대단히 그렇다':
                sheet6['C51'].value = st_1chk164["text"]

        def st_1chkImage171Func():
            if not str(sheet6['C54'].value).strip() == '전혀 그렇지 않다':
                sheet6['C54'].value = st_1chk171["text"]

        def st_1chkImage172Func():
            if not str(sheet6['C54'].value).strip() == '조금 그렇다':
                sheet6['C54'].value = st_1chk172["text"]

        def st_1chkImage173Func():
            if not str(sheet6['C54'].value).strip() == '보통으로 그렇다':
                sheet6['C54'].value = st_1chk173["text"]

        def st_1chkImage174Func():
            if not str(sheet6['C54'].value).strip() == '대단히 그렇다':
                sheet6['C54'].value = st_1chk174["text"]

        def st_1chkImage181Func():
            if not str(sheet6['C57'].value).strip() == '전혀 그렇지 않다':
                sheet6['C57'].value = st_1chk181["text"]

        def st_1chkImage182Func():
            if not str(sheet6['C57'].value).strip() == '조금 그렇다':
                sheet6['C57'].value = st_1chk182["text"]

        def st_1chkImage183Func():
            if not str(sheet6['C57'].value).strip() == '보통으로 그렇다':
                sheet6['C57'].value = st_1chk183["text"]

        def st_1chkImage184Func():
            if not str(sheet6['C57'].value).strip() == '대단히 그렇다':
                sheet6['C57'].value = st_1chk184["text"]

        def st_1chkImage191Func():
            if not str(sheet6['C60'].value).strip() == '전혀 그렇지 않다':
                sheet6['C60'].value = st_1chk181["text"]

        def st_1chkImage192Func():
            if not str(sheet6['C60'].value).strip() == '조금 그렇다':
                sheet6['C60'].value = st_1chk182["text"]

        def st_1chkImage193Func():
            if not str(sheet6['C60'].value).strip() == '보통으로 그렇다':
                sheet6['C60'].value = st_1chk183["text"]

        def st_1chkImage194Func():
            if not str(sheet6['C60'].value).strip() == '대단히 그렇다':
                sheet6['C60'].value = st_1chk184["text"]

        def st_1chkImage201Func():
            if not str(sheet6['C63'].value).strip() == '전혀 그렇지 않다':
                sheet6['C63'].value = st_1chk181["text"]

        def st_1chkImage202Func():
            if not str(sheet6['C63'].value).strip() == '조금 그렇다':
                sheet6['C63'].value = st_1chk182["text"]

        def st_1chkImage203Func():
            if not str(sheet6['C63'].value).strip() == '보통으로 그렇다':
                sheet6['C63'].value = st_1chk183["text"]

        def st_1chkImage204Func():
            if not str(sheet6['C63'].value).strip() == '대단히 그렇다':
                sheet6['C63'].value = st_1chk184["text"]

        frame6St_1BgImg = Image.open("images/st_1bg.png")
        frame6St_1Bg = ImageTk.PhotoImage(frame6St_1BgImg)
        frame6St_1BgLabel = tkinter.Label(
            displayPage.inner, image=frame6St_1Bg)
        frame6St_1BgLabel.image = frame6St_1Bg
        frame6St_1BgLabel.place(x=47, y=420, height=1330)
        frame6St_1BgLabel.pack()

        global st_1chk11
        st_1chk11 = Radiobutton(displayPage.inner, value=4, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType1, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage11Func)
        st_1chk11.deselect()
        global st_1chk12
        st_1chk12 = Radiobutton(displayPage.inner, value=3, text="조금 그렇다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=st_1chkType1, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage12Func)
        st_1chk12.deselect()
        global st_1chk13
        st_1chk13 = Radiobutton(displayPage.inner, value=2, text="보통으로 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=st_1chkType1, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage13Func)
        st_1chk13.deselect()
        global st_1chk14
        st_1chk14 = Radiobutton(displayPage.inner, value=1, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=st_1chkType1, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage14Func)
        st_1chk14.deselect()
        global st_1chk21
        st_1chk21 = Radiobutton(displayPage.inner, value=4, text="전혀 그렇지 않다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType2, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage21Func)
        st_1chk21.deselect()
        global st_1chk22
        st_1chk22 = Radiobutton(displayPage.inner, value=3, text="조금 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=st_1chkType2, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage22Func)
        st_1chk22.deselect()
        global st_1chk23
        st_1chk23 = Radiobutton(displayPage.inner, value=2, text="보통으로 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=st_1chkType2, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage23Func)
        st_1chk23.deselect()
        global st_1chk24
        st_1chk24 = Radiobutton(displayPage.inner, value=1, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=st_1chkType2, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage24Func)
        st_1chk24.deselect()
        global st_1chk31
        st_1chk31 = Radiobutton(displayPage.inner, value=1, text="전혀 그렇지 않다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType3, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage31Func)
        st_1chk31.deselect()
        global st_1chk32
        st_1chk32 = Radiobutton(displayPage.inner, value=2, text="조금 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=st_1chkType3, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage32Func)
        st_1chk32.deselect()
        global st_1chk33
        st_1chk33 = Radiobutton(displayPage.inner, value=3, text="보통으로 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=st_1chkType3, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage33Func)
        st_1chk33.deselect()
        global st_1chk34
        st_1chk34 = Radiobutton(displayPage.inner, value=4, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=st_1chkType3, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage34Func)
        st_1chk34.deselect()
        global st_1chk41
        st_1chk41 = Radiobutton(displayPage.inner, value=1, text="전혀 그렇지 않다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType4, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage41Func)
        st_1chk41.deselect()
        global st_1chk42
        st_1chk42 = Radiobutton(displayPage.inner, value=2, text="조금 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=st_1chkType4, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage42Func)
        st_1chk42.deselect()
        global st_1chk43
        st_1chk43 = Radiobutton(displayPage.inner, value=3, text="보통으로 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=st_1chkType4, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage43Func)
        st_1chk43.deselect()
        global st_1chk44
        st_1chk44 = Radiobutton(displayPage.inner, value=4, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=st_1chkType4, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage44Func)
        st_1chk44.deselect()
        global st_1chk51
        st_1chk51 = Radiobutton(displayPage.inner, value=4, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType5, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage51Func)
        st_1chk51.deselect()
        global st_1chk52
        st_1chk52 = Radiobutton(displayPage.inner, value=3, text="조금 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=st_1chkType5, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage52Func)
        st_1chk52.deselect()
        global st_1chk53
        st_1chk53 = Radiobutton(displayPage.inner, value=2, text="보통으로 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=st_1chkType5, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage53Func)
        st_1chk53.deselect()
        global st_1chk54
        st_1chk54 = Radiobutton(displayPage.inner, value=1, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=st_1chkType5, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage54Func)
        st_1chk54.deselect()
        global st_1chk61
        st_1chk61 = Radiobutton(displayPage.inner, value=1, text="전혀 그렇지 않다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType6, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage61Func)
        st_1chk61.deselect()
        global st_1chk62
        st_1chk62 = Radiobutton(displayPage.inner, value=2, text="조금 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=st_1chkType6, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage62Func)
        st_1chk62.deselect()
        global st_1chk63
        st_1chk63 = Radiobutton(displayPage.inner, value=3, text="보통으로 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=st_1chkType6, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage63Func)
        st_1chk63.deselect()
        global st_1chk64
        st_1chk64 = Radiobutton(displayPage.inner, value=4, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=st_1chkType6, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage64Func)
        st_1chk64.deselect()

        global st_1chk71
        st_1chk71 = Radiobutton(displayPage.inner, value=1, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType7, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage71Func)
        st_1chk71.deselect()
        global st_1chk72
        st_1chk72 = Radiobutton(displayPage.inner, value=2, text="조금 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=st_1chkType7, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage72Func)
        st_1chk72.deselect()
        global st_1chk73
        st_1chk73 = Radiobutton(displayPage.inner, value=3, text="보통으로 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=st_1chkType7, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage73Func)
        st_1chk73.deselect()
        global st_1chk74
        st_1chk74 = Radiobutton(displayPage.inner, value=4, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=st_1chkType7, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage74Func)
        st_1chk74.deselect()

        global st_1chk81
        st_1chk81 = Radiobutton(displayPage.inner, value=4, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType8, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage81Func)
        st_1chk81.deselect()
        global st_1chk82
        st_1chk82 = Radiobutton(displayPage.inner, value=3, text="조금 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=st_1chkType8, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage82Func)
        st_1chk82.deselect()
        global st_1chk83
        st_1chk83 = Radiobutton(displayPage.inner, value=2, text="보통으로 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=st_1chkType8, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage83Func)
        st_1chk83.deselect()
        global st_1chk84
        st_1chk84 = Radiobutton(displayPage.inner, value=1, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=st_1chkType8, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage84Func)
        st_1chk84.deselect()

        global st_1chk91
        st_1chk91 = Radiobutton(displayPage.inner, value=1, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType9, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage91Func)
        st_1chk91.deselect()
        global st_1chk92
        st_1chk92 = Radiobutton(displayPage.inner, value=2, text="조금 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=st_1chkType9, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage92Func)
        st_1chk92.deselect()
        global st_1chk93
        st_1chk93 = Radiobutton(displayPage.inner, value=3, text="보통으로 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=st_1chkType9, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage93Func)
        st_1chk93.deselect()
        global st_1chk94
        st_1chk94 = Radiobutton(displayPage.inner, value=4, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=st_1chkType9, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage94Func)
        st_1chk94.deselect()
        global st_1chk101
        st_1chk101 = Radiobutton(displayPage.inner, value=4, text="전혀 그렇지 않다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType10, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage101Func)
        st_1chk101.deselect()
        global st_1chk102
        st_1chk102 = Radiobutton(displayPage.inner, value=3, text="조금 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_1chkType10, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage102Func)
        st_1chk102.deselect()
        global st_1chk103
        st_1chk103 = Radiobutton(displayPage.inner, value=2, text="보통으로 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_1chkType10, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage103Func)
        st_1chk103.deselect()
        global st_1chk104
        st_1chk104 = Radiobutton(displayPage.inner, value=1, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_1chkType10, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage104Func)
        st_1chk104.deselect()
        global st_1chk111
        st_1chk111 = Radiobutton(displayPage.inner, value=4, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType11, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage111Func)
        st_1chk111.deselect()
        global st_1chk112
        st_1chk112 = Radiobutton(displayPage.inner, value=3, text="조금 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_1chkType11, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage112Func)
        st_1chk112.deselect()
        global st_1chk113
        st_1chk113 = Radiobutton(displayPage.inner, value=2, text="보통으로 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_1chkType11, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage113Func)
        st_1chk113.deselect()
        global st_1chk114
        st_1chk114 = Radiobutton(displayPage.inner, value=1, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_1chkType11, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage114Func)
        st_1chk114.deselect()
        global st_1chk121
        st_1chk121 = Radiobutton(displayPage.inner, value=1, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType12, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage121Func)
        st_1chk121.deselect()
        global st_1chk122
        st_1chk122 = Radiobutton(displayPage.inner, value=2, text="조금 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_1chkType12, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage122Func)
        st_1chk122.deselect()
        global st_1chk123
        st_1chk123 = Radiobutton(displayPage.inner, value=3, text="보통으로 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_1chkType12, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage123Func)
        st_1chk123.deselect()
        global st_1chk124
        st_1chk124 = Radiobutton(displayPage.inner, value=4, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_1chkType12, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage124Func)
        st_1chk124.deselect()
        global st_1chk131
        st_1chk131 = Radiobutton(displayPage.inner, value=1, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType13, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage131Func)
        st_1chk131.deselect()
        global st_1chk132
        st_1chk132 = Radiobutton(displayPage.inner, value=2, text="조금 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_1chkType13, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage132Func)
        st_1chk132.deselect()
        global st_1chk133
        st_1chk133 = Radiobutton(displayPage.inner, value=3, text="보통으로 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_1chkType13, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage133Func)
        st_1chk133.deselect()
        global st_1chk134
        st_1chk134 = Radiobutton(displayPage.inner, value=4, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_1chkType13, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage134Func)
        st_1chk134.deselect()
        global st_1chk141
        st_1chk141 = Radiobutton(displayPage.inner, value=1, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType14, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage141Func)
        st_1chk141.deselect()
        global st_1chk142
        st_1chk142 = Radiobutton(displayPage.inner, value=2, text="조금 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_1chkType14, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage142Func)
        st_1chk142.deselect()
        global st_1chk143
        st_1chk143 = Radiobutton(displayPage.inner, value=3, text="보통으로 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_1chkType14, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage143Func)
        st_1chk143.deselect()
        global st_1chk144
        st_1chk144 = Radiobutton(displayPage.inner, value=4, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_1chkType14, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage144Func)
        st_1chk144.deselect()
        global st_1chk151
        st_1chk151 = Radiobutton(displayPage.inner, value=4, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType15, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage151Func)
        st_1chk151.deselect()
        global st_1chk152
        st_1chk152 = Radiobutton(displayPage.inner, value=3, text="조금 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_1chkType15, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage152Func)
        st_1chk152.deselect()
        global st_1chk153
        st_1chk153 = Radiobutton(displayPage.inner, value=2, text="보통으로 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_1chkType15, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage153Func)
        st_1chk153.deselect()
        global st_1chk154
        st_1chk154 = Radiobutton(displayPage.inner, value=1, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_1chkType15, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage154Func)
        st_1chk154.deselect()
        global st_1chk161
        st_1chk161 = Radiobutton(displayPage.inner, value=4, text="전혀 그렇지 않다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType16, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage161Func)
        st_1chk161.deselect()
        global st_1chk162
        st_1chk162 = Radiobutton(displayPage.inner, value=3, text="조금 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_1chkType16, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage162Func)
        st_1chk162.deselect()
        global st_1chk163
        st_1chk163 = Radiobutton(displayPage.inner, value=2, text="보통으로 그렇다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_1chkType16, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage163Func)
        st_1chk163.deselect()
        global st_1chk164
        st_1chk164 = Radiobutton(displayPage.inner, value=1, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_1chkType16, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage164Func)
        st_1chk164.deselect()
        global st_1chk171
        st_1chk171 = Radiobutton(displayPage.inner, value=1, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType17, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage171Func)
        st_1chk171.deselect()
        global st_1chk172
        st_1chk172 = Radiobutton(displayPage.inner, value=2, text="조금 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_1chkType17, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage172Func)
        st_1chk172.deselect()
        global st_1chk173
        st_1chk173 = Radiobutton(displayPage.inner, value=3, text="보통으로 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_1chkType17, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage173Func)
        st_1chk173.deselect()
        global st_1chk174
        st_1chk174 = Radiobutton(displayPage.inner, value=4, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_1chkType17, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage174Func)
        st_1chk174.deselect()
        global st_1chk181
        st_1chk181 = Radiobutton(displayPage.inner, value=1, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType18, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage181Func)
        st_1chk181.deselect()
        global st_1chk182
        st_1chk182 = Radiobutton(displayPage.inner, value=2, text="조금 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_1chkType18, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage182Func)
        st_1chk182.deselect()
        global st_1chk183
        st_1chk183 = Radiobutton(displayPage.inner, value=3, text="보통으로 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_1chkType18, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage183Func)
        st_1chk183.deselect()
        global st_1chk184
        st_1chk184 = Radiobutton(displayPage.inner, value=4, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_1chkType18, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage184Func)
        st_1chk184.deselect()
        global st_1chk191
        st_1chk191 = Radiobutton(displayPage.inner, value=4, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType19, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage191Func)
        st_1chk191.deselect()
        global st_1chk192
        st_1chk192 = Radiobutton(displayPage.inner, value=3, text="조금 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_1chkType19, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage192Func)
        st_1chk192.deselect()
        global st_1chk193
        st_1chk193 = Radiobutton(displayPage.inner, value=2, text="보통으로 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_1chkType19, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage193Func)
        st_1chk193.deselect()
        global st_1chk194
        st_1chk194 = Radiobutton(displayPage.inner, value=1, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_1chkType19, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage194Func)
        st_1chk194.deselect()
        global st_1chk201
        st_1chk201 = Radiobutton(displayPage.inner, value=4, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType20, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage201Func)
        st_1chk201.deselect()
        global st_1chk202
        st_1chk202 = Radiobutton(displayPage.inner, value=3, text="조금 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_1chkType20, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage202Func)
        st_1chk202.deselect()
        global st_1chk203
        st_1chk203 = Radiobutton(displayPage.inner, value=2, text="보통으로 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_1chkType20, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage203Func)
        st_1chk203.deselect()
        global st_1chk204
        st_1chk204 = Radiobutton(displayPage.inner, value=1, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_1chkType20, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage204Func)
        st_1chk204.deselect()

        st_1chk11.place(x=938, y=472, width=60, height=21)
        st_1chk12.place(x=1014, y=472, width=60, height=21)
        st_1chk13.place(x=1088, y=472, width=60, height=21)
        st_1chk14.place(x=1162, y=472, width=60, height=21)
        st_1chk21.place(x=938, y=512, width=60, height=21)
        st_1chk22.place(x=1014, y=512, width=60, height=21)
        st_1chk23.place(x=1088, y=512, width=60, height=21)
        st_1chk24.place(x=1162, y=512, width=60, height=21)
        st_1chk31.place(x=938, y=551, width=60, height=21)
        st_1chk32.place(x=1014, y=551, width=60, height=21)
        st_1chk33.place(x=1088, y=551, width=60, height=21)
        st_1chk34.place(x=1162, y=551, width=60, height=21)
        st_1chk41.place(x=938, y=589, width=60, height=21)
        st_1chk42.place(x=1014, y=589, width=60, height=21)
        st_1chk43.place(x=1088, y=589, width=60, height=21)
        st_1chk44.place(x=1162, y=589, width=60, height=21)
        st_1chk51.place(x=938, y=628, width=60, height=21)
        st_1chk52.place(x=1014, y=628, width=60, height=21)
        st_1chk53.place(x=1088, y=628, width=60, height=21)
        st_1chk54.place(x=1162, y=628, width=60, height=21)
        st_1chk61.place(x=938, y=667, width=60, height=21)
        st_1chk62.place(x=1014, y=667, width=60, height=21)
        st_1chk63.place(x=1088, y=667, width=60, height=21)
        st_1chk64.place(x=1162, y=667, width=60, height=21)
        st_1chk71.place(x=938, y=704, width=60, height=21)
        st_1chk72.place(x=1014, y=704, width=60, height=21)
        st_1chk73.place(x=1088, y=704, width=60, height=21)
        st_1chk74.place(x=1162, y=704, width=60, height=21)
        st_1chk81.place(x=938, y=742, width=60, height=21)
        st_1chk82.place(x=1014, y=742, width=60, height=21)
        st_1chk83.place(x=1088, y=742, width=60, height=21)
        st_1chk84.place(x=1162, y=742, width=60, height=21)
        st_1chk91.place(x=938, y=781, width=60, height=21)
        st_1chk92.place(x=1014, y=781, width=60, height=21)
        st_1chk93.place(x=1088, y=781, width=60, height=21)
        st_1chk94.place(x=1162, y=781, width=60, height=21)
        st_1chk101.place(x=938, y=821, width=60, height=21)
        st_1chk102.place(x=1014, y=821, width=60, height=21)
        st_1chk103.place(x=1088, y=821, width=60, height=21)
        st_1chk104.place(x=1162, y=821, width=60, height=21)
        st_1chk111.place(x=938, y=860, width=60, height=21)
        st_1chk112.place(x=1014, y=860, width=60, height=21)
        st_1chk113.place(x=1088, y=860, width=60, height=21)
        st_1chk114.place(x=1162, y=860, width=60, height=21)
        st_1chk121.place(x=938, y=899, width=60, height=21)
        st_1chk122.place(x=1014, y=899, width=60, height=21)
        st_1chk123.place(x=1088, y=899, width=60, height=21)
        st_1chk124.place(x=1162, y=899, width=60, height=21)
        st_1chk131.place(x=938, y=938, width=60, height=21)
        st_1chk132.place(x=1014, y=938, width=60, height=21)
        st_1chk133.place(x=1088, y=938, width=60, height=21)
        st_1chk134.place(x=1162, y=938, width=60, height=21)
        st_1chk141.place(x=938, y=977, width=60, height=21)
        st_1chk142.place(x=1014, y=977, width=60, height=21)
        st_1chk143.place(x=1088, y=977, width=60, height=21)
        st_1chk144.place(x=1162, y=977, width=60, height=21)
        st_1chk151.place(x=938, y=1016, width=60, height=21)
        st_1chk152.place(x=1014, y=1016, width=60, height=21)
        st_1chk153.place(x=1088, y=1016, width=60, height=21)
        st_1chk154.place(x=1162, y=1016, width=60, height=21)
        st_1chk161.place(x=938, y=1054, width=60, height=21)
        st_1chk162.place(x=1014, y=1054, width=60, height=21)
        st_1chk163.place(x=1088, y=1054, width=60, height=21)
        st_1chk164.place(x=1162, y=1054, width=60, height=21)
        st_1chk171.place(x=938, y=1092, width=60, height=21)
        st_1chk172.place(x=1014, y=1092, width=60, height=21)
        st_1chk173.place(x=1088, y=1092, width=60, height=21)
        st_1chk174.place(x=1162, y=1092, width=60, height=21)
        st_1chk181.place(x=938, y=1133, width=60, height=21)
        st_1chk182.place(x=1014, y=1133, width=60, height=21)
        st_1chk183.place(x=1088, y=1133, width=60, height=21)
        st_1chk184.place(x=1162, y=1133, width=60, height=21)
        st_1chk191.place(x=938, y=1173, width=60, height=21)
        st_1chk192.place(x=1014, y=1173, width=60, height=21)
        st_1chk193.place(x=1088, y=1173, width=60, height=21)
        st_1chk194.place(x=1162, y=1173, width=60, height=21)
        st_1chk201.place(x=938, y=1213, width=60, height=21)
        st_1chk202.place(x=1014, y=1213, width=60, height=21)
        st_1chk203.place(x=1088, y=1213, width=60, height=21)
        st_1chk204.place(x=1162, y=1213, width=60, height=21)

    def st_1chkFunc2():
        global now
        now = datetime.now()
        global st_1chk
        st_1chk = True
        sheet6['E2'].value = (" % s년 % s월 % s일 % s시 % s분" %
                              (now.year, now.month, now.day, now.hour, now.minute))
        sheet6['K2'].value = idInput.get()

        def st_1chkImage11Func():
            # sumScore()
            if not str(sheet6['C6'].value).strip() == '전혀 그렇지 않다':
                sheet6['C6'].value = st_1chk11["text"]

        def st_1chkImage12Func():
            # sumScore()
            if not str(sheet6['C6'].value).strip() == '조금 그렇다':
                sheet6['C6'].value = st_1chk12["text"]

        def st_1chkImage13Func():
            # sumScore()
            if not str(sheet6['C6'].value).strip() == '보통으로 그렇다':
                sheet6['C6'].value = st_1chk13["text"]

        def st_1chkImage14Func():
            # sumScore()
            if not str(sheet6['C6'].value).strip() == '대단히 그렇다':
                sheet6['C6'].value = st_1chk14["text"]

        def st_1chkImage21Func():
            if not str(sheet6['C9'].value).strip() == '전혀 그렇지 않다':
                sheet6['C9'].value = st_1chk21["text"]

        def st_1chkImage22Func():
            if not str(sheet6['C9'].value).strip() == '조금 그렇다':
                sheet6['C9'].value = st_1chk22["text"]

        def st_1chkImage23Func():
            if not str(sheet6['C9'].value).strip() == '보통으로 그렇다':
                sheet6['C9'].value = st_1chk23["text"]

        def st_1chkImage24Func():
            if not str(sheet6['C9'].value).strip() == '대단히 그렇다':
                sheet6['C9'].value = st_1chk24["text"]

        def st_1chkImage31Func():
            if not str(sheet6['C12'].value).strip() == '전혀 그렇지 않다':
                sheet6['C12'].value = st_1chk31["text"]

        def st_1chkImage32Func():
            if not str(sheet6['C12'].value).strip() == '조금 그렇다':
                sheet6['C12'].value = st_1chk32["text"]

        def st_1chkImage33Func():
            if not str(sheet6['C12'].value).strip() == '보통으로 그렇다':
                sheet6['C12'].value = st_1chk33["text"]

        def st_1chkImage34Func():
            if not str(sheet6['C12'].value).strip() == '대단히 그렇다':
                sheet6['C12'].value = st_1chk34["text"]

        def st_1chkImage41Func():
            if not str(sheet6['C15'].value).strip() == '전혀 그렇지 않다':
                sheet6['C15'].value = st_1chk41["text"]

        def st_1chkImage42Func():
            if not str(sheet6['C15'].value).strip() == '조금 그렇다':
                sheet6['C15'].value = st_1chk42["text"]

        def st_1chkImage43Func():
            if not str(sheet6['C15'].value).strip() == '보통으로 그렇다':
                sheet6['C15'].value = st_1chk43["text"]

        def st_1chkImage44Func():
            if not str(sheet6['C15'].value).strip() == '대단히 그렇다':
                sheet6['C15'].value = st_1chk44["text"]

        def st_1chkImage51Func():
            if not str(sheet6['C18'].value).strip() == '전혀 그렇지 않다':
                sheet6['C18'].value = st_1chk51["text"]

        def st_1chkImage52Func():
            if not str(sheet6['C18'].value).strip() == '조금 그렇다':
                sheet6['C18'].value = st_1chk52["text"]

        def st_1chkImage53Func():
            if not str(sheet6['C18'].value).strip() == '보통으로 그렇다':
                sheet6['C18'].value = st_1chk53["text"]

        def st_1chkImage54Func():
            if not str(sheet6['C18'].value).strip() == '대단히 그렇다':
                sheet6['C18'].value = st_1chk54["text"]

        def st_1chkImage61Func():
            if not str(sheet6['C21'].value).strip() == '전혀 그렇지 않다':
                sheet6['C21'].value = st_1chk61["text"]

        def st_1chkImage62Func():
            if not str(sheet6['C21'].value).strip() == '조금 그렇다':
                sheet6['C21'].value = st_1chk62["text"]

        def st_1chkImage63Func():
            if not str(sheet6['C21'].value).strip() == '보통으로 그렇다':
                sheet6['C21'].value = st_1chk63["text"]

        def st_1chkImage64Func():
            if not str(sheet6['C21'].value).strip() == '대단히 그렇다':
                sheet6['C21'].value = st_1chk64["text"]

        def st_1chkImage71Func():
            if not str(sheet6['C24'].value).strip() == '전혀 그렇지 않다':
                sheet6['C24'].value = st_1chk71["text"]

        def st_1chkImage72Func():
            if not str(sheet6['C24'].value).strip() == '조금 그렇다':
                sheet6['C24'].value = st_1chk72["text"]

        def st_1chkImage73Func():
            if not str(sheet6['C24'].value).strip() == '보통으로 그렇다':
                sheet6['C24'].value = st_1chk73["text"]

        def st_1chkImage74Func():
            if not str(sheet6['C24'].value).strip() == '대단히 그렇다':
                sheet6['C24'].value = st_1chk74["text"]

        def st_1chkImage81Func():
            if not str(sheet6['C27'].value).strip() == '전혀 그렇지 않다':
                sheet6['C27'].value = st_1chk81["text"]

        def st_1chkImage82Func():
            if not str(sheet6['C27'].value).strip() == '조금 그렇다':
                sheet6['C27'].value = st_1chk82["text"]

        def st_1chkImage83Func():
            if not str(sheet6['C27'].value).strip() == '보통으로 그렇다':
                sheet6['C27'].value = st_1chk83["text"]

        def st_1chkImage84Func():
            if not str(sheet6['C27'].value).strip() == '대단히 그렇다':
                sheet6['C27'].value = st_1chk84["text"]

        def st_1chkImage91Func():
            if not str(sheet6['C30'].value).strip() == '전혀 그렇지 않다':
                sheet6['C30'].value = st_1chk91["text"]

        def st_1chkImage92Func():
            if not str(sheet6['C30'].value).strip() == '조금 그렇다':
                sheet6['C30'].value = st_1chk92["text"]

        def st_1chkImage93Func():
            if not str(sheet6['C30'].value).strip() == '보통으로 그렇다':
                sheet6['C30'].value = st_1chk93["text"]

        def st_1chkImage94Func():
            if not str(sheet6['C30'].value).strip() == '대단히 그렇다':
                sheet6['C30'].value = st_1chk94["text"]

        def st_1chkImage101Func():
            if not str(sheet6['C33'].value).strip() == '전혀 그렇지 않다':
                sheet6['C33'].value = st_1chk101["text"]

        def st_1chkImage102Func():
            if not str(sheet6['C33'].value).strip() == '조금 그렇다':
                sheet6['C33'].value = st_1chk102["text"]

        def st_1chkImage103Func():
            if not str(sheet6['C33'].value).strip() == '보통으로 그렇다':
                sheet6['C33'].value = st_1chk103["text"]

        def st_1chkImage104Func():
            if not str(sheet6['C33'].value).strip() == '대단히 그렇다':
                sheet6['C33'].value = st_1chk104["text"]

        def st_1chkImage111Func():
            if not str(sheet6['C36'].value).strip() == '전혀 그렇지 않다':
                sheet6['C36'].value = st_1chk111["text"]

        def st_1chkImage112Func():
            if not str(sheet6['C36'].value).strip() == '조금 그렇다':
                sheet6['C36'].value = st_1chk112["text"]

        def st_1chkImage113Func():
            if not str(sheet6['C36'].value).strip() == '보통으로 그렇다':
                sheet6['C36'].value = st_1chk113["text"]

        def st_1chkImage114Func():
            if not str(sheet6['C36'].value).strip() == '대단히 그렇다':
                sheet6['C36'].value = st_1chk114["text"]

        def st_1chkImage121Func():
            if not str(sheet6['C39'].value).strip() == '전혀 그렇지 않다':
                sheet6['C39'].value = st_1chk121["text"]

        def st_1chkImage122Func():
            if not str(sheet6['C39'].value).strip() == '조금 그렇다':
                sheet6['C39'].value = st_1chk122["text"]

        def st_1chkImage123Func():
            if not str(sheet6['C39'].value).strip() == '보통으로 그렇다':
                sheet6['C39'].value = st_1chk123["text"]

        def st_1chkImage124Func():
            if not str(sheet6['C39'].value).strip() == '대단히 그렇다':
                sheet6['C39'].value = st_1chk124["text"]

        def st_1chkImage131Func():
            if not str(sheet6['C42'].value).strip() == '전혀 그렇지 않다':
                sheet6['C42'].value = st_1chk131["text"]

        def st_1chkImage132Func():
            if not str(sheet6['C42'].value).strip() == '조금 그렇다':
                sheet6['C42'].value = st_1chk132["text"]

        def st_1chkImage133Func():
            if not str(sheet6['C42'].value).strip() == '보통으로 그렇다':
                sheet6['C42'].value = st_1chk133["text"]

        def st_1chkImage134Func():
            if not str(sheet6['C42'].value).strip() == '대단히 그렇다':
                sheet6['C42'].value = st_1chk134["text"]

        def st_1chkImage141Func():
            if not str(sheet6['C45'].value).strip() == '전혀 그렇지 않다':
                sheet6['C45'].value = st_1chk141["text"]

        def st_1chkImage142Func():
            if not str(sheet6['C45'].value).strip() == '조금 그렇다':
                sheet6['C45'].value = st_1chk142["text"]

        def st_1chkImage143Func():
            if not str(sheet6['C45'].value).strip() == '보통으로 그렇다':
                sheet6['C45'].value = st_1chk143["text"]

        def st_1chkImage144Func():
            if not str(sheet6['C45'].value).strip() == '대단히 그렇다':
                sheet6['C45'].value = st_1chk144["text"]

        def st_1chkImage151Func():
            if not str(sheet6['C48'].value).strip() == '전혀 그렇지 않다':
                sheet6['C48'].value = st_1chk151["text"]

        def st_1chkImage152Func():
            if not str(sheet6['C48'].value).strip() == '조금 그렇다':
                sheet6['C48'].value = st_1chk152["text"]

        def st_1chkImage153Func():
            if not str(sheet6['C48'].value).strip() == '보통으로 그렇다':
                sheet6['C48'].value = st_1chk153["text"]

        def st_1chkImage154Func():
            if not str(sheet6['C48'].value).strip() == '대단히 그렇다':
                sheet6['C48'].value = st_1chk154["text"]

        def st_1chkImage161Func():
            if not str(sheet6['C51'].value).strip() == '전혀 그렇지 않다':
                sheet6['C51'].value = st_1chk161["text"]

        def st_1chkImage162Func():
            if not str(sheet6['C51'].value).strip() == '조금 그렇다':
                sheet6['C51'].value = st_1chk162["text"]

        def st_1chkImage163Func():
            if not str(sheet6['C51'].value).strip() == '보통으로 그렇다':
                sheet6['C51'].value = st_1chk163["text"]

        def st_1chkImage164Func():
            if not str(sheet6['C51'].value).strip() == '대단히 그렇다':
                sheet6['C51'].value = st_1chk164["text"]

        def st_1chkImage171Func():
            if not str(sheet6['C54'].value).strip() == '전혀 그렇지 않다':
                sheet6['C54'].value = st_1chk171["text"]

        def st_1chkImage172Func():
            if not str(sheet6['C54'].value).strip() == '조금 그렇다':
                sheet6['C54'].value = st_1chk172["text"]

        def st_1chkImage173Func():
            if not str(sheet6['C54'].value).strip() == '보통으로 그렇다':
                sheet6['C54'].value = st_1chk173["text"]

        def st_1chkImage174Func():
            if not str(sheet6['C54'].value).strip() == '대단히 그렇다':
                sheet6['C54'].value = st_1chk174["text"]

        def st_1chkImage181Func():
            if not str(sheet6['C57'].value).strip() == '전혀 그렇지 않다':
                sheet6['C57'].value = st_1chk181["text"]

        def st_1chkImage182Func():
            if not str(sheet6['C57'].value).strip() == '조금 그렇다':
                sheet6['C57'].value = st_1chk182["text"]

        def st_1chkImage183Func():
            if not str(sheet6['C57'].value).strip() == '보통으로 그렇다':
                sheet6['C57'].value = st_1chk183["text"]

        def st_1chkImage184Func():
            if not str(sheet6['C57'].value).strip() == '대단히 그렇다':
                sheet6['C57'].value = st_1chk184["text"]

        def st_1chkImage191Func():
            if not str(sheet6['C60'].value).strip() == '전혀 그렇지 않다':
                sheet6['C60'].value = st_1chk181["text"]

        def st_1chkImage192Func():
            if not str(sheet6['C60'].value).strip() == '조금 그렇다':
                sheet6['C60'].value = st_1chk182["text"]

        def st_1chkImage193Func():
            if not str(sheet6['C60'].value).strip() == '보통으로 그렇다':
                sheet6['C60'].value = st_1chk183["text"]

        def st_1chkImage194Func():
            if not str(sheet6['C60'].value).strip() == '대단히 그렇다':
                sheet6['C60'].value = st_1chk184["text"]

        def st_1chkImage201Func():
            if not str(sheet6['C63'].value).strip() == '전혀 그렇지 않다':
                sheet6['C63'].value = st_1chk181["text"]

        def st_1chkImage202Func():
            if not str(sheet6['C63'].value).strip() == '조금 그렇다':
                sheet6['C63'].value = st_1chk182["text"]

        def st_1chkImage203Func():
            if not str(sheet6['C63'].value).strip() == '보통으로 그렇다':
                sheet6['C63'].value = st_1chk183["text"]

        def st_1chkImage204Func():
            if not str(sheet6['C63'].value).strip() == '대단히 그렇다':
                sheet6['C63'].value = st_1chk184["text"]

        frame6St_1BgImg = Image.open("images/st_1bg.png")
        frame6St_1Bg = ImageTk.PhotoImage(frame6St_1BgImg)
        frame6St_1BgLabel = tkinter.Label(
            displayPage2.inner, image=frame6St_1Bg)
        frame6St_1BgLabel.image = frame6St_1Bg
        frame6St_1BgLabel.place(x=47, y=420, height=1330)
        frame6St_1BgLabel.pack()

        global st_1chk11
        st_1chk11 = Radiobutton(displayPage2.inner, value=4, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType1, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage11Func)
        st_1chk11.deselect()
        global st_1chk12
        st_1chk12 = Radiobutton(displayPage2.inner, value=3, text="조금 그렇다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=st_1chkType1, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage12Func)
        st_1chk12.deselect()
        global st_1chk13
        st_1chk13 = Radiobutton(displayPage2.inner, value=2, text="보통으로 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=st_1chkType1, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage13Func)
        st_1chk13.deselect()
        global st_1chk14
        st_1chk14 = Radiobutton(displayPage2.inner, value=1, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=st_1chkType1, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage14Func)
        st_1chk14.deselect()
        global st_1chk21
        st_1chk21 = Radiobutton(displayPage2.inner, value=4, text="전혀 그렇지 않다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType2, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage21Func)
        st_1chk21.deselect()
        global st_1chk22
        st_1chk22 = Radiobutton(displayPage2.inner, value=3, text="조금 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=st_1chkType2, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage22Func)
        st_1chk22.deselect()
        global st_1chk23
        st_1chk23 = Radiobutton(displayPage2.inner, value=2, text="보통으로 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=st_1chkType2, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage23Func)
        st_1chk23.deselect()
        global st_1chk24
        st_1chk24 = Radiobutton(displayPage2.inner, value=1, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=st_1chkType2, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage24Func)
        st_1chk24.deselect()
        global st_1chk31
        st_1chk31 = Radiobutton(displayPage2.inner, value=1, text="전혀 그렇지 않다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType3, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage31Func)
        st_1chk31.deselect()
        global st_1chk32
        st_1chk32 = Radiobutton(displayPage2.inner, value=2, text="조금 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=st_1chkType3, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage32Func)
        st_1chk32.deselect()
        global st_1chk33
        st_1chk33 = Radiobutton(displayPage2.inner, value=3, text="보통으로 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=st_1chkType3, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage33Func)
        st_1chk33.deselect()
        global st_1chk34
        st_1chk34 = Radiobutton(displayPage2.inner, value=4, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=st_1chkType3, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage34Func)
        st_1chk34.deselect()
        global st_1chk41
        st_1chk41 = Radiobutton(displayPage2.inner, value=1, text="전혀 그렇지 않다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType4, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage41Func)
        st_1chk41.deselect()
        global st_1chk42
        st_1chk42 = Radiobutton(displayPage2.inner, value=2, text="조금 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=st_1chkType4, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage42Func)
        st_1chk42.deselect()
        global st_1chk43
        st_1chk43 = Radiobutton(displayPage2.inner, value=3, text="보통으로 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=st_1chkType4, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage43Func)
        st_1chk43.deselect()
        global st_1chk44
        st_1chk44 = Radiobutton(displayPage2.inner, value=4, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=st_1chkType4, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage44Func)
        st_1chk44.deselect()
        global st_1chk51
        st_1chk51 = Radiobutton(displayPage2.inner, value=4, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType5, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage51Func)
        st_1chk51.deselect()
        global st_1chk52
        st_1chk52 = Radiobutton(displayPage2.inner, value=3, text="조금 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=st_1chkType5, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage52Func)
        st_1chk52.deselect()
        global st_1chk53
        st_1chk53 = Radiobutton(displayPage2.inner, value=2, text="보통으로 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=st_1chkType5, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage53Func)
        st_1chk53.deselect()
        global st_1chk54
        st_1chk54 = Radiobutton(displayPage2.inner, value=1, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=st_1chkType5, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage54Func)
        st_1chk54.deselect()
        global st_1chk61
        st_1chk61 = Radiobutton(displayPage2.inner, value=1, text="전혀 그렇지 않다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType6, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage61Func)
        st_1chk61.deselect()
        global st_1chk62
        st_1chk62 = Radiobutton(displayPage2.inner, value=2, text="조금 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=st_1chkType6, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage62Func)
        st_1chk62.deselect()
        global st_1chk63
        st_1chk63 = Radiobutton(displayPage2.inner, value=3, text="보통으로 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=st_1chkType6, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage63Func)
        st_1chk63.deselect()
        global st_1chk64
        st_1chk64 = Radiobutton(displayPage2.inner, value=4, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=st_1chkType6, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage64Func)
        st_1chk64.deselect()

        global st_1chk71
        st_1chk71 = Radiobutton(displayPage2.inner, value=1, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType7, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage71Func)
        st_1chk71.deselect()
        global st_1chk72
        st_1chk72 = Radiobutton(displayPage2.inner, value=2, text="조금 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=st_1chkType7, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage72Func)
        st_1chk72.deselect()
        global st_1chk73
        st_1chk73 = Radiobutton(displayPage2.inner, value=3, text="보통으로 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=st_1chkType7, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage73Func)
        st_1chk73.deselect()
        global st_1chk74
        st_1chk74 = Radiobutton(displayPage2.inner, value=4, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=st_1chkType7, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage74Func)
        st_1chk74.deselect()

        global st_1chk81
        st_1chk81 = Radiobutton(displayPage2.inner, value=4, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType8, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage81Func)
        st_1chk81.deselect()
        global st_1chk82
        st_1chk82 = Radiobutton(displayPage2.inner, value=3, text="조금 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=st_1chkType8, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage82Func)
        st_1chk82.deselect()
        global st_1chk83
        st_1chk83 = Radiobutton(displayPage2.inner, value=2, text="보통으로 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=st_1chkType8, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage83Func)
        st_1chk83.deselect()
        global st_1chk84
        st_1chk84 = Radiobutton(displayPage2.inner, value=1, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=st_1chkType8, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage84Func)
        st_1chk84.deselect()

        global st_1chk91
        st_1chk91 = Radiobutton(displayPage2.inner, value=1, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType9, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage91Func)
        st_1chk91.deselect()
        global st_1chk92
        st_1chk92 = Radiobutton(displayPage2.inner, value=2, text="조금 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=st_1chkType9, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage92Func)
        st_1chk92.deselect()
        global st_1chk93
        st_1chk93 = Radiobutton(displayPage2.inner, value=3, text="보통으로 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=st_1chkType9, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage93Func)
        st_1chk93.deselect()
        global st_1chk94
        st_1chk94 = Radiobutton(displayPage2.inner, value=4, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=st_1chkType9, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage94Func)
        st_1chk94.deselect()
        global st_1chk101
        st_1chk101 = Radiobutton(displayPage2.inner, value=4, text="전혀 그렇지 않다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType10, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage101Func)
        st_1chk101.deselect()
        global st_1chk102
        st_1chk102 = Radiobutton(displayPage2.inner, value=3, text="조금 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_1chkType10, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage102Func)
        st_1chk102.deselect()
        global st_1chk103
        st_1chk103 = Radiobutton(displayPage2.inner, value=2, text="보통으로 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_1chkType10, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage103Func)
        st_1chk103.deselect()
        global st_1chk104
        st_1chk104 = Radiobutton(displayPage2.inner, value=1, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_1chkType10, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage104Func)
        st_1chk104.deselect()
        global st_1chk111
        st_1chk111 = Radiobutton(displayPage2.inner, value=4, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType11, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage111Func)
        st_1chk111.deselect()
        global st_1chk112
        st_1chk112 = Radiobutton(displayPage2.inner, value=3, text="조금 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_1chkType11, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage112Func)
        st_1chk112.deselect()
        global st_1chk113
        st_1chk113 = Radiobutton(displayPage2.inner, value=2, text="보통으로 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_1chkType11, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage113Func)
        st_1chk113.deselect()
        global st_1chk114
        st_1chk114 = Radiobutton(displayPage2.inner, value=1, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_1chkType11, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage114Func)
        st_1chk114.deselect()
        global st_1chk121
        st_1chk121 = Radiobutton(displayPage2.inner, value=1, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType12, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage121Func)
        st_1chk121.deselect()
        global st_1chk122
        st_1chk122 = Radiobutton(displayPage2.inner, value=2, text="조금 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_1chkType12, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage122Func)
        st_1chk122.deselect()
        global st_1chk123
        st_1chk123 = Radiobutton(displayPage2.inner, value=3, text="보통으로 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_1chkType12, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage123Func)
        st_1chk123.deselect()
        global st_1chk124
        st_1chk124 = Radiobutton(displayPage2.inner, value=4, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_1chkType12, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage124Func)
        st_1chk124.deselect()
        global st_1chk131
        st_1chk131 = Radiobutton(displayPage2.inner, value=1, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType13, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage131Func)
        st_1chk131.deselect()
        global st_1chk132
        st_1chk132 = Radiobutton(displayPage2.inner, value=2, text="조금 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_1chkType13, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage132Func)
        st_1chk132.deselect()
        global st_1chk133
        st_1chk133 = Radiobutton(displayPage2.inner, value=3, text="보통으로 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_1chkType13, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage133Func)
        st_1chk133.deselect()
        global st_1chk134
        st_1chk134 = Radiobutton(displayPage2.inner, value=4, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_1chkType13, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage134Func)
        st_1chk134.deselect()
        global st_1chk141
        st_1chk141 = Radiobutton(displayPage2.inner, value=1, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType14, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage141Func)
        st_1chk141.deselect()
        global st_1chk142
        st_1chk142 = Radiobutton(displayPage2.inner, value=2, text="조금 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_1chkType14, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage142Func)
        st_1chk142.deselect()
        global st_1chk143
        st_1chk143 = Radiobutton(displayPage2.inner, value=3, text="보통으로 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_1chkType14, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage143Func)
        st_1chk143.deselect()
        global st_1chk144
        st_1chk144 = Radiobutton(displayPage2.inner, value=4, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_1chkType14, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage144Func)
        st_1chk144.deselect()
        global st_1chk151
        st_1chk151 = Radiobutton(displayPage2.inner, value=4, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType15, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage151Func)
        st_1chk151.deselect()
        global st_1chk152
        st_1chk152 = Radiobutton(displayPage2.inner, value=3, text="조금 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_1chkType15, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage152Func)
        st_1chk152.deselect()
        global st_1chk153
        st_1chk153 = Radiobutton(displayPage2.inner, value=2, text="보통으로 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_1chkType15, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage153Func)
        st_1chk153.deselect()
        global st_1chk154
        st_1chk154 = Radiobutton(displayPage2.inner, value=1, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_1chkType15, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage154Func)
        st_1chk154.deselect()
        global st_1chk161
        st_1chk161 = Radiobutton(displayPage2.inner, value=4, text="전혀 그렇지 않다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType16, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage161Func)
        st_1chk161.deselect()
        global st_1chk162
        st_1chk162 = Radiobutton(displayPage2.inner, value=3, text="조금 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_1chkType16, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage162Func)
        st_1chk162.deselect()
        global st_1chk163
        st_1chk163 = Radiobutton(displayPage2.inner, value=2, text="보통으로 그렇다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_1chkType16, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage163Func)
        st_1chk163.deselect()
        global st_1chk164
        st_1chk164 = Radiobutton(displayPage2.inner, value=1, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_1chkType16, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage164Func)
        st_1chk164.deselect()
        global st_1chk171
        st_1chk171 = Radiobutton(displayPage2.inner, value=1, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType17, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage171Func)
        st_1chk171.deselect()
        global st_1chk172
        st_1chk172 = Radiobutton(displayPage2.inner, value=2, text="조금 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_1chkType17, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage172Func)
        st_1chk172.deselect()
        global st_1chk173
        st_1chk173 = Radiobutton(displayPage2.inner, value=3, text="보통으로 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_1chkType17, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage173Func)
        st_1chk173.deselect()
        global st_1chk174
        st_1chk174 = Radiobutton(displayPage2.inner, value=4, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_1chkType17, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage174Func)
        st_1chk174.deselect()
        global st_1chk181
        st_1chk181 = Radiobutton(displayPage2.inner, value=1, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType18, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage181Func)
        st_1chk181.deselect()
        global st_1chk182
        st_1chk182 = Radiobutton(displayPage2.inner, value=2, text="조금 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_1chkType18, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage182Func)
        st_1chk182.deselect()
        global st_1chk183
        st_1chk183 = Radiobutton(displayPage2.inner, value=3, text="보통으로 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_1chkType18, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage183Func)
        st_1chk183.deselect()
        global st_1chk184
        st_1chk184 = Radiobutton(displayPage2.inner, value=4, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_1chkType18, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage184Func)
        st_1chk184.deselect()
        global st_1chk191
        st_1chk191 = Radiobutton(displayPage2.inner, value=4, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType19, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage191Func)
        st_1chk191.deselect()
        global st_1chk192
        st_1chk192 = Radiobutton(displayPage2.inner, value=3, text="조금 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_1chkType19, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage192Func)
        st_1chk192.deselect()
        global st_1chk193
        st_1chk193 = Radiobutton(displayPage2.inner, value=2, text="보통으로 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_1chkType19, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage193Func)
        st_1chk193.deselect()
        global st_1chk194
        st_1chk194 = Radiobutton(displayPage2.inner, value=1, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_1chkType19, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage194Func)
        st_1chk194.deselect()
        global st_1chk201
        st_1chk201 = Radiobutton(displayPage2.inner, value=4, text="전혀 그렇지 않다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211,
                                 selectimage=chkImage212, indicatoron=False, cursor="circle", variable=st_1chkType20, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage201Func)
        st_1chk201.deselect()
        global st_1chk202
        st_1chk202 = Radiobutton(displayPage2.inner, value=3, text="조금 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=st_1chkType20, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage202Func)
        st_1chk202.deselect()
        global st_1chk203
        st_1chk203 = Radiobutton(displayPage2.inner, value=2, text="보통으로 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=st_1chkType20, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage203Func)
        st_1chk203.deselect()
        global st_1chk204
        st_1chk204 = Radiobutton(displayPage2.inner, value=1, text="대단히 그렇다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=st_1chkType20, font=malgungothic13, bd=0, highlightthickness=0, command=st_1chkImage204Func)
        st_1chk204.deselect()

        st_1chk11.place(x=938, y=472+addheight, width=60, height=21)
        st_1chk12.place(x=1014, y=472+addheight, width=60, height=21)
        st_1chk13.place(x=1088, y=472+addheight, width=60, height=21)
        st_1chk14.place(x=1162, y=472+addheight, width=60, height=21)
        st_1chk21.place(x=938, y=512+addheight, width=60, height=21)
        st_1chk22.place(x=1014, y=512+addheight, width=60, height=21)
        st_1chk23.place(x=1088, y=512+addheight, width=60, height=21)
        st_1chk24.place(x=1162, y=512+addheight, width=60, height=21)
        st_1chk31.place(x=938, y=551+addheight, width=60, height=21)
        st_1chk32.place(x=1014, y=551+addheight, width=60, height=21)
        st_1chk33.place(x=1088, y=551+addheight, width=60, height=21)
        st_1chk34.place(x=1162, y=551+addheight, width=60, height=21)
        st_1chk41.place(x=938, y=589+addheight, width=60, height=21)
        st_1chk42.place(x=1014, y=589+addheight, width=60, height=21)
        st_1chk43.place(x=1088, y=589+addheight, width=60, height=21)
        st_1chk44.place(x=1162, y=589+addheight, width=60, height=21)
        st_1chk51.place(x=938, y=628+addheight, width=60, height=21)
        st_1chk52.place(x=1014, y=628+addheight, width=60, height=21)
        st_1chk53.place(x=1088, y=628+addheight, width=60, height=21)
        st_1chk54.place(x=1162, y=628+addheight, width=60, height=21)
        st_1chk61.place(x=938, y=667+addheight, width=60, height=21)
        st_1chk62.place(x=1014, y=667+addheight, width=60, height=21)
        st_1chk63.place(x=1088, y=667+addheight, width=60, height=21)
        st_1chk64.place(x=1162, y=667+addheight, width=60, height=21)
        st_1chk71.place(x=938, y=704+addheight, width=60, height=21)
        st_1chk72.place(x=1014, y=704+addheight, width=60, height=21)
        st_1chk73.place(x=1088, y=704+addheight, width=60, height=21)
        st_1chk74.place(x=1162, y=704+addheight, width=60, height=21)
        st_1chk81.place(x=938, y=742+addheight, width=60, height=21)
        st_1chk82.place(x=1014, y=742+addheight, width=60, height=21)
        st_1chk83.place(x=1088, y=742+addheight, width=60, height=21)
        st_1chk84.place(x=1162, y=742+addheight, width=60, height=21)
        st_1chk91.place(x=938, y=781+addheight, width=60, height=21)
        st_1chk92.place(x=1014, y=781+addheight, width=60, height=21)
        st_1chk93.place(x=1088, y=781+addheight, width=60, height=21)
        st_1chk94.place(x=1162, y=781+addheight, width=60, height=21)
        st_1chk101.place(x=938, y=821+addheight, width=60, height=21)
        st_1chk102.place(x=1014, y=821+addheight, width=60, height=21)
        st_1chk103.place(x=1088, y=821+addheight, width=60, height=21)
        st_1chk104.place(x=1162, y=821+addheight, width=60, height=21)
        st_1chk111.place(x=938, y=860+addheight, width=60, height=21)
        st_1chk112.place(x=1014, y=860+addheight, width=60, height=21)
        st_1chk113.place(x=1088, y=860+addheight, width=60, height=21)
        st_1chk114.place(x=1162, y=860+addheight, width=60, height=21)
        st_1chk121.place(x=938, y=899+addheight, width=60, height=21)
        st_1chk122.place(x=1014, y=899+addheight, width=60, height=21)
        st_1chk123.place(x=1088, y=899+addheight, width=60, height=21)
        st_1chk124.place(x=1162, y=899+addheight, width=60, height=21)
        st_1chk131.place(x=938, y=938+addheight, width=60, height=21)
        st_1chk132.place(x=1014, y=938+addheight, width=60, height=21)
        st_1chk133.place(x=1088, y=938+addheight, width=60, height=21)
        st_1chk134.place(x=1162, y=938+addheight, width=60, height=21)
        st_1chk141.place(x=938, y=977+addheight, width=60, height=21)
        st_1chk142.place(x=1014, y=977+addheight, width=60, height=21)
        st_1chk143.place(x=1088, y=977+addheight, width=60, height=21)
        st_1chk144.place(x=1162, y=977+addheight, width=60, height=21)
        st_1chk151.place(x=938, y=1016+addheight, width=60, height=21)
        st_1chk152.place(x=1014, y=1016+addheight, width=60, height=21)
        st_1chk153.place(x=1088, y=1016+addheight, width=60, height=21)
        st_1chk154.place(x=1162, y=1016+addheight, width=60, height=21)
        st_1chk161.place(x=938, y=1054+addheight, width=60, height=21)
        st_1chk162.place(x=1014, y=1054+addheight, width=60, height=21)
        st_1chk163.place(x=1088, y=1054+addheight, width=60, height=21)
        st_1chk164.place(x=1162, y=1054+addheight, width=60, height=21)
        st_1chk171.place(x=938, y=1092+addheight, width=60, height=21)
        st_1chk172.place(x=1014, y=1092+addheight, width=60, height=21)
        st_1chk173.place(x=1088, y=1092+addheight, width=60, height=21)
        st_1chk174.place(x=1162, y=1092+addheight, width=60, height=21)
        st_1chk181.place(x=938, y=1133+addheight, width=60, height=21)
        st_1chk182.place(x=1014, y=1133+addheight, width=60, height=21)
        st_1chk183.place(x=1088, y=1133+addheight, width=60, height=21)
        st_1chk184.place(x=1162, y=1133+addheight, width=60, height=21)
        st_1chk191.place(x=938, y=1173+addheight, width=60, height=21)
        st_1chk192.place(x=1014, y=1173+addheight, width=60, height=21)
        st_1chk193.place(x=1088, y=1173+addheight, width=60, height=21)
        st_1chk194.place(x=1162, y=1173+addheight, width=60, height=21)
        st_1chk201.place(x=938, y=1213+addheight, width=60, height=21)
        st_1chk202.place(x=1014, y=1213+addheight, width=60, height=21)
        st_1chk203.place(x=1088, y=1213+addheight, width=60, height=21)
        st_1chk204.place(x=1162, y=1213+addheight, width=60, height=21)

    def hamachkFunc():
        global now
        now = datetime.now()
        global hamachk
        hamachk = True
        sheet7['E2'].value = (" % s년 % s월 % s일 % s시 % s분" %
                              (now.year, now.month, now.day, now.hour, now.minute))
        sheet7['K2'].value = idInput.get()

        def hamachkImage11Func():
            # sumScore()
            if not str(sheet7['C6'].value).strip() == '없다':
                sheet7['C6'].value = hamachk11["text"]

        def hamachkImage12Func():
            # sumScore()
            if not str(sheet7['C6'].value).strip() == '약간':
                sheet7['C6'].value = hamachk12["text"]

        def hamachkImage13Func():
            # sumScore()
            if not str(sheet7['C6'].value).strip() == '중간 정도':
                sheet7['C6'].value = hamachk13["text"]

        def hamachkImage14Func():
            # sumScore()
            if not str(sheet7['C6'].value).strip() == '심함':
                sheet7['C6'].value = hamachk14["text"]

        def hamachkImage15Func():
            # sumScore()
            if not str(sheet7['C6'].value).strip() == '매우 심함':
                sheet7['C6'].value = hamachk15["text"]

        def hamachkImage21Func():
            if not str(sheet7['C9'].value).strip() == '없다':
                sheet7['C9'].value = hamachk21["text"]

        def hamachkImage22Func():
            if not str(sheet7['C9'].value).strip() == '약간':
                sheet7['C9'].value = hamachk22["text"]

        def hamachkImage23Func():
            if not str(sheet7['C9'].value).strip() == '중간 정도':
                sheet7['C9'].value = hamachk23["text"]

        def hamachkImage24Func():
            if not str(sheet7['C9'].value).strip() == '심함':
                sheet7['C9'].value = hamachk24["text"]

        def hamachkImage25Func():
            if not str(sheet7['C9'].value).strip() == '매우 심함':
                sheet7['C9'].value = hamachk25["text"]

        def hamachkImage31Func():
            if not str(sheet7['C12'].value).strip() == '없다':
                sheet7['C12'].value = hamachk31["text"]

        def hamachkImage32Func():
            if not str(sheet7['C12'].value).strip() == '약간':
                sheet7['C12'].value = hamachk32["text"]

        def hamachkImage33Func():
            if not str(sheet7['C12'].value).strip() == '중간 정도':
                sheet7['C12'].value = hamachk33["text"]

        def hamachkImage34Func():
            if not str(sheet7['C12'].value).strip() == '심함':
                sheet7['C12'].value = hamachk34["text"]

        def hamachkImage35Func():
            if not str(sheet7['C12'].value).strip() == '매우 심함':
                sheet7['C12'].value = hamachk35["text"]

        def hamachkImage41Func():
            if not str(sheet7['C15'].value).strip() == '없다':
                sheet7['C15'].value = hamachk41["text"]

        def hamachkImage42Func():
            if not str(sheet7['C15'].value).strip() == '약간':
                sheet7['C15'].value = hamachk42["text"]

        def hamachkImage43Func():
            if not str(sheet7['C15'].value).strip() == '중간 정도':
                sheet7['C15'].value = hamachk43["text"]

        def hamachkImage44Func():
            if not str(sheet7['C15'].value).strip() == '심함':
                sheet7['C15'].value = hamachk44["text"]

        def hamachkImage45Func():
            if not str(sheet7['C15'].value).strip() == '매우 심함':
                sheet7['C15'].value = hamachk45["text"]

        def hamachkImage51Func():
            if not str(sheet7['C18'].value).strip() == '없다':
                sheet7['C18'].value = hamachk51["text"]

        def hamachkImage52Func():
            if not str(sheet7['C18'].value).strip() == '약간':
                sheet7['C18'].value = hamachk52["text"]

        def hamachkImage53Func():
            if not str(sheet7['C18'].value).strip() == '중간 정도':
                sheet7['C18'].value = hamachk53["text"]

        def hamachkImage54Func():
            if not str(sheet7['C18'].value).strip() == '심함':
                sheet7['C18'].value = hamachk54["text"]

        def hamachkImage55Func():
            if not str(sheet7['C18'].value).strip() == '매우 심함':
                sheet7['C18'].value = hamachk55["text"]

        def hamachkImage61Func():
            if not str(sheet7['C21'].value).strip() == '없다':
                sheet7['C21'].value = hamachk61["text"]

        def hamachkImage62Func():
            if not str(sheet7['C21'].value).strip() == '약간':
                sheet7['C21'].value = hamachk62["text"]

        def hamachkImage63Func():
            if not str(sheet7['C21'].value).strip() == '중간 정도':
                sheet7['C21'].value = hamachk63["text"]

        def hamachkImage64Func():
            if not str(sheet7['C21'].value).strip() == '심함':
                sheet7['C21'].value = hamachk64["text"]

        def hamachkImage65Func():
            if not str(sheet7['C21'].value).strip() == '매우 심함':
                sheet7['C21'].value = hamachk65["text"]

        def hamachkImage71Func():
            if not str(sheet7['C24'].value).strip() == '없다':
                sheet7['C24'].value = hamachk71["text"]

        def hamachkImage72Func():
            if not str(sheet7['C24'].value).strip() == '약간':
                sheet7['C24'].value = hamachk72["text"]

        def hamachkImage73Func():
            if not str(sheet7['C24'].value).strip() == '중간 정도':
                sheet7['C24'].value = hamachk73["text"]

        def hamachkImage74Func():
            if not str(sheet7['C24'].value).strip() == '심함':
                sheet7['C24'].value = hamachk74["text"]

        def hamachkImage75Func():
            if not str(sheet7['C24'].value).strip() == '매우 심함':
                sheet7['C24'].value = hamachk75["text"]

        def hamachkImage81Func():
            if not str(sheet7['C27'].value).strip() == '없다':
                sheet7['C27'].value = hamachk81["text"]

        def hamachkImage82Func():
            if not str(sheet7['C27'].value).strip() == '약간':
                sheet7['C27'].value = hamachk82["text"]

        def hamachkImage83Func():
            if not str(sheet7['C27'].value).strip() == '중간 정도':
                sheet7['C27'].value = hamachk83["text"]

        def hamachkImage84Func():
            if not str(sheet7['C27'].value).strip() == '심함':
                sheet7['C27'].value = hamachk84["text"]

        def hamachkImage85Func():
            if not str(sheet7['C27'].value).strip() == '매우 심함':
                sheet7['C27'].value = hamachk85["text"]

        def hamachkImage91Func():
            if not str(sheet7['C30'].value).strip() == '없다':
                sheet7['C30'].value = hamachk91["text"]

        def hamachkImage92Func():
            if not str(sheet7['C30'].value).strip() == '약간':
                sheet7['C30'].value = hamachk92["text"]

        def hamachkImage93Func():
            if not str(sheet7['C30'].value).strip() == '중간 정도':
                sheet7['C30'].value = hamachk93["text"]

        def hamachkImage94Func():
            if not str(sheet7['C30'].value).strip() == '심함':
                sheet7['C30'].value = hamachk94["text"]

        def hamachkImage95Func():
            if not str(sheet7['C30'].value).strip() == '매우 심함':
                sheet7['C30'].value = hamachk95["text"]

        def hamachkImage101Func():
            if not str(sheet7['C33'].value).strip() == '없다':
                sheet7['C33'].value = hamachk101["text"]

        def hamachkImage102Func():
            if not str(sheet7['C33'].value).strip() == '약간':
                sheet7['C33'].value = hamachk102["text"]

        def hamachkImage103Func():
            if not str(sheet7['C33'].value).strip() == '중간 정도':
                sheet7['C33'].value = hamachk103["text"]

        def hamachkImage104Func():
            if not str(sheet7['C33'].value).strip() == '심함':
                sheet7['C33'].value = hamachk104["text"]

        def hamachkImage105Func():
            if not str(sheet7['C33'].value).strip() == '매우 심함':
                sheet7['C33'].value = hamachk105["text"]

        def hamachkImage111Func():
            if not str(sheet7['C36'].value).strip() == '없다':
                sheet7['C36'].value = hamachk111["text"]

        def hamachkImage112Func():
            if not str(sheet7['C36'].value).strip() == '약간':
                sheet7['C36'].value = hamachk112["text"]

        def hamachkImage113Func():
            if not str(sheet7['C36'].value).strip() == '중간 정도':
                sheet7['C36'].value = hamachk113["text"]

        def hamachkImage114Func():
            if not str(sheet7['C36'].value).strip() == '심함':
                sheet7['C36'].value = hamachk114["text"]

        def hamachkImage115Func():
            if not str(sheet7['C36'].value).strip() == '매우 심함':
                sheet7['C36'].value = hamachk115["text"]

        def hamachkImage121Func():
            if not str(sheet7['C39'].value).strip() == '없다':
                sheet7['C39'].value = hamachk121["text"]

        def hamachkImage122Func():
            if not str(sheet7['C39'].value).strip() == '약간':
                sheet7['C39'].value = hamachk122["text"]

        def hamachkImage123Func():
            if not str(sheet7['C39'].value).strip() == '중간 정도':
                sheet7['C39'].value = hamachk123["text"]

        def hamachkImage124Func():
            if not str(sheet7['C39'].value).strip() == '심함':
                sheet7['C39'].value = hamachk124["text"]

        def hamachkImage125Func():
            if not str(sheet7['C39'].value).strip() == '매우 심함':
                sheet7['C39'].value = hamachk125["text"]

        def hamachkImage131Func():
            if not str(sheet7['C42'].value).strip() == '없다':
                sheet7['C42'].value = hamachk131["text"]

        def hamachkImage132Func():
            if not str(sheet7['C42'].value).strip() == '약간':
                sheet7['C42'].value = hamachk132["text"]

        def hamachkImage133Func():
            if not str(sheet7['C42'].value).strip() == '중간 정도':
                sheet7['C42'].value = hamachk133["text"]

        def hamachkImage134Func():
            if not str(sheet7['C42'].value).strip() == '심함':
                sheet7['C42'].value = hamachk134["text"]

        def hamachkImage135Func():
            if not str(sheet7['C42'].value).strip() == '매우 심함':
                sheet7['C42'].value = hamachk135["text"]

        frame7HamaBgImg = Image.open("images/hamabg.png")
        frame7HamaBg = ImageTk.PhotoImage(frame7HamaBgImg)
        frame7HamaBgLabel = tkinter.Label(
            displayPage.inner, image=frame7HamaBg)
        frame7HamaBgLabel.image = frame7HamaBg
        frame7HamaBgLabel.place(x=47, y=420, height=1411)
        frame7HamaBgLabel.pack()

        global hamachk11
        hamachk11 = Radiobutton(displayPage.inner, value=0, text="없다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage201,
                                selectimage=chkImage202, indicatoron=False, cursor="circle", variable=hamachkType1, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage11Func)
        hamachk11.deselect()
        global hamachk12
        hamachk12 = Radiobutton(displayPage.inner, value=1, text="약간",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211, selectimage=chkImage212,
                                indicatoron=False, cursor="circle", variable=hamachkType1, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage12Func)
        hamachk12.deselect()
        global hamachk13
        hamachk13 = Radiobutton(displayPage.inner, value=2, text="중간 정도", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=hamachkType1, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage13Func)
        hamachk13.deselect()
        global hamachk14
        hamachk14 = Radiobutton(displayPage.inner, value=3, text="심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=hamachkType1, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage14Func)
        hamachk14.deselect()
        global hamachk15
        hamachk15 = Radiobutton(displayPage.inner, value=4, text="매우 심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=hamachkType1, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage15Func)
        hamachk15.deselect()
        global hamachk21
        hamachk21 = Radiobutton(displayPage.inner, value=0, text="없다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage201,
                                selectimage=chkImage202, indicatoron=False, cursor="circle", variable=hamachkType2, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage21Func)
        hamachk21.deselect()
        global hamachk22
        hamachk22 = Radiobutton(displayPage.inner, value=1, text="약간", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211, selectimage=chkImage212,
                                indicatoron=False, cursor="circle", variable=hamachkType2, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage22Func)
        hamachk22.deselect()
        global hamachk23
        hamachk23 = Radiobutton(displayPage.inner, value=2, text="중간 정도", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=hamachkType2, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage23Func)
        hamachk23.deselect()
        global hamachk24
        hamachk24 = Radiobutton(displayPage.inner, value=3, text="심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=hamachkType2, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage24Func)
        hamachk24.deselect()
        global hamachk25
        hamachk25 = Radiobutton(displayPage.inner, value=4, text="매우 심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=hamachkType2, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage25Func)
        hamachk25.deselect()
        global hamachk31
        hamachk31 = Radiobutton(displayPage.inner, value=0, text="없다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage201,
                                selectimage=chkImage202, indicatoron=False, cursor="circle", variable=hamachkType3, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage31Func)
        hamachk31.deselect()
        global hamachk32
        hamachk32 = Radiobutton(displayPage.inner, value=1, text="약간", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211, selectimage=chkImage212,
                                indicatoron=False, cursor="circle", variable=hamachkType3, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage32Func)
        hamachk32.deselect()
        global hamachk33
        hamachk33 = Radiobutton(displayPage.inner, value=2, text="중간 정도", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=hamachkType3, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage33Func)
        hamachk33.deselect()
        global hamachk34
        hamachk34 = Radiobutton(displayPage.inner, value=3, text="심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=hamachkType3, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage34Func)
        hamachk34.deselect()
        global hamachk35
        hamachk35 = Radiobutton(displayPage.inner, value=4, text="매우 심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=hamachkType3, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage35Func)
        hamachk35.deselect()
        global hamachk41
        hamachk41 = Radiobutton(displayPage.inner, value=0, text="없다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage201,
                                selectimage=chkImage202, indicatoron=False, cursor="circle", variable=hamachkType4, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage41Func)
        hamachk41.deselect()
        global hamachk42
        hamachk42 = Radiobutton(displayPage.inner, value=1, text="약간", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211, selectimage=chkImage212,
                                indicatoron=False, cursor="circle", variable=hamachkType4, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage42Func)
        hamachk42.deselect()
        global hamachk43
        hamachk43 = Radiobutton(displayPage.inner, value=2, text="중간 정도", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=hamachkType4, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage43Func)
        hamachk43.deselect()
        global hamachk44
        hamachk44 = Radiobutton(displayPage.inner, value=3, text="심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=hamachkType4, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage44Func)
        hamachk44.deselect()
        global hamachk45
        hamachk45 = Radiobutton(displayPage.inner, value=4, text="매우 심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=hamachkType4, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage45Func)
        hamachk45.deselect()
        global hamachk51
        hamachk51 = Radiobutton(displayPage.inner, value=0, text="없다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage201,
                                selectimage=chkImage202, indicatoron=False, cursor="circle", variable=hamachkType5, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage51Func)
        hamachk51.deselect()
        global hamachk52
        hamachk52 = Radiobutton(displayPage.inner, value=1, text="약간", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211, selectimage=chkImage212,
                                indicatoron=False, cursor="circle", variable=hamachkType5, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage52Func)
        hamachk52.deselect()
        global hamachk53
        hamachk53 = Radiobutton(displayPage.inner, value=2, text="중간 정도", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=hamachkType5, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage53Func)
        hamachk53.deselect()
        global hamachk54
        hamachk54 = Radiobutton(displayPage.inner, value=3, text="심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=hamachkType5, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage54Func)
        hamachk54.deselect()
        global hamachk55
        hamachk55 = Radiobutton(displayPage.inner, value=4, text="매우 심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=hamachkType5, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage55Func)
        hamachk55.deselect()
        global hamachk61
        hamachk61 = Radiobutton(displayPage.inner, value=0, text="없다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage201,
                                selectimage=chkImage202, indicatoron=False, cursor="circle", variable=hamachkType6, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage61Func)
        hamachk61.deselect()
        global hamachk62
        hamachk62 = Radiobutton(displayPage.inner, value=1, text="약간", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211, selectimage=chkImage212,
                                indicatoron=False, cursor="circle", variable=hamachkType6, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage62Func)
        hamachk62.deselect()
        global hamachk63
        hamachk63 = Radiobutton(displayPage.inner, value=2, text="중간 정도", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=hamachkType6, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage63Func)
        hamachk63.deselect()
        global hamachk64
        hamachk64 = Radiobutton(displayPage.inner, value=3, text="심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=hamachkType6, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage64Func)
        hamachk64.deselect()
        global hamachk65
        hamachk65 = Radiobutton(displayPage.inner, value=4, text="매우 심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=hamachkType6, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage65Func)
        hamachk65.deselect()

        global hamachk71
        hamachk71 = Radiobutton(displayPage.inner, value=0, text="없다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage201,
                                selectimage=chkImage202, indicatoron=False, cursor="circle", variable=hamachkType7, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage71Func)
        hamachk71.deselect()
        global hamachk72
        hamachk72 = Radiobutton(displayPage.inner, value=1, text="약간", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211, selectimage=chkImage212,
                                indicatoron=False, cursor="circle", variable=hamachkType7, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage72Func)
        hamachk72.deselect()
        global hamachk73
        hamachk73 = Radiobutton(displayPage.inner, value=2, text="중간 정도", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=hamachkType7, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage73Func)
        hamachk73.deselect()
        global hamachk74
        hamachk74 = Radiobutton(displayPage.inner, value=3, text="심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=hamachkType7, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage74Func)
        hamachk74.deselect()
        global hamachk75
        hamachk75 = Radiobutton(displayPage.inner, value=4, text="매우 심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=hamachkType7, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage75Func)
        hamachk75.deselect()

        global hamachk81
        hamachk81 = Radiobutton(displayPage.inner, value=0, text="없다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage201,
                                selectimage=chkImage202, indicatoron=False, cursor="circle", variable=hamachkType8, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage81Func)
        hamachk81.deselect()
        global hamachk82
        hamachk82 = Radiobutton(displayPage.inner, value=1, text="약간", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211, selectimage=chkImage212,
                                indicatoron=False, cursor="circle", variable=hamachkType8, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage82Func)
        hamachk82.deselect()
        global hamachk83
        hamachk83 = Radiobutton(displayPage.inner, value=2, text="중간 정도", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=hamachkType8, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage83Func)
        hamachk83.deselect()
        global hamachk84
        hamachk84 = Radiobutton(displayPage.inner, value=3, text="심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=hamachkType8, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage84Func)
        hamachk84.deselect()
        global hamachk85
        hamachk85 = Radiobutton(displayPage.inner, value=4, text="매우 심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=hamachkType8, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage85Func)
        hamachk85.deselect()

        global hamachk91
        hamachk91 = Radiobutton(displayPage.inner, value=0, text="없다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage201,
                                selectimage=chkImage202, indicatoron=False, cursor="circle", variable=hamachkType9, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage91Func)
        hamachk91.deselect()
        global hamachk92
        hamachk92 = Radiobutton(displayPage.inner, value=1, text="약간", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211, selectimage=chkImage212,
                                indicatoron=False, cursor="circle", variable=hamachkType9, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage92Func)
        hamachk92.deselect()
        global hamachk93
        hamachk93 = Radiobutton(displayPage.inner, value=2, text="중간 정도", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=hamachkType9, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage93Func)
        hamachk93.deselect()
        global hamachk94
        hamachk94 = Radiobutton(displayPage.inner, value=3, text="심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=hamachkType9, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage94Func)
        hamachk94.deselect()
        global hamachk95
        hamachk95 = Radiobutton(displayPage.inner, value=4, text="매우 심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=hamachkType9, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage95Func)
        hamachk95.deselect()
        global hamachk101
        hamachk101 = Radiobutton(displayPage.inner, value=0, text="없다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage201,
                                 selectimage=chkImage202, indicatoron=False, cursor="circle", variable=hamachkType10, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage101Func)
        hamachk101.deselect()
        global hamachk102
        hamachk102 = Radiobutton(displayPage.inner, value=1, text="약간", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211, selectimage=chkImage212,
                                 indicatoron=False, cursor="circle", variable=hamachkType10, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage102Func)
        hamachk102.deselect()
        global hamachk103
        hamachk103 = Radiobutton(displayPage.inner, value=2, text="중간 정도", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=hamachkType10, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage103Func)
        hamachk103.deselect()
        global hamachk104
        hamachk104 = Radiobutton(displayPage.inner, value=3, text="심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=hamachkType10, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage104Func)
        hamachk104.deselect()
        global hamachk105
        hamachk105 = Radiobutton(displayPage.inner, value=4, text="매우 심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=hamachkType10, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage105Func)
        hamachk105.deselect()
        global hamachk111
        hamachk111 = Radiobutton(displayPage.inner, value=0, text="없다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage201,
                                 selectimage=chkImage202, indicatoron=False, cursor="circle", variable=hamachkType11, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage111Func)
        hamachk111.deselect()
        global hamachk112
        hamachk112 = Radiobutton(displayPage.inner, value=1, text="약간", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211, selectimage=chkImage212,
                                 indicatoron=False, cursor="circle", variable=hamachkType11, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage112Func)
        hamachk112.deselect()
        global hamachk113
        hamachk113 = Radiobutton(displayPage.inner, value=2, text="중간 정도", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=hamachkType11, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage113Func)
        hamachk113.deselect()
        global hamachk114
        hamachk114 = Radiobutton(displayPage.inner, value=3, text="심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=hamachkType11, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage114Func)
        hamachk114.deselect()
        global hamachk115
        hamachk115 = Radiobutton(displayPage.inner, value=4, text="매우 심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=hamachkType11, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage115Func)
        hamachk115.deselect()
        global hamachk121
        hamachk121 = Radiobutton(displayPage.inner, value=0, text="없다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage201,
                                 selectimage=chkImage202, indicatoron=False, cursor="circle", variable=hamachkType12, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage121Func)
        hamachk121.deselect()
        global hamachk122
        hamachk122 = Radiobutton(displayPage.inner, value=1, text="약간", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211, selectimage=chkImage212,
                                 indicatoron=False, cursor="circle", variable=hamachkType12, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage122Func)
        hamachk122.deselect()
        global hamachk123
        hamachk123 = Radiobutton(displayPage.inner, value=2, text="중간 정도", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=hamachkType12, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage123Func)
        hamachk123.deselect()
        global hamachk124
        hamachk124 = Radiobutton(displayPage.inner, value=3, text="심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=hamachkType12, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage124Func)
        hamachk124.deselect()
        global hamachk125
        hamachk125 = Radiobutton(displayPage.inner, value=4, text="매우 심함",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241,
                                 selectimage=chkImage242, indicatoron=False, cursor="circle", variable=hamachkType12, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage125Func)
        hamachk125.deselect()
        global hamachk131
        hamachk131 = Radiobutton(displayPage.inner, value=0, text="없다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage201,
                                 selectimage=chkImage202, indicatoron=False, cursor="circle", variable=hamachkType13, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage131Func)
        hamachk131.deselect()
        global hamachk132
        hamachk132 = Radiobutton(displayPage.inner, value=1, text="약간", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211, selectimage=chkImage212,
                                 indicatoron=False, cursor="circle", variable=hamachkType13, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage132Func)
        hamachk132.deselect()
        global hamachk133
        hamachk133 = Radiobutton(displayPage.inner, value=2, text="중간 정도", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=hamachkType13, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage133Func)
        hamachk133.deselect()
        global hamachk134
        hamachk134 = Radiobutton(displayPage.inner, value=3, text="심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=hamachkType13, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage134Func)
        hamachk134.deselect()
        global hamachk135
        hamachk135 = Radiobutton(displayPage.inner, value=4, text="매우 심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=hamachkType13, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage135Func)
        hamachk135.deselect()

        hamachk11.place(x=983, y=494, width=48, height=21)
        hamachk12.place(x=1033, y=494, width=48, height=21)
        hamachk13.place(x=1083, y=494, width=48, height=21)
        hamachk14.place(x=1133, y=494, width=48, height=21)
        hamachk15.place(x=1183, y=494, width=48, height=21)
        hamachk21.place(x=983, y=537, width=48, height=21)
        hamachk22.place(x=1033, y=537, width=48, height=21)
        hamachk23.place(x=1083, y=537, width=48, height=21)
        hamachk24.place(x=1133, y=537, width=48, height=21)
        hamachk25.place(x=1183, y=537, width=48, height=21)
        hamachk31.place(x=983, y=580, width=48, height=21)
        hamachk32.place(x=1033, y=580, width=48, height=21)
        hamachk33.place(x=1083, y=580, width=48, height=21)
        hamachk34.place(x=1133, y=580, width=48, height=21)
        hamachk35.place(x=1183, y=580, width=48, height=21)
        hamachk41.place(x=983, y=623, width=48, height=21)
        hamachk42.place(x=1033, y=623, width=48, height=21)
        hamachk43.place(x=1083, y=623, width=48, height=21)
        hamachk44.place(x=1133, y=623, width=48, height=21)
        hamachk45.place(x=1183, y=623, width=48, height=21)
        hamachk51.place(x=983, y=667, width=48, height=21)
        hamachk52.place(x=1033, y=667, width=48, height=21)
        hamachk53.place(x=1083, y=667, width=48, height=21)
        hamachk54.place(x=1133, y=667, width=48, height=21)
        hamachk55.place(x=1183, y=667, width=48, height=21)
        hamachk61.place(x=983, y=710, width=48, height=21)
        hamachk62.place(x=1033, y=710, width=48, height=21)
        hamachk63.place(x=1083, y=710, width=48, height=21)
        hamachk64.place(x=1133, y=710, width=48, height=21)
        hamachk65.place(x=1183, y=710, width=48, height=21)
        hamachk71.place(x=983, y=753, width=48, height=21)
        hamachk72.place(x=1033, y=753, width=48, height=21)
        hamachk73.place(x=1083, y=753, width=48, height=21)
        hamachk74.place(x=1133, y=753, width=48, height=21)
        hamachk75.place(x=1183, y=753, width=48, height=21)
        hamachk81.place(x=983, y=796, width=48, height=21)
        hamachk82.place(x=1033, y=796, width=48, height=21)
        hamachk83.place(x=1083, y=796, width=48, height=21)
        hamachk84.place(x=1133, y=796, width=48, height=21)
        hamachk85.place(x=1183, y=796, width=48, height=21)
        hamachk91.place(x=983, y=839, width=48, height=21)
        hamachk92.place(x=1033, y=839, width=48, height=21)
        hamachk93.place(x=1083, y=839, width=48, height=21)
        hamachk94.place(x=1133, y=839, width=48, height=21)
        hamachk95.place(x=1183, y=839, width=48, height=21)
        hamachk101.place(x=983, y=882, width=48, height=21)
        hamachk102.place(x=1033, y=882, width=48, height=21)
        hamachk103.place(x=1083, y=882, width=48, height=21)
        hamachk104.place(x=1133, y=882, width=48, height=21)
        hamachk105.place(x=1183, y=882, width=48, height=21)
        hamachk111.place(x=983, y=945, width=48, height=21)
        hamachk112.place(x=1033, y=945, width=48, height=21)
        hamachk113.place(x=1083, y=945, width=48, height=21)
        hamachk114.place(x=1133, y=945, width=48, height=21)
        hamachk115.place(x=1183, y=945, width=48, height=21)
        hamachk121.place(x=983, y=1030, width=48, height=21)
        hamachk122.place(x=1033, y=1030, width=48, height=21)
        hamachk123.place(x=1083, y=1030, width=48, height=21)
        hamachk124.place(x=1133, y=1030, width=48, height=21)
        hamachk125.place(x=1183, y=1030, width=48, height=21)
        hamachk131.place(x=983, y=1097, width=48, height=21)
        hamachk132.place(x=1033, y=1097, width=48, height=21)
        hamachk133.place(x=1083, y=1097, width=48, height=21)
        hamachk134.place(x=1133, y=1097, width=48, height=21)
        hamachk135.place(x=1183, y=1097, width=48, height=21)

    def hamachkFunc2():
        global now
        now = datetime.now()
        global hamachk
        hamachk = True
        sheet7['E2'].value = (" % s년 % s월 % s일 % s시 % s분" %
                              (now.year, now.month, now.day, now.hour, now.minute))
        sheet7['K2'].value = idInput.get()

        def hamachkImage11Func():
            # sumScore()
            if not str(sheet7['C6'].value).strip() == '없다':
                sheet7['C6'].value = hamachk11["text"]

        def hamachkImage12Func():
            # sumScore()
            if not str(sheet7['C6'].value).strip() == '약간':
                sheet7['C6'].value = hamachk12["text"]

        def hamachkImage13Func():
            # sumScore()
            if not str(sheet7['C6'].value).strip() == '중간 정도':
                sheet7['C6'].value = hamachk13["text"]

        def hamachkImage14Func():
            # sumScore()
            if not str(sheet7['C6'].value).strip() == '심함':
                sheet7['C6'].value = hamachk14["text"]

        def hamachkImage15Func():
            # sumScore()
            if not str(sheet7['C6'].value).strip() == '매우 심함':
                sheet7['C6'].value = hamachk15["text"]

        def hamachkImage21Func():
            if not str(sheet7['C9'].value).strip() == '없다':
                sheet7['C9'].value = hamachk21["text"]

        def hamachkImage22Func():
            if not str(sheet7['C9'].value).strip() == '약간':
                sheet7['C9'].value = hamachk22["text"]

        def hamachkImage23Func():
            if not str(sheet7['C9'].value).strip() == '중간 정도':
                sheet7['C9'].value = hamachk23["text"]

        def hamachkImage24Func():
            if not str(sheet7['C9'].value).strip() == '심함':
                sheet7['C9'].value = hamachk24["text"]

        def hamachkImage25Func():
            if not str(sheet7['C9'].value).strip() == '매우 심함':
                sheet7['C9'].value = hamachk24["text"]

        def hamachkImage31Func():
            if not str(sheet7['C12'].value).strip() == '없다':
                sheet7['C12'].value = hamachk31["text"]

        def hamachkImage32Func():
            if not str(sheet7['C12'].value).strip() == '약간':
                sheet7['C12'].value = hamachk32["text"]

        def hamachkImage33Func():
            if not str(sheet7['C12'].value).strip() == '중간 정도':
                sheet7['C12'].value = hamachk33["text"]

        def hamachkImage34Func():
            if not str(sheet7['C12'].value).strip() == '심함':
                sheet7['C12'].value = hamachk34["text"]

        def hamachkImage35Func():
            if not str(sheet7['C12'].value).strip() == '매우 심함':
                sheet7['C12'].value = hamachk34["text"]

        def hamachkImage41Func():
            if not str(sheet7['C15'].value).strip() == '없다':
                sheet7['C15'].value = hamachk41["text"]

        def hamachkImage42Func():
            if not str(sheet7['C15'].value).strip() == '약간':
                sheet7['C15'].value = hamachk42["text"]

        def hamachkImage43Func():
            if not str(sheet7['C15'].value).strip() == '중간 정도':
                sheet7['C15'].value = hamachk43["text"]

        def hamachkImage44Func():
            if not str(sheet7['C15'].value).strip() == '심함':
                sheet7['C15'].value = hamachk44["text"]

        def hamachkImage45Func():
            if not str(sheet7['C15'].value).strip() == '매우 심함':
                sheet7['C15'].value = hamachk44["text"]

        def hamachkImage51Func():
            if not str(sheet7['C18'].value).strip() == '없다':
                sheet7['C18'].value = hamachk51["text"]

        def hamachkImage52Func():
            if not str(sheet7['C18'].value).strip() == '약간':
                sheet7['C18'].value = hamachk52["text"]

        def hamachkImage53Func():
            if not str(sheet7['C18'].value).strip() == '중간 정도':
                sheet7['C18'].value = hamachk53["text"]

        def hamachkImage54Func():
            if not str(sheet7['C18'].value).strip() == '심함':
                sheet7['C18'].value = hamachk54["text"]

        def hamachkImage55Func():
            if not str(sheet7['C18'].value).strip() == '매우 심함':
                sheet7['C18'].value = hamachk54["text"]

        def hamachkImage61Func():
            if not str(sheet7['C21'].value).strip() == '없다':
                sheet7['C21'].value = hamachk61["text"]

        def hamachkImage62Func():
            if not str(sheet7['C21'].value).strip() == '약간':
                sheet7['C21'].value = hamachk62["text"]

        def hamachkImage63Func():
            if not str(sheet7['C21'].value).strip() == '중간 정도':
                sheet7['C21'].value = hamachk63["text"]

        def hamachkImage64Func():
            if not str(sheet7['C21'].value).strip() == '심함':
                sheet7['C21'].value = hamachk64["text"]

        def hamachkImage65Func():
            if not str(sheet7['C21'].value).strip() == '매우 심함':
                sheet7['C21'].value = hamachk64["text"]

        def hamachkImage71Func():
            if not str(sheet7['C24'].value).strip() == '없다':
                sheet7['C24'].value = hamachk71["text"]

        def hamachkImage72Func():
            if not str(sheet7['C24'].value).strip() == '약간':
                sheet7['C24'].value = hamachk72["text"]

        def hamachkImage73Func():
            if not str(sheet7['C24'].value).strip() == '중간 정도':
                sheet7['C24'].value = hamachk73["text"]

        def hamachkImage74Func():
            if not str(sheet7['C24'].value).strip() == '심함':
                sheet7['C24'].value = hamachk74["text"]

        def hamachkImage75Func():
            if not str(sheet7['C24'].value).strip() == '매우 심함':
                sheet7['C24'].value = hamachk74["text"]

        def hamachkImage81Func():
            if not str(sheet7['C27'].value).strip() == '없다':
                sheet7['C27'].value = hamachk81["text"]

        def hamachkImage82Func():
            if not str(sheet7['C27'].value).strip() == '약간':
                sheet7['C27'].value = hamachk82["text"]

        def hamachkImage83Func():
            if not str(sheet7['C27'].value).strip() == '중간 정도':
                sheet7['C27'].value = hamachk83["text"]

        def hamachkImage84Func():
            if not str(sheet7['C27'].value).strip() == '심함':
                sheet7['C27'].value = hamachk84["text"]

        def hamachkImage85Func():
            if not str(sheet7['C27'].value).strip() == '매우 심함':
                sheet7['C27'].value = hamachk84["text"]

        def hamachkImage91Func():
            if not str(sheet7['C30'].value).strip() == '없다':
                sheet7['C30'].value = hamachk91["text"]

        def hamachkImage92Func():
            if not str(sheet7['C30'].value).strip() == '약간':
                sheet7['C30'].value = hamachk92["text"]

        def hamachkImage93Func():
            if not str(sheet7['C30'].value).strip() == '중간 정도':
                sheet7['C30'].value = hamachk93["text"]

        def hamachkImage94Func():
            if not str(sheet7['C30'].value).strip() == '심함':
                sheet7['C30'].value = hamachk94["text"]

        def hamachkImage95Func():
            if not str(sheet7['C30'].value).strip() == '매우 심함':
                sheet7['C30'].value = hamachk94["text"]

        def hamachkImage101Func():
            if not str(sheet7['C33'].value).strip() == '없다':
                sheet7['C33'].value = hamachk101["text"]

        def hamachkImage102Func():
            if not str(sheet7['C33'].value).strip() == '약간':
                sheet7['C33'].value = hamachk102["text"]

        def hamachkImage103Func():
            if not str(sheet7['C33'].value).strip() == '중간 정도':
                sheet7['C33'].value = hamachk103["text"]

        def hamachkImage104Func():
            if not str(sheet7['C33'].value).strip() == '심함':
                sheet7['C33'].value = hamachk104["text"]

        def hamachkImage105Func():
            if not str(sheet7['C33'].value).strip() == '매우 심함':
                sheet7['C33'].value = hamachk104["text"]

        def hamachkImage111Func():
            if not str(sheet7['C36'].value).strip() == '없다':
                sheet7['C36'].value = hamachk111["text"]

        def hamachkImage112Func():
            if not str(sheet7['C36'].value).strip() == '약간':
                sheet7['C36'].value = hamachk112["text"]

        def hamachkImage113Func():
            if not str(sheet7['C36'].value).strip() == '중간 정도':
                sheet7['C36'].value = hamachk113["text"]

        def hamachkImage114Func():
            if not str(sheet7['C36'].value).strip() == '심함':
                sheet7['C36'].value = hamachk114["text"]

        def hamachkImage115Func():
            if not str(sheet7['C36'].value).strip() == '매우 심함':
                sheet7['C36'].value = hamachk114["text"]

        def hamachkImage121Func():
            if not str(sheet7['C39'].value).strip() == '없다':
                sheet7['C39'].value = hamachk121["text"]

        def hamachkImage122Func():
            if not str(sheet7['C39'].value).strip() == '약간':
                sheet7['C39'].value = hamachk122["text"]

        def hamachkImage123Func():
            if not str(sheet7['C39'].value).strip() == '중간 정도':
                sheet7['C39'].value = hamachk123["text"]

        def hamachkImage124Func():
            if not str(sheet7['C39'].value).strip() == '심함':
                sheet7['C39'].value = hamachk124["text"]

        def hamachkImage125Func():
            if not str(sheet7['C39'].value).strip() == '매우 심함':
                sheet7['C39'].value = hamachk124["text"]

        def hamachkImage131Func():
            if not str(sheet7['C42'].value).strip() == '없다':
                sheet7['C42'].value = hamachk131["text"]

        def hamachkImage132Func():
            if not str(sheet7['C42'].value).strip() == '약간':
                sheet7['C42'].value = hamachk132["text"]

        def hamachkImage133Func():
            if not str(sheet7['C42'].value).strip() == '중간 정도':
                sheet7['C42'].value = hamachk133["text"]

        def hamachkImage134Func():
            if not str(sheet7['C42'].value).strip() == '심함':
                sheet7['C42'].value = hamachk134["text"]

        def hamachkImage135Func():
            if not str(sheet7['C42'].value).strip() == '매우 심함':
                sheet7['C42'].value = hamachk134["text"]

        frame7HamaBgImg = Image.open("images/hamabg.png")
        frame7HamaBg = ImageTk.PhotoImage(frame7HamaBgImg)
        frame7HamaBgLabel = tkinter.Label(
            displayPage2.inner, image=frame7HamaBg)
        frame7HamaBgLabel.image = frame7HamaBg
        frame7HamaBgLabel.place(x=47, y=420, height=1411)
        frame7HamaBgLabel.pack()

        global hamachk11
        hamachk11 = Radiobutton(displayPage2.inner, value=0, text="없다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage201,
                                selectimage=chkImage202, indicatoron=False, cursor="circle", variable=hamachkType1, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage11Func)
        hamachk11.deselect()
        global hamachk12
        hamachk12 = Radiobutton(displayPage2.inner, value=1, text="약간",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211, selectimage=chkImage212,
                                indicatoron=False, cursor="circle", variable=hamachkType1, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage12Func)
        hamachk12.deselect()
        global hamachk13
        hamachk13 = Radiobutton(displayPage2.inner, value=2, text="중간 정도", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=hamachkType1, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage13Func)
        hamachk13.deselect()
        global hamachk14
        hamachk14 = Radiobutton(displayPage2.inner, value=3, text="심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=hamachkType1, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage14Func)
        hamachk14.deselect()
        global hamachk15
        hamachk15 = Radiobutton(displayPage2.inner, value=4, text="매우 심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=hamachkType1, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage15Func)
        hamachk15.deselect()
        global hamachk21
        hamachk21 = Radiobutton(displayPage2.inner, value=0, text="없다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage201,
                                selectimage=chkImage202, indicatoron=False, cursor="circle", variable=hamachkType2, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage21Func)
        hamachk21.deselect()
        global hamachk22
        hamachk22 = Radiobutton(displayPage2.inner, value=1, text="약간", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211, selectimage=chkImage212,
                                indicatoron=False, cursor="circle", variable=hamachkType2, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage22Func)
        hamachk22.deselect()
        global hamachk23
        hamachk23 = Radiobutton(displayPage2.inner, value=2, text="중간 정도", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=hamachkType2, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage23Func)
        hamachk23.deselect()
        global hamachk24
        hamachk24 = Radiobutton(displayPage2.inner, value=3, text="심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=hamachkType2, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage24Func)
        hamachk24.deselect()
        global hamachk25
        hamachk25 = Radiobutton(displayPage2.inner, value=4, text="매우 심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=hamachkType2, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage25Func)
        hamachk25.deselect()
        global hamachk31
        hamachk31 = Radiobutton(displayPage2.inner, value=0, text="없다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage201,
                                selectimage=chkImage202, indicatoron=False, cursor="circle", variable=hamachkType3, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage31Func)
        hamachk31.deselect()
        global hamachk32
        hamachk32 = Radiobutton(displayPage2.inner, value=1, text="약간", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211, selectimage=chkImage212,
                                indicatoron=False, cursor="circle", variable=hamachkType3, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage32Func)
        hamachk32.deselect()
        global hamachk33
        hamachk33 = Radiobutton(displayPage2.inner, value=2, text="중간 정도", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=hamachkType3, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage33Func)
        hamachk33.deselect()
        global hamachk34
        hamachk34 = Radiobutton(displayPage2.inner, value=3, text="심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=hamachkType3, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage34Func)
        hamachk34.deselect()
        global hamachk35
        hamachk35 = Radiobutton(displayPage2.inner, value=4, text="매우 심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=hamachkType3, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage35Func)
        hamachk35.deselect()
        global hamachk41
        hamachk41 = Radiobutton(displayPage2.inner, value=0, text="없다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage201,
                                selectimage=chkImage202, indicatoron=False, cursor="circle", variable=hamachkType4, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage41Func)
        hamachk41.deselect()
        global hamachk42
        hamachk42 = Radiobutton(displayPage2.inner, value=1, text="약간", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211, selectimage=chkImage212,
                                indicatoron=False, cursor="circle", variable=hamachkType4, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage42Func)
        hamachk42.deselect()
        global hamachk43
        hamachk43 = Radiobutton(displayPage2.inner, value=2, text="중간 정도", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=hamachkType4, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage43Func)
        hamachk43.deselect()
        global hamachk44
        hamachk44 = Radiobutton(displayPage2.inner, value=3, text="심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=hamachkType4, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage44Func)
        hamachk44.deselect()
        global hamachk45
        hamachk45 = Radiobutton(displayPage2.inner, value=4, text="매우 심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=hamachkType4, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage45Func)
        hamachk45.deselect()
        global hamachk51
        hamachk51 = Radiobutton(displayPage2.inner, value=0, text="없다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage201,
                                selectimage=chkImage202, indicatoron=False, cursor="circle", variable=hamachkType5, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage51Func)
        hamachk51.deselect()
        global hamachk52
        hamachk52 = Radiobutton(displayPage2.inner, value=1, text="약간", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211, selectimage=chkImage212,
                                indicatoron=False, cursor="circle", variable=hamachkType5, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage52Func)
        hamachk52.deselect()
        global hamachk53
        hamachk53 = Radiobutton(displayPage2.inner, value=2, text="중간 정도", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=hamachkType5, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage53Func)
        hamachk53.deselect()
        global hamachk54
        hamachk54 = Radiobutton(displayPage2.inner, value=3, text="심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=hamachkType5, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage54Func)
        hamachk54.deselect()
        global hamachk55
        hamachk55 = Radiobutton(displayPage2.inner, value=4, text="매우 심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=hamachkType5, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage55Func)
        hamachk55.deselect()
        global hamachk61
        hamachk61 = Radiobutton(displayPage2.inner, value=0, text="없다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage201,
                                selectimage=chkImage202, indicatoron=False, cursor="circle", variable=hamachkType6, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage61Func)
        hamachk61.deselect()
        global hamachk62
        hamachk62 = Radiobutton(displayPage2.inner, value=1, text="약간", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211, selectimage=chkImage212,
                                indicatoron=False, cursor="circle", variable=hamachkType6, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage62Func)
        hamachk62.deselect()
        global hamachk63
        hamachk63 = Radiobutton(displayPage2.inner, value=2, text="중간 정도", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=hamachkType6, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage63Func)
        hamachk63.deselect()
        global hamachk64
        hamachk64 = Radiobutton(displayPage2.inner, value=3, text="심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=hamachkType6, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage64Func)
        hamachk64.deselect()
        global hamachk65
        hamachk65 = Radiobutton(displayPage2.inner, value=4, text="매우 심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=hamachkType6, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage65Func)
        hamachk65.deselect()

        global hamachk71
        hamachk71 = Radiobutton(displayPage2.inner, value=0, text="없다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage201,
                                selectimage=chkImage202, indicatoron=False, cursor="circle", variable=hamachkType7, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage71Func)
        hamachk71.deselect()
        global hamachk72
        hamachk72 = Radiobutton(displayPage2.inner, value=1, text="약간", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211, selectimage=chkImage212,
                                indicatoron=False, cursor="circle", variable=hamachkType7, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage72Func)
        hamachk72.deselect()
        global hamachk73
        hamachk73 = Radiobutton(displayPage2.inner, value=2, text="중간 정도", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=hamachkType7, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage73Func)
        hamachk73.deselect()
        global hamachk74
        hamachk74 = Radiobutton(displayPage2.inner, value=3, text="심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=hamachkType7, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage74Func)
        hamachk74.deselect()
        global hamachk75
        hamachk75 = Radiobutton(displayPage2.inner, value=4, text="매우 심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=hamachkType7, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage75Func)
        hamachk75.deselect()

        global hamachk81
        hamachk81 = Radiobutton(displayPage2.inner, value=0, text="없다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage201,
                                selectimage=chkImage202, indicatoron=False, cursor="circle", variable=hamachkType8, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage81Func)
        hamachk81.deselect()
        global hamachk82
        hamachk82 = Radiobutton(displayPage2.inner, value=1, text="약간", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211, selectimage=chkImage212,
                                indicatoron=False, cursor="circle", variable=hamachkType8, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage82Func)
        hamachk82.deselect()
        global hamachk83
        hamachk83 = Radiobutton(displayPage2.inner, value=2, text="중간 정도", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=hamachkType8, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage83Func)
        hamachk83.deselect()
        global hamachk84
        hamachk84 = Radiobutton(displayPage2.inner, value=3, text="심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=hamachkType8, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage84Func)
        hamachk84.deselect()
        global hamachk85
        hamachk85 = Radiobutton(displayPage2.inner, value=4, text="매우 심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=hamachkType8, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage85Func)
        hamachk85.deselect()

        global hamachk91
        hamachk91 = Radiobutton(displayPage2.inner, value=0, text="없다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage201,
                                selectimage=chkImage202, indicatoron=False, cursor="circle", variable=hamachkType9, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage91Func)
        hamachk91.deselect()
        global hamachk92
        hamachk92 = Radiobutton(displayPage2.inner, value=1, text="약간", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211, selectimage=chkImage212,
                                indicatoron=False, cursor="circle", variable=hamachkType9, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage92Func)
        hamachk92.deselect()
        global hamachk93
        hamachk93 = Radiobutton(displayPage2.inner, value=2, text="중간 정도", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                indicatoron=False, cursor="circle", variable=hamachkType9, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage93Func)
        hamachk93.deselect()
        global hamachk94
        hamachk94 = Radiobutton(displayPage2.inner, value=3, text="심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                indicatoron=False, cursor="circle", variable=hamachkType9, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage94Func)
        hamachk94.deselect()
        global hamachk95
        hamachk95 = Radiobutton(displayPage2.inner, value=4, text="매우 심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                indicatoron=False, cursor="circle", variable=hamachkType9, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage95Func)
        hamachk95.deselect()
        global hamachk101
        hamachk101 = Radiobutton(displayPage2.inner, value=0, text="없다", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage201,
                                 selectimage=chkImage202, indicatoron=False, cursor="circle", variable=hamachkType10, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage101Func)
        hamachk101.deselect()
        global hamachk102
        hamachk102 = Radiobutton(displayPage2.inner, value=1, text="약간", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211, selectimage=chkImage212,
                                 indicatoron=False, cursor="circle", variable=hamachkType10, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage102Func)
        hamachk102.deselect()
        global hamachk103
        hamachk103 = Radiobutton(displayPage2.inner, value=2, text="중간 정도", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=hamachkType10, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage103Func)
        hamachk103.deselect()
        global hamachk104
        hamachk104 = Radiobutton(displayPage2.inner, value=3, text="심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=hamachkType10, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage104Func)
        hamachk104.deselect()
        global hamachk105
        hamachk105 = Radiobutton(displayPage2.inner, value=4, text="매우 심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=hamachkType10, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage105Func)
        hamachk105.deselect()
        global hamachk111
        hamachk111 = Radiobutton(displayPage2.inner, value=0, text="없다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage201,
                                 selectimage=chkImage202, indicatoron=False, cursor="circle", variable=hamachkType11, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage111Func)
        hamachk111.deselect()
        global hamachk112
        hamachk112 = Radiobutton(displayPage2.inner, value=1, text="약간", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211, selectimage=chkImage212,
                                 indicatoron=False, cursor="circle", variable=hamachkType11, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage112Func)
        hamachk112.deselect()
        global hamachk113
        hamachk113 = Radiobutton(displayPage2.inner, value=2, text="중간 정도", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=hamachkType11, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage113Func)
        hamachk113.deselect()
        global hamachk114
        hamachk114 = Radiobutton(displayPage2.inner, value=3, text="심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=hamachkType11, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage114Func)
        hamachk114.deselect()
        global hamachk115
        hamachk115 = Radiobutton(displayPage2.inner, value=4, text="매우 심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=hamachkType11, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage115Func)
        hamachk115.deselect()
        global hamachk121
        hamachk121 = Radiobutton(displayPage2.inner, value=0, text="없다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage201,
                                 selectimage=chkImage202, indicatoron=False, cursor="circle", variable=hamachkType12, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage121Func)
        hamachk121.deselect()
        global hamachk122
        hamachk122 = Radiobutton(displayPage2.inner, value=1, text="약간", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211, selectimage=chkImage212,
                                 indicatoron=False, cursor="circle", variable=hamachkType12, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage122Func)
        hamachk122.deselect()
        global hamachk123
        hamachk123 = Radiobutton(displayPage2.inner, value=2, text="중간 정도", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=hamachkType12, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage123Func)
        hamachk123.deselect()
        global hamachk124
        hamachk124 = Radiobutton(displayPage2.inner, value=3, text="심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=hamachkType12, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage124Func)
        hamachk124.deselect()
        global hamachk125
        hamachk125 = Radiobutton(displayPage2.inner, value=4, text="매우 심함",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241,
                                 selectimage=chkImage242, indicatoron=False, cursor="circle", variable=hamachkType12, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage125Func)
        hamachk125.deselect()
        global hamachk131
        hamachk131 = Radiobutton(displayPage2.inner, value=0, text="없다",  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage201,
                                 selectimage=chkImage202, indicatoron=False, cursor="circle", variable=hamachkType13, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage131Func)
        hamachk131.deselect()
        global hamachk132
        hamachk132 = Radiobutton(displayPage2.inner, value=1, text="약간", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage211, selectimage=chkImage212,
                                 indicatoron=False, cursor="circle", variable=hamachkType13, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage132Func)
        hamachk132.deselect()
        global hamachk133
        hamachk133 = Radiobutton(displayPage2.inner, value=2, text="중간 정도", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage221, selectimage=chkImage222,
                                 indicatoron=False, cursor="circle", variable=hamachkType13, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage133Func)
        hamachk133.deselect()
        global hamachk134
        hamachk134 = Radiobutton(displayPage2.inner, value=3, text="심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage231, selectimage=chkImage232,
                                 indicatoron=False, cursor="circle", variable=hamachkType13, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage134Func)
        hamachk134.deselect()
        global hamachk135
        hamachk135 = Radiobutton(displayPage2.inner, value=4, text="매우 심함", background="#FFFFFF", activebackground="#FFFFFF", image=chkImage241, selectimage=chkImage242,
                                 indicatoron=False, cursor="circle", variable=hamachkType13, font=malgungothic13, bd=0, highlightthickness=0, command=hamachkImage135Func)
        hamachk135.deselect()

        hamachk11.place(x=983, y=494+addheight, width=48, height=21)
        hamachk12.place(x=1033, y=494+addheight, width=48, height=21)
        hamachk13.place(x=1083, y=494+addheight, width=48, height=21)
        hamachk14.place(x=1133, y=494+addheight, width=48, height=21)
        hamachk15.place(x=1183, y=494+addheight, width=48, height=21)
        hamachk21.place(x=983, y=537+addheight, width=48, height=21)
        hamachk22.place(x=1033, y=537+addheight, width=48, height=21)
        hamachk23.place(x=1083, y=537+addheight, width=48, height=21)
        hamachk24.place(x=1133, y=537+addheight, width=48, height=21)
        hamachk25.place(x=1183, y=537+addheight, width=48, height=21)
        hamachk31.place(x=983, y=580+addheight, width=48, height=21)
        hamachk32.place(x=1033, y=580+addheight, width=48, height=21)
        hamachk33.place(x=1083, y=580+addheight, width=48, height=21)
        hamachk34.place(x=1133, y=580+addheight, width=48, height=21)
        hamachk35.place(x=1183, y=580+addheight, width=48, height=21)
        hamachk41.place(x=983, y=623+addheight, width=48, height=21)
        hamachk42.place(x=1033, y=623+addheight, width=48, height=21)
        hamachk43.place(x=1083, y=623+addheight, width=48, height=21)
        hamachk44.place(x=1133, y=623+addheight, width=48, height=21)
        hamachk45.place(x=1183, y=623+addheight, width=48, height=21)
        hamachk51.place(x=983, y=667+addheight, width=48, height=21)
        hamachk52.place(x=1033, y=667+addheight, width=48, height=21)
        hamachk53.place(x=1083, y=667+addheight, width=48, height=21)
        hamachk54.place(x=1133, y=667+addheight, width=48, height=21)
        hamachk55.place(x=1183, y=667+addheight, width=48, height=21)
        hamachk61.place(x=983, y=710+addheight, width=48, height=21)
        hamachk62.place(x=1033, y=710+addheight, width=48, height=21)
        hamachk63.place(x=1083, y=710+addheight, width=48, height=21)
        hamachk64.place(x=1133, y=710+addheight, width=48, height=21)
        hamachk65.place(x=1183, y=710+addheight, width=48, height=21)
        hamachk71.place(x=983, y=753+addheight, width=48, height=21)
        hamachk72.place(x=1033, y=753+addheight, width=48, height=21)
        hamachk73.place(x=1083, y=753+addheight, width=48, height=21)
        hamachk74.place(x=1133, y=753+addheight, width=48, height=21)
        hamachk75.place(x=1183, y=753+addheight, width=48, height=21)
        hamachk81.place(x=983, y=796+addheight, width=48, height=21)
        hamachk82.place(x=1033, y=796+addheight, width=48, height=21)
        hamachk83.place(x=1083, y=796+addheight, width=48, height=21)
        hamachk84.place(x=1133, y=796+addheight, width=48, height=21)
        hamachk85.place(x=1183, y=796+addheight, width=48, height=21)
        hamachk91.place(x=983, y=839+addheight, width=48, height=21)
        hamachk92.place(x=1033, y=839+addheight, width=48, height=21)
        hamachk93.place(x=1083, y=839+addheight, width=48, height=21)
        hamachk94.place(x=1133, y=839+addheight, width=48, height=21)
        hamachk95.place(x=1183, y=839+addheight, width=48, height=21)
        hamachk101.place(x=983, y=882+addheight, width=48, height=21)
        hamachk102.place(x=1033, y=882+addheight, width=48, height=21)
        hamachk103.place(x=1083, y=882+addheight, width=48, height=21)
        hamachk104.place(x=1133, y=882+addheight, width=48, height=21)
        hamachk105.place(x=1183, y=882+addheight, width=48, height=21)
        hamachk111.place(x=983, y=945+addheight, width=48, height=21)
        hamachk112.place(x=1033, y=945+addheight, width=48, height=21)
        hamachk113.place(x=1083, y=945+addheight, width=48, height=21)
        hamachk114.place(x=1133, y=945+addheight, width=48, height=21)
        hamachk115.place(x=1183, y=945+addheight, width=48, height=21)
        hamachk121.place(x=983, y=1030+addheight, width=48, height=21)
        hamachk122.place(x=1033, y=1030+addheight, width=48, height=21)
        hamachk123.place(x=1083, y=1030+addheight, width=48, height=21)
        hamachk124.place(x=1133, y=1030+addheight, width=48, height=21)
        hamachk125.place(x=1183, y=1030+addheight, width=48, height=21)
        hamachk131.place(x=983, y=1097+addheight, width=48, height=21)
        hamachk132.place(x=1033, y=1097+addheight, width=48, height=21)
        hamachk133.place(x=1083, y=1097+addheight, width=48, height=21)
        hamachk134.place(x=1133, y=1097+addheight, width=48, height=21)
        hamachk135.place(x=1183, y=1097+addheight, width=48, height=21)

    def hamdchkFunc():
        global now
        now = datetime.now()
        global hamdchk
        hamdchk = True
        sheet8['E2'].value = (" % s년 % s월 % s일 % s시 % s분" %
                              (now.year, now.month, now.day, now.hour, now.minute))
        sheet8['K2'].value = idInput.get()

        def hamdchkImage11Func():
            if not str(sheet8['C11'].value).strip() == 0:
                sheet8['C11'].value = str(int(hamdchkType1.get())+1) + "."

        def hamdchkImage12Func():
            if not str(sheet8['C11'].value).strip() == 1:
                sheet8['C11'].value = str(int(hamdchkType1.get())+1) + "."

        def hamdchkImage13Func():
            if not str(sheet8['C11'].value).strip() == 2:
                sheet8['C11'].value = str(int(hamdchkType1.get())+1) + "."

        def hamdchkImage14Func():
            if not str(sheet8['C11'].value).strip() == 3:
                sheet8['C11'].value = str(int(hamdchkType1.get())+1) + "."

        def hamdchkImage15Func():
            if not str(sheet8['C11'].value).strip() == 4:
                sheet8['C11'].value = str(int(hamdchkType1.get())+1) + "."

        def hamdchkImage21Func():
            if not str(sheet8['C19'].value).strip() == 0:
                sheet8['C19'].value = str(int(hamdchkType2.get())+1) + "."

        def hamdchkImage22Func():
            if not str(sheet8['C19'].value).strip() == 1:
                sheet8['C19'].value = str(int(hamdchkType2.get())+1) + "."

        def hamdchkImage23Func():
            if not str(sheet8['C19'].value).strip() == 2:
                sheet8['C19'].value = str(int(hamdchkType2.get())+1) + "."

        def hamdchkImage24Func():
            if not str(sheet8['C19'].value).strip() == 3:
                sheet8['C19'].value = str(int(hamdchkType2.get())+1) + "."

        def hamdchkImage25Func():
            if not str(sheet8['C19'].value).strip() == 4:
                sheet8['C19'].value = str(int(hamdchkType2.get())+1) + "."

        def hamdchkImage31Func():
            if not str(sheet8['C25'].value).strip() == 0:
                sheet8['C25'].value = str(int(hamdchkType3.get())+1) + "."

        def hamdchkImage32Func():
            if not str(sheet8['C25'].value).strip() == 1:
                sheet8['C25'].value = str(int(hamdchkType3.get())+1) + "."

        def hamdchkImage33Func():
            if not str(sheet8['C25'].value).strip() == 2:
                sheet8['C25'].value = str(int(hamdchkType3.get())+1) + "."

        def hamdchkImage41Func():
            if not str(sheet8['C31'].value).strip() == 0:
                sheet8['C31'].value = str(int(hamdchkType4.get())+1) + "."

        def hamdchkImage42Func():
            if not str(sheet8['C31'].value).strip() == 1:
                sheet8['C31'].value = str(int(hamdchkType4.get())+1) + "."

        def hamdchkImage43Func():
            if not str(sheet8['C31'].value).strip() == 2:
                sheet8['C31'].value = str(int(hamdchkType4.get())+1) + "."

        def hamdchkImage51Func():
            if not str(sheet8['C37'].value).strip() == 0:
                sheet8['C37'].value = str(int(hamdchkType5.get())+1) + "."

        def hamdchkImage52Func():
            if not str(sheet8['C37'].value).strip() == 1:
                sheet8['C37'].value = str(int(hamdchkType5.get())+1) + "."

        def hamdchkImage53Func():
            if not str(sheet8['C37'].value).strip() == 2:
                sheet8['C37'].value = str(int(hamdchkType5.get())+1) + "."

        def hamdchkImage61Func():
            if not str(sheet8['C45'].value).strip() == 0:
                sheet8['C45'].value = str(int(hamdchkType6.get())+1) + "."

        def hamdchkImage62Func():
            if not str(sheet8['C45'].value).strip() == 1:
                sheet8['C45'].value = str(int(hamdchkType6.get())+1) + "."

        def hamdchkImage63Func():
            if not str(sheet8['C45'].value).strip() == 2:
                sheet8['C45'].value = str(int(hamdchkType6.get())+1) + "."

        def hamdchkImage64Func():
            if not str(sheet8['C45'].value).strip() == 3:
                sheet8['C45'].value = str(int(hamdchkType6.get())+1) + "."

        def hamdchkImage65Func():
            if not str(sheet8['C45'].value).strip() == 4:
                sheet8['C45'].value = str(int(hamdchkType6.get())+1) + "."

        def hamdchkImage71Func():
            if not str(sheet8['C53'].value).strip() == 0:
                sheet8['C53'].value = str(int(hamdchkType7.get())+1) + "."

        def hamdchkImage72Func():
            if not str(sheet8['C53'].value).strip() == 1:
                sheet8['C53'].value = str(int(hamdchkType7.get())+1) + "."

        def hamdchkImage73Func():
            if not str(sheet8['C53'].value).strip() == 2:
                sheet8['C53'].value = str(int(hamdchkType7.get())+1) + "."

        def hamdchkImage74Func():
            if not str(sheet8['C53'].value).strip() == 3:
                sheet8['C53'].value = str(int(hamdchkType7.get())+1) + "."

        def hamdchkImage75Func():
            if not str(sheet8['C53'].value).strip() == 4:
                sheet8['C53'].value = str(int(hamdchkType7.get())+1) + "."

        def hamdchkImage81Func():
            if not str(sheet8['C59'].value).strip() == 0:
                sheet8['C59'].value = str(int(hamdchkType8.get())+1) + "."

        def hamdchkImage82Func():
            if not str(sheet8['C59'].value).strip() == 1:
                sheet8['C59'].value = str(int(hamdchkType8.get())+1) + "."

        def hamdchkImage83Func():
            if not str(sheet8['C59'].value).strip() == 2:
                sheet8['C59'].value = str(int(hamdchkType8.get())+1) + "."

        def hamdchkImage91Func():
            if not str(sheet8['C65'].value).strip() == 0:
                sheet8['C65'].value = str(int(hamdchkType9.get())+1) + "."

        def hamdchkImage92Func():
            if not str(sheet8['C65'].value).strip() == 1:
                sheet8['C65'].value = str(int(hamdchkType9.get())+1) + "."

        def hamdchkImage93Func():
            if not str(sheet8['C65'].value).strip() == 2:
                sheet8['C65'].value = str(int(hamdchkType9.get())+1) + "."

        def hamdchkImage101Func():
            if not str(sheet8['C71'].value).strip() == 0:
                sheet8['C71'].value = str(int(hamdchkType10.get())+1) + "."

        def hamdchkImage102Func():
            if not str(sheet8['C71'].value).strip() == 1:
                sheet8['C71'].value = str(int(hamdchkType10.get())+1) + "."

        def hamdchkImage103Func():
            if not str(sheet8['C71'].value).strip() == 2:
                sheet8['C71'].value = str(int(hamdchkType10.get())+1) + "."

        def hamdchkImage111Func():
            if not str(sheet8['C79'].value).strip() == 0:
                sheet8['C79'].value = str(int(hamdchkType11.get())+1) + "."

        def hamdchkImage112Func():
            if not str(sheet8['C79'].value).strip() == 1:
                sheet8['C79'].value = str(int(hamdchkType11.get())+1) + "."

        def hamdchkImage113Func():
            if not str(sheet8['C79'].value).strip() == 2:
                sheet8['C79'].value = str(int(hamdchkType11.get())+1) + "."

        def hamdchkImage114Func():
            if not str(sheet8['C79'].value).strip() == 3:
                sheet8['C79'].value = str(int(hamdchkType11.get())+1) + "."

        def hamdchkImage115Func():
            if not str(sheet8['C79'].value).strip() == 4:
                sheet8['C79'].value = str(int(hamdchkType11.get())+1) + "."

        def hamdchkImage121Func():
            if not str(sheet8['C85'].value).strip() == 0:
                sheet8['C85'].value = str(int(hamdchkType12.get())+1) + "."

        def hamdchkImage122Func():
            if not str(sheet8['C85'].value).strip() == 1:
                sheet8['C85'].value = str(int(hamdchkType12.get())+1) + "."

        def hamdchkImage123Func():
            if not str(sheet8['C85'].value).strip() == 2:
                sheet8['C85'].value = str(int(hamdchkType12.get())+1) + "."

        frame8HamdBgImg = Image.open("images/hamdbg.png")
        frame8HamdBg = ImageTk.PhotoImage(frame8HamdBgImg)
        frame8HamdBgLabel = tkinter.Label(
            displayPage.inner, image=frame8HamdBg)
        frame8HamdBgLabel.image = frame8HamdBg
        frame8HamdBgLabel.place(x=47, y=420, height=2430)
        frame8HamdBgLabel.pack()

        global hamdchk11
        hamdchk11 = Radiobutton(displayPage.inner, value=0,  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                selectimage=chkImage12, indicatoron=False, cursor="circle", variable=hamdchkType1, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage11Func)
        hamdchk11.deselect()
        global hamdchk12
        hamdchk12 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType1, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage12Func)
        hamdchk12.deselect()
        global hamdchk13
        hamdchk13 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType1, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage13Func)
        hamdchk13.deselect()
        global hamdchk14
        hamdchk14 = Radiobutton(displayPage.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType1, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage14Func)
        hamdchk14.deselect()
        global hamdchk15
        hamdchk15 = Radiobutton(displayPage.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType1, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage15Func)
        hamdchk15.deselect()
        global hamdchk21
        hamdchk21 = Radiobutton(displayPage.inner, value=0,  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                selectimage=chkImage12, indicatoron=False, cursor="circle", variable=hamdchkType2, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage21Func)
        hamdchk21.deselect()
        global hamdchk22
        hamdchk22 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType2, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage22Func)
        hamdchk22.deselect()
        global hamdchk23
        hamdchk23 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType2, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage23Func)
        hamdchk23.deselect()
        global hamdchk24
        hamdchk24 = Radiobutton(displayPage.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType2, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage24Func)
        hamdchk24.deselect()
        global hamdchk25
        hamdchk25 = Radiobutton(displayPage.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType2, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage25Func)
        hamdchk25.deselect()
        global hamdchk31
        hamdchk31 = Radiobutton(displayPage.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                selectimage=chkImage12, indicatoron=False, cursor="circle", variable=hamdchkType3, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage31Func)
        hamdchk31.deselect()
        global hamdchk32
        hamdchk32 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType3, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage32Func)
        hamdchk32.deselect()
        global hamdchk33
        hamdchk33 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType3, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage33Func)
        hamdchk33.deselect()
        global hamdchk41
        hamdchk41 = Radiobutton(displayPage.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                selectimage=chkImage12, indicatoron=False, cursor="circle", variable=hamdchkType4, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage41Func)
        hamdchk41.deselect()
        global hamdchk42
        hamdchk42 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType4, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage42Func)
        hamdchk42.deselect()
        global hamdchk43
        hamdchk43 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType4, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage43Func)
        hamdchk43.deselect()
        global hamdchk51
        hamdchk51 = Radiobutton(displayPage.inner, value=0,  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                selectimage=chkImage12, indicatoron=False, cursor="circle", variable=hamdchkType5, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage51Func)
        hamdchk51.deselect()
        global hamdchk52
        hamdchk52 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType5, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage52Func)
        hamdchk52.deselect()
        global hamdchk53
        hamdchk53 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType5, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage53Func)
        hamdchk53.deselect()
        global hamdchk61
        hamdchk61 = Radiobutton(displayPage.inner, value=0,  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                selectimage=chkImage12, indicatoron=False, cursor="circle", variable=hamdchkType6, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage61Func)
        hamdchk61.deselect()
        global hamdchk62
        hamdchk62 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType6, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage62Func)
        hamdchk62.deselect()
        global hamdchk63
        hamdchk63 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType6, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage63Func)
        hamdchk63.deselect()
        global hamdchk64
        hamdchk64 = Radiobutton(displayPage.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType6, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage64Func)
        hamdchk64.deselect()
        global hamdchk65
        hamdchk65 = Radiobutton(displayPage.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType6, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage65Func)
        hamdchk65.deselect()
        global hamdchk71
        hamdchk71 = Radiobutton(displayPage.inner, value=0,  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                selectimage=chkImage12, indicatoron=False, cursor="circle", variable=hamdchkType7, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage71Func)
        hamdchk71.deselect()
        global hamdchk72
        hamdchk72 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType7, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage72Func)
        hamdchk72.deselect()
        global hamdchk73
        hamdchk73 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType7, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage73Func)
        hamdchk73.deselect()
        global hamdchk74
        hamdchk74 = Radiobutton(displayPage.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType7, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage74Func)
        hamdchk74.deselect()
        global hamdchk75
        hamdchk75 = Radiobutton(displayPage.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType7, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage75Func)
        hamdchk75.deselect()
        global hamdchk81
        hamdchk81 = Radiobutton(displayPage.inner, value=0,  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                selectimage=chkImage12, indicatoron=False, cursor="circle", variable=hamdchkType8, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage81Func)
        hamdchk81.deselect()
        global hamdchk82
        hamdchk82 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType8, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage82Func)
        hamdchk82.deselect()
        global hamdchk83
        hamdchk83 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType8, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage83Func)
        hamdchk83.deselect()
        global hamdchk91
        hamdchk91 = Radiobutton(displayPage.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                selectimage=chkImage12, indicatoron=False, cursor="circle", variable=hamdchkType9, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage91Func)
        hamdchk91.deselect()
        global hamdchk92
        hamdchk92 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType9, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage92Func)
        hamdchk92.deselect()
        global hamdchk93
        hamdchk93 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType9, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage93Func)
        hamdchk93.deselect()
        global hamdchk101
        hamdchk101 = Radiobutton(displayPage.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=hamdchkType10, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage101Func)
        hamdchk101.deselect()
        global hamdchk102
        hamdchk102 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=hamdchkType10, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage102Func)
        hamdchk102.deselect()
        global hamdchk103
        hamdchk103 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=hamdchkType10, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage103Func)
        hamdchk103.deselect()
        global hamdchk111
        hamdchk111 = Radiobutton(displayPage.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=hamdchkType11, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage111Func)
        hamdchk111.deselect()
        global hamdchk112
        hamdchk112 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=hamdchkType11, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage112Func)
        hamdchk112.deselect()
        global hamdchk113
        hamdchk113 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=hamdchkType11, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage113Func)
        hamdchk113.deselect()
        global hamdchk114
        hamdchk114 = Radiobutton(displayPage.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=hamdchkType11, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage114Func)
        hamdchk114.deselect()
        global hamdchk115
        hamdchk115 = Radiobutton(displayPage.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=hamdchkType11, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage115Func)
        hamdchk115.deselect()
        global hamdchk121
        hamdchk121 = Radiobutton(displayPage.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=hamdchkType12, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage121Func)
        hamdchk121.deselect()
        global hamdchk122
        hamdchk122 = Radiobutton(displayPage.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=hamdchkType12, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage122Func)
        hamdchk122.deselect()
        global hamdchk123
        hamdchk123 = Radiobutton(displayPage.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=hamdchkType12, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage123Func)
        hamdchk123.deselect()

        hamdchk11.place(x=160, y=243, width=21, height=21)
        hamdchk12.place(x=160, y=275, width=21, height=21)
        hamdchk13.place(x=160, y=305, width=21, height=21)
        hamdchk14.place(x=160, y=336, width=21, height=21)
        hamdchk15.place(x=160, y=366, width=21, height=21)
        hamdchk21.place(x=160, y=460, width=21, height=21)
        hamdchk22.place(x=160, y=491, width=21, height=21)
        hamdchk23.place(x=160, y=522, width=21, height=21)
        hamdchk24.place(x=160, y=553, width=21, height=21)
        hamdchk25.place(x=160, y=585, width=21, height=21)
        hamdchk31.place(x=160, y=677, width=21, height=21)
        hamdchk32.place(x=160, y=708, width=21, height=21)
        hamdchk33.place(x=160, y=740, width=21, height=21)
        hamdchk41.place(x=160, y=832, width=21, height=21)
        hamdchk42.place(x=160, y=863, width=21, height=21)
        hamdchk43.place(x=160, y=895, width=21, height=21)
        hamdchk51.place(x=160, y=986, width=21, height=21)
        hamdchk52.place(x=160, y=1018, width=21, height=21)
        hamdchk53.place(x=160, y=1049, width=21, height=21)
        hamdchk61.place(x=160, y=1142, width=21, height=21)
        hamdchk62.place(x=160, y=1173, width=21, height=21)
        hamdchk63.place(x=160, y=1203, width=21, height=21)
        hamdchk64.place(x=160, y=1234, width=21, height=21)
        hamdchk65.place(x=160, y=1265, width=21, height=21)
        hamdchk71.place(x=160, y=1364, width=21, height=21)
        hamdchk72.place(x=160, y=1396, width=21, height=21)
        hamdchk73.place(x=160, y=1427, width=21, height=21)
        hamdchk74.place(x=160, y=1458, width=21, height=21)
        hamdchk75.place(x=160, y=1489, width=21, height=21)
        hamdchk81.place(x=160, y=1581, width=21, height=21)
        hamdchk82.place(x=160, y=1612, width=21, height=21)
        hamdchk83.place(x=160, y=1644, width=21, height=21)
        hamdchk91.place(x=160, y=1736, width=21, height=21)
        hamdchk92.place(x=160, y=1768, width=21, height=21)
        hamdchk93.place(x=160, y=1799, width=21, height=21)
        hamdchk101.place(x=160, y=1892, width=21, height=21)
        hamdchk102.place(x=160, y=1923, width=21, height=21)
        hamdchk103.place(x=160, y=1955, width=21, height=21)
        hamdchk111.place(x=160, y=2047, width=21, height=21)
        hamdchk112.place(x=160, y=2078, width=21, height=21)
        hamdchk113.place(x=160, y=2108, width=21, height=21)
        hamdchk114.place(x=160, y=2140, width=21, height=21)
        hamdchk115.place(x=160, y=2171, width=21, height=21)
        hamdchk121.place(x=160, y=2265, width=21, height=21)
        hamdchk122.place(x=160, y=2296, width=21, height=21)
        hamdchk123.place(x=160, y=2326, width=21, height=21)

    def hamdchkFunc2():
        global now
        now = datetime.now()
        global hamdchk
        hamdchk = True
        sheet8['E2'].value = (" % s년 % s월 % s일 % s시 % s분" %
                              (now.year, now.month, now.day, now.hour, now.minute))
        sheet8['K2'].value = idInput.get()

        def hamdchkImage11Func():
            if not str(sheet8['C11'].value).strip() == 0:
                sheet8['C11'].value = str(int(hamdchkType1.get())+1) + "."

        def hamdchkImage12Func():
            if not str(sheet8['C11'].value).strip() == 1:
                sheet8['C11'].value = str(int(hamdchkType1.get())+1) + "."

        def hamdchkImage13Func():
            if not str(sheet8['C11'].value).strip() == 2:
                sheet8['C11'].value = str(int(hamdchkType1.get())+1) + "."

        def hamdchkImage14Func():
            if not str(sheet8['C11'].value).strip() == 3:
                sheet8['C11'].value = str(int(hamdchkType1.get())+1) + "."

        def hamdchkImage15Func():
            if not str(sheet8['C11'].value).strip() == 4:
                sheet8['C11'].value = str(int(hamdchkType1.get())+1) + "."

        def hamdchkImage21Func():
            if not str(sheet8['C19'].value).strip() == 0:
                sheet8['C19'].value = str(int(hamdchkType2.get())+1) + "."

        def hamdchkImage22Func():
            if not str(sheet8['C19'].value).strip() == 1:
                sheet8['C19'].value = str(int(hamdchkType2.get())+1) + "."

        def hamdchkImage23Func():
            if not str(sheet8['C19'].value).strip() == 2:
                sheet8['C19'].value = str(int(hamdchkType2.get())+1) + "."

        def hamdchkImage24Func():
            if not str(sheet8['C19'].value).strip() == 3:
                sheet8['C19'].value = str(int(hamdchkType2.get())+1) + "."

        def hamdchkImage25Func():
            if not str(sheet8['C19'].value).strip() == 4:
                sheet8['C19'].value = str(int(hamdchkType2.get())+1) + "."

        def hamdchkImage31Func():
            if not str(sheet8['C25'].value).strip() == 0:
                sheet8['C25'].value = str(int(hamdchkType3.get())+1) + "."

        def hamdchkImage32Func():
            if not str(sheet8['C25'].value).strip() == 1:
                sheet8['C25'].value = str(int(hamdchkType3.get())+1) + "."

        def hamdchkImage33Func():
            if not str(sheet8['C25'].value).strip() == 2:
                sheet8['C25'].value = str(int(hamdchkType3.get())+1) + "."

        def hamdchkImage41Func():
            if not str(sheet8['C31'].value).strip() == 0:
                sheet8['C31'].value = str(int(hamdchkType4.get())+1) + "."

        def hamdchkImage42Func():
            if not str(sheet8['C31'].value).strip() == 1:
                sheet8['C31'].value = str(int(hamdchkType4.get())+1) + "."

        def hamdchkImage43Func():
            if not str(sheet8['C31'].value).strip() == 2:
                sheet8['C31'].value = str(int(hamdchkType4.get())+1) + "."

        def hamdchkImage51Func():
            if not str(sheet8['C37'].value).strip() == 0:
                sheet8['C37'].value = str(int(hamdchkType5.get())+1) + "."

        def hamdchkImage52Func():
            if not str(sheet8['C37'].value).strip() == 1:
                sheet8['C37'].value = str(int(hamdchkType5.get())+1) + "."

        def hamdchkImage53Func():
            if not str(sheet8['C37'].value).strip() == 2:
                sheet8['C37'].value = str(int(hamdchkType5.get())+1) + "."

        def hamdchkImage61Func():
            if not str(sheet8['C45'].value).strip() == 0:
                sheet8['C45'].value = str(int(hamdchkType6.get())+1) + "."

        def hamdchkImage62Func():
            if not str(sheet8['C45'].value).strip() == 1:
                sheet8['C45'].value = str(int(hamdchkType6.get())+1) + "."

        def hamdchkImage63Func():
            if not str(sheet8['C45'].value).strip() == 2:
                sheet8['C45'].value = str(int(hamdchkType6.get())+1) + "."

        def hamdchkImage64Func():
            if not str(sheet8['C45'].value).strip() == 3:
                sheet8['C45'].value = str(int(hamdchkType6.get())+1) + "."

        def hamdchkImage65Func():
            if not str(sheet8['C45'].value).strip() == 4:
                sheet8['C45'].value = str(int(hamdchkType6.get())+1) + "."

        def hamdchkImage71Func():
            if not str(sheet8['C53'].value).strip() == 0:
                sheet8['C53'].value = str(int(hamdchkType7.get())+1) + "."

        def hamdchkImage72Func():
            if not str(sheet8['C53'].value).strip() == 1:
                sheet8['C53'].value = str(int(hamdchkType7.get())+1) + "."

        def hamdchkImage73Func():
            if not str(sheet8['C53'].value).strip() == 2:
                sheet8['C53'].value = str(int(hamdchkType7.get())+1) + "."

        def hamdchkImage74Func():
            if not str(sheet8['C53'].value).strip() == 3:
                sheet8['C53'].value = str(int(hamdchkType7.get())+1) + "."

        def hamdchkImage75Func():
            if not str(sheet8['C53'].value).strip() == 4:
                sheet8['C53'].value = str(int(hamdchkType7.get())+1) + "."

        def hamdchkImage81Func():
            if not str(sheet8['C59'].value).strip() == 0:
                sheet8['C59'].value = str(int(hamdchkType8.get())+1) + "."

        def hamdchkImage82Func():
            if not str(sheet8['C59'].value).strip() == 1:
                sheet8['C59'].value = str(int(hamdchkType8.get())+1) + "."

        def hamdchkImage83Func():
            if not str(sheet8['C59'].value).strip() == 2:
                sheet8['C59'].value = str(int(hamdchkType8.get())+1) + "."

        def hamdchkImage91Func():
            if not str(sheet8['C65'].value).strip() == 0:
                sheet8['C65'].value = str(int(hamdchkType9.get())+1) + "."

        def hamdchkImage92Func():
            if not str(sheet8['C65'].value).strip() == 1:
                sheet8['C65'].value = str(int(hamdchkType9.get())+1) + "."

        def hamdchkImage93Func():
            if not str(sheet8['C65'].value).strip() == 2:
                sheet8['C65'].value = str(int(hamdchkType9.get())+1) + "."

        def hamdchkImage101Func():
            if not str(sheet8['C71'].value).strip() == 0:
                sheet8['C71'].value = str(int(hamdchkType10.get())+1) + "."

        def hamdchkImage102Func():
            if not str(sheet8['C71'].value).strip() == 1:
                sheet8['C71'].value = str(int(hamdchkType10.get())+1) + "."

        def hamdchkImage103Func():
            if not str(sheet8['C71'].value).strip() == 2:
                sheet8['C71'].value = str(int(hamdchkType10.get())+1) + "."

        def hamdchkImage111Func():
            if not str(sheet8['C79'].value).strip() == 0:
                sheet8['C79'].value = str(int(hamdchkType11.get())+1) + "."

        def hamdchkImage112Func():
            if not str(sheet8['C79'].value).strip() == 1:
                sheet8['C79'].value = str(int(hamdchkType11.get())+1) + "."

        def hamdchkImage113Func():
            if not str(sheet8['C79'].value).strip() == 2:
                sheet8['C79'].value = str(int(hamdchkType11.get())+1) + "."

        def hamdchkImage114Func():
            if not str(sheet8['C79'].value).strip() == 3:
                sheet8['C79'].value = str(int(hamdchkType11.get())+1) + "."

        def hamdchkImage115Func():
            if not str(sheet8['C79'].value).strip() == 4:
                sheet8['C79'].value = str(int(hamdchkType11.get())+1) + "."

        def hamdchkImage121Func():
            if not str(sheet8['C85'].value).strip() == 0:
                sheet8['C85'].value = str(int(hamdchkType12.get())+1) + "."

        def hamdchkImage122Func():
            if not str(sheet8['C85'].value).strip() == 1:
                sheet8['C85'].value = str(int(hamdchkType12.get())+1) + "."

        def hamdchkImage123Func():
            if not str(sheet8['C85'].value).strip() == 2:
                sheet8['C85'].value = str(int(hamdchkType12.get())+1) + "."

        frame8HamdBgImg = Image.open("images/hamdbg.png")
        frame8HamdBg = ImageTk.PhotoImage(frame8HamdBgImg)
        frame8HamdBgLabel = tkinter.Label(
            displayPage2.inner, image=frame8HamdBg)
        frame8HamdBgLabel.image = frame8HamdBg
        frame8HamdBgLabel.place(x=47, y=420, height=2430)
        frame8HamdBgLabel.pack()

        global hamdchk11
        hamdchk11 = Radiobutton(displayPage2.inner, value=0,  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                selectimage=chkImage12, indicatoron=False, cursor="circle", variable=hamdchkType1, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage11Func)
        hamdchk11.deselect()
        global hamdchk12
        hamdchk12 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType1, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage12Func)
        hamdchk12.deselect()
        global hamdchk13
        hamdchk13 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType1, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage13Func)
        hamdchk13.deselect()
        global hamdchk14
        hamdchk14 = Radiobutton(displayPage2.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType1, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage14Func)
        hamdchk14.deselect()
        global hamdchk15
        hamdchk15 = Radiobutton(displayPage2.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType1, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage15Func)
        hamdchk15.deselect()
        global hamdchk21
        hamdchk21 = Radiobutton(displayPage2.inner, value=0,  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                selectimage=chkImage12, indicatoron=False, cursor="circle", variable=hamdchkType2, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage21Func)
        hamdchk21.deselect()
        global hamdchk22
        hamdchk22 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType2, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage22Func)
        hamdchk22.deselect()
        global hamdchk23
        hamdchk23 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType2, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage23Func)
        hamdchk23.deselect()
        global hamdchk24
        hamdchk24 = Radiobutton(displayPage2.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType2, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage24Func)
        hamdchk24.deselect()
        global hamdchk25
        hamdchk25 = Radiobutton(displayPage2.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType2, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage25Func)
        hamdchk25.deselect()
        global hamdchk31
        hamdchk31 = Radiobutton(displayPage2.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                selectimage=chkImage12, indicatoron=False, cursor="circle", variable=hamdchkType3, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage31Func)
        hamdchk31.deselect()
        global hamdchk32
        hamdchk32 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType3, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage32Func)
        hamdchk32.deselect()
        global hamdchk33
        hamdchk33 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType3, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage33Func)
        hamdchk33.deselect()
        global hamdchk41
        hamdchk41 = Radiobutton(displayPage2.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                selectimage=chkImage12, indicatoron=False, cursor="circle", variable=hamdchkType4, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage41Func)
        hamdchk41.deselect()
        global hamdchk42
        hamdchk42 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType4, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage42Func)
        hamdchk42.deselect()
        global hamdchk43
        hamdchk43 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType4, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage43Func)
        hamdchk43.deselect()
        global hamdchk51
        hamdchk51 = Radiobutton(displayPage2.inner, value=0,  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                selectimage=chkImage12, indicatoron=False, cursor="circle", variable=hamdchkType5, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage51Func)
        hamdchk51.deselect()
        global hamdchk52
        hamdchk52 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType5, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage52Func)
        hamdchk52.deselect()
        global hamdchk53
        hamdchk53 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType5, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage53Func)
        hamdchk53.deselect()
        global hamdchk61
        hamdchk61 = Radiobutton(displayPage2.inner, value=0,  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                selectimage=chkImage12, indicatoron=False, cursor="circle", variable=hamdchkType6, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage61Func)
        hamdchk61.deselect()
        global hamdchk62
        hamdchk62 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType6, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage62Func)
        hamdchk62.deselect()
        global hamdchk63
        hamdchk63 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType6, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage63Func)
        hamdchk63.deselect()
        global hamdchk64
        hamdchk64 = Radiobutton(displayPage2.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType6, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage64Func)
        hamdchk64.deselect()
        global hamdchk65
        hamdchk65 = Radiobutton(displayPage2.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType6, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage65Func)
        hamdchk65.deselect()
        global hamdchk71
        hamdchk71 = Radiobutton(displayPage2.inner, value=0,  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                selectimage=chkImage12, indicatoron=False, cursor="circle", variable=hamdchkType7, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage71Func)
        hamdchk71.deselect()
        global hamdchk72
        hamdchk72 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType7, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage72Func)
        hamdchk72.deselect()
        global hamdchk73
        hamdchk73 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType7, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage73Func)
        hamdchk73.deselect()
        global hamdchk74
        hamdchk74 = Radiobutton(displayPage2.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType7, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage74Func)
        hamdchk74.deselect()
        global hamdchk75
        hamdchk75 = Radiobutton(displayPage2.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType7, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage75Func)
        hamdchk75.deselect()
        global hamdchk81
        hamdchk81 = Radiobutton(displayPage2.inner, value=0,  background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                selectimage=chkImage12, indicatoron=False, cursor="circle", variable=hamdchkType8, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage81Func)
        hamdchk81.deselect()
        global hamdchk82
        hamdchk82 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType8, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage82Func)
        hamdchk82.deselect()
        global hamdchk83
        hamdchk83 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType8, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage83Func)
        hamdchk83.deselect()
        global hamdchk91
        hamdchk91 = Radiobutton(displayPage2.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11,
                                selectimage=chkImage12, indicatoron=False, cursor="circle", variable=hamdchkType9, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage91Func)
        hamdchk91.deselect()
        global hamdchk92
        hamdchk92 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType9, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage92Func)
        hamdchk92.deselect()
        global hamdchk93
        hamdchk93 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                indicatoron=False, cursor="circle", variable=hamdchkType9, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage93Func)
        hamdchk93.deselect()
        global hamdchk101
        hamdchk101 = Radiobutton(displayPage2.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=hamdchkType10, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage101Func)
        hamdchk101.deselect()
        global hamdchk102
        hamdchk102 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=hamdchkType10, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage102Func)
        hamdchk102.deselect()
        global hamdchk103
        hamdchk103 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=hamdchkType10, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage103Func)
        hamdchk103.deselect()
        global hamdchk111
        hamdchk111 = Radiobutton(displayPage2.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=hamdchkType11, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage111Func)
        hamdchk111.deselect()
        global hamdchk112
        hamdchk112 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=hamdchkType11, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage112Func)
        hamdchk112.deselect()
        global hamdchk113
        hamdchk113 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=hamdchkType11, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage113Func)
        hamdchk113.deselect()
        global hamdchk114
        hamdchk114 = Radiobutton(displayPage2.inner, value=3, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=hamdchkType11, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage114Func)
        hamdchk114.deselect()
        global hamdchk115
        hamdchk115 = Radiobutton(displayPage2.inner, value=4, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=hamdchkType11, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage115Func)
        hamdchk115.deselect()
        global hamdchk121
        hamdchk121 = Radiobutton(displayPage2.inner, value=0, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=hamdchkType12, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage121Func)
        hamdchk121.deselect()
        global hamdchk122
        hamdchk122 = Radiobutton(displayPage2.inner, value=1, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=hamdchkType12, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage122Func)
        hamdchk122.deselect()
        global hamdchk123
        hamdchk123 = Radiobutton(displayPage2.inner, value=2, background="#FFFFFF", activebackground="#FFFFFF", image=chkImage11, selectimage=chkImage12,
                                 indicatoron=False, cursor="circle", variable=hamdchkType12, font=malgungothic13, bd=0, highlightthickness=0, command=hamdchkImage123Func)
        hamdchk123.deselect()

        hamdchk11.place(x=160, y=243+addheight, width=21, height=21)
        hamdchk12.place(x=160, y=275+addheight, width=21, height=21)
        hamdchk13.place(x=160, y=305+addheight, width=21, height=21)
        hamdchk14.place(x=160, y=336+addheight, width=21, height=21)
        hamdchk15.place(x=160, y=366+addheight, width=21, height=21)
        hamdchk21.place(x=160, y=460+addheight, width=21, height=21)
        hamdchk22.place(x=160, y=491+addheight, width=21, height=21)
        hamdchk23.place(x=160, y=522+addheight, width=21, height=21)
        hamdchk24.place(x=160, y=553+addheight, width=21, height=21)
        hamdchk25.place(x=160, y=585+addheight, width=21, height=21)
        hamdchk31.place(x=160, y=677+addheight, width=21, height=21)
        hamdchk32.place(x=160, y=708+addheight, width=21, height=21)
        hamdchk33.place(x=160, y=740+addheight, width=21, height=21)
        hamdchk41.place(x=160, y=832+addheight, width=21, height=21)
        hamdchk42.place(x=160, y=863+addheight, width=21, height=21)
        hamdchk43.place(x=160, y=895+addheight, width=21, height=21)
        hamdchk51.place(x=160, y=986+addheight, width=21, height=21)
        hamdchk52.place(x=160, y=1018+addheight, width=21, height=21)
        hamdchk53.place(x=160, y=1049+addheight, width=21, height=21)
        hamdchk61.place(x=160, y=1142+addheight, width=21, height=21)
        hamdchk62.place(x=160, y=1173+addheight, width=21, height=21)
        hamdchk63.place(x=160, y=1203+addheight, width=21, height=21)
        hamdchk64.place(x=160, y=1234+addheight, width=21, height=21)
        hamdchk65.place(x=160, y=1265+addheight, width=21, height=21)
        hamdchk71.place(x=160, y=1364+addheight, width=21, height=21)
        hamdchk72.place(x=160, y=1396+addheight, width=21, height=21)
        hamdchk73.place(x=160, y=1427+addheight, width=21, height=21)
        hamdchk74.place(x=160, y=1458+addheight, width=21, height=21)
        hamdchk75.place(x=160, y=1489+addheight, width=21, height=21)
        hamdchk81.place(x=160, y=1581+addheight, width=21, height=21)
        hamdchk82.place(x=160, y=1612+addheight, width=21, height=21)
        hamdchk83.place(x=160, y=1644+addheight, width=21, height=21)
        hamdchk91.place(x=160, y=1736+addheight, width=21, height=21)
        hamdchk92.place(x=160, y=1768+addheight, width=21, height=21)
        hamdchk93.place(x=160, y=1799+addheight, width=21, height=21)
        hamdchk101.place(x=160, y=1892+addheight, width=21, height=21)
        hamdchk102.place(x=160, y=1923+addheight, width=21, height=21)
        hamdchk103.place(x=160, y=1955+addheight, width=21, height=21)
        hamdchk111.place(x=160, y=2047+addheight, width=21, height=21)
        hamdchk112.place(x=160, y=2078+addheight, width=21, height=21)
        hamdchk113.place(x=160, y=2108+addheight, width=21, height=21)
        hamdchk114.place(x=160, y=2140+addheight, width=21, height=21)
        hamdchk115.place(x=160, y=2171+addheight, width=21, height=21)
        hamdchk121.place(x=160, y=2265+addheight, width=21, height=21)
        hamdchk122.place(x=160, y=2296+addheight, width=21, height=21)
        hamdchk123.place(x=160, y=2326+addheight, width=21, height=21)

    def allchkFunc():
        
        wbb = openpyxl.load_workbook('//DESKTOP/ptest//xlsm//'+nameInput.get()+'['+idInput.get()+'].xlsm', keep_vba=True)
        wss = wbb.active
        sheet00 = wbb['Ptest']
        altersheet = wbb['변화']
        interviewsheet = wbb["인터뷰"]
        addedVar2 = altersheet['AB4'].value + 1
        altersheet['AB4'].value = addedVar2
        avgVar = altersheet['AC4'].value + 1
        altersheet['AC4'].value = avgVar
        intVar = altersheet['AD4'].value + 1
        altersheet['AD4'].value = intVar
        # if hamachk3 == False and (phqchk3 == True or cdichk3 == True or bdichk3 == True or snapchk3 == True or st_2chk == True or st_1chk == True or hamachk == True or hamdchk == True):
        
        global gotten_list
        gotten_list = []
        global cellCriteria
        # global sumscore
        # global sumscore2
        if phqchk2 == True:
            if phqchkType1.get() == "":
                messagebox.showinfo('마음정원', "'PHQ-9' 1번 문항을 체크해주십시오.")
            elif phqchkType2.get() == "":
                messagebox.showinfo('마음정원', "'PHQ-9' 2번 문항을 체크해주십시오.")
            elif phqchkType3.get() == "":
                messagebox.showinfo('마음정원', "'PHQ-9' 3번 문항을 체크해주십시오.")
            elif phqchkType4.get() == "":
                messagebox.showinfo('마음정원', "'PHQ-9' 4번 문항을 체크해주십시오.")
            elif phqchkType5.get() == "":
                messagebox.showinfo('마음정원', "'PHQ-9' 5번 문항을 체크해주십시오.")
            elif phqchkType6.get() == "":
                messagebox.showinfo('마음정원', "'PHQ-9' 6번 문항을 체크해주십시오.")
            elif phqchkType7.get() == "":
                messagebox.showinfo('마음정원', "'PHQ-9' 7번 문항을 체크해주십시오.")
            elif phqchkType8.get() == "":
                messagebox.showinfo('마음정원', "'PHQ-9' 8번 문항을 체크해주십시오.")
            elif phqlastchkType.get() == "":
                messagebox.showinfo('마음정원', "'PHQ-9' 마지막 문항을 체크해주십시오.")
            if phqchkType1.get() == "" or phqchkType2.get() == "" or phqchkType3.get() == "" or phqchkType4.get() == "" or phqchkType5.get() == "" or phqchkType6.get() == "" or phqchkType7.get() == "" or phqchkType8.get() == "" or phqlastchkType.get() == "":
                return

            sumscore = int(phqchkType1.get()) + int(phqchkType2.get()) + int(phqchkType3.get()) + int(phqchkType4.get()) + int(phqchkType5.get()) + int(phqchkType6.get()) + int(phqchkType7.get()) + int(phqchkType8.get())

            phqrltScore1 = str(phqchkType1.get())
            phqrltScore2 = str(phqchkType2.get())
            phqrltScore3 = str(phqchkType3.get())
            phqrltScore4 = str(phqchkType4.get())
            phqrltScore5 = str(phqchkType5.get())
            phqrltScore6 = str(phqchkType6.get())
            phqrltScore7 = str(phqchkType7.get())
            phqrltScore8 = str(phqchkType8.get())
            sheet1['E6'] = phqrltScore1 + " 점"
            sheet1['E9'] = phqrltScore2 + " 점"
            sheet1['E12'] = phqrltScore3 + " 점"
            sheet1['E15'] = phqrltScore4 + " 점"
            sheet1['E18'] = phqrltScore5 + " 점"
            sheet1['E21'] = phqrltScore6 + " 점"
            sheet1['E24'] = phqrltScore7 + " 점"
            sheet1['E27'] = phqrltScore8 + " 점"
            sheet1['E30'] = "X"
            sheet1['N3'] = sumscore
            sheet1['K3'] = nameInput.get()
            min_col, min_row, max_col, max_row = range_boundaries('A1')
            
            # now = datetime.now()
            # for row in sheet0.iter_rows(min_row=cellCriteria):
            #     for col in row:
            #         col.value = None
            #         col._style = None
            wss.insert_rows(2, 33)
            for row in sheet00.iter_rows(1, 33, 1, 15):
                for cell in row:
                    cell.value = wb['PHQ'].cell(cell.row, cell.column).value
                    cell._style = wb['PHQ'].cell(
                        cell.row, cell.column)._style
            altersheet['C6'].value = sheet1['N3'].value
        
        if cdichk2 == True:
            if cdichkType1.get() == "":
                messagebox.showinfo('마음정원', "'CDI' 1번 문항을 체크해주십시오.")
            elif cdichkType2.get() == "":
                messagebox.showinfo('마음정원', "'CDI' 2번 문항을 체크해주십시오.")
            elif cdichkType3.get() == "":
                messagebox.showinfo('마음정원', "'CDI' 3번 문항을 체크해주십시오.")
            elif cdichkType4.get() == "":
                messagebox.showinfo('마음정원', "'CDI' 4번 문항을 체크해주십시오.")
            elif cdichkType5.get() == "":
                messagebox.showinfo('마음정원', "'CDI' 5번 문항을 체크해주십시오.")
            elif cdichkType6.get() == "":
                messagebox.showinfo('마음정원', "'CDI' 6번 문항을 체크해주십시오.")
            elif cdichkType7.get() == "":
                messagebox.showinfo('마음정원', "'CDI' 7번 문항을 체크해주십시오.")
            elif cdichkType8.get() == "":
                messagebox.showinfo('마음정원', "'CDI' 8번 문항을 체크해주십시오.")
            elif cdichkType9.get() == "":
                messagebox.showinfo('마음정원', "'CDI' 9번 문항을 체크해주십시오.")
            elif cdichkType10.get() == "":
                messagebox.showinfo('마음정원', "'CDI' 10번 문항을 체크해주십시오.")
            elif cdichkType11.get() == "":
                messagebox.showinfo('마음정원', "'CDI' 11번 문항을 체크해주십시오.")
            elif cdichkType12.get() == "":
                messagebox.showinfo('마음정원', "'CDI' 12번 문항을 체크해주십시오.")
            elif cdichkType13.get() == "":
                messagebox.showinfo('마음정원', "'CDI' 13번 문항을 체크해주십시오.")
            elif cdichkType14.get() == "":
                messagebox.showinfo('마음정원', "'CDI' 14번 문항을 체크해주십시오.")
            elif cdichkType15.get() == "":
                messagebox.showinfo('마음정원', "'CDI' 15번 문항을 체크해주십시오.")
            elif cdichkType16.get() == "":
                messagebox.showinfo('마음정원', "'CDI' 16번 문항을 체크해주십시오.")
            elif cdichkType17.get() == "":
                messagebox.showinfo('마음정원', "'CDI' 17번 문항을 체크해주십시오.")
            elif cdichkType18.get() == "":
                messagebox.showinfo('마음정원', "'CDI' 18번 문항을 체크해주십시오.")
            elif cdichkType19.get() == "":
                messagebox.showinfo('마음정원', "'CDI' 19번 문항을 체크해주십시오.")
            elif cdichkType20.get() == "":
                messagebox.showinfo('마음정원', "'CDI' 20번 문항을 체크해주십시오.")
            elif cdichkType21.get() == "":
                messagebox.showinfo('마음정원', "'CDI' 21번 문항을 체크해주십시오.")
            elif cdichkType22.get() == "":
                messagebox.showinfo('마음정원', "'CDI' 22번 문항을 체크해주십시오.")
            elif cdichkType23.get() == "":
                messagebox.showinfo('마음정원', "'CDI' 23번 문항을 체크해주십시오.")
            elif cdichkType24.get() == "":
                messagebox.showinfo('마음정원', "'CDI' 24번 문항을 체크해주십시오.")
            elif cdichkType25.get() == "":
                messagebox.showinfo('마음정원', "'CDI' 25번 문항을 체크해주십시오.")
            elif cdichkType26.get() == "":
                messagebox.showinfo('마음정원', "'CDI' 26번 문항을 체크해주십시오.")
            elif cdichkType27.get() == "":
                messagebox.showinfo('마음정원', "'CDI' 마지막 문항을 체크해주십시오.")

            if cdichkType1.get() == "" or cdichkType2.get() == "" or cdichkType3.get() == "" or cdichkType4.get() == "" or cdichkType5.get() == "" or cdichkType6.get() == "" or cdichkType7.get() == "" or cdichkType8.get() == "" or cdichkType9.get() == "" or cdichkType10.get() == "" or cdichkType11.get() == "" or cdichkType12.get() == "" or cdichkType13.get() == "" or cdichkType14.get() == "" or cdichkType15.get() == "" or cdichkType16.get() == "" or cdichkType17.get() == "" or cdichkType18.get() == "" or cdichkType19.get() == "" or cdichkType20.get() == "" or cdichkType21.get() == "" or cdichkType22.get() == "" or cdichkType23.get() == "" or cdichkType24.get() == "" or cdichkType25.get() == "" or cdichkType26.get() == "" or cdichkType27.get() == "":
                return

            sumscore2 = int(cdichkType1.get()) + int(cdichkType2.get()) + int(cdichkType3.get()) + int(cdichkType4.get()) + int(cdichkType5.get()) + int(cdichkType6.get()) + int(cdichkType7.get()) + int(cdichkType8.get()) + int(cdichkType9.get()) + int(cdichkType10.get()) + int(cdichkType11.get()) + int(cdichkType12.get()) + int(cdichkType13.get()) + int(
                cdichkType14.get()) + int(cdichkType15.get()) + int(cdichkType16.get()) + int(cdichkType17.get()) + int(cdichkType18.get()) + int(cdichkType19.get()) + int(cdichkType20.get()) + int(cdichkType21.get()) + int(cdichkType22.get()) + int(cdichkType23.get()) + int(cdichkType24.get()) + int(cdichkType25.get()) + int(cdichkType26.get()) + int(cdichkType27.get())

            cdirltScore1 = str(cdichkType1.get())
            cdirltScore2 = str(cdichkType2.get())
            cdirltScore3 = str(cdichkType3.get())
            cdirltScore4 = str(cdichkType4.get())
            cdirltScore5 = str(cdichkType5.get())
            cdirltScore6 = str(cdichkType6.get())
            cdirltScore7 = str(cdichkType7.get())
            cdirltScore8 = str(cdichkType8.get())
            cdirltScore9 = str(cdichkType9.get())
            cdirltScore10 = str(cdichkType10.get())
            cdirltScore11 = str(cdichkType11.get())
            cdirltScore12 = str(cdichkType12.get())
            cdirltScore13 = str(cdichkType13.get())
            cdirltScore14 = str(cdichkType14.get())
            cdirltScore15 = str(cdichkType15.get())
            cdirltScore16 = str(cdichkType16.get())
            cdirltScore17 = str(cdichkType17.get())
            cdirltScore18 = str(cdichkType18.get())
            cdirltScore19 = str(cdichkType19.get())
            cdirltScore20 = str(cdichkType20.get())
            cdirltScore21 = str(cdichkType21.get())
            cdirltScore22 = str(cdichkType22.get())
            cdirltScore23 = str(cdichkType23.get())
            cdirltScore24 = str(cdichkType24.get())
            cdirltScore25 = str(cdichkType25.get())
            cdirltScore26 = str(cdichkType26.get())
            cdirltScore27 = str(cdichkType27.get())
            sheet2['E7'] = cdirltScore1 + " 점"
            sheet2['E12'] = cdirltScore2 + " 점"
            sheet2['E17'] = cdirltScore3 + " 점"
            sheet2['E22'] = cdirltScore4 + " 점"
            sheet2['E27'] = cdirltScore5 + " 점"
            sheet2['E32'] = cdirltScore6 + " 점"
            sheet2['E37'] = cdirltScore7 + " 점"
            sheet2['E42'] = cdirltScore8 + " 점"
            sheet2['E47'] = cdirltScore9 + " 점"
            sheet2['E52'] = cdirltScore10 + " 점"
            sheet2['E57'] = cdirltScore11 + " 점"
            sheet2['E62'] = cdirltScore12 + " 점"
            sheet2['E67'] = cdirltScore13 + " 점"
            sheet2['E72'] = cdirltScore14 + " 점"
            sheet2['E77'] = cdirltScore15 + " 점"
            sheet2['E82'] = cdirltScore16 + " 점"
            sheet2['E87'] = cdirltScore17 + " 점"
            sheet2['E92'] = cdirltScore18 + " 점"
            sheet2['E97'] = cdirltScore19 + " 점"
            sheet2['E102'] = cdirltScore20 + " 점"
            sheet2['E107'] = cdirltScore21 + " 점"
            sheet2['E112'] = cdirltScore22 + " 점"
            sheet2['E117'] = cdirltScore23 + " 점"
            sheet2['E122'] = cdirltScore24 + " 점"
            sheet2['E127'] = cdirltScore25 + " 점"
            sheet2['E132'] = cdirltScore26 + " 점"
            sheet2['E137'] = cdirltScore27 + " 점"
            sheet2['N3'] = sumscore2
            sheet2['K3'] = nameInput.get()
            min_col, min_row, max_col, max_row = range_boundaries('A1')
            # now = datetime.now()
            # for row in sheet0.iter_rows(min_row=cellCriteria):
            #     for col in row:
            #         col.value = None
            #         col._style = None
            wss.insert_rows(2, 140)
            for row in sheet00.iter_rows(1, 140, 1, 15):
                for cell in row:
                    cell.value = wb['CDI'].cell(cell.row, cell.column).value
                    cell._style = wb['CDI'].cell(
                        cell.row, cell.column)._style
            altersheet['D6'].value = sheet2['N3'].value
        if bdichk2 == True:
            if bdichkType1.get() == "":
                messagebox.showinfo('마음정원', "'BDI' 1번 문항을 체크해주십시오.")
            elif bdichkType2.get() == "":
                messagebox.showinfo('마음정원', "'BDI' 2번 문항을 체크해주십시오.")
            elif bdichkType3.get() == "":
                messagebox.showinfo('마음정원', "'BDI' 3번 문항을 체크해주십시오.")
            elif bdichkType4.get() == "":
                messagebox.showinfo('마음정원', "'BDI' 4번 문항을 체크해주십시오.")
            elif bdichkType5.get() == "":
                messagebox.showinfo('마음정원', "'BDI' 5번 문항을 체크해주십시오.")
            elif bdichkType6.get() == "":
                messagebox.showinfo('마음정원', "'BDI' 6번 문항을 체크해주십시오.")
            elif bdichkType7.get() == "":
                messagebox.showinfo('마음정원', "'BDI' 7번 문항을 체크해주십시오.")
            elif bdichkType8.get() == "":
                messagebox.showinfo('마음정원', "'BDI' 8번 문항을 체크해주십시오.")
            elif bdichkType9.get() == "":
                messagebox.showinfo('마음정원', "'BDI' 9번 문항을 체크해주십시오.")
            elif bdichkType10.get() == "":
                messagebox.showinfo('마음정원', "'BDI' 10번 문항을 체크해주십시오.")
            elif bdichkType11.get() == "":
                messagebox.showinfo('마음정원', "'BDI' 11번 문항을 체크해주십시오.")
            elif bdichkType12.get() == "":
                messagebox.showinfo('마음정원', "'BDI' 12번 문항을 체크해주십시오.")
            elif bdichkType13.get() == "":
                messagebox.showinfo('마음정원', "'BDI' 13번 문항을 체크해주십시오.")
            elif bdichkType14.get() == "":
                messagebox.showinfo('마음정원', "'BDI' 14번 문항을 체크해주십시오.")
            elif bdichkType15.get() == "":
                messagebox.showinfo('마음정원', "'BDI' 15번 문항을 체크해주십시오.")
            elif bdichkType16.get() == "":
                messagebox.showinfo('마음정원', "'BDI' 16번 문항을 체크해주십시오.")
            elif bdichkType17.get() == "":
                messagebox.showinfo('마음정원', "'BDI' 17번 문항을 체크해주십시오.")
            elif bdichkType18.get() == "":
                messagebox.showinfo('마음정원', "'BDI' 18번 문항을 체크해주십시오.")
            elif bdichkType19.get() == "":
                messagebox.showinfo('마음정원', "'BDI' 19번 문항을 체크해주십시오.")
            elif bdichkType20.get() == "":
                messagebox.showinfo('마음정원', "'BDI' 20번 문항을 체크해주십시오.")
            elif bdichkType21.get() == "":
                messagebox.showinfo('마음정원', "'BDI' 마지막 문항을 체크해주십시오.")
            if bdichkType1.get() == "" or bdichkType2.get() == "" or bdichkType3.get() == "" or bdichkType4.get() == "" or bdichkType5.get() == "" or bdichkType6.get() == "" or bdichkType7.get() == "" or bdichkType8.get() == "" or bdichkType9.get() == "" or bdichkType10.get() == "" or bdichkType11.get() == "" or bdichkType12.get() == "" or bdichkType13.get() == "" or bdichkType14.get() == "" or bdichkType15.get() == "" or bdichkType16.get() == "" or bdichkType17.get() == "" or bdichkType18.get() == "" or bdichkType19.get() == "" or bdichkType20.get() == "" or bdichkType21.get() == "":
                return


            sumscore3 = int(bdichkType1.get()) + int(bdichkType2.get()) + int(bdichkType3.get()) + int(bdichkType4.get()) + int(bdichkType5.get()) + int(bdichkType6.get()) + int(bdichkType7.get()) + int(bdichkType8.get()) + int(bdichkType9.get()) + int(bdichkType10.get()) + int(bdichkType11.get()) + int(bdichkType12.get()) + int(bdichkType13.get()) + int(
                bdichkType14.get()) + int(bdichkType15.get()) + int(bdichkType16.get()) + int(bdichkType17.get()) + int(bdichkType18.get()) + int(bdichkType19.get()) + int(bdichkType20.get()) + int(bdichkType21.get())

            bdirltScore1 = str(bdichkType1.get())
            bdirltScore2 = str(bdichkType2.get())
            bdirltScore3 = str(bdichkType3.get())
            bdirltScore4 = str(bdichkType4.get())
            bdirltScore5 = str(bdichkType5.get())
            bdirltScore6 = str(bdichkType6.get())
            bdirltScore7 = str(bdichkType7.get())
            bdirltScore8 = str(bdichkType8.get())
            bdirltScore9 = str(bdichkType9.get())
            bdirltScore10 = str(bdichkType10.get())
            bdirltScore11 = str(bdichkType11.get())
            bdirltScore12 = str(bdichkType12.get())
            bdirltScore13 = str(bdichkType13.get())
            bdirltScore14 = str(bdichkType14.get())
            bdirltScore15 = str(bdichkType15.get())
            bdirltScore16 = str(bdichkType16.get())
            bdirltScore17 = str(bdichkType17.get())
            bdirltScore18 = str(bdichkType18.get())
            bdirltScore19 = str(bdichkType19.get())
            bdirltScore20 = str(bdichkType20.get())
            bdirltScore21 = str(bdichkType21.get())
            sheet3['E9'] = bdirltScore1 + " 점"
            sheet3['E15'] = bdirltScore2 + " 점"
            sheet3['E21'] = bdirltScore3 + " 점"
            sheet3['E27'] = bdirltScore4 + " 점"
            sheet3['E33'] = bdirltScore5 + " 점"
            sheet3['E39'] = bdirltScore6 + " 점"
            sheet3['E45'] = bdirltScore7 + " 점"
            sheet3['E51'] = bdirltScore8 + " 점"
            sheet3['E57'] = bdirltScore9 + " 점"
            sheet3['E63'] = bdirltScore10 + " 점"
            sheet3['E69'] = bdirltScore11 + " 점"
            sheet3['E75'] = bdirltScore12 + " 점"
            sheet3['E81'] = bdirltScore13 + " 점"
            sheet3['E87'] = bdirltScore14 + " 점"
            sheet3['E93'] = bdirltScore15 + " 점"
            sheet3['E99'] = bdirltScore16 + " 점"
            sheet3['E105'] = bdirltScore17 + " 점"
            sheet3['E111'] = bdirltScore18 + " 점"
            sheet3['E118'] = bdirltScore19 + " 점"
            sheet3['E124'] = bdirltScore20 + " 점"
            sheet3['E130'] = bdirltScore21 + " 점"
            sheet3['M3'] = sumscore3
            sheet3['K3'] = nameInput.get()
            min_col, min_row, max_col, max_row = range_boundaries(
                'A1')
            # now = datetime.now()
            # for row in sheet0.iter_rows(min_row=cellCriteria):
            #     for col in row:
            #         col.value = None
            #         col._style = None
            wss.insert_rows(2, 133)
            for row in sheet00.iter_rows(1, 133, 1, 15):
                for cell in row:
                    cell.value = wb['BDI'].cell(cell.row, cell.column).value
                    cell._style = wb['BDI'].cell(
                        cell.row, cell.column)._style
            altersheet['E6'].value = sheet3['M3'].value
        if snapchk2 == True:
            if snapchkType1.get() == "":
                messagebox.showinfo('마음정원', "'SNAP-IV' 1번 문항을 체크해주십시오.")
            elif snapchkType2.get() == "":
                messagebox.showinfo('마음정원', "'SNAP-IV' 2번 문항을 체크해주십시오.")
            elif snapchkType3.get() == "":
                messagebox.showinfo('마음정원', "'SNAP-IV' 3번 문항을 체크해주십시오.")
            elif snapchkType4.get() == "":
                messagebox.showinfo('마음정원', "'SNAP-IV' 4번 문항을 체크해주십시오.")
            elif snapchkType5.get() == "":
                messagebox.showinfo('마음정원', "'SNAP-IV' 5번 문항을 체크해주십시오.")
            elif snapchkType6.get() == "":
                messagebox.showinfo('마음정원', "'SNAP-IV' 6번 문항을 체크해주십시오.")
            elif snapchkType7.get() == "":
                messagebox.showinfo('마음정원', "'SNAP-IV' 7번 문항을 체크해주십시오.")
            elif snapchkType8.get() == "":
                messagebox.showinfo('마음정원', "'SNAP-IV' 8번 문항을 체크해주십시오.")
            elif snapchkType9.get() == "":
                messagebox.showinfo('마음정원', "'SNAP-IV' 9번 문항을 체크해주십시오.")
            elif snapchkType10.get() == "":
                messagebox.showinfo('마음정원', "'SNAP-IV' 10번 문항을 체크해주십시오.")
            elif snapchkType11.get() == "":
                messagebox.showinfo('마음정원', "'SNAP-IV' 11번 문항을 체크해주십시오.")
            elif snapchkType12.get() == "":
                messagebox.showinfo('마음정원', "'SNAP-IV' 12번 문항을 체크해주십시오.")
            elif snapchkType13.get() == "":
                messagebox.showinfo('마음정원', "'SNAP-IV' 13번 문항을 체크해주십시오.")
            elif snapchkType14.get() == "":
                messagebox.showinfo('마음정원', "'SNAP-IV' 14번 문항을 체크해주십시오.")
            elif snapchkType15.get() == "":
                messagebox.showinfo('마음정원', "'SNAP-IV' 15번 문항을 체크해주십시오.")
            elif snapchkType16.get() == "":
                messagebox.showinfo('마음정원', "'SNAP-IV' 16번 문항을 체크해주십시오.")
            elif snapchkType17.get() == "":
                messagebox.showinfo('마음정원', "'SNAP-IV' 17번 문항을 체크해주십시오.")
            elif snapchkType18.get() == "":
                messagebox.showinfo('마음정원', "마지막 문항을 체크해주십시오.")
            if snapchkType1.get() == "" or snapchkType2.get() == "" or snapchkType3.get() == "" or snapchkType4.get() == "" or snapchkType5.get() == "" or snapchkType6.get() == "" or snapchkType7.get() == "" or snapchkType8.get() == "" or snapchkType9.get() == "" or snapchkType10.get() == "" or snapchkType11.get() == "" or snapchkType12.get() == "" or snapchkType13.get() == "" or snapchkType14.get() == "" or snapchkType15.get() == "" or snapchkType16.get() == "" or snapchkType17.get() == "" or snapchkType18.get() == "":
                return

            sumscore4 = int(snapchkType1.get()) + int(snapchkType2.get()) + int(snapchkType3.get()) + int(snapchkType4.get()) + int(snapchkType5.get()) + int(snapchkType6.get()) + int(snapchkType7.get()) + int(snapchkType8.get()) + int(snapchkType9.get()) + int(snapchkType10.get()) + int(snapchkType11.get()) + int(snapchkType12.get()) + int(snapchkType13.get()) + int(
                snapchkType14.get()) + int(snapchkType15.get()) + int(snapchkType16.get()) + int(snapchkType17.get()) + int(snapchkType18.get())

            snaprltScore1 = str(snapchkType1.get())
            snaprltScore2 = str(snapchkType2.get())
            snaprltScore3 = str(snapchkType3.get())
            snaprltScore4 = str(snapchkType4.get())
            snaprltScore5 = str(snapchkType5.get())
            snaprltScore6 = str(snapchkType6.get())
            snaprltScore7 = str(snapchkType7.get())
            snaprltScore8 = str(snapchkType8.get())
            snaprltScore9 = str(snapchkType9.get())
            snaprltScore10 = str(snapchkType10.get())
            snaprltScore11 = str(snapchkType11.get())
            snaprltScore12 = str(snapchkType12.get())
            snaprltScore13 = str(snapchkType13.get())
            snaprltScore14 = str(snapchkType14.get())
            snaprltScore15 = str(snapchkType15.get())
            snaprltScore16 = str(snapchkType16.get())
            snaprltScore17 = str(snapchkType17.get())
            snaprltScore18 = str(snapchkType18.get())
            sheet4['E6'] = snaprltScore1 + " 점"
            sheet4['E9'] = snaprltScore2 + " 점"
            sheet4['E12'] = snaprltScore3 + " 점"
            sheet4['E15'] = snaprltScore4 + " 점"
            sheet4['E18'] = snaprltScore5 + " 점"
            sheet4['E21'] = snaprltScore6 + " 점"
            sheet4['E24'] = snaprltScore7 + " 점"
            sheet4['E27'] = snaprltScore8 + " 점"
            sheet4['E30'] = snaprltScore9 + " 점"
            sheet4['E33'] = snaprltScore10 + " 점"
            sheet4['E36'] = snaprltScore11 + " 점"
            sheet4['E39'] = snaprltScore12 + " 점"
            sheet4['E42'] = snaprltScore13 + " 점"
            sheet4['E45'] = snaprltScore14 + " 점"
            sheet4['E48'] = snaprltScore15 + " 점"
            sheet4['E51'] = snaprltScore16 + " 점"
            sheet4['E54'] = snaprltScore17 + " 점"
            sheet4['E57'] = snaprltScore18 + " 점"
            sheet4['M3'] = sumscore4
            sheet4['K3'] = nameInput.get()

            min_col, min_row, max_col, max_row = range_boundaries(
                'A1')
            # now = datetime.now()
            # for row in sheet0.iter_rows(min_row=cellCriteria):
            #     for col in row:
            #         col.value = None
            #         col._style = None
            wss.insert_rows(2, 60)
            for row in sheet00.iter_rows(1, 60, 1, 15):
                for cell in row:
                    cell.value = wb['SNAP'].cell(cell.row, cell.column).value
                    cell._style = wb['SNAP'].cell(
                        cell.row, cell.column)._style
            altersheet['F6'].value = sheet4['M3'].value
        if st_2chk2 == True:
            if st_2chkType1.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-2' 1번 문항을 체크해주십시오.")
            elif st_2chkType2.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-2' 2번 문항을 체크해주십시오.")
            elif st_2chkType3.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-2' 3번 문항을 체크해주십시오.")
            elif st_2chkType4.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-2' 4번 문항을 체크해주십시오.")
            elif st_2chkType5.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-2' 5번 문항을 체크해주십시오.")
            elif st_2chkType6.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-2' 6번 문항을 체크해주십시오.")
            elif st_2chkType7.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-2' 7번 문항을 체크해주십시오.")
            elif st_2chkType8.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-2' 8번 문항을 체크해주십시오.")
            elif st_2chkType9.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-2' 9번 문항을 체크해주십시오.")
            elif st_2chkType10.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-2' 10번 문항을 체크해주십시오.")
            elif st_2chkType11.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-2' 11번 문항을 체크해주십시오.")
            elif st_2chkType12.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-2' 12번 문항을 체크해주십시오.")
            elif st_2chkType13.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-2' 13번 문항을 체크해주십시오.")
            elif st_2chkType14.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-2' 14번 문항을 체크해주십시오.")
            elif st_2chkType15.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-2' 15번 문항을 체크해주십시오.")
            elif st_2chkType16.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-2' 16번 문항을 체크해주십시오.")
            elif st_2chkType17.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-2' 17번 문항을 체크해주십시오.")
            elif st_2chkType18.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-2' 18번 문항을 체크해주십시오.")
            elif st_2chkType19.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-2' 19번 문항을 체크해주십시오.")
            elif st_2chkType20.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-2' 마지막 문항을 체크해주십시오.")
            if st_2chkType1.get() == "" or st_2chkType2.get() == "" or st_2chkType3.get() == "" or st_2chkType4.get() == "" or st_2chkType5.get() == "" or st_2chkType6.get() == "" or st_2chkType7.get() == "" or st_2chkType8.get() == "" or st_2chkType9.get() == "" or st_2chkType10.get() == "" or st_2chkType11.get() == "" or st_2chkType12.get() == "" or st_2chkType13.get() == "" or st_2chkType14.get() == "" or st_2chkType15.get() == "" or st_2chkType16.get() == "" or st_2chkType17.get() == "" or st_2chkType18.get() == "" or st_2chkType19.get() == "" or st_2chkType20.get() == "":
                return

            sumscore5 = int(st_2chkType1.get()) + int(st_2chkType2.get()) + int(st_2chkType3.get()) + int(st_2chkType4.get()) + int(st_2chkType5.get()) + int(st_2chkType6.get()) + int(st_2chkType7.get()) + int(st_2chkType8.get()) + int(st_2chkType9.get()) + int(st_2chkType10.get()) + int(st_2chkType11.get()) + int(st_2chkType12.get()) + int(st_2chkType13.get()) + int(
                st_2chkType14.get()) + int(st_2chkType15.get()) + int(st_2chkType16.get()) + int(st_2chkType17.get()) + int(st_2chkType18.get()) + int(st_2chkType19.get()) + int(st_2chkType20.get())

            st_2rltScore1 = str(st_2chkType1.get())
            st_2rltScore2 = str(st_2chkType2.get())
            st_2rltScore3 = str(st_2chkType3.get())
            st_2rltScore4 = str(st_2chkType4.get())
            st_2rltScore5 = str(st_2chkType5.get())
            st_2rltScore6 = str(st_2chkType6.get())
            st_2rltScore7 = str(st_2chkType7.get())
            st_2rltScore8 = str(st_2chkType8.get())
            st_2rltScore9 = str(st_2chkType9.get())
            st_2rltScore10 = str(st_2chkType10.get())
            st_2rltScore11 = str(st_2chkType11.get())
            st_2rltScore12 = str(st_2chkType12.get())
            st_2rltScore13 = str(st_2chkType13.get())
            st_2rltScore14 = str(st_2chkType14.get())
            st_2rltScore15 = str(st_2chkType15.get())
            st_2rltScore16 = str(st_2chkType16.get())
            st_2rltScore17 = str(st_2chkType17.get())
            st_2rltScore18 = str(st_2chkType18.get())
            st_2rltScore19 = str(st_2chkType19.get())
            st_2rltScore20 = str(st_2chkType20.get())
            sheet5['E6'] = st_2rltScore1 + " 점"
            sheet5['E9'] = st_2rltScore2 + " 점"
            sheet5['E12'] = st_2rltScore3 + " 점"
            sheet5['E15'] = st_2rltScore4 + " 점"
            sheet5['E18'] = st_2rltScore5 + " 점"
            sheet5['E21'] = st_2rltScore6 + " 점"
            sheet5['E24'] = st_2rltScore7 + " 점"
            sheet5['E27'] = st_2rltScore8 + " 점"
            sheet5['E30'] = st_2rltScore9 + " 점"
            sheet5['E33'] = st_2rltScore10 + " 점"
            sheet5['E36'] = st_2rltScore11 + " 점"
            sheet5['E39'] = st_2rltScore12 + " 점"
            sheet5['E42'] = st_2rltScore13 + " 점"
            sheet5['E45'] = st_2rltScore14 + " 점"
            sheet5['E48'] = st_2rltScore15 + " 점"
            sheet5['E51'] = st_2rltScore16 + " 점"
            sheet5['E54'] = st_2rltScore17 + " 점"
            sheet5['E57'] = st_2rltScore18 + " 점"
            sheet5['E60'] = st_2rltScore19 + " 점"
            sheet5['E63'] = st_2rltScore20 + " 점"
            sheet5['M3'] = sumscore5
            sheet5['K3'] = nameInput.get()

            min_col, min_row, max_col, max_row = range_boundaries(
                'A1')
            # now = datetime.now()
            # for row in sheet0.iter_rows(min_row=cellCriteria):
            #     for col in row:
            #         col.value = None
            #         col._style = None
            wss.insert_rows(2, 66)
            for row in sheet00.iter_rows(1, 66, 1, 15):
                for cell in row:
                    cell.value = wb['ST_2'].cell(cell.row, cell.column).value
                    cell._style = wb['ST_2'].cell(
                        cell.row, cell.column)._style
            altersheet['G6'].value = sheet5['M3'].value
        if st_1chk2 == True:
            if st_1chkType1.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-1' 1번 문항을 체크해주십시오.")
            elif st_1chkType2.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-1' 2번 문항을 체크해주십시오.")
            elif st_1chkType3.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-1' 3번 문항을 체크해주십시오.")
            elif st_1chkType4.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-1' 4번 문항을 체크해주십시오.")
            elif st_1chkType5.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-1' 5번 문항을 체크해주십시오.")
            elif st_1chkType6.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-1' 6번 문항을 체크해주십시오.")
            elif st_1chkType7.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-1' 7번 문항을 체크해주십시오.")
            elif st_1chkType8.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-1' 8번 문항을 체크해주십시오.")
            elif st_1chkType9.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-1' 9번 문항을 체크해주십시오.")
            elif st_1chkType10.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-1' 10번 문항을 체크해주십시오.")
            elif st_1chkType11.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-1' 11번 문항을 체크해주십시오.")
            elif st_1chkType12.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-1' 12번 문항을 체크해주십시오.")
            elif st_1chkType13.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-1' 13번 문항을 체크해주십시오.")
            elif st_1chkType14.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-1' 14번 문항을 체크해주십시오.")
            elif st_1chkType15.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-1' 15번 문항을 체크해주십시오.")
            elif st_1chkType16.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-1' 16번 문항을 체크해주십시오.")
            elif st_1chkType17.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-1' 17번 문항을 체크해주십시오.")
            elif st_1chkType18.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-1' 18번 문항을 체크해주십시오.")
            elif st_1chkType19.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-1' 19번 문항을 체크해주십시오.")
            elif st_1chkType20.get() == "":
                messagebox.showinfo('마음정원', "'STAI-X-1' 마지막 문항을 체크해주십시오.")
            if st_1chkType1.get() == "" or st_1chkType2.get() == "" or st_1chkType3.get() == "" or st_1chkType4.get() == "" or st_1chkType5.get() == "" or st_1chkType6.get() == "" or st_1chkType7.get() == "" or st_1chkType8.get() == "" or st_1chkType9.get() == "" or st_1chkType10.get() == "" or st_1chkType11.get() == "" or st_1chkType12.get() == "" or st_1chkType13.get() == "" or st_1chkType14.get() == "" or st_1chkType15.get() == "" or st_1chkType16.get() == "" or st_1chkType17.get() == "" or st_1chkType18.get() == "" or st_1chkType19.get() == "" or st_1chkType20.get() == "":
                return

            sumscore6 = int(st_1chkType1.get()) + int(st_1chkType2.get()) + int(st_1chkType3.get()) + int(st_1chkType4.get()) + int(st_1chkType5.get()) + int(st_1chkType6.get()) + int(st_1chkType7.get()) + int(st_1chkType8.get()) + int(st_1chkType9.get()) + int(st_1chkType10.get()) + int(st_1chkType11.get()) + int(st_1chkType12.get()) + int(st_1chkType13.get()) + int(
                st_1chkType14.get()) + int(st_1chkType15.get()) + int(st_1chkType16.get()) + int(st_1chkType17.get()) + int(st_1chkType18.get()) + int(st_1chkType19.get()) + int(st_1chkType20.get())

            st_1rltScore1 = str(st_1chkType1.get())
            st_1rltScore2 = str(st_1chkType2.get())
            st_1rltScore3 = str(st_1chkType3.get())
            st_1rltScore4 = str(st_1chkType4.get())
            st_1rltScore5 = str(st_1chkType5.get())
            st_1rltScore6 = str(st_1chkType6.get())
            st_1rltScore7 = str(st_1chkType7.get())
            st_1rltScore8 = str(st_1chkType8.get())
            st_1rltScore9 = str(st_1chkType9.get())
            st_1rltScore10 = str(st_1chkType10.get())
            st_1rltScore11 = str(st_1chkType11.get())
            st_1rltScore12 = str(st_1chkType12.get())
            st_1rltScore13 = str(st_1chkType13.get())
            st_1rltScore14 = str(st_1chkType14.get())
            st_1rltScore15 = str(st_1chkType15.get())
            st_1rltScore16 = str(st_1chkType16.get())
            st_1rltScore17 = str(st_1chkType17.get())
            st_1rltScore18 = str(st_1chkType18.get())
            st_1rltScore19 = str(st_1chkType19.get())
            st_1rltScore20 = str(st_1chkType20.get())
            sheet6['E6'] = st_1rltScore1 + " 점"
            sheet6['E9'] = st_1rltScore2 + " 점"
            sheet6['E12'] = st_1rltScore3 + " 점"
            sheet6['E15'] = st_1rltScore4 + " 점"
            sheet6['E18'] = st_1rltScore5 + " 점"
            sheet6['E21'] = st_1rltScore6 + " 점"
            sheet6['E24'] = st_1rltScore7 + " 점"
            sheet6['E27'] = st_1rltScore8 + " 점"
            sheet6['E30'] = st_1rltScore9 + " 점"
            sheet6['E33'] = st_1rltScore10 + " 점"
            sheet6['E36'] = st_1rltScore11 + " 점"
            sheet6['E39'] = st_1rltScore12 + " 점"
            sheet6['E42'] = st_1rltScore13 + " 점"
            sheet6['E45'] = st_1rltScore14 + " 점"
            sheet6['E48'] = st_1rltScore15 + " 점"
            sheet6['E51'] = st_1rltScore16 + " 점"
            sheet6['E54'] = st_1rltScore17 + " 점"
            sheet6['E57'] = st_1rltScore18 + " 점"
            sheet6['E60'] = st_1rltScore19 + " 점"
            sheet6['E63'] = st_1rltScore20 + " 점"
            sheet6['M3'] = sumscore6
            sheet6['K3'] = nameInput.get()

            min_col, min_row, max_col, max_row = range_boundaries(
                'A1')
            # now = datetime.now()
            # for row in sheet0.iter_rows(min_row=cellCriteria):
            #     for col in row:
            #         col.value = None
            #         col._style = None
            wss.insert_rows(2, 66)
            for row in sheet00.iter_rows(1, 66, 1, 15):
                for cell in row:
                    cell.value = wb['ST_1'].cell(cell.row, cell.column).value
                    cell._style = wb['ST_1'].cell(
                        cell.row, cell.column)._style
            altersheet['H6'].value = sheet6['M3'].value

        if hamachk2 == True:
            interviewsheet.insert_rows(23, 1)
            interviewsheet['D24'].value = 0
            interviewsheet['E24'].value = 0
            interviewsheet['F24'].value = 0
            interviewsheet['G24'].value = 0
            interviewsheet['H24'].value = 0
            interviewsheet['I24'].value = 0
            addedVar = altersheet['AA4'].value - 1
            altersheet['AA4'].value = addedVar
            if hamachkType1.get() == "":
                messagebox.showinfo('마음정원', "'HAM A' 1번 문항을 체크해주십시오.")
            elif hamachkType2.get() == "":
                messagebox.showinfo('마음정원', "'HAM A' 2번 문항을 체크해주십시오.")
            elif hamachkType3.get() == "":
                messagebox.showinfo('마음정원', "'HAM A' 3번 문항을 체크해주십시오.")
            elif hamachkType4.get() == "":
                messagebox.showinfo('마음정원', "'HAM A' 4번 문항을 체크해주십시오.")
            elif hamachkType5.get() == "":
                messagebox.showinfo('마음정원', "'HAM A' 5번 문항을 체크해주십시오.")
            elif hamachkType6.get() == "":
                messagebox.showinfo('마음정원', "'HAM A' 6번 문항을 체크해주십시오.")
            elif hamachkType7.get() == "":
                messagebox.showinfo('마음정원', "'HAM A' 7번 문항을 체크해주십시오.")
            elif hamachkType8.get() == "":
                messagebox.showinfo('마음정원', "'HAM A' 8번 문항을 체크해주십시오.")
            elif hamachkType9.get() == "":
                messagebox.showinfo('마음정원', "'HAM A' 9번 문항을 체크해주십시오.")
            elif hamachkType10.get() == "":
                messagebox.showinfo('마음정원', "'HAM A' 10번 문항을 체크해주십시오.")
            elif hamachkType11.get() == "":
                messagebox.showinfo('마음정원', "'HAM A' 11번 문항을 체크해주십시오.")
            elif hamachkType12.get() == "":
                messagebox.showinfo('마음정원', "'HAM A' 12번 문항을 체크해주십시오.")
            elif hamachkType13.get() == "":
                messagebox.showinfo('마음정원', "'HAM A' 마지막 문항을 체크해주십시오.")
            if hamachkType1.get() == "" or hamachkType2.get() == "" or hamachkType3.get() == "" or hamachkType4.get() == "" or hamachkType5.get() == "" or hamachkType6.get() == "" or hamachkType7.get() == "" or hamachkType8.get() == "" or hamachkType9.get() == "" or hamachkType10.get() == "" or hamachkType11.get() == "" or hamachkType12.get() == "" or hamachkType13.get() == "":
                return


            sumscore7 = int(hamachkType1.get()) + int(hamachkType2.get()) + int(hamachkType3.get()) + int(hamachkType4.get()) + int(hamachkType5.get()) + int(hamachkType6.get()) + int(
            	hamachkType7.get()) + int(hamachkType8.get()) + int(hamachkType9.get()) + int(hamachkType10.get()) + int(hamachkType11.get()) + int(hamachkType12.get()) + int(hamachkType13.get())

             
            hamarltScore1 = str(hamachkType1.get())
            hamarltScore2 = str(hamachkType2.get())
            hamarltScore3 = str(hamachkType3.get())
            hamarltScore4 = str(hamachkType4.get())
            hamarltScore5 = str(hamachkType5.get())
            hamarltScore6 = str(hamachkType6.get())
            hamarltScore7 = str(hamachkType7.get())
            hamarltScore8 = str(hamachkType8.get())
            hamarltScore9 = str(hamachkType9.get())
            hamarltScore10 = str(hamachkType10.get())
            hamarltScore11 = str(hamachkType11.get())
            hamarltScore12 = str(hamachkType12.get())
            hamarltScore13 = str(hamachkType13.get())
            min_col, min_row, max_col, max_row = range_boundaries('A1')
            
            # for row, row_cells in enumerate(altersheet, min_row):
            #     for column, cell in enumerate(row_cells, min_col):
            #         wb['변화'].cell(row=row, column=column+1)._style = copy(cell._style)
            #         wb['변화'].cell(row=row, column=column+1).value = copy(cell.value)
            sheet7['E6'] = hamarltScore1 + " 점"
            sheet7['E9'] = hamarltScore2 + " 점"
            sheet7['E12'] = hamarltScore3 + " 점"
            sheet7['E15'] = hamarltScore4 + " 점"
            sheet7['E18'] = hamarltScore5 + " 점"
            sheet7['E21'] = hamarltScore6 + " 점"
            sheet7['E24'] = hamarltScore7 + " 점"
            sheet7['E27'] = hamarltScore8 + " 점"
            sheet7['E30'] = hamarltScore9 + " 점"
            sheet7['E33'] = hamarltScore10 + " 점"
            sheet7['E36'] = hamarltScore11 + " 점"
            sheet7['E39'] = hamarltScore12 + " 점"
            sheet7['E42'] = hamarltScore13 + " 점"
            sheet7['E47'] = "="+"OFFSET(인터뷰!$D$24,COUNT(인터뷰!D:D)" + str(addedVar)+",)"
            sheet7['C47'] = "="+"OFFSET(인터뷰!$D$24,COUNT(인터뷰!D:D)" + str(addedVar)+',)&"."'
            sheet7['N3'] = ("="+"IFERROR(OFFSET(인터뷰!$D$24,COUNT(인터뷰!D:D)"+str(addedVar)+",)+"+str(sumscore7)+',('+str(sumscore7)+'))')
            sheet7['K3'] = nameInput.get()
            min_col, min_row, max_col, max_row = range_boundaries(
                'A1')

            # now = datetime.now()
            # for row in sheet0.iter_rows(min_row=cellCriteria):
            #     for col in row:
            #         col.value = None
            #         col._style = None
            # ws.insert_rows(2, 50)

            for row, row_cells in enumerate(wb['HAMA'], min_row):
                for column, cell in enumerate(row_cells, min_col):

                    ws.cell(row=row, column=column).value = copy(cell.value)
                    ws.cell(row=row, column=column)._style = copy(cell._style)
                    ws.cell(row=row, column=column).number_format = copy(cell.number_format)
                    ws.cell(row=row, column=column).fill = copy(cell.fill)
                    ws.cell(row=row, column=column).border =  copy(cell.border)
                    ws.cell(row=row, column=column).alignment = copy(cell.alignment)
                    ws.cell(row=row, column=column).font = copy(cell.font)
            wss.insert_rows(2, 50)

            for row in sheet00.iter_rows(1, 50, 1, 15):
                for cell in row:
                    cell.value = wb['HAMA'].cell(cell.row, cell.column).value
                    cell._style = wb['HAMA'].cell(
                        cell.row, cell.column)._style
            interviewsheet['B24'].value = ("% s년 % s월 % s일 % s시 % s분" % (now.year, now.month, now.day, now.hour, now.minute))
            altersheet['I6'].value = ("="+"IFERROR(OFFSET(인터뷰!$D$24,COUNT(인터뷰!D:D)"+str(addedVar)+",)+"+str((sumscore7))+',('+str(sumscore7)+'))')
            interviewsheet['B24']._style = interviewsheet['AA2']._style
            interviewsheet['C24']._style = interviewsheet['AB2']._style
            interviewsheet['D24']._style = interviewsheet['AC2']._style
            interviewsheet['E24']._style = interviewsheet['AD2']._style
            interviewsheet['F24']._style = interviewsheet['AE2']._style
            interviewsheet['G24']._style = interviewsheet['AF2']._style
            interviewsheet['H24']._style = interviewsheet['AG2']._style
            interviewsheet['I24']._style = interviewsheet['AH2']._style
            interviewsheet['J24']._style = interviewsheet['AI2']._style
            
        if hamdchk2 == True:
            if hamachk2 == False and (phqchk3 == True or cdichk3 == True or bdichk3 == True or snapchk3 == True or st_2chk == True or st_1chk == True or hamachk == True or hamdchk == True):
                interviewsheet.insert_rows(23, 1)
                interviewsheet['D24'].value = 0
                interviewsheet['E24'].value = 0
                interviewsheet['F24'].value = 0
                interviewsheet['G24'].value = 0
                interviewsheet['H24'].value = 0
                interviewsheet['I24'].value = 0
                interviewsheet['B24'].value = ("% s년 % s월 % s일 % s시 % s분" % (now.year, now.month, now.day, now.hour, now.minute))
                addedVar = altersheet['AA4'].value - 1
                altersheet['AA4'].value = addedVar
                
                interviewsheet['B24']._style = interviewsheet['AA2']._style
                interviewsheet['C24']._style = interviewsheet['AB2']._style
                interviewsheet['D24']._style = interviewsheet['AC2']._style
                interviewsheet['E24']._style = interviewsheet['AD2']._style
                interviewsheet['F24']._style = interviewsheet['AE2']._style
                interviewsheet['G24']._style = interviewsheet['AF2']._style
                interviewsheet['H24']._style = interviewsheet['AG2']._style
                interviewsheet['I24']._style = interviewsheet['AH2']._style
                interviewsheet['J24']._style = interviewsheet['AI2']._style
            if hamdchkType1.get() == "":
                messagebox.showinfo('마음정원', "'HAM D' 1번 문항을 체크해주십시오.")
            elif hamdchkType2.get() == "":
                messagebox.showinfo('마음정원', "'HAM D' 2번 문항을 체크해주십시오.")
            elif hamdchkType3.get() == "":
                messagebox.showinfo('마음정원', "'HAM D' 3번 문항을 체크해주십시오.")
            elif hamdchkType4.get() == "":
                messagebox.showinfo('마음정원', "'HAM D' 4번 문항을 체크해주십시오.")
            elif hamdchkType5.get() == "":
                messagebox.showinfo('마음정원', "'HAM D' 5번 문항을 체크해주십시오.")
            elif hamdchkType6.get() == "":
                messagebox.showinfo('마음정원', "'HAM D' 6번 문항을 체크해주십시오.")
            elif hamdchkType7.get() == "":
                messagebox.showinfo('마음정원', "'HAM D' 7번 문항을 체크해주십시오.")
            elif hamdchkType8.get() == "":
                messagebox.showinfo('마음정원', "'HAM D' 8번 문항을 체크해주십시오.")
            elif hamdchkType9.get() == "":
                messagebox.showinfo('마음정원', "'HAM D' 9번 문항을 체크해주십시오.")
            elif hamdchkType10.get() == "":
                messagebox.showinfo('마음정원', "'HAM D' 10번 문항을 체크해주십시오.")
            elif hamdchkType11.get() == "":
                messagebox.showinfo('마음정원', "'HAM D' 11번 문항을 체크해주십시오.")
            elif hamdchkType12.get() == "":
                messagebox.showinfo('마음정원', "'HAM D' 마지막 문항을 체크해주십시오.")
            if hamdchkType1.get() == "" or hamdchkType2.get() == "" or hamdchkType3.get() == "" or hamdchkType4.get() == "" or hamdchkType5.get() == "" or hamdchkType6.get() == "" or hamdchkType7.get() == "" or hamdchkType8.get() == "" or hamdchkType9.get() == "" or hamdchkType10.get() == "" or hamdchkType11.get() == "" or hamdchkType12.get() == "":
                return

            sumscore8 = int(hamdchkType1.get()) + int(hamdchkType2.get()) + int(hamdchkType3.get()) + int(hamdchkType4.get()) + int(hamdchkType5.get()) + int(hamdchkType6.get()) + \
                int(hamdchkType7.get()) + int(hamdchkType8.get()) + int(hamdchkType9.get()) + \
                int(hamdchkType10.get()) + \
                int(hamdchkType11.get()) + int(hamdchkType12.get())


            hamdrltScore1 = str(hamdchkType1.get())
            hamdrltScore2 = str(hamdchkType2.get())
            hamdrltScore3 = str(hamdchkType3.get())
            hamdrltScore4 = str(hamdchkType4.get())
            hamdrltScore5 = str(hamdchkType5.get())
            hamdrltScore6 = str(hamdchkType6.get())
            hamdrltScore7 = str(hamdchkType7.get())
            hamdrltScore8 = str(hamdchkType8.get())
            hamdrltScore9 = str(hamdchkType9.get())
            hamdrltScore10 = str(hamdchkType10.get())
            hamdrltScore11 = str(hamdchkType11.get())
            hamdrltScore12 = str(hamdchkType12.get())
            sheet8['E11'] = hamdrltScore1 + " 점"
            sheet8['E19'] = hamdrltScore2 + " 점"
            sheet8['E25'] = hamdrltScore3 + " 점"
            sheet8['E31'] = hamdrltScore4 + " 점"
            sheet8['E37'] = hamdrltScore5 + " 점"
            sheet8['E45'] = hamdrltScore6 + " 점"
            sheet8['E53'] = hamdrltScore7 + " 점"
            sheet8['E59'] = hamdrltScore8 + " 점"
            sheet8['E65'] = hamdrltScore9 + " 점"
            sheet8['E71'] = hamdrltScore10 + " 점"
            sheet8['E79'] = hamdrltScore11 + " 점"
            sheet8['E85'] = hamdrltScore12 + " 점"
            sheet8['E95'] = "="+"OFFSET(인터뷰!$E$24,COUNT(인터뷰!E:E)" + str(addedVar)+",)"
            sheet8['E103'] = "="+"OFFSET(인터뷰!$F$24,COUNT(인터뷰!F:F)" + str(addedVar)+",)"
            sheet8['E111'] = "="+"OFFSET(인터뷰!$G$24,COUNT(인터뷰!G:G)" + str(addedVar)+",)"
            sheet8['E119'] = "="+"OFFSET(인터뷰!$H$24,COUNT(인터뷰!H:H)" + str(addedVar)+",)"
            sheet8['E125'] = "="+"OFFSET(인터뷰!$I$24,COUNT(인터뷰!I:I)" + str(addedVar)+",)"
            sheet8['C95'] = "="+"OFFSET(인터뷰!$E$24,COUNT(인터뷰!E:E)" + str(addedVar)+',)&"."'
            sheet8['C103'] = "="+"OFFSET(인터뷰!$F$24,COUNT(인터뷰!F:F)" + str(addedVar)+',)&"."'
            sheet8['C111'] = "="+"OFFSET(인터뷰!$G$24,COUNT(인터뷰!G:G)" + str(addedVar)+',)&"."'
            sheet8['C119'] = "="+"OFFSET(인터뷰!$H$24,COUNT(인터뷰!H:H)" + str(addedVar)+',)&"."'
            sheet8['C125'] = "="+"OFFSET(인터뷰!$I$24,COUNT(인터뷰!I:I)" + str(addedVar)+',)&"."'
            sheet8['N3'] = ("="+"IFERROR((OFFSET(인터뷰!$E$24,COUNT(인터뷰!E:E)"+str(addedVar)+",))+(OFFSET(인터뷰!$F$24,COUNT(인터뷰!F:F)"+str(addedVar)+",))+(OFFSET(인터뷰!$G$24,COUNT(인터뷰!G:G)"+str(
            	addedVar)+",))+(OFFSET(인터뷰!$H$24,COUNT(인터뷰!H:H)"+str(addedVar)+",))+(OFFSET(인터뷰!$I$24,COUNT(인터뷰!I:I)"+str(addedVar)+",))+"+str(sumscore8)+",("+str(sumscore8)+"))")
            sheet8['K3'] = nameInput.get()
            min_col, min_row, max_col, max_row = range_boundaries(
                'A1')

            ws.insert_rows(2, 128)
            for row, row_cells in enumerate(wb['HAMD'], min_row):
                for column, cell in enumerate(row_cells, min_col):
                    
                    ws.cell(row=row, column=column).value = copy(cell.value)
                    ws.cell(row=row, column=column)._style = copy(cell._style)
                    ws.cell(row=row, column=column).number_format = copy(
                    	cell.number_format)
                    ws.cell(row=row, column=column).fill = copy(cell.fill)
                    ws.cell(row=row, column=column).border = copy(cell.border)
                    ws.cell(row=row, column=column).alignment = copy(
                    	cell.alignment)
                    ws.cell(row=row, column=column).font = copy(cell.font)
            wss.insert_rows(2, 128)
            
            for row in sheet00.iter_rows(1, 128, 1, 15):
                for cell in row:
                    cell.value = wb['HAMD'].cell(cell.row, cell.column).value
                    cell._style = wb['HAMD'].cell(
                        cell.row, cell.column)._style
            altersheet['J6'].value = ("="+"=IFERROR((OFFSET(인터뷰!$E$24,COUNT(인터뷰!E:E)"+str(addedVar)+",))+(OFFSET(인터뷰!$F$24,COUNT(인터뷰!F:F)"+str(addedVar)+",))+(OFFSET(인터뷰!$G$24,COUNT(인터뷰!G:G)"+str(
            	addedVar)+",))+(OFFSET(인터뷰!$H$24,COUNT(인터뷰!H:H)"+str(addedVar)+",))+(OFFSET(인터뷰!$I$24,COUNT(인터뷰!I:I)"+str(addedVar)+",))+"+str(sumscore8)+",("+str(sumscore8)+"))")
        
        for row in altersheet.iter_rows(min_row,altersheet.max_row):
            altersheet['C'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($C$"+str(min_row+6)+':$C$'+str(
                altersheet.max_row)+'," 점","")))/COUNTA($C'+str(min_row+6)+':$C'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
            altersheet['D'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($D$"+str(min_row+6)+':$D$'+str(
                altersheet.max_row)+'," 점","")))/COUNTA($D'+str(min_row+6)+':$D'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
            altersheet['E'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($E$"+str(min_row+6)+':$E$'+str(
                altersheet.max_row)+'," 점","")))/COUNTA($E'+str(min_row+6)+':$E'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
            altersheet['F'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($F$"+str(min_row+6)+':$F$'+str(
                altersheet.max_row)+'," 점","")))/COUNTA($F'+str(min_row+6)+':$F'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
            altersheet['G'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($G$"+str(min_row+6)+':$G$'+str(
                altersheet.max_row)+'," 점","")))/COUNTA($G'+str(min_row+6)+':$G'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
            altersheet['H'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($H$"+str(min_row+6)+':$H$'+str(
                altersheet.max_row)+'," 점","")))/COUNTA($H'+str(min_row+6)+':$H'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
            altersheet['I'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($I$"+str(min_row+6)+':$I$'+str(
                altersheet.max_row)+'," 점","")))/COUNTA($I'+str(min_row+6)+':$I'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
            altersheet['J'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($J$"+str(min_row+6)+':$J$'+str(
                altersheet.max_row)+'," 점","")))/COUNTA($J'+str(min_row+6)+':$J'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'

            
        altersheet['K2'].value = wss['K2'].value
        altersheet['K3'].value = wss['K3'].value
        altersheet['B6'].value = ("% s년 % s월 % s일 % s시 % s분" %
                              (now.year, now.month, now.day, now.hour, now.minute))

        for row in range(7,altersheet.max_row):
            altersheet.row_dimensions[row].height = 21

        altersheet['L6'].value = "=SUM(INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-2),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-3),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-4),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-5),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-6),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-7),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-8),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-9),2)))) &"+'" 점"'
        altersheet.insert_rows(6, 1)


        print("interview max row->"+str(interviewsheet.max_row))

        altersheet['B7']._style = altersheet['AA2']._style
        altersheet['C7']._style = altersheet['AB2']._style
        altersheet['D7']._style = altersheet['AC2']._style
        altersheet['E7']._style = altersheet['AD2']._style
        altersheet['F7']._style = altersheet['AE2']._style
        altersheet['G7']._style = altersheet['AF2']._style
        altersheet['H7']._style = altersheet['AG2']._style
        altersheet['I6']._style = altersheet['AH2']._style
        altersheet['J6']._style = altersheet['AI2']._style




        altersheet['K7']._style = altersheet['AJ2']._style
        altersheet['L7']._style = altersheet['AK2']._style
        
        altersheet['K7'].value = "=TRUNC(SUM(INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-1),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-2),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-3),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-4),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-5),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-6),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-7),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-8),2))))/COUNT(INDIRECT(ADDRESS(ROW(),COLUMN()-8)),INDIRECT(ADDRESS(ROW(),COLUMN()-7)),INDIRECT(ADDRESS(ROW(),COLUMN()-6)),INDIRECT(ADDRESS(ROW(),COLUMN()-5)),INDIRECT(ADDRESS(ROW(),COLUMN()-4)),INDIRECT(ADDRESS(ROW(),COLUMN()-3)),INDIRECT(ADDRESS(ROW(),COLUMN()-2)),INDIRECT(ADDRESS(ROW(),COLUMN()-1))),1)&"+'" 점"'
        
        try:
            wbb.save('//DESKTOP/ptest//xlsm//'+nameInput.get() +
                    '['+idInput.get()+'].xlsm')
            messagebox.showinfo('마음정원', "제출이 완료 되었습니다. 수고하셨습니다.")
            wb.close()
            wbb.close()
            root.quit()
        except:
            messagebox.showinfo('마음정원', "메인 컴퓨터 엑셀파일이 열려있습니다.")

class selFrames:

    srtchkType1 = IntVar()
    srtchkType2 = IntVar()
    srtchkType3 = IntVar()
    srtchkType4 = IntVar()
    srtchkType5 = IntVar()
    srtchkType6 = IntVar()
    srtchkType7 = IntVar()
    srtchkType8 = IntVar()

    image_up = tk.PhotoImage(
        file='images/btnnormal.png')
    image_down = tk.PhotoImage(
        file='images/btnclicked.png')

    def sumTab():
        criteria = ["PHQ", "CDI", "BDI", "SNAP",
                    "ST_2", "ST_1", "HAMA", "HAMD"]

        global a
        a = ""
        if selFrames.srtchkType1.get() == 1:
            a = " / " + criteria[0]
        if selFrames.srtchkType2.get() == 1:
            a += " / " + criteria[1]
        if selFrames.srtchkType3.get() == 1:
            a += " / " + criteria[2]
        if selFrames.srtchkType4.get() == 1:
            a += " / " + criteria[3]
        if selFrames.srtchkType5.get() == 1:
            a += " / " + criteria[4]
        if selFrames.srtchkType6.get() == 1:
            a += " / " + criteria[5]
        if selFrames.srtchkType7.get() == 1:
            a += " / " + criteria[6]
        if selFrames.srtchkType8.get() == 1:
            a += " / " + criteria[7]

    def checkNameAndNum():
        
        
        
        now = datetime.now()
        global cellCriteria
        cellCriteria = 0

        cache_file = ('//DESKTOP/ptest//xlsm//'+nameInput.get()+'['+idInput.get()+'].xlsm')
        if not os.path.exists(cache_file):
            wb.save('//DESKTOP/ptest//xlsm//'+nameInput.get()+'['+idInput.get()+'].xlsm')
        wbb = openpyxl.load_workbook('//DESKTOP/ptest//xlsm//'+nameInput.get()+'['+idInput.get()+'].xlsm', keep_vba=True)
        try:
            wbb.save('//DESKTOP/ptest//xlsm//'+nameInput.get()+'['+idInput.get()+'].xlsm')    
        except:
            messagebox.showinfo('마음정원', "메인 컴퓨터 엑셀파일이 열려있습니다.")
            wb.close()
            root.quit()
      


        namecomparison2 = wb['Ptest']
        idcomparison2 = wb['Ptest']


        namecomparison2 = wb['Ptest']
        idcomparison2 = wb['Ptest']
        namecomparison2['K3'].value = nameInput.get()
        idcomparison2['K2'].value = idInput.get()


        if len(limitName.get()) == 0:
            messagebox.showinfo('마음정원', '이름을 입력해주십시오.')
            nameInput.focus_set()
            return
        elif len(limitid.get()) == 0:
            messagebox.showinfo('마음정원', 'ID를 입력해주십시오.')
            idInput.focus_set()
            return

        if selFrames.srtchkType1.get() + selFrames.srtchkType2.get() + selFrames.srtchkType3.get() + selFrames.srtchkType4.get() + selFrames.srtchkType5.get() + selFrames.srtchkType6.get() + selFrames.srtchkType7.get() + selFrames.srtchkType8.get() == 0:
            messagebox.showinfo('마음정원', "유형 체크가 필요합니다.")
            return

        if selFrames.srtchkType1.get() + selFrames.srtchkType2.get() + selFrames.srtchkType3.get() + selFrames.srtchkType4.get() + selFrames.srtchkType5.get() + selFrames.srtchkType6.get() + selFrames.srtchkType7.get() + selFrames.srtchkType8.get() == 1:
            try:
                global displayPage
                displayPage = VerticalScrolledFrame(root)
                wbb = openpyxl.load_workbook('//DESKTOP/ptest//xlsm//'+nameInput.get()+'['+idInput.get()+'].xlsm', keep_vba=True)
                wss = wbb.active
                sheet00 = wbb['Ptest']
                altersheet = wbb['변화']
                interviewsheet = wbb["인터뷰"]
                avgVar = altersheet['AC4'].value + 1
                altersheet['AC4'].value = avgVar
                intVar = altersheet['AD4'].value + 1
                altersheet['AD4'].value = intVar

            except:
                messagebox.showinfo('마음정원', "메인 컴퓨터에 저장되어 있던 엑셀파일이 활성시트로 선택되어 있지 않은 상태에서 저장되어 있습니다. 오류가 발생한 시트에서 불필요한 행을 제거 후 활성 시트로 선택하여 다시 저장 후 시도해주시기 바랍니다.")
                return
            if (selFrames.srtchkType1.get()) == 1:
                mainContent.phqchkFunc()
                
                def btnPhq():

                    if phqchk == True:
                        if phqchkType1.get() == "":
                            messagebox.showinfo('마음정원', "1번 문항을 체크해주십시오.")
                        elif phqchkType2.get() == "":
                            messagebox.showinfo('마음정원', "2번 문항을 체크해주십시오.")
                        elif phqchkType3.get() == "":
                            messagebox.showinfo('마음정원', "3번 문항을 체크해주십시오.")
                        elif phqchkType4.get() == "":
                            messagebox.showinfo('마음정원', "4번 문항을 체크해주십시오.")
                        elif phqchkType5.get() == "":
                            messagebox.showinfo('마음정원', "5번 문항을 체크해주십시오.")
                        elif phqchkType6.get() == "":
                            messagebox.showinfo('마음정원', "6번 문항을 체크해주십시오.")
                        elif phqchkType7.get() == "":
                            messagebox.showinfo('마음정원', "7번 문항을 체크해주십시오.")
                        elif phqchkType8.get() == "":
                            messagebox.showinfo('마음정원', "8번 문항을 체크해주십시오.")
                        elif phqlastchkType.get() == "":
                            messagebox.showinfo('마음정원', "마지막 문항을 체크해주십시오.")
                        if phqchkType1.get() == "" or phqchkType2.get() == "" or phqchkType3.get() == "" or phqchkType4.get() == "" or phqchkType5.get() == "" or phqchkType6.get() == "" or phqchkType7.get() == "" or phqchkType8.get() == "" or phqlastchkType.get() == "":
                            return

                        sumscore = int(phqchkType1.get()) + int(phqchkType2.get()) + int(phqchkType3.get()) + int(phqchkType4.get()) + \
                            int(phqchkType5.get()) + int(phqchkType6.get()) + int(phqchkType7.get()) + \
                            int(phqchkType8.get())
                        phqrltScore1 = str(phqchkType1.get())
                        phqrltScore2 = str(phqchkType2.get())
                        phqrltScore3 = str(phqchkType3.get())
                        phqrltScore4 = str(phqchkType4.get())
                        phqrltScore5 = str(phqchkType5.get())
                        phqrltScore6 = str(phqchkType6.get())
                        phqrltScore7 = str(phqchkType7.get())
                        phqrltScore8 = str(phqchkType8.get())
                        sheet1['E6'] = phqrltScore1 + " 점"
                        sheet1['E9'] = phqrltScore2 + " 점"
                        sheet1['E12'] = phqrltScore3 + " 점"
                        sheet1['E15'] = phqrltScore4 + " 점"
                        sheet1['E18'] = phqrltScore5 + " 점"
                        sheet1['E21'] = phqrltScore6 + " 점"
                        sheet1['E24'] = phqrltScore7 + " 점"
                        sheet1['E27'] = phqrltScore8 + " 점"
                        sheet1['E30'] = "X"
                        sheet1['N3'] = sumscore
                        sheet1['K3'] = nameInput.get()
                        min_col, min_row, max_col, max_row = range_boundaries(
                            'A1')
                        wss.insert_rows(2, 33)
                        for row in sheet00.iter_rows(1, 33, 1, 15):
                            for cell in row:
                                cell.value = wb['PHQ'].cell(cell.row, cell.column).value
                                cell._style = wb['PHQ'].cell(
                                    cell.row, cell.column)._style

                        altersheet['C6'].value = sheet1['N3'].value
                        for row in range(7,altersheet.max_row):
                            altersheet.row_dimensions[row].height = 21
                        # interviewsheet['B24'].value = ("% s년 % s월 % s일 % s시 % s분" % (now.year, now.month, now.day, now.hour, now.minute))
                        altersheet['B6'].value = ("% s년 % s월 % s일 % s시 % s분" % (
                            now.year, now.month, now.day, now.hour, now.minute))
                        altersheet['AA5'].value = "=SUM(IF(ISBLANK(C"+str(
                            6+min_row)+"),0,VALUE(LEFT(C"+str(6+min_row)+",2)))) &"+'" 점"'
                        altersheet['L6'].value = "=SUM(INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-2),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-3),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-4),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-5),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-6),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-7),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-8),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-9),2)))) &"+'" 점"'
                        altersheet.insert_rows(6, 1)

                        altersheet['B7']._style = altersheet['AA2']._style
                        altersheet['C7']._style = altersheet['AB2']._style
                        altersheet['D7']._style = altersheet['AC2']._style
                        altersheet['E7']._style = altersheet['AD2']._style
                        altersheet['F7']._style = altersheet['AE2']._style
                        altersheet['G7']._style = altersheet['AF2']._style
                        altersheet['H7']._style = altersheet['AG2']._style
                        altersheet['I6']._style = altersheet['AH2']._style
                        altersheet['J6']._style = altersheet['AI2']._style


                        altersheet['K7']._style = altersheet['AJ2']._style
                        altersheet['L7']._style = altersheet['AK2']._style
                        altersheet['K7'].value = "=TRUNC(SUM(INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-1),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-2),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-3),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-4),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-5),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-6),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-7),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-8),2))))/COUNT(INDIRECT(ADDRESS(ROW(),COLUMN()-8)),INDIRECT(ADDRESS(ROW(),COLUMN()-7)),INDIRECT(ADDRESS(ROW(),COLUMN()-6)),INDIRECT(ADDRESS(ROW(),COLUMN()-5)),INDIRECT(ADDRESS(ROW(),COLUMN()-4)),INDIRECT(ADDRESS(ROW(),COLUMN()-3)),INDIRECT(ADDRESS(ROW(),COLUMN()-2)),INDIRECT(ADDRESS(ROW(),COLUMN()-1))),1)&"+'" 점"'
        

                        try:
                            wbb.save('//DESKTOP/ptest//xlsm//'+nameInput.get()+'['+idInput.get()+'].xlsm')
                            messagebox.showinfo('마음정원', "제출이 완료 되었습니다. 수고하셨습니다.")
                            wbb.close()
                            root.quit()
                        except:
                            messagebox.showinfo('마음정원', "메인 컴퓨터 엑셀파일이 열려있습니다.")
                min_col, min_row, max_col, max_row = range_boundaries('A1')
                for row in altersheet.iter_rows(min_row, altersheet.max_row):
                    altersheet['C'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($C$"+str(min_row+6)+':$C$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($C'+str(min_row+6)+':$C'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['D'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($D$"+str(min_row+6)+':$D$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($D'+str(min_row+6)+':$D'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['E'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($E$"+str(min_row+6)+':$E$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($E'+str(min_row+6)+':$E'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['F'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($F$"+str(min_row+6)+':$F$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($F'+str(min_row+6)+':$F'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['G'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($G$"+str(min_row+6)+':$G$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($G'+str(min_row+6)+':$G'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['H'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($H$"+str(min_row+6)+':$H$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($H'+str(min_row+6)+':$H'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['I'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($I$"+str(min_row+6)+':$I$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($I'+str(min_row+6)+':$I'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['J'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($J$"+str(min_row+6)+':$J$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($J'+str(min_row+6)+':$J'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                submit = Button(displayPage.inner, image=addsheetImg,
                                activebackground="#f0f0f0", command=btnPhq, borderwidth=0)
                submit.place(relx=0.5, anchor="center",
                             width=366, height=125)
                submit.pack(ipady=30, side=BOTTOM)
                submit.focus_set()
                selFrames.mainHide()
                now = datetime.now()  # now 함수는 시간마다 계산해서 처리해야 하므로 여러개 배치
                notebook.add2(displayPage, text="% s년 % s월 % s일 % s시" %
                              (now.year, now.month, now.day, now.hour) + a)
                selFrames.mainAdd()

            if (selFrames.srtchkType2.get()) == 1:
                mainContent.cdichkFunc()

                def btnCdi():
                    if cdichk == True:
                        if cdichkType1.get() == "":
                            messagebox.showinfo('마음정원', "1번 문항을 체크해주십시오.")
                        elif cdichkType2.get() == "":
                            messagebox.showinfo('마음정원', "2번 문항을 체크해주십시오.")
                        elif cdichkType3.get() == "":
                            messagebox.showinfo('마음정원', "3번 문항을 체크해주십시오.")
                        elif cdichkType4.get() == "":
                            messagebox.showinfo('마음정원', "4번 문항을 체크해주십시오.")
                        elif cdichkType5.get() == "":
                            messagebox.showinfo('마음정원', "5번 문항을 체크해주십시오.")
                        elif cdichkType6.get() == "":
                            messagebox.showinfo('마음정원', "6번 문항을 체크해주십시오.")
                        elif cdichkType7.get() == "":
                            messagebox.showinfo('마음정원', "7번 문항을 체크해주십시오.")
                        elif cdichkType8.get() == "":
                            messagebox.showinfo('마음정원', "8번 문항을 체크해주십시오.")
                        elif cdichkType9.get() == "":
                            messagebox.showinfo('마음정원', "9번 문항을 체크해주십시오.")
                        elif cdichkType10.get() == "":
                            messagebox.showinfo('마음정원', "10번 문항을 체크해주십시오.")
                        elif cdichkType11.get() == "":
                            messagebox.showinfo('마음정원', "11번 문항을 체크해주십시오.")
                        elif cdichkType12.get() == "":
                            messagebox.showinfo('마음정원', "12번 문항을 체크해주십시오.")
                        elif cdichkType13.get() == "":
                            messagebox.showinfo('마음정원', "13번 문항을 체크해주십시오.")
                        elif cdichkType14.get() == "":
                            messagebox.showinfo('마음정원', "14번 문항을 체크해주십시오.")
                        elif cdichkType15.get() == "":
                            messagebox.showinfo('마음정원', "15번 문항을 체크해주십시오.")
                        elif cdichkType16.get() == "":
                            messagebox.showinfo('마음정원', "16번 문항을 체크해주십시오.")
                        elif cdichkType17.get() == "":
                            messagebox.showinfo('마음정원', "17번 문항을 체크해주십시오.")
                        elif cdichkType18.get() == "":
                            messagebox.showinfo('마음정원', "18번 문항을 체크해주십시오.")
                        elif cdichkType19.get() == "":
                            messagebox.showinfo('마음정원', "19번 문항을 체크해주십시오.")
                        elif cdichkType20.get() == "":
                            messagebox.showinfo('마음정원', "20번 문항을 체크해주십시오.")
                        elif cdichkType21.get() == "":
                            messagebox.showinfo('마음정원', "21번 문항을 체크해주십시오.")
                        elif cdichkType22.get() == "":
                            messagebox.showinfo('마음정원', "22번 문항을 체크해주십시오.")
                        elif cdichkType23.get() == "":
                            messagebox.showinfo('마음정원', "23번 문항을 체크해주십시오.")
                        elif cdichkType24.get() == "":
                            messagebox.showinfo('마음정원', "24번 문항을 체크해주십시오.")
                        elif cdichkType25.get() == "":
                            messagebox.showinfo('마음정원', "25번 문항을 체크해주십시오.")
                        elif cdichkType26.get() == "":
                            messagebox.showinfo('마음정원', "26번 문항을 체크해주십시오.")
                        elif cdichkType27.get() == "":
                            messagebox.showinfo('마음정원', "마지막 문항을 체크해주십시오.")
                        if cdichkType1.get() == "" or cdichkType2.get() == "" or cdichkType3.get() == "" or cdichkType4.get() == "" or cdichkType5.get() == "" or cdichkType6.get() == "" or cdichkType7.get() == "" or cdichkType8.get() == "" or cdichkType9.get() == "" or cdichkType10.get() == "" or cdichkType11.get() == "" or cdichkType12.get() == "" or cdichkType13.get() == "" or cdichkType14.get() == "" or cdichkType15.get() == "" or cdichkType16.get() == "" or cdichkType17.get() == "" or cdichkType18.get() == "" or cdichkType19.get() == "" or cdichkType20.get() == "" or cdichkType21.get() == "" or cdichkType22.get() == "" or cdichkType23.get() == "" or cdichkType24.get() == "" or cdichkType25.get() == "" or cdichkType26.get() == "" or cdichkType27.get() == "":
                            return
                        sumscore2 = int(cdichkType1.get()) + int(cdichkType2.get()) + int(cdichkType3.get()) + int(cdichkType4.get()) + int(cdichkType5.get()) + int(cdichkType6.get()) + int(cdichkType7.get()) + int(cdichkType8.get()) + int(cdichkType9.get()) + int(cdichkType10.get()) + int(cdichkType11.get()) + int(cdichkType12.get()) + int(cdichkType13.get()) + int(
                            cdichkType14.get()) + int(cdichkType15.get()) + int(cdichkType16.get()) + int(cdichkType17.get()) + int(cdichkType18.get()) + int(cdichkType19.get()) + int(cdichkType20.get()) + int(cdichkType21.get()) + int(cdichkType22.get()) + int(cdichkType23.get()) + int(cdichkType24.get()) + int(cdichkType25.get()) + int(cdichkType26.get()) + int(cdichkType27.get())
                        cdirltScore1 = str(cdichkType1.get())
                        cdirltScore2 = str(cdichkType2.get())
                        cdirltScore3 = str(cdichkType3.get())
                        cdirltScore4 = str(cdichkType4.get())
                        cdirltScore5 = str(cdichkType5.get())
                        cdirltScore6 = str(cdichkType6.get())
                        cdirltScore7 = str(cdichkType7.get())
                        cdirltScore8 = str(cdichkType8.get())
                        cdirltScore9 = str(cdichkType9.get())
                        cdirltScore10 = str(cdichkType10.get())
                        cdirltScore11 = str(cdichkType11.get())
                        cdirltScore12 = str(cdichkType12.get())
                        cdirltScore13 = str(cdichkType13.get())
                        cdirltScore14 = str(cdichkType14.get())
                        cdirltScore15 = str(cdichkType15.get())
                        cdirltScore16 = str(cdichkType16.get())
                        cdirltScore17 = str(cdichkType17.get())
                        cdirltScore18 = str(cdichkType18.get())
                        cdirltScore19 = str(cdichkType19.get())
                        cdirltScore20 = str(cdichkType20.get())
                        cdirltScore21 = str(cdichkType21.get())
                        cdirltScore22 = str(cdichkType22.get())
                        cdirltScore23 = str(cdichkType23.get())
                        cdirltScore24 = str(cdichkType24.get())
                        cdirltScore25 = str(cdichkType25.get())
                        cdirltScore26 = str(cdichkType26.get())
                        cdirltScore27 = str(cdichkType27.get())
                        sheet2['E7'] = cdirltScore1 + " 점"
                        sheet2['E12'] = cdirltScore2 + " 점"
                        sheet2['E17'] = cdirltScore3 + " 점"
                        sheet2['E22'] = cdirltScore4 + " 점"
                        sheet2['E27'] = cdirltScore5 + " 점"
                        sheet2['E32'] = cdirltScore6 + " 점"
                        sheet2['E37'] = cdirltScore7 + " 점"
                        sheet2['E42'] = cdirltScore8 + " 점"
                        sheet2['E47'] = cdirltScore9 + " 점"
                        sheet2['E52'] = cdirltScore10 + " 점"
                        sheet2['E57'] = cdirltScore11 + " 점"
                        sheet2['E62'] = cdirltScore12 + " 점"
                        sheet2['E67'] = cdirltScore13 + " 점"
                        sheet2['E72'] = cdirltScore14 + " 점"
                        sheet2['E77'] = cdirltScore15 + " 점"
                        sheet2['E82'] = cdirltScore16 + " 점"
                        sheet2['E87'] = cdirltScore17 + " 점"
                        sheet2['E92'] = cdirltScore18 + " 점"
                        sheet2['E97'] = cdirltScore19 + " 점"
                        sheet2['E102'] = cdirltScore20 + " 점"
                        sheet2['E107'] = cdirltScore21 + " 점"
                        sheet2['E112'] = cdirltScore22 + " 점"
                        sheet2['E117'] = cdirltScore23 + " 점"
                        sheet2['E122'] = cdirltScore24 + " 점"
                        sheet2['E127'] = cdirltScore25 + " 점"
                        sheet2['E132'] = cdirltScore26 + " 점"
                        sheet2['E137'] = cdirltScore27 + " 점"
                        sheet2['N3'] = sumscore2
                        sheet2['K3'] = nameInput.get()
                        min_col, min_row, max_col, max_row = range_boundaries(
                            'A1')

                        wss.insert_rows(2, 140)
                        for row in sheet00.iter_rows(1, 140, 1, 15):
                            for cell in row:
                                cell.value = wb['CDI'].cell(cell.row, cell.column).value
                                cell._style = wb['CDI'].cell(
                                    cell.row, cell.column)._style                        
                        altersheet['D6'].value = sheet2['N3'].value
                        for row in range(7,altersheet.max_row):
                            altersheet.row_dimensions[row].height = 21
                        # interviewsheet['B24'].value = ("% s년 % s월 % s일 % s시 % s분" % (now.year, now.month, now.day, now.hour, now.minute))
                        altersheet['B6'].value = ("% s년 % s월 % s일 % s시 % s분" % (now.year, now.month, now.day, now.hour, now.minute))
                        altersheet['L6'].value = "=SUM(INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-2),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-3),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-4),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-5),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-6),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-7),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-8),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-9),2)))) &"+'" 점"'
                        altersheet.insert_rows(6, 1)
                        altersheet['B7']._style = altersheet['AA2']._style
                        altersheet['C7']._style = altersheet['AB2']._style
                        altersheet['D7']._style = altersheet['AC2']._style
                        altersheet['E7']._style = altersheet['AD2']._style
                        altersheet['F7']._style = altersheet['AE2']._style
                        altersheet['G7']._style = altersheet['AF2']._style
                        altersheet['H7']._style = altersheet['AG2']._style
                        altersheet['I6']._style = altersheet['AH2']._style
                        altersheet['J6']._style = altersheet['AI2']._style

                        altersheet['K7']._style = altersheet['AJ2']._style
                        altersheet['L7']._style = altersheet['AK2']._style
                        altersheet['K7'].value = "=TRUNC(SUM(INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-1),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-2),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-3),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-4),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-5),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-6),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-7),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-8),2))))/COUNT(INDIRECT(ADDRESS(ROW(),COLUMN()-8)),INDIRECT(ADDRESS(ROW(),COLUMN()-7)),INDIRECT(ADDRESS(ROW(),COLUMN()-6)),INDIRECT(ADDRESS(ROW(),COLUMN()-5)),INDIRECT(ADDRESS(ROW(),COLUMN()-4)),INDIRECT(ADDRESS(ROW(),COLUMN()-3)),INDIRECT(ADDRESS(ROW(),COLUMN()-2)),INDIRECT(ADDRESS(ROW(),COLUMN()-1))),1)&"+'" 점"'
        

                        try:
                            wbb.save('//DESKTOP/ptest//xlsm//'+nameInput.get()+'['+idInput.get()+'].xlsm')
                            messagebox.showinfo('마음정원', "제출이 완료 되었습니다. 수고하셨습니다.")
                            wbb.close()
                            root.quit()
                        except:
                            messagebox.showinfo('마음정원', "메인 컴퓨터 엑셀파일이 열려있습니다.")
                min_col, min_row, max_col, max_row = range_boundaries('A1')
                for row in altersheet.iter_rows(min_row,altersheet.max_row):
                    altersheet['C'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($C$"+str(min_row+6)+':$C$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($C'+str(min_row+6)+':$C'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['D'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($D$"+str(min_row+6)+':$D$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($D'+str(min_row+6)+':$D'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['E'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($E$"+str(min_row+6)+':$E$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($E'+str(min_row+6)+':$E'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['F'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($F$"+str(min_row+6)+':$F$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($F'+str(min_row+6)+':$F'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['G'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($G$"+str(min_row+6)+':$G$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($G'+str(min_row+6)+':$G'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['H'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($H$"+str(min_row+6)+':$H$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($H'+str(min_row+6)+':$H'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['I'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($I$"+str(min_row+6)+':$I$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($I'+str(min_row+6)+':$I'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['J'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($J$"+str(min_row+6)+':$J$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($J'+str(min_row+6)+':$J'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                submit = Button(displayPage.inner, image=addsheetImg,
                                activebackground="#f0f0f0", command=btnCdi, borderwidth=0)
                submit.place(relx=0.5, anchor="center",
                             width=366, height=125)
                submit.pack(ipady=30, side=BOTTOM)
                submit.focus_set()
                selFrames.mainHide()
                now = datetime.now()
                notebook.add2(displayPage, text="% s년 % s월 % s일 % s시" %
                              (now.year, now.month, now.day, now.hour) + a)
                selFrames.mainAdd()

            if (selFrames.srtchkType3.get()) == 1:
                mainContent.bdichkFunc()

                def btnBdi():

                    if bdichk == True:
                        if bdichkType1.get() == "":
                            messagebox.showinfo('마음정원', "1번 문항을 체크해주십시오.")
                        elif bdichkType2.get() == "":
                            messagebox.showinfo('마음정원', "2번 문항을 체크해주십시오.")
                        elif bdichkType3.get() == "":
                            messagebox.showinfo('마음정원', "3번 문항을 체크해주십시오.")
                        elif bdichkType4.get() == "":
                            messagebox.showinfo('마음정원', "4번 문항을 체크해주십시오.")
                        elif bdichkType5.get() == "":
                            messagebox.showinfo('마음정원', "5번 문항을 체크해주십시오.")
                        elif bdichkType6.get() == "":
                            messagebox.showinfo('마음정원', "6번 문항을 체크해주십시오.")
                        elif bdichkType7.get() == "":
                            messagebox.showinfo('마음정원', "7번 문항을 체크해주십시오.")
                        elif bdichkType8.get() == "":
                            messagebox.showinfo('마음정원', "8번 문항을 체크해주십시오.")
                        elif bdichkType9.get() == "":
                            messagebox.showinfo('마음정원', "9번 문항을 체크해주십시오.")
                        elif bdichkType10.get() == "":
                            messagebox.showinfo('마음정원', "10번 문항을 체크해주십시오.")
                        elif bdichkType11.get() == "":
                            messagebox.showinfo('마음정원', "11번 문항을 체크해주십시오.")
                        elif bdichkType12.get() == "":
                            messagebox.showinfo('마음정원', "12번 문항을 체크해주십시오.")
                        elif bdichkType13.get() == "":
                            messagebox.showinfo('마음정원', "13번 문항을 체크해주십시오.")
                        elif bdichkType14.get() == "":
                            messagebox.showinfo('마음정원', "14번 문항을 체크해주십시오.")
                        elif bdichkType15.get() == "":
                            messagebox.showinfo('마음정원', "15번 문항을 체크해주십시오.")
                        elif bdichkType16.get() == "":
                            messagebox.showinfo('마음정원', "16번 문항을 체크해주십시오.")
                        elif bdichkType17.get() == "":
                            messagebox.showinfo('마음정원', "17번 문항을 체크해주십시오.")
                        elif bdichkType18.get() == "":
                            messagebox.showinfo('마음정원', "18번 문항을 체크해주십시오.")
                        elif bdichkType19.get() == "":
                            messagebox.showinfo('마음정원', "19번 문항을 체크해주십시오.")
                        elif bdichkType20.get() == "":
                            messagebox.showinfo('마음정원', "20번 문항을 체크해주십시오.")
                        elif bdichkType21.get() == "":
                            messagebox.showinfo('마음정원', "마지막 문항을 체크해주십시오.")
                        if bdichkType1.get() == "" or bdichkType2.get() == "" or bdichkType3.get() == "" or bdichkType4.get() == "" or bdichkType5.get() == "" or bdichkType6.get() == "" or bdichkType7.get() == "" or bdichkType8.get() == "" or bdichkType9.get() == "" or bdichkType10.get() == "" or bdichkType11.get() == "" or bdichkType12.get() == "" or bdichkType13.get() == "" or bdichkType14.get() == "" or bdichkType15.get() == "" or bdichkType16.get() == "" or bdichkType17.get() == "" or bdichkType18.get() == "" or bdichkType19.get() == "" or bdichkType20.get() == "" or bdichkType21.get() == "":
                            return
                        sumscore3 = int(bdichkType1.get()) + int(bdichkType2.get()) + int(bdichkType3.get()) + int(bdichkType4.get()) + int(bdichkType5.get()) + int(bdichkType6.get()) + int(bdichkType7.get()) + int(bdichkType8.get()) + int(bdichkType9.get()) + int(bdichkType10.get()) + int(bdichkType11.get()) + int(bdichkType12.get()) + int(bdichkType13.get()) + int(
                            bdichkType14.get()) + int(bdichkType15.get()) + int(bdichkType16.get()) + int(bdichkType17.get()) + int(bdichkType18.get()) + int(bdichkType19.get()) + int(bdichkType20.get()) + int(bdichkType21.get())
                        bdirltScore1 = str(bdichkType1.get())
                        bdirltScore2 = str(bdichkType2.get())
                        bdirltScore3 = str(bdichkType3.get())
                        bdirltScore4 = str(bdichkType4.get())
                        bdirltScore5 = str(bdichkType5.get())
                        bdirltScore6 = str(bdichkType6.get())
                        bdirltScore7 = str(bdichkType7.get())
                        bdirltScore8 = str(bdichkType8.get())
                        bdirltScore9 = str(bdichkType9.get())
                        bdirltScore10 = str(bdichkType10.get())
                        bdirltScore11 = str(bdichkType11.get())
                        bdirltScore12 = str(bdichkType12.get())
                        bdirltScore13 = str(bdichkType13.get())
                        bdirltScore14 = str(bdichkType14.get())
                        bdirltScore15 = str(bdichkType15.get())
                        bdirltScore16 = str(bdichkType16.get())
                        bdirltScore17 = str(bdichkType17.get())
                        bdirltScore18 = str(bdichkType18.get())
                        bdirltScore19 = str(bdichkType19.get())
                        bdirltScore20 = str(bdichkType20.get())
                        bdirltScore21 = str(bdichkType21.get())
                        sheet3['E9'] = bdirltScore1 + " 점"
                        sheet3['E15'] = bdirltScore2 + " 점"
                        sheet3['E21'] = bdirltScore3 + " 점"
                        sheet3['E27'] = bdirltScore4 + " 점"
                        sheet3['E33'] = bdirltScore5 + " 점"
                        sheet3['E39'] = bdirltScore6 + " 점"
                        sheet3['E45'] = bdirltScore7 + " 점"
                        sheet3['E51'] = bdirltScore8 + " 점"
                        sheet3['E57'] = bdirltScore9 + " 점"
                        sheet3['E63'] = bdirltScore10 + " 점"
                        sheet3['E69'] = bdirltScore11 + " 점"
                        sheet3['E75'] = bdirltScore12 + " 점"
                        sheet3['E81'] = bdirltScore13 + " 점"
                        sheet3['E87'] = bdirltScore14 + " 점"
                        sheet3['E93'] = bdirltScore15 + " 점"
                        sheet3['E99'] = bdirltScore16 + " 점"
                        sheet3['E105'] = bdirltScore17 + " 점"
                        sheet3['E111'] = bdirltScore18 + " 점"
                        sheet3['E118'] = bdirltScore19 + " 점"
                        sheet3['E124'] = bdirltScore20 + " 점"
                        sheet3['E130'] = bdirltScore21 + " 점"
                        sheet3['M3'] = sumscore3
                        sheet3['K3'] = nameInput.get()
                        min_col, min_row, max_col, max_row = range_boundaries(
                            'A1')

                        wss.insert_rows(2, 133)
                        for row in sheet00.iter_rows(1, 133, 1, 15):
                            for cell in row:
                                cell.value = wb['BDI'].cell(cell.row, cell.column).value
                                cell._style = wb['BDI'].cell(
                                    cell.row, cell.column)._style
                        altersheet['E6'].value = sheet3['M3'].value
                        for row in range(7,altersheet.max_row):
                            altersheet.row_dimensions[row].height = 21
                        # interviewsheet['B24'].value = ("% s년 % s월 % s일 % s시 % s분" % (now.year, now.month, now.day, now.hour, now.minute))
                        altersheet['B6'].value = ("% s년 % s월 % s일 % s시 % s분" % (now.year, now.month, now.day, now.hour, now.minute))
                        altersheet['L6'].value = "=SUM(INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-2),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-3),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-4),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-5),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-6),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-7),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-8),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-9),2)))) &"+'" 점"'
                        altersheet.insert_rows(6, 1)
                        altersheet['B7']._style = altersheet['AA2']._style
                        altersheet['C7']._style = altersheet['AB2']._style
                        altersheet['D7']._style = altersheet['AC2']._style
                        altersheet['E7']._style = altersheet['AD2']._style
                        altersheet['F7']._style = altersheet['AE2']._style
                        altersheet['G7']._style = altersheet['AF2']._style
                        altersheet['H7']._style = altersheet['AG2']._style
                        altersheet['I6']._style = altersheet['AH2']._style
                        altersheet['J6']._style = altersheet['AI2']._style

                        altersheet['K7']._style = altersheet['AJ2']._style
                        altersheet['L7']._style = altersheet['AK2']._style
                        altersheet['K7'].value = "=TRUNC(SUM(INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-1),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-2),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-3),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-4),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-5),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-6),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-7),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-8),2))))/COUNT(INDIRECT(ADDRESS(ROW(),COLUMN()-8)),INDIRECT(ADDRESS(ROW(),COLUMN()-7)),INDIRECT(ADDRESS(ROW(),COLUMN()-6)),INDIRECT(ADDRESS(ROW(),COLUMN()-5)),INDIRECT(ADDRESS(ROW(),COLUMN()-4)),INDIRECT(ADDRESS(ROW(),COLUMN()-3)),INDIRECT(ADDRESS(ROW(),COLUMN()-2)),INDIRECT(ADDRESS(ROW(),COLUMN()-1))),1)&"+'" 점"'
        
                        try:
                            wbb.save('//DESKTOP/ptest//xlsm//'+nameInput.get()+'['+idInput.get()+'].xlsm')
                            messagebox.showinfo('마음정원', "제출이 완료 되었습니다. 수고하셨습니다.")
                            wbb.close()
                            root.quit()
                        except:
                            messagebox.showinfo('마음정원', "메인 컴퓨터 엑셀파일이 열려있습니다.")
                min_col, min_row, max_col, max_row = range_boundaries('A1')
                for row in altersheet.iter_rows(min_row, altersheet.max_row):
                    altersheet['C'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($C$"+str(min_row+6)+':$C$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($C'+str(min_row+6)+':$C'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['D'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($D$"+str(min_row+6)+':$D$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($D'+str(min_row+6)+':$D'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['E'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($E$"+str(min_row+6)+':$E$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($E'+str(min_row+6)+':$E'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['F'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($F$"+str(min_row+6)+':$F$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($F'+str(min_row+6)+':$F'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['G'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($G$"+str(min_row+6)+':$G$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($G'+str(min_row+6)+':$G'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['H'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($H$"+str(min_row+6)+':$H$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($H'+str(min_row+6)+':$H'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['I'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($I$"+str(min_row+6)+':$I$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($I'+str(min_row+6)+':$I'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['J'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($J$"+str(min_row+6)+':$J$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($J'+str(min_row+6)+':$J'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                submit = Button(displayPage.inner, image=addsheetImg,
                                activebackground="#f0f0f0", command=btnBdi, borderwidth=0)
                submit.place(relx=0.5, anchor="center",
                             width=366, height=125)
                submit.pack(ipady=30, side=BOTTOM)
                submit.focus_set()
                selFrames.mainHide()
                now = datetime.now()
                notebook.add2(displayPage, text="% s년 % s월 % s일 % s시" %
                              (now.year, now.month, now.day, now.hour) + a)
                selFrames.mainAdd()

            if (selFrames.srtchkType4.get()) == 1:
                mainContent.snapchkFunc()

                def btnSnap():
                    if snapchk == True:
                        if snapchkType1.get() == "":
                            messagebox.showinfo('마음정원', "1번 문항을 체크해주십시오.")
                        elif snapchkType2.get() == "":
                            messagebox.showinfo('마음정원', "2번 문항을 체크해주십시오.")
                        elif snapchkType3.get() == "":
                            messagebox.showinfo('마음정원', "3번 문항을 체크해주십시오.")
                        elif snapchkType4.get() == "":
                            messagebox.showinfo('마음정원', "4번 문항을 체크해주십시오.")
                        elif snapchkType5.get() == "":
                            messagebox.showinfo('마음정원', "5번 문항을 체크해주십시오.")
                        elif snapchkType6.get() == "":
                            messagebox.showinfo('마음정원', "6번 문항을 체크해주십시오.")
                        elif snapchkType7.get() == "":
                            messagebox.showinfo('마음정원', "7번 문항을 체크해주십시오.")
                        elif snapchkType8.get() == "":
                            messagebox.showinfo('마음정원', "8번 문항을 체크해주십시오.")
                        elif snapchkType9.get() == "":
                            messagebox.showinfo('마음정원', "9번 문항을 체크해주십시오.")
                        elif snapchkType10.get() == "":
                            messagebox.showinfo('마음정원', "10번 문항을 체크해주십시오.")
                        elif snapchkType11.get() == "":
                            messagebox.showinfo('마음정원', "11번 문항을 체크해주십시오.")
                        elif snapchkType12.get() == "":
                            messagebox.showinfo('마음정원', "12번 문항을 체크해주십시오.")
                        elif snapchkType13.get() == "":
                            messagebox.showinfo('마음정원', "13번 문항을 체크해주십시오.")
                        elif snapchkType14.get() == "":
                            messagebox.showinfo('마음정원', "14번 문항을 체크해주십시오.")
                        elif snapchkType15.get() == "":
                            messagebox.showinfo('마음정원', "15번 문항을 체크해주십시오.")
                        elif snapchkType16.get() == "":
                            messagebox.showinfo('마음정원', "16번 문항을 체크해주십시오.")
                        elif snapchkType17.get() == "":
                            messagebox.showinfo('마음정원', "17번 문항을 체크해주십시오.")
                        elif snapchkType18.get() == "":
                            messagebox.showinfo('마음정원', "마지막 문항을 체크해주십시오.")
                        if snapchkType1.get() == "" or snapchkType2.get() == "" or snapchkType3.get() == "" or snapchkType4.get() == "" or snapchkType5.get() == "" or snapchkType6.get() == "" or snapchkType7.get() == "" or snapchkType8.get() == "" or snapchkType9.get() == "" or snapchkType10.get() == "" or snapchkType11.get() == "" or snapchkType12.get() == "" or snapchkType13.get() == "" or snapchkType14.get() == "" or snapchkType15.get() == "" or snapchkType16.get() == "" or snapchkType17.get() == "" or snapchkType18.get() == "":
                            return
                        sumscore4 = int(snapchkType1.get()) + int(snapchkType2.get()) + int(snapchkType3.get()) + int(snapchkType4.get()) + int(snapchkType5.get()) + int(snapchkType6.get()) + int(snapchkType7.get()) + int(snapchkType8.get()) + int(snapchkType9.get()) + int(snapchkType10.get()) + int(snapchkType11.get()) + int(snapchkType12.get()) + int(snapchkType13.get()) + int(
                            snapchkType14.get()) + int(snapchkType15.get()) + int(snapchkType16.get()) + int(snapchkType17.get()) + int(snapchkType18.get())
                        snaprltScore1 = str(snapchkType1.get())
                        snaprltScore2 = str(snapchkType2.get())
                        snaprltScore3 = str(snapchkType3.get())
                        snaprltScore4 = str(snapchkType4.get())
                        snaprltScore5 = str(snapchkType5.get())
                        snaprltScore6 = str(snapchkType6.get())
                        snaprltScore7 = str(snapchkType7.get())
                        snaprltScore8 = str(snapchkType8.get())
                        snaprltScore9 = str(snapchkType9.get())
                        snaprltScore10 = str(snapchkType10.get())
                        snaprltScore11 = str(snapchkType11.get())
                        snaprltScore12 = str(snapchkType12.get())
                        snaprltScore13 = str(snapchkType13.get())
                        snaprltScore14 = str(snapchkType14.get())
                        snaprltScore15 = str(snapchkType15.get())
                        snaprltScore16 = str(snapchkType16.get())
                        snaprltScore17 = str(snapchkType17.get())
                        snaprltScore18 = str(snapchkType18.get())
                        sheet4['E6'] = snaprltScore1 + " 점"
                        sheet4['E9'] = snaprltScore2 + " 점"
                        sheet4['E12'] = snaprltScore3 + " 점"
                        sheet4['E15'] = snaprltScore4 + " 점"
                        sheet4['E18'] = snaprltScore5 + " 점"
                        sheet4['E21'] = snaprltScore6 + " 점"
                        sheet4['E24'] = snaprltScore7 + " 점"
                        sheet4['E27'] = snaprltScore8 + " 점"
                        sheet4['E30'] = snaprltScore9 + " 점"
                        sheet4['E33'] = snaprltScore10 + " 점"
                        sheet4['E36'] = snaprltScore11 + " 점"
                        sheet4['E39'] = snaprltScore12 + " 점"
                        sheet4['E42'] = snaprltScore13 + " 점"
                        sheet4['E45'] = snaprltScore14 + " 점"
                        sheet4['E48'] = snaprltScore15 + " 점"
                        sheet4['E51'] = snaprltScore16 + " 점"
                        sheet4['E54'] = snaprltScore17 + " 점"
                        sheet4['E57'] = snaprltScore18 + " 점"
                        sheet4['M3'] = sumscore4
                        sheet4['K3'] = nameInput.get()
                        min_col, min_row, max_col, max_row = range_boundaries(
                            'A1')

                        wss.insert_rows(2, 60)
                        for row in sheet00.iter_rows(1, 60, 1, 15):
                            for cell in row:
                                cell.value = wb['SNAP'].cell(cell.row, cell.column).value
                                cell._style = wb['SNAP'].cell(
                                    cell.row, cell.column)._style
                        altersheet['F6'].value = sheet4['M3'].value
                        for row in range(7,altersheet.max_row):
                            altersheet.row_dimensions[row].height = 21
                        # interviewsheet['B24'].value = ("% s년 % s월 % s일 % s시 % s분" % (now.year, now.month, now.day, now.hour, now.minute))
                        altersheet['B6'].value = ("% s년 % s월 % s일 % s시 % s분" % (now.year, now.month, now.day, now.hour, now.minute))
                        altersheet['L6'].value = "=SUM(INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-2),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-3),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-4),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-5),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-6),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-7),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-8),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-9),2)))) &"+'" 점"'
                        altersheet.insert_rows(6, 1)
                        altersheet['B7']._style = altersheet['AA2']._style
                        altersheet['C7']._style = altersheet['AB2']._style
                        altersheet['D7']._style = altersheet['AC2']._style
                        altersheet['E7']._style = altersheet['AD2']._style
                        altersheet['F7']._style = altersheet['AE2']._style
                        altersheet['G7']._style = altersheet['AF2']._style
                        altersheet['H7']._style = altersheet['AG2']._style
                        altersheet['I6']._style = altersheet['AH2']._style
                        altersheet['J6']._style = altersheet['AI2']._style

                        altersheet['K7']._style = altersheet['AJ2']._style
                        altersheet['L7']._style = altersheet['AK2']._style
                        altersheet['K7'].value = "=TRUNC(SUM(INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-1),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-2),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-3),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-4),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-5),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-6),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-7),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-8),2))))/COUNT(INDIRECT(ADDRESS(ROW(),COLUMN()-8)),INDIRECT(ADDRESS(ROW(),COLUMN()-7)),INDIRECT(ADDRESS(ROW(),COLUMN()-6)),INDIRECT(ADDRESS(ROW(),COLUMN()-5)),INDIRECT(ADDRESS(ROW(),COLUMN()-4)),INDIRECT(ADDRESS(ROW(),COLUMN()-3)),INDIRECT(ADDRESS(ROW(),COLUMN()-2)),INDIRECT(ADDRESS(ROW(),COLUMN()-1))),1)&"+'" 점"'
                        try:
                            wbb.save('//DESKTOP/ptest//xlsm//'+nameInput.get()+'['+idInput.get()+'].xlsm')
                            messagebox.showinfo('마음정원', "제출이 완료 되었습니다. 수고하셨습니다.")
                            wbb.close()
                            root.quit()
                        except:
                            messagebox.showinfo('마음정원', "메인 컴퓨터 엑셀파일이 열려있습니다.")
                min_col, min_row, max_col, max_row = range_boundaries('A1')
                for row in altersheet.iter_rows(min_row,altersheet.max_row):
                    altersheet['C'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($C$"+str(min_row+6)+':$C$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($C'+str(min_row+6)+':$C'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['D'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($D$"+str(min_row+6)+':$D$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($D'+str(min_row+6)+':$D'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['E'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($E$"+str(min_row+6)+':$E$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($E'+str(min_row+6)+':$E'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['F'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($F$"+str(min_row+6)+':$F$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($F'+str(min_row+6)+':$F'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['G'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($G$"+str(min_row+6)+':$G$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($G'+str(min_row+6)+':$G'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['H'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($H$"+str(min_row+6)+':$H$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($H'+str(min_row+6)+':$H'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['I'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($I$"+str(min_row+6)+':$I$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($I'+str(min_row+6)+':$I'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['J'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($J$"+str(min_row+6)+':$J$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($J'+str(min_row+6)+':$J'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                submit = Button(displayPage.inner, image=addsheetImg,
                                activebackground="#f0f0f0", command=btnSnap, borderwidth=0)
                submit.place(relx=0.5, anchor="center",
                             width=366, height=125)
                submit.pack(ipady=30, side=BOTTOM)
                submit.focus_set()
                selFrames.mainHide()
                now = datetime.now()
                notebook.add2(displayPage, text="% s년 % s월 % s일 % s시" %
                              (now.year, now.month, now.day, now.hour) + a)
                selFrames.mainAdd()
            if (selFrames.srtchkType5.get()) == 1:
                mainContent.st_2chkFunc()

                def btnSt_2():
                    if st_2chk == True:
                        if st_2chkType1.get() == "":
                            messagebox.showinfo('마음정원', "1번 문항을 체크해주십시오.")
                        elif st_2chkType2.get() == "":
                            messagebox.showinfo('마음정원', "2번 문항을 체크해주십시오.")
                        elif st_2chkType3.get() == "":
                            messagebox.showinfo('마음정원', "3번 문항을 체크해주십시오.")
                        elif st_2chkType4.get() == "":
                            messagebox.showinfo('마음정원', "4번 문항을 체크해주십시오.")
                        elif st_2chkType5.get() == "":
                            messagebox.showinfo('마음정원', "5번 문항을 체크해주십시오.")
                        elif st_2chkType6.get() == "":
                            messagebox.showinfo('마음정원', "6번 문항을 체크해주십시오.")
                        elif st_2chkType7.get() == "":
                            messagebox.showinfo('마음정원', "7번 문항을 체크해주십시오.")
                        elif st_2chkType8.get() == "":
                            messagebox.showinfo('마음정원', "8번 문항을 체크해주십시오.")
                        elif st_2chkType9.get() == "":
                            messagebox.showinfo('마음정원', "9번 문항을 체크해주십시오.")
                        elif st_2chkType10.get() == "":
                            messagebox.showinfo('마음정원', "10번 문항을 체크해주십시오.")
                        elif st_2chkType11.get() == "":
                            messagebox.showinfo('마음정원', "11번 문항을 체크해주십시오.")
                        elif st_2chkType12.get() == "":
                            messagebox.showinfo('마음정원', "12번 문항을 체크해주십시오.")
                        elif st_2chkType13.get() == "":
                            messagebox.showinfo('마음정원', "13번 문항을 체크해주십시오.")
                        elif st_2chkType14.get() == "":
                            messagebox.showinfo('마음정원', "14번 문항을 체크해주십시오.")
                        elif st_2chkType15.get() == "":
                            messagebox.showinfo('마음정원', "15번 문항을 체크해주십시오.")
                        elif st_2chkType16.get() == "":
                            messagebox.showinfo('마음정원', "16번 문항을 체크해주십시오.")
                        elif st_2chkType17.get() == "":
                            messagebox.showinfo('마음정원', "17번 문항을 체크해주십시오.")
                        elif st_2chkType18.get() == "":
                            messagebox.showinfo('마음정원', "18번 문항을 체크해주십시오.")
                        elif st_2chkType19.get() == "":
                            messagebox.showinfo('마음정원', "19번 문항을 체크해주십시오.")
                        elif st_2chkType20.get() == "":
                            messagebox.showinfo('마음정원', "마지막 문항을 체크해주십시오.")
                        if st_2chkType1.get() == "" or st_2chkType2.get() == "" or st_2chkType3.get() == "" or st_2chkType4.get() == "" or st_2chkType5.get() == "" or st_2chkType6.get() == "" or st_2chkType7.get() == "" or st_2chkType8.get() == "" or st_2chkType9.get() == "" or st_2chkType10.get() == "" or st_2chkType11.get() == "" or st_2chkType12.get() == "" or st_2chkType13.get() == "" or st_2chkType14.get() == "" or st_2chkType15.get() == "" or st_2chkType16.get() == "" or st_2chkType17.get() == "" or st_2chkType18.get() == "" or st_2chkType19.get() == "" or st_2chkType20.get() == "":
                            return
                        sumscore5 = int(st_2chkType1.get()) + int(st_2chkType2.get()) + int(st_2chkType3.get()) + int(st_2chkType4.get()) + int(st_2chkType5.get()) + int(st_2chkType6.get()) + int(st_2chkType7.get()) + int(st_2chkType8.get()) + int(st_2chkType9.get()) + int(st_2chkType10.get()) + int(st_2chkType11.get()) + int(st_2chkType12.get()) + int(st_2chkType13.get()) + int(
                            st_2chkType14.get()) + int(st_2chkType15.get()) + int(st_2chkType16.get()) + int(st_2chkType17.get()) + int(st_2chkType18.get()) + int(st_2chkType19.get()) + int(st_2chkType20.get())
                        st_2rltScore1 = str(st_2chkType1.get())
                        st_2rltScore2 = str(st_2chkType2.get())
                        st_2rltScore3 = str(st_2chkType3.get())
                        st_2rltScore4 = str(st_2chkType4.get())
                        st_2rltScore5 = str(st_2chkType5.get())
                        st_2rltScore6 = str(st_2chkType6.get())
                        st_2rltScore7 = str(st_2chkType7.get())
                        st_2rltScore8 = str(st_2chkType8.get())
                        st_2rltScore9 = str(st_2chkType9.get())
                        st_2rltScore10 = str(st_2chkType10.get())
                        st_2rltScore11 = str(st_2chkType11.get())
                        st_2rltScore12 = str(st_2chkType12.get())
                        st_2rltScore13 = str(st_2chkType13.get())
                        st_2rltScore14 = str(st_2chkType14.get())
                        st_2rltScore15 = str(st_2chkType15.get())
                        st_2rltScore16 = str(st_2chkType16.get())
                        st_2rltScore17 = str(st_2chkType17.get())
                        st_2rltScore18 = str(st_2chkType18.get())
                        st_2rltScore19 = str(st_2chkType19.get())
                        st_2rltScore20 = str(st_2chkType20.get())
                        sheet5['E6'] = st_2rltScore1 + " 점"
                        sheet5['E9'] = st_2rltScore2 + " 점"
                        sheet5['E12'] = st_2rltScore3 + " 점"
                        sheet5['E15'] = st_2rltScore4 + " 점"
                        sheet5['E18'] = st_2rltScore5 + " 점"
                        sheet5['E21'] = st_2rltScore6 + " 점"
                        sheet5['E24'] = st_2rltScore7 + " 점"
                        sheet5['E27'] = st_2rltScore8 + " 점"
                        sheet5['E30'] = st_2rltScore9 + " 점"
                        sheet5['E33'] = st_2rltScore10 + " 점"
                        sheet5['E36'] = st_2rltScore11 + " 점"
                        sheet5['E39'] = st_2rltScore12 + " 점"
                        sheet5['E42'] = st_2rltScore13 + " 점"
                        sheet5['E45'] = st_2rltScore14 + " 점"
                        sheet5['E48'] = st_2rltScore15 + " 점"
                        sheet5['E51'] = st_2rltScore16 + " 점"
                        sheet5['E54'] = st_2rltScore17 + " 점"
                        sheet5['E57'] = st_2rltScore18 + " 점"
                        sheet5['E60'] = st_2rltScore19 + " 점"
                        sheet5['E63'] = st_2rltScore20 + " 점"
                        sheet5['M3'] = sumscore5
                        sheet5['K3'] = nameInput.get()
                        min_col, min_row, max_col, max_row = range_boundaries(
                            'A1')

                        wss.insert_rows(2, 66)
                        for row in sheet00.iter_rows(1, 66, 1, 15):
                            for cell in row:
                                cell.value = wb['ST_2'].cell(cell.row, cell.column).value
                                cell._style = wb['ST_2'].cell(
                                    cell.row, cell.column)._style
                        altersheet['G6'].value = sheet5['M3'].value
                        for row in range(7, altersheet.max_row):
                            altersheet.row_dimensions[row].height = 21
                        # interviewsheet['B24'].value = ("% s년 % s월 % s일 % s시 % s분" % (now.year, now.month, now.day, now.hour, now.minute))
                        altersheet['B6'].value = ("% s년 % s월 % s일 % s시 % s분" % (now.year, now.month, now.day, now.hour, now.minute))
                        altersheet['L6'].value = "=SUM(INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-2),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-3),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-4),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-5),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-6),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-7),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-8),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-9),2)))) &"+'" 점"'
                        altersheet.insert_rows(6, 1)
                        altersheet['B7']._style = altersheet['AA2']._style
                        altersheet['C7']._style = altersheet['AB2']._style
                        altersheet['D7']._style = altersheet['AC2']._style
                        altersheet['E7']._style = altersheet['AD2']._style
                        altersheet['F7']._style = altersheet['AE2']._style
                        altersheet['G7']._style = altersheet['AF2']._style
                        altersheet['H7']._style = altersheet['AG2']._style
                        altersheet['I6']._style = altersheet['AH2']._style
                        altersheet['J6']._style = altersheet['AI2']._style

                        altersheet['K7']._style = altersheet['AJ2']._style
                        altersheet['L7']._style = altersheet['AK2']._style
                        altersheet['K7'].value = "=TRUNC(SUM(INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-1),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-2),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-3),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-4),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-5),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-6),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-7),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-8),2))))/COUNT(INDIRECT(ADDRESS(ROW(),COLUMN()-8)),INDIRECT(ADDRESS(ROW(),COLUMN()-7)),INDIRECT(ADDRESS(ROW(),COLUMN()-6)),INDIRECT(ADDRESS(ROW(),COLUMN()-5)),INDIRECT(ADDRESS(ROW(),COLUMN()-4)),INDIRECT(ADDRESS(ROW(),COLUMN()-3)),INDIRECT(ADDRESS(ROW(),COLUMN()-2)),INDIRECT(ADDRESS(ROW(),COLUMN()-1))),1)&"+'" 점"'
                        try:
                            wbb.save('//DESKTOP/ptest//xlsm//'+nameInput.get()+'['+idInput.get()+'].xlsm')
                            messagebox.showinfo('마음정원', "제출이 완료 되었습니다. 수고하셨습니다.")
                            wbb.close()
                            root.quit()
                        except:
                            messagebox.showinfo('마음정원', "메인 컴퓨터 엑셀파일이 열려있습니다.")
                min_col, min_row, max_col, max_row = range_boundaries('A1')
                for row in altersheet.iter_rows(min_row,altersheet.max_row):
                    altersheet['C'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($C$"+str(min_row+6)+':$C$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($C'+str(min_row+6)+':$C'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['D'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($D$"+str(min_row+6)+':$D$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($D'+str(min_row+6)+':$D'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['E'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($E$"+str(min_row+6)+':$E$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($E'+str(min_row+6)+':$E'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['F'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($F$"+str(min_row+6)+':$F$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($F'+str(min_row+6)+':$F'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['G'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($G$"+str(min_row+6)+':$G$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($G'+str(min_row+6)+':$G'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['H'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($H$"+str(min_row+6)+':$H$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($H'+str(min_row+6)+':$H'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['I'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($I$"+str(min_row+6)+':$I$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($I'+str(min_row+6)+':$I'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['J'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($J$"+str(min_row+6)+':$J$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($J'+str(min_row+6)+':$J'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                submit = Button(displayPage.inner, image=addsheetImg,
                                activebackground="#f0f0f0", command=btnSt_2, borderwidth=0)
                submit.place(relx=0.5, anchor="center",
                             width=366, height=125)
                submit.pack(ipady=30, side=BOTTOM)
                submit.focus_set()
                selFrames.mainHide()
                now = datetime.now()
                notebook.add2(displayPage, text="% s년 % s월 % s일 % s시" %
                              (now.year, now.month, now.day, now.hour) + a)
                selFrames.mainAdd()
            if (selFrames.srtchkType6.get()) == 1:
                mainContent.st_1chkFunc()

                def btnSt_1():
                    if st_1chk == True:
                        if st_1chkType1.get() == "":
                            messagebox.showinfo('마음정원', "1번 문항을 체크해주십시오.")
                        elif st_1chkType2.get() == "":
                            messagebox.showinfo('마음정원', "2번 문항을 체크해주십시오.")
                        elif st_1chkType3.get() == "":
                            messagebox.showinfo('마음정원', "3번 문항을 체크해주십시오.")
                        elif st_1chkType4.get() == "":
                            messagebox.showinfo('마음정원', "4번 문항을 체크해주십시오.")
                        elif st_1chkType5.get() == "":
                            messagebox.showinfo('마음정원', "5번 문항을 체크해주십시오.")
                        elif st_1chkType6.get() == "":
                            messagebox.showinfo('마음정원', "6번 문항을 체크해주십시오.")
                        elif st_1chkType7.get() == "":
                            messagebox.showinfo('마음정원', "7번 문항을 체크해주십시오.")
                        elif st_1chkType8.get() == "":
                            messagebox.showinfo('마음정원', "8번 문항을 체크해주십시오.")
                        elif st_1chkType9.get() == "":
                            messagebox.showinfo('마음정원', "9번 문항을 체크해주십시오.")
                        elif st_1chkType10.get() == "":
                            messagebox.showinfo('마음정원', "10번 문항을 체크해주십시오.")
                        elif st_1chkType11.get() == "":
                            messagebox.showinfo('마음정원', "11번 문항을 체크해주십시오.")
                        elif st_1chkType12.get() == "":
                            messagebox.showinfo('마음정원', "12번 문항을 체크해주십시오.")
                        elif st_1chkType13.get() == "":
                            messagebox.showinfo('마음정원', "13번 문항을 체크해주십시오.")
                        elif st_1chkType14.get() == "":
                            messagebox.showinfo('마음정원', "14번 문항을 체크해주십시오.")
                        elif st_1chkType15.get() == "":
                            messagebox.showinfo('마음정원', "15번 문항을 체크해주십시오.")
                        elif st_1chkType16.get() == "":
                            messagebox.showinfo('마음정원', "16번 문항을 체크해주십시오.")
                        elif st_1chkType17.get() == "":
                            messagebox.showinfo('마음정원', "17번 문항을 체크해주십시오.")
                        elif st_1chkType18.get() == "":
                            messagebox.showinfo('마음정원', "18번 문항을 체크해주십시오.")
                        elif st_1chkType19.get() == "":
                            messagebox.showinfo('마음정원', "19번 문항을 체크해주십시오.")
                        elif st_1chkType20.get() == "":
                            messagebox.showinfo('마음정원', "마지막 문항을 체크해주십시오.")
                        if st_1chkType1.get() == "" or st_1chkType2.get() == "" or st_1chkType3.get() == "" or st_1chkType4.get() == "" or st_1chkType5.get() == "" or st_1chkType6.get() == "" or st_1chkType7.get() == "" or st_1chkType8.get() == "" or st_1chkType9.get() == "" or st_1chkType10.get() == "" or st_1chkType11.get() == "" or st_1chkType12.get() == "" or st_1chkType13.get() == "" or st_1chkType14.get() == "" or st_1chkType15.get() == "" or st_1chkType16.get() == "" or st_1chkType17.get() == "" or st_1chkType18.get() == "" or st_1chkType19.get() == "" or st_1chkType20.get() == "":
                            return
                        sumscore6 = int(st_1chkType1.get()) + int(st_1chkType2.get()) + int(st_1chkType3.get()) + int(st_1chkType4.get()) + int(st_1chkType5.get()) + int(st_1chkType6.get()) + int(st_1chkType7.get()) + int(st_1chkType8.get()) + int(st_1chkType9.get()) + int(st_1chkType10.get()) + int(st_1chkType11.get()) + int(st_1chkType12.get()) + int(st_1chkType13.get()) + int(
                            st_1chkType14.get()) + int(st_1chkType15.get()) + int(st_1chkType16.get()) + int(st_1chkType17.get()) + int(st_1chkType18.get()) + int(st_1chkType19.get()) + int(st_1chkType20.get())
                        st_1rltScore1 = str(st_1chkType1.get())
                        st_1rltScore2 = str(st_1chkType2.get())
                        st_1rltScore3 = str(st_1chkType3.get())
                        st_1rltScore4 = str(st_1chkType4.get())
                        st_1rltScore5 = str(st_1chkType5.get())
                        st_1rltScore6 = str(st_1chkType6.get())
                        st_1rltScore7 = str(st_1chkType7.get())
                        st_1rltScore8 = str(st_1chkType8.get())
                        st_1rltScore9 = str(st_1chkType9.get())
                        st_1rltScore10 = str(st_1chkType10.get())
                        st_1rltScore11 = str(st_1chkType11.get())
                        st_1rltScore12 = str(st_1chkType12.get())
                        st_1rltScore13 = str(st_1chkType13.get())
                        st_1rltScore14 = str(st_1chkType14.get())
                        st_1rltScore15 = str(st_1chkType15.get())
                        st_1rltScore16 = str(st_1chkType16.get())
                        st_1rltScore17 = str(st_1chkType17.get())
                        st_1rltScore18 = str(st_1chkType18.get())
                        st_1rltScore19 = str(st_1chkType19.get())
                        st_1rltScore20 = str(st_1chkType20.get())
                        sheet6['E6'] = st_1rltScore1 + " 점"
                        sheet6['E9'] = st_1rltScore2 + " 점"
                        sheet6['E12'] = st_1rltScore3 + " 점"
                        sheet6['E15'] = st_1rltScore4 + " 점"
                        sheet6['E18'] = st_1rltScore5 + " 점"
                        sheet6['E21'] = st_1rltScore6 + " 점"
                        sheet6['E24'] = st_1rltScore7 + " 점"
                        sheet6['E27'] = st_1rltScore8 + " 점"
                        sheet6['E30'] = st_1rltScore9 + " 점"
                        sheet6['E33'] = st_1rltScore10 + " 점"
                        sheet6['E36'] = st_1rltScore11 + " 점"
                        sheet6['E39'] = st_1rltScore12 + " 점"
                        sheet6['E42'] = st_1rltScore13 + " 점"
                        sheet6['E45'] = st_1rltScore14 + " 점"
                        sheet6['E48'] = st_1rltScore15 + " 점"
                        sheet6['E51'] = st_1rltScore16 + " 점"
                        sheet6['E54'] = st_1rltScore17 + " 점"
                        sheet6['E57'] = st_1rltScore18 + " 점"
                        sheet6['E60'] = st_1rltScore19 + " 점"
                        sheet6['E63'] = st_1rltScore20 + " 점"
                        sheet6['M3'] = sumscore6
                        sheet6['K3'] = nameInput.get()
                        min_col, min_row, max_col, max_row = range_boundaries(
                            'A1')

                        wss.insert_rows(2, 66)
                        for row in sheet00.iter_rows(1, 66, 1, 15):
                            for cell in row:
                                cell.value = wb['ST_1'].cell(cell.row, cell.column).value
                                cell._style = wb['ST_1'].cell(
                                    cell.row, cell.column)._style
                        altersheet['H6'].value = sheet6['M3'].value
                        for row in range(7,altersheet.max_row):
                            altersheet.row_dimensions[row].height = 21
                        # interviewsheet['B24'].value = ("% s년 % s월 % s일 % s시 % s분" % (now.year, now.month, now.day, now.hour, now.minute))
                        altersheet['B6'].value = ("% s년 % s월 % s일 % s시 % s분" % (now.year, now.month, now.day, now.hour, now.minute))
                        altersheet['L6'].value = "=SUM(INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-2),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-3),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-4),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-5),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-6),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-7),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-8),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-9),2)))) &"+'" 점"'
                        altersheet.insert_rows(6, 1)
                        altersheet['B7']._style = altersheet['AA2']._style
                        altersheet['C7']._style = altersheet['AB2']._style
                        altersheet['D7']._style = altersheet['AC2']._style
                        altersheet['E7']._style = altersheet['AD2']._style
                        altersheet['F7']._style = altersheet['AE2']._style
                        altersheet['G7']._style = altersheet['AF2']._style
                        altersheet['H7']._style = altersheet['AG2']._style
                        altersheet['I6']._style = altersheet['AH2']._style
                        altersheet['J6']._style = altersheet['AI2']._style

                        altersheet['K7']._style = altersheet['AJ2']._style
                        altersheet['L7']._style = altersheet['AK2']._style
                        altersheet['K7'].value = "=TRUNC(SUM(INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-1),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-2),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-3),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-4),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-5),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-6),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-7),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-8),2))))/COUNT(INDIRECT(ADDRESS(ROW(),COLUMN()-8)),INDIRECT(ADDRESS(ROW(),COLUMN()-7)),INDIRECT(ADDRESS(ROW(),COLUMN()-6)),INDIRECT(ADDRESS(ROW(),COLUMN()-5)),INDIRECT(ADDRESS(ROW(),COLUMN()-4)),INDIRECT(ADDRESS(ROW(),COLUMN()-3)),INDIRECT(ADDRESS(ROW(),COLUMN()-2)),INDIRECT(ADDRESS(ROW(),COLUMN()-1))),1)&"+'" 점"'
                        try:
                            wbb.save('//DESKTOP/ptest//xlsm//'+nameInput.get()+'['+idInput.get()+'].xlsm')
                            messagebox.showinfo('마음정원', "제출이 완료 되었습니다. 수고하셨습니다.")
                            wbb.close()
                            root.quit()
                        except:
                            messagebox.showinfo('마음정원', "메인 컴퓨터 엑셀파일이 열려있습니다.")
                min_col, min_row, max_col, max_row = range_boundaries('A1')
                for row in altersheet.iter_rows(min_row, altersheet.max_row):
                    altersheet['C'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($C$"+str(min_row+6)+':$C$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($C'+str(min_row+6)+':$C'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['D'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($D$"+str(min_row+6)+':$D$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($D'+str(min_row+6)+':$D'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['E'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($E$"+str(min_row+6)+':$E$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($E'+str(min_row+6)+':$E'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['F'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($F$"+str(min_row+6)+':$F$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($F'+str(min_row+6)+':$F'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['G'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($G$"+str(min_row+6)+':$G$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($G'+str(min_row+6)+':$G'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['H'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($H$"+str(min_row+6)+':$H$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($H'+str(min_row+6)+':$H'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['I'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($I$"+str(min_row+6)+':$I$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($I'+str(min_row+6)+':$I'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['J'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($J$"+str(min_row+6)+':$J$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($J'+str(min_row+6)+':$J'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                submit = Button(displayPage.inner, image=addsheetImg,
                                activebackground="#f0f0f0", command=btnSt_1, borderwidth=0)
                submit.place(relx=0.5, anchor="center",
                             width=366, height=125)
                submit.pack(ipady=30, side=BOTTOM)
                submit.focus_set()
                selFrames.mainHide()
                now = datetime.now()
                notebook.add2(displayPage, text="% s년 % s월 % s일 % s시" %
                              (now.year, now.month, now.day, now.hour) + a)
                selFrames.mainAdd()

            if (selFrames.srtchkType7.get()) == 1:
                mainContent.hamachkFunc()
                addedVar = altersheet['AA4'].value - 1
                altersheet['AA4'].value = addedVar
                def btnHama():
                    if hamachk == True:
                        if hamachkType1.get() == "":
                            messagebox.showinfo('마음정원', "1번 문항을 체크해주십시오.")
                        elif hamachkType2.get() == "":
                            messagebox.showinfo('마음정원', "2번 문항을 체크해주십시오.")
                        elif hamachkType3.get() == "":
                            messagebox.showinfo('마음정원', "3번 문항을 체크해주십시오.")
                        elif hamachkType4.get() == "":
                            messagebox.showinfo('마음정원', "4번 문항을 체크해주십시오.")
                        elif hamachkType5.get() == "":
                            messagebox.showinfo('마음정원', "5번 문항을 체크해주십시오.")
                        elif hamachkType6.get() == "":
                            messagebox.showinfo('마음정원', "6번 문항을 체크해주십시오.")
                        elif hamachkType7.get() == "":
                            messagebox.showinfo('마음정원', "7번 문항을 체크해주십시오.")
                        elif hamachkType8.get() == "":
                            messagebox.showinfo('마음정원', "8번 문항을 체크해주십시오.")
                        elif hamachkType9.get() == "":
                            messagebox.showinfo('마음정원', "9번 문항을 체크해주십시오.")
                        elif hamachkType10.get() == "":
                            messagebox.showinfo('마음정원', "10번 문항을 체크해주십시오.")
                        elif hamachkType11.get() == "":
                            messagebox.showinfo('마음정원', "11번 문항을 체크해주십시오.")
                        elif hamachkType12.get() == "":
                            messagebox.showinfo('마음정원', "12번 문항을 체크해주십시오.")
                        elif hamachkType13.get() == "":
                            messagebox.showinfo('마음정원', "마지막 문항을 체크해주십시오.")
                        if hamachkType1.get() == "" or hamachkType2.get() == "" or hamachkType3.get() == "" or hamachkType4.get() == "" or hamachkType5.get() == "" or hamachkType6.get() == "" or hamachkType7.get() == "" or hamachkType8.get() == "" or hamachkType9.get() == "" or hamachkType10.get() == "" or hamachkType11.get() == "" or hamachkType12.get() == "" or hamachkType13.get() == "":
                            return
                        
                        sumscore7 = int(hamachkType1.get()) + int(hamachkType2.get()) + int(hamachkType3.get()) + int(hamachkType4.get()) + int(hamachkType5.get()) + int(hamachkType6.get()) + int(hamachkType7.get()) + int(hamachkType8.get()) + int(hamachkType9.get()) + int(hamachkType10.get()) + int(hamachkType11.get()) + int(hamachkType12.get()) + int(hamachkType13.get())
                        
                        hamarltScore1 = str(hamachkType1.get())
                        hamarltScore2 = str(hamachkType2.get())
                        hamarltScore3 = str(hamachkType3.get())
                        hamarltScore4 = str(hamachkType4.get())
                        hamarltScore5 = str(hamachkType5.get())
                        hamarltScore6 = str(hamachkType6.get())
                        hamarltScore7 = str(hamachkType7.get())
                        hamarltScore8 = str(hamachkType8.get())
                        hamarltScore9 = str(hamachkType9.get())
                        hamarltScore10 = str(hamachkType10.get())
                        hamarltScore11 = str(hamachkType11.get())
                        hamarltScore12 = str(hamachkType12.get())
                        hamarltScore13 = str(hamachkType13.get())
                        sheet7['E6'] = hamarltScore1 + " 점"
                        sheet7['E9'] = hamarltScore2 + " 점"
                        sheet7['E12'] = hamarltScore3 + " 점"
                        sheet7['E15'] = hamarltScore4 + " 점"
                        sheet7['E18'] = hamarltScore5 + " 점"
                        sheet7['E21'] = hamarltScore6 + " 점"
                        sheet7['E24'] = hamarltScore7 + " 점"
                        sheet7['E27'] = hamarltScore8 + " 점"
                        sheet7['E30'] = hamarltScore9 + " 점"
                        sheet7['E33'] = hamarltScore10 + " 점"
                        sheet7['E36'] = hamarltScore11 + " 점"
                        sheet7['E39'] = hamarltScore12 + " 점"
                        sheet7['E42'] = hamarltScore13 + " 점"
                        sheet7['E47'] = "="+"OFFSET(인터뷰!$D$24,COUNT(인터뷰!D:D)" + str(addedVar)+",)"
                        sheet7['C47'] = "="+"OFFSET(인터뷰!$D$24,COUNT(인터뷰!D:D)" + str(addedVar)+',)&"."'            
                        sheet7['N3'] = ("="+"IFERROR(OFFSET(인터뷰!$D$24,COUNT(인터뷰!D:D)"+str(
                            addedVar)+",)+"+str(sumscore7)+',('+str(sumscore7)+'))')
                        sheet7['K3'] = nameInput.get()
                        min_col, min_row, max_col, max_row = range_boundaries(
                            'A1')

                        wss.insert_rows(2, 50)
                        for row in sheet00.iter_rows(1, 50, 1, 15):
                            for cell in row:
                                cell.value = wb['HAMA'].cell(cell.row, cell.column).value
                                cell._style = wb['HAMA'].cell(
                                    cell.row, cell.column)._style
                        for row in range(7,altersheet.max_row):
                            altersheet.row_dimensions[row].height = 21
                        interviewsheet.insert_rows(23, 1)
                        interviewsheet['D24'].value = 0
                        interviewsheet['E24'].value = 0
                        interviewsheet['F24'].value = 0
                        interviewsheet['G24'].value = 0
                        interviewsheet['H24'].value = 0
                        interviewsheet['I24'].value = 0
                        interviewsheet['B24'].value = ("% s년 % s월 % s일 % s시 % s분" % (now.year, now.month, now.day, now.hour, now.minute))
                        altersheet['B6'].value = ("% s년 % s월 % s일 % s시 % s분" % (now.year, now.month, now.day, now.hour, now.minute))
                        altersheet['L6'].value = "=SUM(INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-2),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-3),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-4),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-5),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-6),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-7),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-8),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-9),2)))) &"+'" 점"'
                        altersheet['I6'].value = ("="+"IFERROR(OFFSET(인터뷰!$D$24,COUNT(인터뷰!D:D)"+str(
                            addedVar)+",)+"+str((sumscore7))+',('+str(sumscore7)+'))')
                        
                        altersheet.insert_rows(6, 1)
                        altersheet['B7']._style = altersheet['AA2']._style
                        altersheet['C7']._style = altersheet['AB2']._style
                        altersheet['D7']._style = altersheet['AC2']._style
                        altersheet['E7']._style = altersheet['AD2']._style
                        altersheet['F7']._style = altersheet['AE2']._style
                        altersheet['G7']._style = altersheet['AF2']._style
                        altersheet['H7']._style = altersheet['AG2']._style
                        altersheet['I6']._style = altersheet['AH2']._style
                        altersheet['J6']._style = altersheet['AI2']._style

                        interviewsheet['B24']._style = interviewsheet['AA2']._style
                        interviewsheet['C24']._style = interviewsheet['AB2']._style
                        interviewsheet['D24']._style = interviewsheet['AC2']._style
                        interviewsheet['E24']._style = interviewsheet['AD2']._style
                        interviewsheet['F24']._style = interviewsheet['AE2']._style
                        interviewsheet['G24']._style = interviewsheet['AF2']._style
                        interviewsheet['H24']._style = interviewsheet['AG2']._style
                        interviewsheet['I24']._style = interviewsheet['AH2']._style
                        interviewsheet['J24']._style = interviewsheet['AI2']._style


                        altersheet['K7']._style = altersheet['AJ2']._style
                        altersheet['L7']._style = altersheet['AK2']._style
                        altersheet['K7'].value = "=TRUNC(SUM(INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-1),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-2),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-3),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-4),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-5),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-6),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-7),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-8),2))))/COUNT(INDIRECT(ADDRESS(ROW(),COLUMN()-8)),INDIRECT(ADDRESS(ROW(),COLUMN()-7)),INDIRECT(ADDRESS(ROW(),COLUMN()-6)),INDIRECT(ADDRESS(ROW(),COLUMN()-5)),INDIRECT(ADDRESS(ROW(),COLUMN()-4)),INDIRECT(ADDRESS(ROW(),COLUMN()-3)),INDIRECT(ADDRESS(ROW(),COLUMN()-2)),INDIRECT(ADDRESS(ROW(),COLUMN()-1))),1)&"+'" 점"'

                        try:
                            wbb.save('//DESKTOP/ptest//xlsm//'+nameInput.get()+'['+idInput.get()+'].xlsm')
                            messagebox.showinfo('마음정원', "제출이 완료 되었습니다. 수고하셨습니다.")
                            wbb.close()
                            root.quit()
                        except:
                            messagebox.showinfo('마음정원', "메인 컴퓨터 엑셀파일이 열려있습니다.")
                min_col, min_row, max_col, max_row = range_boundaries('A1')
                for row in altersheet.iter_rows(min_row, altersheet.max_row):
                    altersheet['C'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($C$"+str(min_row+6)+':$C$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($C'+str(min_row+6)+':$C'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['D'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($D$"+str(min_row+6)+':$D$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($D'+str(min_row+6)+':$D'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['E'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($E$"+str(min_row+6)+':$E$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($E'+str(min_row+6)+':$E'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['F'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($F$"+str(min_row+6)+':$F$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($F'+str(min_row+6)+':$F'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['G'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($G$"+str(min_row+6)+':$G$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($G'+str(min_row+6)+':$G'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['H'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($H$"+str(min_row+6)+':$H$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($H'+str(min_row+6)+':$H'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['I'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($I$"+str(min_row+6)+':$I$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($I'+str(min_row+6)+':$I'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['J'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($J$"+str(min_row+6)+':$J$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($J'+str(min_row+6)+':$J'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                submit = Button(displayPage.inner, image=addsheetImg,
                                activebackground="#f0f0f0", command=btnHama, borderwidth=0)
                submit.place(relx=0.5, anchor="center", width=366, height=125)
                submit.pack(ipady=30, side=BOTTOM)
                submit.focus_set()
                selFrames.mainHide()
                now = datetime.now()
                notebook.add2(displayPage, text="% s년 % s월 % s일 % s시" %
                              (now.year, now.month, now.day, now.hour) + a)
                selFrames.mainAdd()

            if (selFrames.srtchkType8.get()) == 1:
                mainContent.hamdchkFunc()
                addedVar = altersheet['AA4'].value - 1
                altersheet['AA4'].value = addedVar
                def btnHamd():
                    if hamdchk == True:
                        if hamdchkType1.get() == "":
                            messagebox.showinfo('마음정원', "1번 문항을 체크해주십시오.")
                        elif hamdchkType2.get() == "":
                            messagebox.showinfo('마음정원', "2번 문항을 체크해주십시오.")
                        elif hamdchkType3.get() == "":
                            messagebox.showinfo('마음정원', "3번 문항을 체크해주십시오.")
                        elif hamdchkType4.get() == "":
                            messagebox.showinfo('마음정원', "4번 문항을 체크해주십시오.")
                        elif hamdchkType5.get() == "":
                            messagebox.showinfo('마음정원', "5번 문항을 체크해주십시오.")
                        elif hamdchkType6.get() == "":
                            messagebox.showinfo('마음정원', "6번 문항을 체크해주십시오.")
                        elif hamdchkType7.get() == "":
                            messagebox.showinfo('마음정원', "7번 문항을 체크해주십시오.")
                        elif hamdchkType8.get() == "":
                            messagebox.showinfo('마음정원', "8번 문항을 체크해주십시오.")
                        elif hamdchkType9.get() == "":
                            messagebox.showinfo('마음정원', "9번 문항을 체크해주십시오.")
                        elif hamdchkType10.get() == "":
                            messagebox.showinfo('마음정원', "10번 문항을 체크해주십시오.")
                        elif hamdchkType11.get() == "":
                            messagebox.showinfo('마음정원', "11번 문항을 체크해주십시오.")
                        elif hamdchkType12.get() == "":
                            messagebox.showinfo('마음정원', "'마지막 문항을 체크해주십시오.")
                        if hamdchkType1.get() == "" or hamdchkType2.get() == "" or hamdchkType3.get() == "" or hamdchkType4.get() == "" or hamdchkType5.get() == "" or hamdchkType6.get() == "" or hamdchkType7.get() == "" or hamdchkType8.get() == "" or hamdchkType9.get() == "" or hamdchkType10.get() == "" or hamdchkType11.get() == "" or hamdchkType12.get() == "":
                            return
                        
                        sumscore8 = int(hamdchkType1.get()) + int(hamdchkType2.get()) + int(hamdchkType3.get()) + int(hamdchkType4.get()) + int(hamdchkType5.get()) + int(hamdchkType6.get(
                        )) + int(hamdchkType7.get()) + int(hamdchkType8.get()) + int(hamdchkType9.get()) + int(hamdchkType10.get()) + int(hamdchkType11.get()) + int(hamdchkType12.get())

                        hamdrltScore1 = str(hamdchkType1.get())
                        hamdrltScore2 = str(hamdchkType2.get())
                        hamdrltScore3 = str(hamdchkType3.get())
                        hamdrltScore4 = str(hamdchkType4.get())
                        hamdrltScore5 = str(hamdchkType5.get())
                        hamdrltScore6 = str(hamdchkType6.get())
                        hamdrltScore7 = str(hamdchkType7.get())
                        hamdrltScore8 = str(hamdchkType8.get())
                        hamdrltScore9 = str(hamdchkType9.get())
                        hamdrltScore10 = str(hamdchkType10.get())
                        hamdrltScore11 = str(hamdchkType11.get())
                        hamdrltScore12 = str(hamdchkType12.get())
                        sheet8['E11'] = hamdrltScore1 + " 점"
                        sheet8['E19'] = hamdrltScore2 + " 점"
                        sheet8['E25'] = hamdrltScore3 + " 점"
                        sheet8['E31'] = hamdrltScore4 + " 점"
                        sheet8['E37'] = hamdrltScore5 + " 점"
                        sheet8['E45'] = hamdrltScore6 + " 점"
                        sheet8['E53'] = hamdrltScore7 + " 점"
                        sheet8['E59'] = hamdrltScore8 + " 점"
                        sheet8['E65'] = hamdrltScore9 + " 점"
                        sheet8['E71'] = hamdrltScore10 + " 점"
                        sheet8['E79'] = hamdrltScore11 + " 점"
                        sheet8['E85'] = hamdrltScore12 + " 점"
                        sheet8['E95'] = "="+"OFFSET(인터뷰!$E$24,COUNT(인터뷰!E:E)" + str(addedVar)+",)"
                        sheet8['E103'] = "="+"OFFSET(인터뷰!$F$24,COUNT(인터뷰!F:F)" + str(addedVar)+",)"
                        sheet8['E111'] = "="+"OFFSET(인터뷰!$G$24,COUNT(인터뷰!G:G)" + str(addedVar)+",)"
                        sheet8['E119'] = "="+"OFFSET(인터뷰!$H$24,COUNT(인터뷰!H:H)" + str(addedVar)+",)"
                        sheet8['E125'] = "="+"OFFSET(인터뷰!$I$24,COUNT(인터뷰!I:I)" + str(addedVar)+",)"
                        sheet8['C95'] = "="+"OFFSET(인터뷰!$E$24,COUNT(인터뷰!E:E)" + str(addedVar)+',)&"."'
                        sheet8['C103'] = "="+"OFFSET(인터뷰!$F$24,COUNT(인터뷰!F:F)" + str(addedVar)+',)&"."'
                        sheet8['C111'] = "="+"OFFSET(인터뷰!$G$24,COUNT(인터뷰!G:G)" + str(addedVar)+',)&"."'
                        sheet8['C119'] = "="+"OFFSET(인터뷰!$H$24,COUNT(인터뷰!H:H)" + str(addedVar)+',)&"."'
                        sheet8['C125'] = "="+"OFFSET(인터뷰!$I$24,COUNT(인터뷰!I:I)" + str(addedVar)+',)&"."'            
                        sheet8['N3'] = ("="+"IFERROR((OFFSET(인터뷰!$E$24,COUNT(인터뷰!E:E)"+str(addedVar)+",))+(OFFSET(인터뷰!$F$24,COUNT(인터뷰!F:F)"+str(addedVar)+",))+(OFFSET(인터뷰!$G$24,COUNT(인터뷰!G:G)"+str(addedVar)+",))+(OFFSET(인터뷰!$H$24,COUNT(인터뷰!H:H)"+str(addedVar)+",))+(OFFSET(인터뷰!$I$24,COUNT(인터뷰!I:I)"+str(addedVar)+",))+"+str(sumscore8)+",("+str(sumscore8)+"))")
                        sheet8['K3'] = nameInput.get()
                        min_col, min_row, max_col, max_row = range_boundaries(
                            'A1')

                        wss.insert_rows(2, 128)

                        for row in sheet00.iter_rows(1, 128, 1, 15):
                            for cell in row:
                                cell.value = wb['HAMD'].cell(cell.row, cell.column).value
                                cell._style = wb['HAMD'].cell(
                                    cell.row, cell.column)._style

                        for row in range(7, altersheet.max_row):
                            altersheet.row_dimensions[row].height = 21
                        interviewsheet.insert_rows(23, 1)
                        interviewsheet['D24'].value = 0
                        interviewsheet['E24'].value = 0
                        interviewsheet['F24'].value = 0
                        interviewsheet['G24'].value = 0
                        interviewsheet['H24'].value = 0
                        interviewsheet['I24'].value = 0
                        interviewsheet['B24'].value = ("% s년 % s월 % s일 % s시 % s분" % (now.year, now.month, now.day, now.hour, now.minute))
                        altersheet['B6'].value = ("% s년 % s월 % s일 % s시 % s분" % (
                            now.year, now.month, now.day, now.hour, now.minute))
                        altersheet['L6'].value = "=SUM(INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-2),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-3),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-4),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-5),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-6),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-7),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-8),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-9),2)))) &"+'" 점"'
                        altersheet['J6'].value = ("="+"=IFERROR((OFFSET(인터뷰!$E$24,COUNT(인터뷰!E:E)"+str(addedVar)+",))+(OFFSET(인터뷰!$F$24,COUNT(인터뷰!F:F)"+str(addedVar)+",))+(OFFSET(인터뷰!$G$24,COUNT(인터뷰!G:G)"+str(addedVar)+",))+(OFFSET(인터뷰!$H$24,COUNT(인터뷰!H:H)"+str(addedVar)+",))+(OFFSET(인터뷰!$I$24,COUNT(인터뷰!I:I)"+str(addedVar)+",))+"+str(sumscore8)+",("+str(sumscore8)+"))")
                        altersheet.insert_rows(6, 1)
                        altersheet['B7']._style = altersheet['AA2']._style
                        altersheet['C7']._style = altersheet['AB2']._style
                        altersheet['D7']._style = altersheet['AC2']._style
                        altersheet['E7']._style = altersheet['AD2']._style
                        altersheet['F7']._style = altersheet['AE2']._style
                        altersheet['G7']._style = altersheet['AF2']._style
                        altersheet['H7']._style = altersheet['AG2']._style
                        altersheet['I6']._style = altersheet['AH2']._style
                        altersheet['J6']._style = altersheet['AI2']._style

                        interviewsheet['B24']._style = interviewsheet['AA2']._style
                        interviewsheet['C24']._style = interviewsheet['AB2']._style
                        interviewsheet['D24']._style = interviewsheet['AC2']._style
                        interviewsheet['E24']._style = interviewsheet['AD2']._style
                        interviewsheet['F24']._style = interviewsheet['AE2']._style
                        interviewsheet['G24']._style = interviewsheet['AF2']._style
                        interviewsheet['H24']._style = interviewsheet['AG2']._style
                        interviewsheet['I24']._style = interviewsheet['AH2']._style
                        interviewsheet['J24']._style = interviewsheet['AI2']._style


                        altersheet['K7']._style = altersheet['AJ2']._style
                        altersheet['L7']._style = altersheet['AK2']._style
                        altersheet['K7'].value = "=TRUNC(SUM(INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-1),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-2),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-3),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-4),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-5),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-6),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-7),2))),INDIRECT(ADDRESS(ROW(),LEFT((COLUMN()-8),2))))/COUNT(INDIRECT(ADDRESS(ROW(),COLUMN()-8)),INDIRECT(ADDRESS(ROW(),COLUMN()-7)),INDIRECT(ADDRESS(ROW(),COLUMN()-6)),INDIRECT(ADDRESS(ROW(),COLUMN()-5)),INDIRECT(ADDRESS(ROW(),COLUMN()-4)),INDIRECT(ADDRESS(ROW(),COLUMN()-3)),INDIRECT(ADDRESS(ROW(),COLUMN()-2)),INDIRECT(ADDRESS(ROW(),COLUMN()-1))),1)&"+'" 점"'


                        try:
                            wbb.save('//DESKTOP/ptest//xlsm//'+nameInput.get()+'['+idInput.get()+'].xlsm')
                            messagebox.showinfo('마음정원', "제출이 완료 되었습니다. 수고하셨습니다.")
                            wbb.close()
                            root.quit()
                        except:
                            messagebox.showinfo('마음정원', "메인 컴퓨터 엑셀파일이 열려있습니다.")
                min_col, min_row, max_col, max_row = range_boundaries('A1')
                for row in altersheet.iter_rows(min_row,altersheet.max_row):
                    altersheet['C'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($C$"+str(min_row+6)+':$C$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($C'+str(min_row+6)+':$C'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['D'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($D$"+str(min_row+6)+':$D$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($D'+str(min_row+6)+':$D'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['E'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($E$"+str(min_row+6)+':$E$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($E'+str(min_row+6)+':$E'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['F'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($F$"+str(min_row+6)+':$F$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($F'+str(min_row+6)+':$F'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['G'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($G$"+str(min_row+6)+':$G$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($G'+str(min_row+6)+':$G'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['H'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($H$"+str(min_row+6)+':$H$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($H'+str(min_row+6)+':$H'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['I'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($I$"+str(min_row+6)+':$I$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($I'+str(min_row+6)+':$I'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                    altersheet['J'+str(altersheet.max_row)].value = "="+"IFERROR(TRUNC(SUMPRODUCT(--(0&SUBSTITUTE($J$"+str(min_row+6)+':$J$'+str(
                    	altersheet.max_row)+'," 점","")))/COUNTA($J'+str(min_row+6)+':$J'+str(altersheet.max_row)+'),1)&' + '"' + ' 점' + '"'+',"")'
                submit = Button(displayPage.inner, image=addsheetImg,
                                activebackground="#f0f0f0", command=btnHamd, borderwidth=0)
                submit.place(relx=0.5, anchor="center", width=366, height=125)
                submit.pack(ipady=30, side=BOTTOM)
                submit.focus_set()
                selFrames.mainHide()
                now = datetime.now()
                notebook.add2(displayPage, text="% s년 % s월 % s일 % s시" %
                              (now.year, now.month, now.day, now.hour) + a)
                selFrames.mainAdd()

        else:
            global displayPage2
            displayPage2 = VerticalScrolledFrame(root)
            global addheight
            addheight = 0
            global phqchk2
            phqchk2 = False
            global cdichk2
            cdichk2 = False
            global bdichk2
            bdichk2 = False
            global snapchk2
            snapchk2 = False
            global st_2chk2
            st_2chk2 = False
            global st_1chk2
            st_1chk2 = False
            global hamachk2
            hamachk2 = False
            global hamdchk2
            hamdchk2 = False
            cellCriteria = 0
            # phqchk3 변수 불린 값으로 함수에 재진입 하여 프로퍼티에 등록하는 것을 차단
            if (selFrames.srtchkType1.get()) == 1 and phqchk3 == True:
                cellCriteria += 33
                mainContent.phqchkFunc2()
                phqchk2 = True
                if phqchk2 == True:
                    addheight = 1009
                selFrames.phqchk3 = False
            if (selFrames.srtchkType2.get()) == 1 and cdichk3 == True:
                cellCriteria += 140
                mainContent.cdichkFunc2()
                cdichk2 = True
                if cdichk2 == True:
                    addheight += 3558
                selFrames.cdichk3 = False
            if (selFrames.srtchkType3.get()) == 1 and bdichk3 == True:
                cellCriteria += 133
                mainContent.bdichkFunc2()
                bdichk2 = True
                if bdichk2 == True:
                    addheight += 3510
                selFrames.bdichk3 = False
            if (selFrames.srtchkType4.get()) == 1 and snapchk3 == True:
                cellCriteria += 60
                mainContent.snapchkFunc2()
                snapchk2 = True
                if snapchk2 == True:
                    addheight += 1141
                selFrames.snapchk3 = False
            if (selFrames.srtchkType5.get()) == 1 and st_2chk3 == True:
                cellCriteria += 66
                mainContent.st_2chkFunc2()
                st_2chk2 = True
                if st_2chk2 == True:
                    addheight += 1342
                selFrames.st_2chk3 = False

            if (selFrames.srtchkType6.get()) == 1 and st_1chk3 == True:
                cellCriteria += 66
                mainContent.st_1chkFunc2()
                st_1chk2 = True
                if st_1chk2 == True:
                    addheight += 1345
                selFrames.st_1chk3 = False
                
            if (selFrames.srtchkType7.get()) == 1 and hamachk3 == True:
                cellCriteria += 50
                mainContent.hamachkFunc2()
                hamachk2 = True
                if hamachk2 == True:
                    addheight += 1248
                selFrames.hamachk3 = False

            if (selFrames.srtchkType8.get()) == 1 and hamdchk3 == True:
                cellCriteria += 128
                mainContent.hamdchkFunc2()
                hamdchk2 = True
                if hamdchk2 == True:
                    addheight += 2451
                selFrames.hamdchk3 = False
                
        submit = Button(displayPage2.inner, image=addsheetImg,
                        activebackground="#f0f0f0", command=mainContent.allchkFunc, borderwidth=0)
        submit.place(relx=0.5, anchor="center",
                     width=366, height=125)
        submit.pack(ipady=30, side=BOTTOM)
        submit.focus_set()

        selFrames.mainHide()
        now = datetime.now()
        notebook.add2(displayPage2, text="% s-% s-%s" %
                      (now.year, now.month, now.day) + a)
        selFrames.mainAdd()
            
    def mainAdd():
        notebook.add3(END, startframe, text="평가")
        # notebook.add3(END, frame0Alteration, text="변화")

    def mainHide():

        # notebook.forget(frame0Alteration)
        notebook.forget(startframe)




# tab_names = []
# for i in notebook.tabs():
#         tab_names.append(notebook.tab(i, "text"))
# abc = (tab_names[1]['text'])
# abc.startswith(1, 2)
enterpng = tk.PhotoImage(
    file='images/enter.png')
addSheet = Button(startframe, image=enterpng, text="확인", activebackground="#FFFFFF",
                  command=selFrames.checkNameAndNum, borderwidth=0)
addSheet.place(x=935, y=370, width=162, height=125)
addSheet.focus_set()
addSheet.bind("<Return>", selFrames.checkNameAndNum)

startchk1 = Checkbutton(
    startframe, compound=tk.LEFT, background="#FFFFFF", activebackground="#FFFFFF", image=selFrames.image_up, selectimage=selFrames.image_down, text=" PHQ", indicatoron=False, cursor="circle", variable=selFrames.srtchkType1, font=malgungothic13, bd=0, command=selFrames.sumTab)
startchk2 = Checkbutton(startframe, compound=tk.LEFT, background="#FFFFFF", activebackground="#FFFFFF", image=selFrames.image_up, selectimage=selFrames.image_down, text=" CDI", indicatoron=False, cursor="circle", variable=selFrames.srtchkType2,
                        font=malgungothic13, bd=0, command=selFrames.sumTab)
startchk3 = Checkbutton(
    startframe, compound=tk.LEFT, background="#FFFFFF", activebackground="#FFFFFF", image=selFrames.image_up, selectimage=selFrames.image_down, text=" BDI", indicatoron=False, cursor="circle", variable=selFrames.srtchkType3, font=malgungothic13, bd=0, command=selFrames.sumTab)
startchk4 = Checkbutton(
    startframe, compound=tk.LEFT, background="#FFFFFF", activebackground="#FFFFFF", image=selFrames.image_up, selectimage=selFrames.image_down, text=" SNAP", indicatoron=False, cursor="circle", variable=selFrames.srtchkType4, font=malgungothic13, bd=0, command=selFrames.sumTab)
startchk5 = Checkbutton(
    startframe, compound=tk.LEFT, background="#FFFFFF", activebackground="#FFFFFF", image=selFrames.image_up, selectimage=selFrames.image_down, text=" ST_2", indicatoron=False, cursor="circle", variable=selFrames.srtchkType5, font=malgungothic13, bd=0, command=selFrames.sumTab)
startchk6 = Checkbutton(
    startframe, compound=tk.LEFT, background="#FFFFFF", activebackground="#FFFFFF", image=selFrames.image_up, selectimage=selFrames.image_down, text=" ST_1", indicatoron=False, cursor="circle", variable=selFrames.srtchkType6, font=malgungothic13, bd=0, command=selFrames.sumTab)
startchk7 = Checkbutton(
    startframe, compound=tk.LEFT, background="#FFFFFF", activebackground="#FFFFFF", image=selFrames.image_up, selectimage=selFrames.image_down, text=" HAMA", indicatoron=False, cursor="circle", variable=selFrames.srtchkType7, font=malgungothic13, bd=0, command=selFrames.sumTab)
startchk8 = Checkbutton(startframe, background="#FFFFFF", activebackground="#FFFFFF", compound=tk.LEFT, image=selFrames.image_up, selectimage=selFrames.image_down, text=" HAMD", selectcolor="#FFFFFF", indicatoron=False, cursor="circle", variable=selFrames.srtchkType8,
                        font=malgungothic13, bd=0, command=selFrames.sumTab)

startchk1.place(width=130, x=360, y=369)
startchk2.place(width=130, x=494, y=369)
startchk3.place(width=130, x=628, y=369)
startchk4.place(width=130, x=784, y=369)
startchk5.place(width=130, x=359, y=436)
startchk6.place(width=130, x=494, y=436)
startchk7.place(width=130, x=633, y=436)
startchk8.place(width=130, x=784, y=436)
startchk1.deselect()
startchk2.deselect()
startchk3.deselect()
startchk4.deselect()
startchk5.deselect()
startchk6.deselect()
startchk7.deselect()
startchk8.deselect()


def nextTab():
    x = 0
    x += 1
    for i in range(1, 10000):
        if notebook.index(CURRENT) == i:
            i += 1
            notebook.select(i)
            return

    notebook.select(x)


def prevTab():
    for i in range(1, 10000):
        if notebook.index(CURRENT) == i:
            i -= 1
            notebook.select(i)
            return
    notebook.select(x)


prevButton = Button(root, text=" \u276E", font=malgungothic9,
                    command=prevTab, borderwidth=0.3)
prevButton.place(x=28, y=864, width=31, height=31)
nextButton = Button(root, text=" \u276F", font=malgungothic9,
                    command=nextTab, borderwidth=0.3)
nextButton.place(x=55, y=864, width=31, height=31)

root.mainloop()
